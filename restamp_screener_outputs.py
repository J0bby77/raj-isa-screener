#!/usr/bin/env python3
"""
restamp_screener_outputs.py — WP-M2 (29-Jul-2026).

THE GAP THIS CLOSES
-------------------
The monthly pre-run consumes the SUMMARY tab of each weekly screener workbook. Those workbooks are
generated at screen time under the calibration config LIVE AT THAT MOMENT. When a calibration
changes between the last screen and the next pre-run, the pre-run silently consumes candidates
ranked under the OLD config. On 29-Jul-2026 FORWARD_AXIS_BUCKET_WEIGHTS changed from
{margin .30, price .70} to thirds; the next natural screen is 07-Aug but the pre-run is 01-Aug.

This module restamps existing workbooks under the CURRENT config without re-running the screens.
Every input the forward axis needs is already in the SCORES tab for EVERY row (not just SUMMARY),
so the axis is fully recomputable offline. Only the source-score FV/analyst inputs need a live
fetch, and only for rows that could plausibly clear SUMMARY_SOURCE_FLOOR.

PHASES
  1  offline : read SCORES -> canonical rows -> 1m momentum from price cache -> cross-sectional
               percentile -> new forward_axis. Reports the fetch shortlist.
  2  fetch   : batched yfinance for shortlist rows lacking price/target/analyst inputs.
  3  restamp : source score + doors -> re-select SUMMARY -> write workbook + diff report.
"""
from __future__ import annotations
import argparse, json, os, sys, math

SCORES_MAP = {
    "Ticker": "ticker", "Company": "company", "Final Status": "final_status",
    "Part A Total": "part_a_score", "Part B Total": "part_b_score", "Grand Total": "total_score",
    "Fwd Axis /100": "forward_axis_score_old",
    "EPS Trend": "score_f_eps_trend", "EPS Mom %": "eps_trend_mom_pct",
    "Mgn Traj": "score_f_margin_traj", "Mgn d-pp": "margin_traj_delta_pp",
    "Rev Est": "score_f_rev_est", "Rev Est %": "rev_est_fwd_pct",
    "Price Mom": "score_f_price_mom_old", "Price 12-1m %": "price_mom_12_1m_pct",
    "Est Rev (B)": "score_b_est_rev", "Rev Runway": "revision_runway", "Stage": "revision_stage",
    "ND/EBITDA": "score_nd_ebitda_raw", "FCF Pos (HG)": "score_fcf_pos_raw",
}
SUMMARY_MAP = {
    "Ticker": "ticker", "Current Price": "current_price", "Target Price": "target_price_mean",
    "Analyst Rating": "analyst_rating", "# Analysts": "num_analysts",
    "Est Rev Direction": "est_rev_direction", "Sector": "sector", "Industry": "industry",
    "Sector Bucket": "sector_bucket", "Net Debt/EBITDA": "net_debt_ebitda",
    "FCF Positive Yrs": "fcf_positive_years", "Op Margin Trend": "op_margin_trend",
    "Div Payout/FCF": "div_payout_fcf", "Trailing P/E": "trailing_pe",
    "P/E vs 3yr Avg": "val_hist_pe_premium_disc",
}


def _f(v):
    try:
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return None
        s = str(v).replace("$", "").replace("£", "").replace("%", "").replace(",", "").replace("x", "").strip()
        if s in ("", "N/A", "None", "-", "n/a"):
            return None
        return float(s)
    except (TypeError, ValueError):
        return None


def _header_row(ws, limit=8):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "Ticker" in cells:
            return i, cells
        if i > limit:
            break
    return None, None


def read_workbook(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for sheet, cmap in (("SCORES", SCORES_MAP), ("SUMMARY", SUMMARY_MAP)):
        if sheet not in wb.sheetnames:
            out[sheet] = {}
            continue
        ws = wb[sheet]
        hi, hdr = _header_row(ws)
        if hi is None:
            out[sheet] = {}
            continue
        idx = {h: i for i, h in enumerate(hdr)}
        recs = {}
        for j, row in enumerate(ws.iter_rows(values_only=True)):
            if j <= hi or not row or not row[0]:
                continue
            t = str(row[0]).strip()
            if not t or t.lower().startswith(("note", "total", "—")):
                continue
            rec = {}
            for h, canon in cmap.items():
                if h in idx and idx[h] < len(row):
                    rec[canon] = row[idx[h]]
            recs[t] = rec
        out[sheet] = recs
    wb.close()
    return out


NUMERIC = {"part_a_score", "part_b_score", "total_score", "score_f_eps_trend", "score_f_margin_traj",
           "score_f_rev_est", "score_b_est_rev", "revision_runway", "price_mom_12_1m_pct",
           "current_price", "target_price_mean", "num_analysts", "net_debt_ebitda",
           "fcf_positive_years", "op_margin_trend", "div_payout_fcf", "trailing_pe",
           "forward_axis_score_old", "eps_trend_mom_pct", "margin_traj_delta_pp", "rev_est_fwd_pct"}


def build_rows(wbdata):
    rows = []
    summ = wbdata.get("SUMMARY", {})
    for t, rec in wbdata.get("SCORES", {}).items():
        r = dict(rec)
        r["ticker"] = t
        r.update({k: v for k, v in summ.get(t, {}).items() if v not in (None, "")})
        for k in list(r):
            if k in NUMERIC:
                r[k] = _f(r[k])
        rows.append(r)
    return rows


def attach_short_momentum(rows, price_csv, asof):
    """price_mom_1m_pct from a cached close matrix (exactly what the screen would have computed)."""
    import pandas as pd
    px = pd.read_csv(price_csv, index_col=0, parse_dates=True).sort_index()
    px = px.loc[px.index <= pd.Timestamp(asof)]
    n_ok = 0
    for r in rows:
        t = r.get("ticker")
        if t in px.columns:
            ser = px[t].dropna()
            if len(ser) >= 25:
                r["price_mom_1m_pct"] = round((float(ser.iloc[-1]) / float(ser.iloc[-22]) - 1) * 100, 1)
                n_ok += 1
                continue
        r["price_mom_1m_pct"] = None
    return n_ok


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--group", required=True)
    ap.add_argument("--asof", required=True, help="screen as-of date (YYYY-MM-DD)")
    ap.add_argument("--price-cache", default="/tmp/all_px.csv")
    ap.add_argument("--out-json", default=None)
    ap.add_argument("--shm", default="/tmp/pylibs")
    ap.add_argument("--fetch-floor", type=float, default=55.0,
                    help="new forward-axis floor above which a live FV/analyst fetch is worthwhile")
    a = ap.parse_args()
    if a.shm and os.path.isdir(a.shm):
        sys.path.insert(0, a.shm)
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    import screener_core as sc

    wbd = read_workbook(a.workbook)
    rows = build_rows(wbd)
    rankable = [r for r in rows if str(r.get("final_status", "")).strip() == "CANDIDATE_RANKABLE"]
    n_mom = attach_short_momentum(rankable, a.price_cache, a.asof)
    for r in rankable:
        r["momentum_state"] = sc._momentum_state(r.get("price_mom_12_1m_pct"), r.get("price_mom_1m_pct"))
    sc.apply_cross_sectional_momentum(rankable)

    have_fv = [r for r in rankable if r.get("current_price") and r.get("target_price_mean")]
    shortlist = [r["ticker"] for r in rankable
                 if (r.get("forward_axis_score") or 0) >= a.fetch_floor
                 and not (r.get("current_price") and r.get("target_price_mean"))]
    res = {"group": a.group, "asof": a.asof, "workbook": os.path.basename(a.workbook),
           "rankable": len(rankable), "short_mom_resolved": n_mom,
           "fv_inputs_present": len(have_fv), "fetch_shortlist": sorted(shortlist),
           "rows": rankable}
    if a.out_json:
        with open(a.out_json, "w") as f:
            json.dump(res, f, default=str)
    print(f"{a.group:10s} rankable={len(rankable):4d} 1m_mom={n_mom:4d} "
          f"fv_present={len(have_fv):3d} need_fetch={len(shortlist):4d}")


if __name__ == "__main__":
    main()
