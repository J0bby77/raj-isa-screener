#!/usr/bin/env python3
"""
energy_screener.py  --  ISA Energy Growth Stock Screener
Version: 1.0  |  2026-05-31

Screens a curated watchlist of energy/clean-tech companies against an
empirically calibrated energy-specific scorecard. Designed to run as a
scheduled task (4th Sunday monthly, 09:00).

Thresholds calibrated May 2026 from 25-company financial data pull.
Investment universe: ~28 companies across US, EU, UK, Canada.

Usage:
    python3 energy_screener.py \\
        --date 2026-06-22 \\
        --outputs /path/to/outputs/ \\
        --inv-dir "/path/to/Investment Analysis/"

Outputs (to inv-dir — permanent):
    {YYYYMMDD}_ENERGY_full_data.csv
    {YYYYMMDD}_ENERGY_gate_results.csv
    {YYYYMMDD}_ENERGY_retrospective.md
    {YYYYMMDD}_ENERGY_run_qa.csv
    Growth Stock Analysis ENERGY W-e DD-Mon-YY.xlsx

Imports shared utilities from screener_core.py (must be in same directory).
If screener_core import fails, falls back to inline implementations.
"""

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1: IMPORTS & CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
import os
try:
    import isa_env_guard  # noqa  (disk guardrail: forces temp + yfinance cache onto tmpfs /dev/shm)
except Exception:
    pass
import sys
import json
import csv
import math
import time
import logging
import argparse
import traceback
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path

import requests
import pandas as pd
import numpy as np
import os as _os, sys as _sys
_sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
import scoring_config as _cfg  # SINGLE SOURCE OF TRUTH

try:
    import yfinance as yf
except ImportError:
    print("ERROR: yfinance not installed. Run: pip install yfinance openpyxl requests -q")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

# ── Try to import shared utilities from screener_core ────────────────────────
_SCREENER_CORE_AVAILABLE = False
try:
    import importlib.util
    _sc_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "screener_core.py")
    if os.path.exists(_sc_path):
        spec = importlib.util.spec_from_file_location("screener_core", _sc_path)
        _sc = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(_sc)
        get_stmt_value   = _sc.get_stmt_value
        get_stmt_series  = _sc.get_stmt_series
        safe_float       = _sc.safe_float
        compute_cagr     = _sc.compute_cagr
        apply_pence_correction = _sc.apply_pence_correction
        pence_divisor    = _sc.pence_divisor
        compute_ebitda   = _sc.compute_ebitda
        compute_roic     = _sc.compute_roic
        compute_forward_axis = _sc.compute_forward_axis   # shared forward axis (Part 3 §13) — energy parity
        _SCREENER_CORE_AVAILABLE = True
except Exception as _e:
    pass  # fallback inline below

if not _SCREENER_CORE_AVAILABLE:
    # ── Inline fallback implementations ──────────────────────────────────────
    def safe_float(v, default=None):
        try:
            f = float(v)
            return default if (f != f) else f  # NaN check
        except (TypeError, ValueError):
            return default

    def get_stmt_value(stmt_df, row_keys, col_idx=0):
        if stmt_df is None or stmt_df.empty:
            return None
        cols = stmt_df.columns
        if col_idx >= len(cols):
            return None
        col = cols[col_idx]
        for k in row_keys:
            if k in stmt_df.index:
                v = stmt_df.at[k, col]
                f = safe_float(v)
                if f is not None:
                    return f
        return None

    def get_stmt_series(stmt_df, row_keys, max_periods=5):
        if stmt_df is None or stmt_df.empty:
            return []
        results = []
        for col_idx, col in enumerate(stmt_df.columns[:max_periods]):
            year = col.year if hasattr(col, "year") else int(str(col)[:4])
            for k in row_keys:
                if k in stmt_df.index:
                    v = stmt_df.at[k, col]
                    f = safe_float(v)
                    if f is not None:
                        results.append((year, f))
                        break
        return results

    def compute_cagr(start_val, end_val, years):
        if None in (start_val, end_val, years) or years <= 0:
            return None
        if start_val <= 0 or end_val <= 0:
            return None
        try:
            return (end_val / start_val) ** (1.0 / years) - 1
        except Exception:
            return None

    def pence_divisor(info):
        currency = (info.get("currency") or "").strip()
        price = safe_float(info.get("currentPrice") or info.get("regularMarketPrice"))
        if currency == "GBp":
            return 100
        if currency == "GBP" and price is not None and price > 500:
            return 100
        return 1

    def apply_pence_correction(info):
        div = pence_divisor(info)
        return {
            "current_price": safe_float(info.get("currentPrice") or info.get("regularMarketPrice"), 0) / div,
            "target_mean":   (safe_float(info.get("targetMeanPrice")) or 0) / div,
            "target_high":   (safe_float(info.get("targetHighPrice")) or 0) / div,
            "target_low":    (safe_float(info.get("targetLowPrice")) or 0) / div,
            "low_52wk":      (safe_float(info.get("fiftyTwoWeekLow")) or 0) / div,
            "high_52wk":     (safe_float(info.get("fiftyTwoWeekHigh")) or 0) / div,
            "pence_div":     div,
        }

    def compute_ebitda(income_stmt, info, col_idx=0):
        ebitda = get_stmt_value(income_stmt, ["EBITDA", "Normalized EBITDA"], col_idx)
        if ebitda is not None:
            return ebitda, "EBITDA_DIRECT"
        ebitda = safe_float(info.get("ebitda"))
        if ebitda is not None:
            return ebitda, "EBITDA_INFO"
        ebit = get_stmt_value(income_stmt, ["EBIT", "Ebit", "Operating Income"], col_idx)
        da = get_stmt_value(income_stmt, ["Depreciation Amortization Depletion",
                                          "Reconciled Depreciation", "Depreciation"], col_idx)
        if ebit is not None and da is not None:
            return ebit + abs(da), "EBITDA_DERIVED"
        return None, "EBITDA_UNRESOLVED"

    def compute_roic(income_stmt, balance_sheet, col_idx=0):
        ebit = get_stmt_value(income_stmt, ["EBIT", "Ebit"], col_idx)
        tax_rate_raw = get_stmt_value(income_stmt, ["Tax Rate For Calcs"], col_idx)
        tax_rate = min(max(safe_float(tax_rate_raw) or 0.21, 0.0), 0.5)
        invested_cap = get_stmt_value(balance_sheet, ["Invested Capital"], col_idx)
        if ebit is not None and invested_cap and invested_cap > 0:
            nopat = ebit * (1 - tax_rate)
            return nopat / invested_cap, "ROIC_DIRECT"
        equity = get_stmt_value(balance_sheet, ["Common Stock Equity", "Stockholders Equity"], col_idx)
        debt   = get_stmt_value(balance_sheet, ["Total Debt"], col_idx)
        if ebit is not None and equity and debt is not None:
            ic = (equity or 0) + (debt or 0)
            if ic > 0:
                return ebit * (1 - tax_rate) / ic, "ROIC_DERIVED"
        return None, "ROIC_UNRESOLVED"


# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("energy_screener")

# ── Scoring constants (empirically calibrated May 2026) ──────────────────────

# Hard gates
GATE1_MIN_REVENUE        = 50_000_000    # $50M TTM revenue
GATE2_MIN_REV_GROWTH_TTM = 0.00          # TTM revenue growth > 0%
GATE2_FWD_GROWTH_ALT     = 0.08          # OR forward revenue growth > 8%
# Gate 3: EBITDA must be > 0 (no gross margin gate for energy)

# Part A thresholds (max 20 pts — 10 metrics × 2pts each)
A_REV_GROWTH_TTM_STRONG  = 0.20   # ≥20% = 2pts
A_REV_GROWTH_TTM_ACCEPT  = 0.08   # 8-19% = 1pt; <8% = 0pts
A_REV_CAGR_STRONG        = 0.12   # ≥12% 3yr CAGR = 2pts
A_REV_CAGR_ACCEPT        = 0.05   # 5-11% = 1pt
A_EBITDA_MARGIN_STRONG   = 0.20   # ≥20% = 2pts
A_EBITDA_MARGIN_ACCEPT   = 0.05   # 5-19% = 1pt
A_GROSS_MARGIN_STRONG    = 0.25   # ≥25% = 2pts (scoring only — NOT a gate)
A_GROSS_MARGIN_ACCEPT    = 0.12   # 12-24% = 1pt
A_EBITDA_GROWTH_STRONG   = 0.20   # ≥20% YoY = 2pts
A_EBITDA_GROWTH_ACCEPT   = 0.05   # 5-19% = 1pt
A_ROE_STRONG             = 0.15   # ≥15% = 2pts
A_ROE_ACCEPT             = 0.05   # 5-14% = 1pt
A_REV_SCALE_LARGE        = 1_000_000_000   # >$1B = 2pts
A_REV_SCALE_MID          = 200_000_000     # $200M-$1B = 1pt
A_FWD_GROWTH_STRONG      = 0.25   # ≥25% fwd = 2pts
A_FWD_GROWTH_ACCEPT      = 0.10   # 10-24% = 1pt
A_CAPEX_INTENSITY_STRONG = 0.08   # ≥8% capex/revenue = 2pts (investing for growth)
A_CAPEX_INTENSITY_ACCEPT = 0.03   # 3-7% = 1pt
# FCF: positive TTM = 2pts; near breakeven (neg but >-10% margin) = 1pt; deeply negative = 0pts
A_FCF_NEAR_BREAKEVEN     = -0.10  # FCF margin > -10% = "near breakeven"

# Part A classification
PART_A_STRONG_GROWTH     = _cfg.ENERGY_PART_A_STRONG       # 14 (scoring_config)
PART_A_ACCEPTABLE        = _cfg.ENERGY_PART_A_ACCEPTABLE   # 8
# <8 = Not Growth

# Part B thresholds (max 16 pts — 8 metrics × 2pts each)
B_EV_EBITDA_STRONG       = 15.0   # ≤15x = 2pts (value)
B_EV_EBITDA_ACCEPT       = 35.0   # 15-35x = 1pt
B_ND_EBITDA_STRONG       = 2.0    # ≤2x net debt/EBITDA = 2pts
B_ND_EBITDA_ACCEPT       = 5.0    # 2-5x = 1pt; >5x = 0pts
B_ANALYST_BUY_STRONG     = 2.5    # mean recommendation ≤2.5 = Strong Buy signal
B_ANALYST_BUY_ACCEPT     = 3.2    # 2.5-3.2 = Hold/Mild Buy
B_UPSIDE_STRONG          = 0.20   # ≥20% upside vs consensus target = 2pts
B_UPSIDE_ACCEPT          = 0.05   # 5-19% = 1pt
B_52WK_POS_STRONG        = 0.80   # ≥80% of 52w range = 2pts (momentum)
B_52WK_POS_ACCEPT        = 0.50   # 50-79% = 1pt
B_FWD_PE_STRONG          = 20.0   # ≤20x forward PE = 2pts
B_FWD_PE_ACCEPT          = 40.0   # 20-40x = 1pt
B_EPS_GROWTH_STRONG      = 0.20   # ≥20% fwd EPS growth = 2pts
B_EPS_GROWTH_ACCEPT      = 0.08   # 8-19% = 1pt
B_ANALYST_COUNT_STRONG   = 10     # ≥10 analysts = 2pts (well-covered)
B_ANALYST_COUNT_ACCEPT   = 5      # 5-9 = 1pt

# Part B classification
PART_B_STRONG_BUY        = _cfg.ENERGY_PART_B_STRONG       # 11 (scoring_config)
PART_B_WATCH             = _cfg.ENERGY_PART_B_WATCH        # 6
# <6 = Avoid

# Status codes
STATUS_STRONG_BUY    = "ENERGY_STRONG_BUY"
STATUS_WATCH         = "ENERGY_WATCH"
STATUS_ACCEPTABLE    = "ENERGY_ACCEPTABLE"
STATUS_NOT_GROWTH    = "ENERGY_NOT_GROWTH"
STATUS_GATE_FAIL     = "ENERGY_GATE_FAIL"
STATUS_DATA_ISSUE    = "ENERGY_DATA_ISSUE"

FETCH_WORKERS        = 4
FETCH_SLEEP          = 0.3   # seconds between individual fetches


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2: WATCHLIST LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_watchlist(inv_dir):
    """Load energy_watchlist.json. Returns list of dicts."""
    path = os.path.join(inv_dir, "energy_watchlist.json")
    if not os.path.exists(path):
        raise FileNotFoundError(f"energy_watchlist.json not found at: {path}")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    log.info(f"Watchlist loaded: {len(data)} companies")
    return data


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3: DATA FETCHING
# ─────────────────────────────────────────────────────────────────────────────

def _fetch_info(ticker_sym):
    """Fetch ticker.info for one ticker. Returns (sym, info|None, err|None)."""
    try:
        tk = yf.Ticker(ticker_sym)
        info = tk.info or {}
        if len(info) < 5:
            return ticker_sym, None, "empty_info"
        return ticker_sym, info, None
    except Exception as e:
        return ticker_sym, None, str(e)


def _fetch_statements(ticker_sym):
    """Fetch annual statements for one ticker."""
    try:
        tk = yf.Ticker(ticker_sym)
        return ticker_sym, {
            "income_stmt":   tk.income_stmt,
            "cashflow":      tk.cashflow,
            "balance_sheet": tk.balance_sheet,
            "quarterly_income_stmt": tk.quarterly_income_stmt,   # forward axis: margin trajectory
            "eps_trend":         tk.eps_trend,                   # forward axis: eps_trend momentum + revision stage
            "growth_estimates":  tk.growth_estimates,            # forward axis: forward revenue growth
            "history":           tk.history(period="1y"),        # forward axis: price momentum
        }, None
    except Exception as e:
        return ticker_sym, None, str(e)


def fetch_all_data(tickers, workers=FETCH_WORKERS, sleep=FETCH_SLEEP):
    """
    Fetch info + statements for all tickers in parallel (small batches).
    Returns (info_map, stmt_map, fetch_errors).
    """
    info_map   = {}
    stmt_map   = {}
    fetch_errs = []

    log.info(f"Fetching yfinance info for {len(tickers)} tickers...")
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_fetch_info, t): t for t in tickers}
        for fut in as_completed(futures):
            sym, info, err = fut.result()
            if info:
                info_map[sym] = info
            else:
                fetch_errs.append({"ticker": sym, "stage": "info", "error": err or "no_data"})
                log.warning(f"  Info fetch failed: {sym} — {err}")
            time.sleep(sleep)

    log.info(f"Fetching statements for {len(tickers)} tickers...")
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_fetch_statements, t): t for t in tickers}
        for fut in as_completed(futures):
            sym, stmts, err = fut.result()
            if stmts:
                stmt_map[sym] = stmts
            else:
                fetch_errs.append({"ticker": sym, "stage": "statements", "error": err or "no_data"})
                log.warning(f"  Statements fetch failed: {sym} — {err}")
            time.sleep(sleep)

    log.info(f"Data fetch complete: {len(info_map)} info | {len(stmt_map)} statements | {len(fetch_errs)} errors")
    return info_map, stmt_map, fetch_errs


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4: ENERGY GATE FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────
# Gate 1: Revenue > $50M TTM (excludes pre-revenue speculative names)
# Gate 2: Revenue growth > 0% TTM OR forward revenue growth > 8%
#         — separates growing from declining/flat energy co's
# Gate 3: EBITDA > 0 (excludes cash-burning pre-revenue plays)
# NOTE: NO gross margin gate. Gross margin is a scoring metric only.
# ─────────────────────────────────────────────────────────────────────────────

def _get_ttm_revenue(income_stmt, info):
    """Return TTM revenue from income_stmt (preferred) or info fallback."""
    rev_series = get_stmt_series(income_stmt, [
        "Total Revenue", "Revenue",
        "Net Revenue", "Sales"
    ], max_periods=1)
    if rev_series:
        return rev_series[0][1]
    return safe_float(info.get("totalRevenue") or info.get("revenue"))


def energy_gate1(info, income_stmt):
    """Revenue > $50M TTM."""
    rev = _get_ttm_revenue(income_stmt, info)
    if rev is None:
        return None, "GATE_DATA_UNRESOLVED: revenue missing", "GATE_DATA_UNRESOLVED"
    if rev >= GATE1_MIN_REVENUE:
        return True, "", ""
    return False, f"Revenue ${rev/1e6:.1f}M below $50M threshold", "Gate 1"


def energy_gate2(info, income_stmt):
    """
    Revenue growth > 0% TTM OR forward revenue growth > 8%.
    Replaces the gross margin gate — revenue momentum is the key energy growth signal.
    """
    # TTM growth from income_stmt (most recent vs prior year)
    rev_series = get_stmt_series(income_stmt, [
        "Total Revenue", "Revenue", "Net Revenue", "Sales"
    ], max_periods=2)
    ttm_growth = None
    if len(rev_series) >= 2:
        rev_new, rev_old = rev_series[0][1], rev_series[1][1]
        if rev_old and rev_old > 0:
            ttm_growth = (rev_new - rev_old) / rev_old

    # Forward growth from info (revenueGrowth = TTM per yfinance, earningsGrowth = forward proxy)
    fwd_growth = safe_float(info.get("revenueGrowth"))  # yfinance TTM; use as leading indicator

    # Pass if either criterion met
    if ttm_growth is not None and ttm_growth > GATE2_MIN_REV_GROWTH_TTM:
        return True, "", "", ttm_growth, fwd_growth
    if fwd_growth is not None and fwd_growth > GATE2_FWD_GROWTH_ALT:
        return True, "", "", ttm_growth, fwd_growth

    # Both missing = unresolved
    if ttm_growth is None and fwd_growth is None:
        return None, "GATE_DATA_UNRESOLVED: no revenue growth data", "GATE_DATA_UNRESOLVED", None, None

    ttm_str = f"{ttm_growth*100:.1f}%" if ttm_growth is not None else "N/A"
    fwd_str = f"{fwd_growth*100:.1f}%" if fwd_growth is not None else "N/A"
    return False, f"Revenue declining: TTM {ttm_str}, Fwd {fwd_str}", "Gate 2", ttm_growth, fwd_growth


def energy_gate3(income_stmt, info):
    """EBITDA must be positive. No gross margin gate."""
    ebitda, ebitda_label = compute_ebitda(income_stmt, info)
    if ebitda is None:
        return None, "GATE_DATA_UNRESOLVED: EBITDA unresolved", "GATE_DATA_UNRESOLVED", None, None
    if ebitda > 0:
        return True, "", "", ebitda, ebitda_label
    return False, f"EBITDA negative: {ebitda/1e6:.1f}M", "Gate 3", ebitda, ebitda_label


def apply_energy_gates(ticker_sym, info, income_stmt, watchlist_entry):
    """
    Apply all 3 energy gates. Returns gate result dict.
    Keys: gate_pass, gate_code, gate_reason, ttm_growth, fwd_growth, ebitda, ebitda_label
    """
    # Gate 1 — Revenue scale
    g1, reason1, code1 = energy_gate1(info, income_stmt)
    if g1 is None:
        return {"gate_pass": None, "gate_code": code1, "gate_reason": reason1}
    if not g1:
        return {"gate_pass": False, "gate_code": code1, "gate_reason": reason1}

    # Gate 2 — Revenue momentum
    g2, reason2, code2, ttm_growth, fwd_growth = energy_gate2(info, income_stmt)
    if g2 is None:
        return {"gate_pass": None, "gate_code": code2, "gate_reason": reason2,
                "ttm_growth": None, "fwd_growth": None}
    if not g2:
        return {"gate_pass": False, "gate_code": code2, "gate_reason": reason2,
                "ttm_growth": ttm_growth, "fwd_growth": fwd_growth}

    # Gate 3 — EBITDA positive
    g3, reason3, code3, ebitda, ebitda_label = energy_gate3(income_stmt, info)
    if g3 is None:
        return {"gate_pass": None, "gate_code": code3, "gate_reason": reason3,
                "ttm_growth": ttm_growth, "fwd_growth": fwd_growth}
    if not g3:
        return {"gate_pass": False, "gate_code": code3, "gate_reason": reason3,
                "ttm_growth": ttm_growth, "fwd_growth": fwd_growth,
                "ebitda": ebitda, "ebitda_label": ebitda_label}

    return {
        "gate_pass": True, "gate_code": "PASS", "gate_reason": "",
        "ttm_growth": ttm_growth, "fwd_growth": fwd_growth,
        "ebitda": ebitda, "ebitda_label": ebitda_label,
    }


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 5: PART A SCORING — GROWTH QUALITY (10 metrics, max 20 pts)
# ─────────────────────────────────────────────────────────────────────────────

def score_part_a(ticker_sym, info, income_stmt, cashflow, balance_sheet, gate_data):
    """
    Energy Part A: 10 growth quality metrics, each scored 0/1/2.
    Max = 20 pts. Strong Growth ≥14, Acceptable 8-13, Not Growth <8.
    Returns dict with per-metric scores and total.
    """
    out = {"ticker": ticker_sym}
    unresolved = []

    # ── Metric 1: Revenue Growth TTM (from gate data, already computed) ──────
    ttm_growth = gate_data.get("ttm_growth")
    if ttm_growth is None:
        # Recompute from income_stmt
        rev_series = get_stmt_series(income_stmt, ["Total Revenue", "Revenue", "Net Revenue"], max_periods=2)
        if len(rev_series) >= 2:
            r0, r1 = rev_series[0][1], rev_series[1][1]
            ttm_growth = (r0 - r1) / r1 if r1 and r1 > 0 else None
    if ttm_growth is None:
        out["score_rev_growth_ttm"] = 0
        out["rev_growth_ttm"] = None
        unresolved.append("rev_growth_ttm")
    elif ttm_growth >= A_REV_GROWTH_TTM_STRONG:
        out["score_rev_growth_ttm"] = 2
        out["rev_growth_ttm"] = ttm_growth
    elif ttm_growth >= A_REV_GROWTH_TTM_ACCEPT:
        out["score_rev_growth_ttm"] = 1
        out["rev_growth_ttm"] = ttm_growth
    else:
        out["score_rev_growth_ttm"] = 0
        out["rev_growth_ttm"] = ttm_growth

    # ── Metric 2: Revenue CAGR 3yr ───────────────────────────────────────────
    rev_series_full = get_stmt_series(income_stmt, ["Total Revenue", "Revenue", "Net Revenue"], max_periods=4)
    rev_cagr = None
    if len(rev_series_full) >= 3:
        rev_cagr = compute_cagr(rev_series_full[2][1], rev_series_full[0][1], 3)
    elif len(rev_series_full) == 2:
        rev_cagr = compute_cagr(rev_series_full[1][1], rev_series_full[0][1], 2)
    out["rev_cagr"] = rev_cagr
    if rev_cagr is None:
        out["score_rev_cagr"] = 0
        unresolved.append("rev_cagr")
    elif rev_cagr >= A_REV_CAGR_STRONG:
        out["score_rev_cagr"] = 2
    elif rev_cagr >= A_REV_CAGR_ACCEPT:
        out["score_rev_cagr"] = 1
    else:
        out["score_rev_cagr"] = 0

    # ── Metric 3: EBITDA Margin ───────────────────────────────────────────────
    ebitda = gate_data.get("ebitda")
    ebitda_label = gate_data.get("ebitda_label", "")
    if ebitda is None:
        ebitda, ebitda_label = compute_ebitda(income_stmt, info)

    rev_ttm = _get_ttm_revenue(income_stmt, info)
    ebitda_margin = None
    if ebitda is not None and rev_ttm and rev_ttm > 0:
        ebitda_margin = ebitda / rev_ttm
    out["ebitda"] = ebitda
    out["ebitda_label"] = ebitda_label
    out["ebitda_margin"] = ebitda_margin
    out["rev_ttm"] = rev_ttm

    if ebitda_margin is None:
        out["score_ebitda_margin"] = 0
        unresolved.append("ebitda_margin")
    elif ebitda_margin >= A_EBITDA_MARGIN_STRONG:
        out["score_ebitda_margin"] = 2
    elif ebitda_margin >= A_EBITDA_MARGIN_ACCEPT:
        out["score_ebitda_margin"] = 1
    else:
        out["score_ebitda_margin"] = 0

    # ── Metric 4: Gross Margin (scoring only — NOT a gate for energy) ─────────
    gross_profit = get_stmt_value(income_stmt, ["Gross Profit", "GrossProfit"])
    gross_margin = None
    if gross_profit is not None and rev_ttm and rev_ttm > 0:
        gross_margin = gross_profit / rev_ttm
    if gross_margin is None:
        gross_margin = safe_float(info.get("grossMargins"))
    out["gross_margin"] = gross_margin

    if gross_margin is None:
        out["score_gross_margin"] = 0
        unresolved.append("gross_margin")
    elif gross_margin >= A_GROSS_MARGIN_STRONG:
        out["score_gross_margin"] = 2
    elif gross_margin >= A_GROSS_MARGIN_ACCEPT:
        out["score_gross_margin"] = 1
    else:
        out["score_gross_margin"] = 0

    # ── Metric 5: EBITDA Growth YoY ──────────────────────────────────────────
    ebitda_now  = compute_ebitda(income_stmt, info, col_idx=0)[0]
    ebitda_prev = compute_ebitda(income_stmt, info, col_idx=1)[0]
    ebitda_growth = None
    if ebitda_now is not None and ebitda_prev is not None and ebitda_prev > 0:
        ebitda_growth = (ebitda_now - ebitda_prev) / ebitda_prev
    out["ebitda_growth"] = ebitda_growth

    if ebitda_growth is None:
        out["score_ebitda_growth"] = 0
        unresolved.append("ebitda_growth")
    elif ebitda_growth >= A_EBITDA_GROWTH_STRONG:
        out["score_ebitda_growth"] = 2
    elif ebitda_growth >= A_EBITDA_GROWTH_ACCEPT:
        out["score_ebitda_growth"] = 1
    else:
        out["score_ebitda_growth"] = 0

    # ── Metric 6: FCF Status ─────────────────────────────────────────────────
    # FCF positive = 2pts; near breakeven (margin > -10%) = 1pt; deeply negative = 0pts
    fcf = get_stmt_value(cashflow, ["Free Cash Flow"])
    if fcf is None:
        ocf  = get_stmt_value(cashflow, ["Operating Cash Flow", "Cash Flow From Continuing Operating Activities"])
        capx = get_stmt_value(cashflow, ["Capital Expenditure", "Capital Expenditures",
                                          "Purchase Of Plant And Equipment", "Purchase Of Ppe"])
        if ocf is not None and capx is not None:
            fcf = ocf + capx  # capex stored as negative in yfinance
    out["fcf"] = fcf

    fcf_margin = fcf / rev_ttm if (fcf is not None and rev_ttm and rev_ttm > 0) else None
    out["fcf_margin"] = fcf_margin

    if fcf is None:
        out["score_fcf"] = 0
        unresolved.append("fcf")
    elif fcf > 0:
        out["score_fcf"] = 2
    elif fcf_margin is not None and fcf_margin > A_FCF_NEAR_BREAKEVEN:
        out["score_fcf"] = 1  # near breakeven (capex cycle)
    else:
        out["score_fcf"] = 0

    # ── Metric 7: ROE ─────────────────────────────────────────────────────────
    roe = safe_float(info.get("returnOnEquity"))
    out["roe"] = roe

    if roe is None:
        out["score_roe"] = 0
        unresolved.append("roe")
    elif roe >= A_ROE_STRONG:
        out["score_roe"] = 2
    elif roe >= A_ROE_ACCEPT:
        out["score_roe"] = 1
    else:
        out["score_roe"] = 0

    # ── Metric 8: Capex Intensity (capex/revenue = investment signal) ─────────
    # For energy, growing capex as % of revenue signals capacity expansion
    capx_val = get_stmt_value(cashflow, [
        "Capital Expenditure", "Capital Expenditures",
        "Purchase Of Plant And Equipment", "Purchase Of Ppe"
    ])
    capex_intensity = None
    if capx_val is not None and rev_ttm and rev_ttm > 0:
        capex_intensity = abs(capx_val) / rev_ttm  # capex is negative, take abs
    out["capex_intensity"] = capex_intensity

    if capex_intensity is None:
        out["score_capex"] = 0
        unresolved.append("capex_intensity")
    elif capex_intensity >= A_CAPEX_INTENSITY_STRONG:
        out["score_capex"] = 2   # heavy investment = growth signal for energy
    elif capex_intensity >= A_CAPEX_INTENSITY_ACCEPT:
        out["score_capex"] = 1
    else:
        out["score_capex"] = 0

    # ── Metric 9: Revenue Scale ───────────────────────────────────────────────
    out["rev_scale"] = rev_ttm

    if rev_ttm is None:
        out["score_rev_scale"] = 0
    elif rev_ttm >= A_REV_SCALE_LARGE:
        out["score_rev_scale"] = 2
    elif rev_ttm >= A_REV_SCALE_MID:
        out["score_rev_scale"] = 1
    else:
        out["score_rev_scale"] = 0

    # ── Metric 10: Forward Growth (earnings or revenue) ───────────────────────
    fwd_growth = gate_data.get("fwd_growth") or safe_float(info.get("revenueGrowth"))
    eps_growth  = safe_float(info.get("earningsGrowth"))
    # Use better of the two as forward growth signal
    best_fwd = None
    if fwd_growth is not None and eps_growth is not None:
        best_fwd = max(fwd_growth, eps_growth)
    else:
        best_fwd = fwd_growth or eps_growth
    out["fwd_growth"] = best_fwd

    if best_fwd is None:
        out["score_fwd_growth"] = 0
        unresolved.append("fwd_growth")
    elif best_fwd >= A_FWD_GROWTH_STRONG:
        out["score_fwd_growth"] = 2
    elif best_fwd >= A_FWD_GROWTH_ACCEPT:
        out["score_fwd_growth"] = 1
    else:
        out["score_fwd_growth"] = 0

    # ── Total Part A ──────────────────────────────────────────────────────────
    score_keys = [
        "score_rev_growth_ttm", "score_rev_cagr", "score_ebitda_margin",
        "score_gross_margin", "score_ebitda_growth", "score_fcf",
        "score_roe", "score_capex", "score_rev_scale", "score_fwd_growth"
    ]
    part_a_total = sum(out.get(k, 0) for k in score_keys)
    out["part_a_score"]   = part_a_total
    out["part_a_max"]     = 20
    out["part_a_unresolved"] = len(unresolved)
    out["unresolved_metrics"] = unresolved

    if part_a_total >= PART_A_STRONG_GROWTH:
        out["part_a_grade"] = "Strong Growth"
    elif part_a_total >= PART_A_ACCEPTABLE:
        out["part_a_grade"] = "Acceptable Growth"
    else:
        out["part_a_grade"] = "Not Growth"

    return out


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 6: PART B SCORING — VALUATION & BUY SIGNAL (8 metrics, max 16 pts)
# ─────────────────────────────────────────────────────────────────────────────

def score_part_b(ticker_sym, info, income_stmt, balance_sheet, part_a_data):
    """
    Energy Part B: 8 valuation/buy-signal metrics, each 0/1/2.
    Max = 16 pts. Strong Buy ≥11, Watch 6-10, Avoid <6.
    Only scored for Part A Strong Growth (≥14).
    """
    out = {"ticker": ticker_sym}
    unresolved = []

    prices = apply_pence_correction(info)
    current_price = prices["current_price"]
    target_mean   = prices["target_mean"]
    high_52wk     = prices["high_52wk"]
    low_52wk      = prices["low_52wk"]

    # ── Metric 1: EV/EBITDA ──────────────────────────────────────────────────
    ev_ebitda = safe_float(info.get("enterpriseToEbitda"))
    out["ev_ebitda"] = ev_ebitda

    if ev_ebitda is None:
        out["score_ev_ebitda"] = 0
        unresolved.append("ev_ebitda")
    elif ev_ebitda <= B_EV_EBITDA_STRONG:
        out["score_ev_ebitda"] = 2
    elif ev_ebitda <= B_EV_EBITDA_ACCEPT:
        out["score_ev_ebitda"] = 1
    else:
        out["score_ev_ebitda"] = 0

    # ── Metric 2: Net Debt / EBITDA ──────────────────────────────────────────
    ebitda = part_a_data.get("ebitda")
    total_debt = get_stmt_value(balance_sheet, ["Total Debt", "Long Term Debt And Capital Lease Obligation"])
    cash = safe_float(info.get("totalCash")) or get_stmt_value(
        balance_sheet, ["Cash And Cash Equivalents", "Cash Cash Equivalents And Short Term Investments"]
    )
    nd_ebitda = None
    if total_debt is not None and ebitda and ebitda > 0:
        net_debt = (total_debt or 0) - (cash or 0)
        nd_ebitda = net_debt / ebitda
    out["net_debt"] = (total_debt or 0) - (cash or 0) if total_debt is not None else None
    out["nd_ebitda"] = nd_ebitda

    if nd_ebitda is None:
        out["score_nd_ebitda"] = 0
        unresolved.append("nd_ebitda")
    elif nd_ebitda <= B_ND_EBITDA_STRONG:
        out["score_nd_ebitda"] = 2
    elif nd_ebitda <= B_ND_EBITDA_ACCEPT:
        out["score_nd_ebitda"] = 1
    else:
        out["score_nd_ebitda"] = 0

    # ── Metric 3: Analyst Recommendation ─────────────────────────────────────
    rec_mean = safe_float(info.get("recommendationMean"))
    analyst_count = safe_float(info.get("numberOfAnalystOpinions"))
    out["recommendation_mean"] = rec_mean
    out["analyst_count"] = analyst_count

    if rec_mean is None:
        out["score_analyst_rec"] = 0
        unresolved.append("analyst_rec")
    elif rec_mean <= B_ANALYST_BUY_STRONG:
        out["score_analyst_rec"] = 2
    elif rec_mean <= B_ANALYST_BUY_ACCEPT:
        out["score_analyst_rec"] = 1
    else:
        out["score_analyst_rec"] = 0

    # ── Metric 4: Price vs Analyst Target (Upside) ───────────────────────────
    upside = None
    if current_price and target_mean and current_price > 0:
        upside = (target_mean - current_price) / current_price
    out["current_price"] = current_price
    out["target_mean"] = target_mean
    out["upside_pct"] = upside

    if upside is None:
        out["score_upside"] = 0
        unresolved.append("upside")
    elif upside >= B_UPSIDE_STRONG:
        out["score_upside"] = 2
    elif upside >= B_UPSIDE_ACCEPT:
        out["score_upside"] = 1
    else:
        out["score_upside"] = 0

    # ── Metric 5: 52-Week Position (momentum) ────────────────────────────────
    pos_52wk = None
    range_52wk = high_52wk - low_52wk if high_52wk and low_52wk else None
    if current_price and range_52wk and range_52wk > 0 and low_52wk:
        pos_52wk = (current_price - low_52wk) / range_52wk
    out["pos_52wk"] = pos_52wk
    out["high_52wk"] = high_52wk
    out["low_52wk"] = low_52wk

    if pos_52wk is None:
        out["score_52wk"] = 0
        unresolved.append("52wk_pos")
    elif pos_52wk >= B_52WK_POS_STRONG:
        out["score_52wk"] = 2
    elif pos_52wk >= B_52WK_POS_ACCEPT:
        out["score_52wk"] = 1
    else:
        out["score_52wk"] = 0

    # ── Metric 6: Forward P/E ─────────────────────────────────────────────────
    fwd_pe = safe_float(info.get("forwardPE"))
    out["fwd_pe"] = fwd_pe

    if fwd_pe is None or fwd_pe <= 0:
        out["score_fwd_pe"] = 0
        if fwd_pe is None:
            unresolved.append("fwd_pe")
    elif fwd_pe <= B_FWD_PE_STRONG:
        out["score_fwd_pe"] = 2
    elif fwd_pe <= B_FWD_PE_ACCEPT:
        out["score_fwd_pe"] = 1
    else:
        out["score_fwd_pe"] = 0

    # ── Metric 7: Forward EPS Growth ─────────────────────────────────────────
    eps_growth = safe_float(info.get("earningsGrowth"))
    out["eps_growth"] = eps_growth

    if eps_growth is None:
        out["score_eps_growth"] = 0
        unresolved.append("eps_growth")
    elif eps_growth >= B_EPS_GROWTH_STRONG:
        out["score_eps_growth"] = 2
    elif eps_growth >= B_EPS_GROWTH_ACCEPT:
        out["score_eps_growth"] = 1
    else:
        out["score_eps_growth"] = 0

    # ── Metric 8: Analyst Coverage (count) ───────────────────────────────────
    ac = int(analyst_count) if analyst_count is not None else None
    out["analyst_count"] = ac

    if ac is None:
        out["score_analyst_count"] = 0
        unresolved.append("analyst_count")
    elif ac >= B_ANALYST_COUNT_STRONG:
        out["score_analyst_count"] = 2
    elif ac >= B_ANALYST_COUNT_ACCEPT:
        out["score_analyst_count"] = 1
    else:
        out["score_analyst_count"] = 0

    # ── Total Part B ──────────────────────────────────────────────────────────
    if getattr(_cfg, "ENERGY_VALUATION_PARITY", False):
        # Part 2 §F parity: growth-ADJUST valuation (high multiple justified by growth not penalised) +
        # DROP the stale 52-week-position metric. Growth from yfinance info (earnings/revenue growth).
        _g = safe_float(info.get("earningsGrowth")) or safe_float(info.get("revenueGrowth"))
        if _g and _g > 0:
            _gp = _g * 100.0
            _ev = safe_float(out.get("ev_ebitda"))
            _fp = safe_float(out.get("fwd_pe"))
            if _ev and _ev > 0:
                out["score_ev_ebitda"] = 2 if (_ev / _gp) < 1.0 else (1 if (_ev / _gp) < 2.0 else 0)
            if _fp and _fp > 0:
                out["score_fwd_pe"] = 2 if (_fp / _gp) < 1.0 else (1 if (_fp / _gp) < 2.0 else 0)
        score_keys_b = ["score_ev_ebitda", "score_nd_ebitda", "score_analyst_rec",
                        "score_upside", "score_fwd_pe", "score_eps_growth", "score_analyst_count"]  # 52wk dropped
        part_b_max = 14
    else:
        score_keys_b = ["score_ev_ebitda", "score_nd_ebitda", "score_analyst_rec",
                        "score_upside", "score_52wk", "score_fwd_pe",
                        "score_eps_growth", "score_analyst_count"]
        part_b_max = 16
    part_b_total = sum(out.get(k, 0) or 0 for k in score_keys_b)
    out["part_b_score"]      = part_b_total
    out["part_b_max"]        = part_b_max
    out["part_b_unresolved"] = len(unresolved)

    if part_b_total >= PART_B_STRONG_BUY:
        out["part_b_grade"] = "Strong Buy"
    elif part_b_total >= PART_B_WATCH:
        out["part_b_grade"] = "Watch"
    else:
        out["part_b_grade"] = "Avoid"

    return out


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 7: FINAL STATUS & ROW ASSEMBLY
# ─────────────────────────────────────────────────────────────────────────────

def _final_status(gate_data, part_a_data, part_b_data):
    """Determine final status string."""
    if not gate_data.get("gate_pass") and gate_data.get("gate_pass") is not None:
        return STATUS_GATE_FAIL
    if gate_data.get("gate_pass") is None:
        return STATUS_DATA_ISSUE
    a = part_a_data.get("part_a_score", 0) if part_a_data else 0
    b = part_b_data.get("part_b_score", 0) if part_b_data else 0
    if a >= PART_A_STRONG_GROWTH and b >= PART_B_STRONG_BUY:
        return STATUS_STRONG_BUY
    if a >= PART_A_STRONG_GROWTH:
        return STATUS_WATCH
    if a >= PART_A_ACCEPTABLE:
        return STATUS_ACCEPTABLE
    return STATUS_NOT_GROWTH


def build_output_row(entry, info, gate_data, part_a_data, part_b_data):
    """Assemble a single flat dict for the full_data CSV."""
    ticker = entry["ticker"]
    company = entry["company"]
    sector = entry.get("sector", "")
    sub_type = entry.get("sub_type", "")
    region = entry.get("region", "")

    final_status = _final_status(gate_data, part_a_data, part_b_data)
    a_total = (part_a_data or {}).get("part_a_score")
    b_total = (part_b_data or {}).get("part_b_score")
    total = (a_total or 0) + (b_total or 0) if (a_total is not None and b_total is not None) else None

    row = {
        # Identity
        "ticker":         ticker,
        "company":        company,
        "sector":         sector,
        "sub_type":       sub_type,
        "region":         region,
        "final_status":   final_status,
        # Gate
        "gate_code":      gate_data.get("gate_code", ""),
        "gate_reason":    gate_data.get("gate_reason", ""),
        # Scores
        "part_a_score":   a_total,
        "part_b_score":   b_total,
        "total_score":    total,
        # Part A metrics
        "rev_growth_ttm":    _pct_fmt(part_a_data, "rev_growth_ttm"),
        "rev_cagr":          _pct_fmt(part_a_data, "rev_cagr"),
        "ebitda_margin":     _pct_fmt(part_a_data, "ebitda_margin"),
        "gross_margin":      _pct_fmt(part_a_data, "gross_margin"),
        "ebitda_growth":     _pct_fmt(part_a_data, "ebitda_growth"),
        "fcf_margin":        _pct_fmt(part_a_data, "fcf_margin"),
        "roe":               _pct_fmt(part_a_data, "roe"),
        "capex_intensity":   _pct_fmt(part_a_data, "capex_intensity"),
        "rev_ttm_bn":        _bn_fmt(part_a_data, "rev_ttm"),
        "fwd_growth":        _pct_fmt(part_a_data, "fwd_growth"),
        # Part A individual scores
        "sa_rev_growth_ttm": _score_fmt(part_a_data, "score_rev_growth_ttm"),
        "sa_rev_cagr":       _score_fmt(part_a_data, "score_rev_cagr"),
        "sa_ebitda_margin":  _score_fmt(part_a_data, "score_ebitda_margin"),
        "sa_gross_margin":   _score_fmt(part_a_data, "score_gross_margin"),
        "sa_ebitda_growth":  _score_fmt(part_a_data, "score_ebitda_growth"),
        "sa_fcf":            _score_fmt(part_a_data, "score_fcf"),
        "sa_roe":            _score_fmt(part_a_data, "score_roe"),
        "sa_capex":          _score_fmt(part_a_data, "score_capex"),
        "sa_rev_scale":      _score_fmt(part_a_data, "score_rev_scale"),
        "sa_fwd_growth":     _score_fmt(part_a_data, "score_fwd_growth"),
        # Part B metrics
        "ev_ebitda":         _mult_fmt(part_b_data, "ev_ebitda"),
        "nd_ebitda":         _mult_fmt(part_b_data, "nd_ebitda"),
        "analyst_rec":       _val_fmt(part_b_data, "recommendation_mean"),
        "upside_pct":        _pct_fmt(part_b_data, "upside_pct"),
        "pos_52wk":          _pct_fmt(part_b_data, "pos_52wk"),
        "fwd_pe":            _val_fmt(part_b_data, "fwd_pe"),
        "eps_growth":        _pct_fmt(part_b_data, "eps_growth"),
        "analyst_count":     _val_fmt(part_b_data, "analyst_count"),
        # Part B individual scores
        "sb_ev_ebitda":      _score_fmt(part_b_data, "score_ev_ebitda"),
        "sb_nd_ebitda":      _score_fmt(part_b_data, "score_nd_ebitda"),
        "sb_analyst_rec":    _score_fmt(part_b_data, "score_analyst_rec"),
        "sb_upside":         _score_fmt(part_b_data, "score_upside"),
        "sb_52wk":           _score_fmt(part_b_data, "score_52wk"),
        "sb_fwd_pe":         _score_fmt(part_b_data, "score_fwd_pe"),
        "sb_eps_growth":     _score_fmt(part_b_data, "score_eps_growth"),
        "sb_analyst_count":  _score_fmt(part_b_data, "score_analyst_count"),
        # Prices
        "current_price":     _val_fmt(part_b_data, "current_price"),
        "target_mean":       _val_fmt(part_b_data, "target_mean"),
        "high_52wk":         _val_fmt(part_b_data, "high_52wk"),
        "low_52wk":          _val_fmt(part_b_data, "low_52wk"),
        # Info
        "analyst_rating":    (info or {}).get("recommendationKey", ""),
        "currency":          (info or {}).get("currency", ""),
        "notes":             entry.get("notes", ""),
    }
    return row


def _pct_fmt(d, key):
    if not d:
        return ""
    v = d.get(key)
    if v is None:
        return ""
    try:
        return f"{float(v) * 100:.2f}%"
    except Exception:
        return str(v)


def _bn_fmt(d, key):
    if not d:
        return ""
    v = d.get(key)
    if v is None:
        return ""
    try:
        return f"{float(v) / 1e9:.2f}"
    except Exception:
        return str(v)


def _mult_fmt(d, key):
    if not d:
        return ""
    v = d.get(key)
    if v is None:
        return ""
    try:
        return f"{float(v):.1f}x"
    except Exception:
        return str(v)


def _val_fmt(d, key):
    if not d:
        return ""
    v = d.get(key)
    if v is None:
        return ""
    return str(v)


def _score_fmt(d, key):
    if not d:
        return ""
    v = d.get(key)
    return str(v) if v is not None else ""


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 8: CSV OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

FULL_DATA_COLUMNS = [
    "ticker", "company", "sector", "sub_type", "region", "final_status",
    "gate_code", "gate_reason",
    "part_a_score", "part_b_score", "total_score",
    "rev_growth_ttm", "rev_cagr", "ebitda_margin", "gross_margin",
    "ebitda_growth", "fcf_margin", "roe", "capex_intensity", "rev_ttm_bn", "fwd_growth",
    "sa_rev_growth_ttm", "sa_rev_cagr", "sa_ebitda_margin", "sa_gross_margin",
    "sa_ebitda_growth", "sa_fcf", "sa_roe", "sa_capex", "sa_rev_scale", "sa_fwd_growth",
    "ev_ebitda", "nd_ebitda", "analyst_rec", "upside_pct", "pos_52wk",
    "fwd_pe", "eps_growth", "analyst_count",
    "sb_ev_ebitda", "sb_nd_ebitda", "sb_analyst_rec", "sb_upside", "sb_52wk",
    "sb_fwd_pe", "sb_eps_growth", "sb_analyst_count",
    "current_price", "target_mean", "high_52wk", "low_52wk",
    "analyst_rating", "currency", "notes",
]

GATE_COLUMNS = [
    "ticker", "company", "sector", "sub_type", "region",
    "gate_code", "gate_reason", "final_status",
]


def write_csv(path, rows, columns):
    """Write list of dicts to CSV with specified column order."""
    if not rows:
        rows = [{}]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9: EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

def _cell_style(ws, row, col, value, bold=False, bg=None, color=None,
                align="left", number_format=None, border=False):
    """Apply value and styles to a single cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if bold:
        cell.font = Font(bold=True, color=color or "000000")
    elif color:
        cell.font = Font(color=color)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if number_format:
        cell.number_format = number_format
    if border:
        thin = Side(style="thin", color="D0D0D0")
        cell.border = Border(bottom=thin)
    return cell


def build_excel(full_rows, gate_rows, run_date_str, inv_dir, group="ENERGY"):
    """
    Build a 5-tab Excel workbook:
    SUMMARY | WATCHLIST | SCORES | EXCLUSIONS | DIAGNOSTICS
    """
    if openpyxl is None:
        log.error("openpyxl not available — skipping Excel build")
        return None

    DARK_BLUE  = "1A3A6B"
    DARK_GREEN = "1A6B2A"
    AMBER      = "D4830A"
    LIGHT_BLUE = "EBF0FA"
    LIGHT_GREEN= "EBF7EE"
    LIGHT_AMBER= "FDF3E3"
    WHITE      = "FFFFFF"
    ROW_ALT    = "F5F7FA"

    wb = openpyxl.Workbook()

    # ── Tab 1: SUMMARY — Strong Buys ranked by Total Score ───────────────────
    ws = wb.active
    ws.title = "SUMMARY"
    strong_buys = sorted(
        [r for r in full_rows if r.get("final_status") == STATUS_STRONG_BUY],
        key=lambda r: -(safe_float(r.get("total_score")) or 0)
    )

    ws.merge_cells("A1:P1")
    title_cell = ws["A1"]
    title_cell.value = f"ISA Energy Growth Stock Analysis — {run_date_str} | Strong Buys (Part A ≥14 AND Part B ≥11)"
    title_cell.font = Font(bold=True, size=13, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:P2")
    sb_count = len(strong_buys)
    ws["A2"].value = f"{sb_count} Strong {'Buy' if sb_count==1 else 'Buys'} | Part A max 20 | Part B max 16 | Total max 36"
    ws["A2"].font = Font(bold=True, size=11, color=DARK_BLUE)
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.row_dimensions[2].height = 18

    summary_headers = [
        "Rank", "Ticker", "Company", "Sector", "Sub-Type", "Region",
        "Part A\n(/20)", "Part B\n(/16)", "Total\n(/36)",
        "Rev Growth\nTTM", "EBITDA\nMargin", "Gross\nMargin",
        "EBITDA\nGrowth", "ROE", "EV/EBITDA", "Upside vs\nTarget",
        "Fwd Axis\n(/100)"   # forward axis (Part 3 §13) — growth/energy parity; display only
    ]
    for c, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "A4"

    for rank, row in enumerate(strong_buys, 1):
        r = 3 + rank
        bg = WHITE if rank % 2 else ROW_ALT
        a_score = safe_float(row.get("part_a_score"))
        b_score = safe_float(row.get("part_b_score"))
        total   = safe_float(row.get("total_score"))

        values = [
            rank, row.get("ticker"), row.get("company"),
            row.get("sector"), row.get("sub_type"), row.get("region"),
            a_score, b_score, total,
            row.get("rev_growth_ttm"), row.get("ebitda_margin"), row.get("gross_margin"),
            row.get("ebitda_growth"), row.get("roe"), row.get("ev_ebitda"), row.get("upside_pct"),
            row.get("forward_axis_score"),
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=9)
            if c == 9 and total is not None:  # Total score — bold green
                cell.font = Font(bold=True, size=10, color=DARK_GREEN)
            if c in (2, 3):
                cell.alignment = Alignment(horizontal="left", vertical="center")

    col_widths = [5, 8, 24, 14, 18, 6, 8, 8, 8, 11, 11, 10, 10, 8, 10, 11, 9]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── Tab 2: WATCHLIST — All companies with scores ─────────────────────────
    ws2 = wb.create_sheet("WATCHLIST")
    wl_headers = [
        "Ticker", "Company", "Sector", "Sub-Type", "Region", "Status",
        "Part A", "Part B", "Total", "Gate Code", "Gate Reason",
        "Rev Growth TTM", "EBITDA Margin", "Gross Margin", "FCF Margin",
        "ROE", "EV/EBITDA", "Upside %", "Current Price", "Target", "Notes"
    ]
    for c, h in enumerate(wl_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, size=9)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center")

    sorted_all = sorted(
        full_rows,
        key=lambda r: (
            0 if r.get("final_status") == STATUS_STRONG_BUY else
            1 if r.get("final_status") == STATUS_WATCH else
            2 if r.get("final_status") == STATUS_ACCEPTABLE else
            3 if r.get("final_status") == STATUS_NOT_GROWTH else
            4 if r.get("final_status") == STATUS_GATE_FAIL else 5,
            -(safe_float(r.get("total_score")) or 0)
        )
    )

    status_colours = {
        STATUS_STRONG_BUY: LIGHT_GREEN,
        STATUS_WATCH:      LIGHT_BLUE,
        STATUS_ACCEPTABLE: LIGHT_BLUE,
        STATUS_NOT_GROWTH: "FFF8F0",
        STATUS_GATE_FAIL:  LIGHT_AMBER,
        STATUS_DATA_ISSUE: LIGHT_AMBER,
    }

    for r_idx, row in enumerate(sorted_all, 2):
        bg = status_colours.get(row.get("final_status"), WHITE)
        vals = [
            row.get("ticker"), row.get("company"), row.get("sector"),
            row.get("sub_type"), row.get("region"), row.get("final_status"),
            row.get("part_a_score"), row.get("part_b_score"), row.get("total_score"),
            row.get("gate_code"), row.get("gate_reason"),
            row.get("rev_growth_ttm"), row.get("ebitda_margin"), row.get("gross_margin"),
            row.get("fcf_margin"), row.get("roe"), row.get("ev_ebitda"),
            row.get("upside_pct"), row.get("current_price"), row.get("target_mean"),
            row.get("notes"),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=r_idx, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left" if c > 2 else "center", wrap_text=(c == 21))
            cell.font = Font(size=9)

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 6
    ws2.column_dimensions["F"].width = 20
    for col_letter in ["G", "H", "I"]:
        ws2.column_dimensions[col_letter].width = 8
    ws2.column_dimensions["U"].width = 35
    ws2.freeze_panes = "A2"

    # ── Tab 3: SCORES — Per-metric breakdown ──────────────────────────────────
    ws3 = wb.create_sheet("SCORES")
    score_headers = [
        "Ticker", "Company", "Part A", "Part B", "Total",
        # Part A
        "A1:RevGrow", "A2:RevCAGR", "A3:EBITDA_M", "A4:GrossM",
        "A5:EBITDA_G", "A6:FCF", "A7:ROE", "A8:Capex", "A9:Scale", "A10:FwdGrow",
        # Part B
        "B1:EV/EBITDA", "B2:ND/EBITDA", "B3:AnalystRec", "B4:Upside",
        "B5:52wk", "B6:FwdPE", "B7:EPS_G", "B8:Coverage",
        # Forward axis (Part 3 §13) — growth/energy parity, display only (not in selection until flagged)
        "Fwd/100", "F:EPStr", "F:Marg", "F:RevEst", "F:PxMom", "RevStage",
    ]
    for c, h in enumerate(score_headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, size=9)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center")

    scored_rows = [r for r in sorted_all if r.get("part_a_score") is not None]
    for r_idx, row in enumerate(scored_rows, 2):
        vals = [
            row.get("ticker"), row.get("company"),
            row.get("part_a_score"), row.get("part_b_score"), row.get("total_score"),
            row.get("sa_rev_growth_ttm"), row.get("sa_rev_cagr"), row.get("sa_ebitda_margin"),
            row.get("sa_gross_margin"), row.get("sa_ebitda_growth"), row.get("sa_fcf"),
            row.get("sa_roe"), row.get("sa_capex"), row.get("sa_rev_scale"), row.get("sa_fwd_growth"),
            row.get("sb_ev_ebitda"), row.get("sb_nd_ebitda"), row.get("sb_analyst_rec"),
            row.get("sb_upside"), row.get("sb_52wk"), row.get("sb_fwd_pe"),
            row.get("sb_eps_growth"), row.get("sb_analyst_count"),
            row.get("forward_axis_score"), row.get("score_f_eps_trend"),
            row.get("score_f_margin_traj"), row.get("score_f_rev_est"),
            row.get("score_f_price_mom"), row.get("revision_stage"),
        ]
        bg = WHITE if r_idx % 2 else ROW_ALT
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=r_idx, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(size=9)
            # Colour individual scores: 2=green, 1=amber, 0=light red
            if c >= 6 and v is not None:
                try:
                    iv = int(v)
                    if iv == 2:
                        cell.font = Font(size=9, bold=True, color=DARK_GREEN)
                    elif iv == 0:
                        cell.font = Font(size=9, color="C0392B")
                except Exception:
                    pass
    ws3.freeze_panes = "A2"
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 24

    # ── Tab 4: EXCLUSIONS ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("EXCLUSIONS")
    excl_headers = ["Ticker", "Company", "Sector", "Sub-Type", "Region", "Gate", "Reason"]
    for c, h in enumerate(excl_headers, 1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, size=9)
        cell.fill = PatternFill("solid", fgColor=AMBER)
        cell.alignment = Alignment(horizontal="center")

    excl_rows = [r for r in sorted_all if r.get("final_status") in (STATUS_GATE_FAIL, STATUS_DATA_ISSUE)]
    for r_idx, row in enumerate(excl_rows, 2):
        bg = WHITE if r_idx % 2 else LIGHT_AMBER
        vals = [
            row.get("ticker"), row.get("company"), row.get("sector"),
            row.get("sub_type"), row.get("region"),
            row.get("gate_code"), row.get("gate_reason"),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws4.cell(row=r_idx, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left" if c > 1 else "center", wrap_text=(c == 7))
            cell.font = Font(size=9)
    ws4.column_dimensions["B"].width = 24
    ws4.column_dimensions["G"].width = 45

    # ── Tab 5: DIAGNOSTICS ────────────────────────────────────────────────────
    ws5 = wb.create_sheet("DIAGNOSTICS")
    ws5["A1"] = "ENERGY SCREENER DIAGNOSTICS"
    ws5["A1"].font = Font(bold=True, size=12, color=WHITE)
    ws5["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)

    diag_data = [
        ("Run Date", run_date_str),
        ("Group", "ENERGY"),
        ("Watchlist Size", len(full_rows) + len(gate_rows)),
        ("Gate Passers Scored", len([r for r in full_rows if r.get("part_a_score") is not None])),
        ("Strong Buys", len(strong_buys)),
        ("Watch", len([r for r in full_rows if r.get("final_status") == STATUS_WATCH])),
        ("Acceptable Growth", len([r for r in full_rows if r.get("final_status") == STATUS_ACCEPTABLE])),
        ("Not Growth", len([r for r in full_rows if r.get("final_status") == STATUS_NOT_GROWTH])),
        ("Gate Fails", len([r for r in full_rows if r.get("final_status") == STATUS_GATE_FAIL])),
        ("Data Issues", len([r for r in full_rows if r.get("final_status") == STATUS_DATA_ISSUE])),
        ("Part A Threshold (Strong)", f">={PART_A_STRONG_GROWTH}/20"),
        ("Part B Threshold (Strong Buy)", f">={PART_B_STRONG_BUY}/16"),
        ("screener_core imported", str(_SCREENER_CORE_AVAILABLE)),
        ("Thresholds calibrated", "May 2026 — 25 company empirical pull"),
    ]
    for r_idx, (k, v) in enumerate(diag_data, 3):
        ws5.cell(row=r_idx, column=1, value=k).font = Font(bold=True, size=9)
        ws5.cell(row=r_idx, column=2, value=str(v)).font = Font(size=9)
    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 30

    # ── Save ──────────────────────────────────────────────────────────────────
    date_obj = datetime.strptime(run_date_str, "%Y%m%d") if len(run_date_str) == 8 else datetime.now()
    display_date = date_obj.strftime("%d-%b-%y")
    filename = f"Growth Stock Analysis ENERGY W-e {display_date}.xlsx"
    output_path = os.path.join(inv_dir, filename)
    wb.save(output_path)
    log.info(f"Excel saved: {filename}")
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 10: RETROSPECTIVE WRITING
# ─────────────────────────────────────────────────────────────────────────────

def write_retrospective(inv_dir, run_date, full_rows, gate_rows, fetch_errors, run_qa):
    """Write {YYYYMMDD}_ENERGY_retrospective.md to Investment Analysis folder."""
    path = os.path.join(inv_dir, f"{run_date}_ENERGY_retrospective.md")

    strong_buys  = [r for r in full_rows if r.get("final_status") == STATUS_STRONG_BUY]
    gate_fails   = [r for r in full_rows if r.get("final_status") == STATUS_GATE_FAIL]
    data_issues  = [r for r in full_rows if r.get("final_status") == STATUS_DATA_ISSUE]
    total_watched = len(full_rows)

    date_display = datetime.strptime(run_date, "%Y%m%d").strftime("%d %b %Y") if len(run_date) == 8 else run_date

    # Build retrospective items based on run
    items = []

    if fetch_errors:
        failed_tickers = [e["ticker"] for e in fetch_errors[:5]]
        items.append((
            "yfinance Fetch Failures",
            f"{len(fetch_errors)} fetch error(s) this run: {', '.join(failed_tickers)}. "
            f"Consider adding these to a retry list or verifying ticker symbols in energy_watchlist.json. "
            f"Impact: completeness.",
            "M"
        ))

    unresolved_fields = []
    for r in full_rows:
        if r.get("part_a_score") is not None:
            for col in ["rev_growth_ttm", "ebitda_margin", "fwd_growth", "ev_ebitda", "nd_ebitda"]:
                if not r.get(col):
                    unresolved_fields.append(col)
    if unresolved_fields:
        from collections import Counter
        field_counts = Counter(unresolved_fields)
        top_fields = ", ".join(f"{f}({c})" for f, c in field_counts.most_common(3))
        items.append((
            "Unresolved Metrics — Fallback Sourcing",
            f"Most commonly unresolved fields this run: {top_fields}. "
            f"Consider adding Finnhub API calls as fallback for these specific metrics. "
            f"Impact: accuracy.",
            "M"
        ))

    if len(strong_buys) == 0:
        items.append((
            "No Strong Buys — Threshold Review",
            f"Zero Strong Buys this run with current thresholds (Part A ≥{PART_A_STRONG_GROWTH}/20 AND Part B ≥{PART_B_STRONG_BUY}/16). "
            f"Consider whether thresholds need recalibration or if this reflects genuine market conditions. "
            f"Review Watch list for near-misses. Impact: effectiveness.",
            "L"
        ))

    if not items:
        items.append((
            "Watchlist Review — Quarterly Cadence",
            f"Run completed cleanly. Schedule quarterly watchlist review: verify index membership changes, "
            f"add new energy growth names (check Siemens Energy ENR.DE sub-indices, new nuclear developers), "
            f"remove companies that have lost growth credentials. Impact: completeness.",
            "S"
        ))

    content = f"""# ISA Energy Growth Stock Analysis — Retrospective
**Run date:** {date_display}
**Group:** ENERGY
**Watchlist:** {total_watched} companies screened

---

## Run Summary

| Metric | Count |
|--------|-------|
| Total watchlist | {total_watched} |
| Strong Buys (A≥{PART_A_STRONG_GROWTH} AND B≥{PART_B_STRONG_BUY}) | {len(strong_buys)} |
| Watch (A≥{PART_A_STRONG_GROWTH}, B<{PART_B_STRONG_BUY}) | {len([r for r in full_rows if r.get("final_status")==STATUS_WATCH])} |
| Acceptable Growth (A {PART_A_ACCEPTABLE}-{PART_A_STRONG_GROWTH-1}) | {len([r for r in full_rows if r.get("final_status")==STATUS_ACCEPTABLE])} |
| Not Growth (A<{PART_A_ACCEPTABLE}) | {len([r for r in full_rows if r.get("final_status")==STATUS_NOT_GROWTH])} |
| Gate Fails | {len(gate_fails)} |
| Data Issues | {len(data_issues)} |
| Fetch Errors | {len(fetch_errors)} |

## Strong Buys This Run

{'None identified.' if not strong_buys else chr(10).join(f"- **{r['ticker']}** ({r['company']}): A={r.get('part_a_score')}/20, B={r.get('part_b_score')}/16, Total={r.get('total_score')}/36" for r in strong_buys)}

## Gate Exclusions

{'None.' if not gate_fails else chr(10).join(f"- **{r['ticker']}** ({r['company']}): {r.get('gate_code')} — {r.get('gate_reason','')}" for r in gate_fails)}

---

## Recommendations

"""
    for i, (title, body, impact) in enumerate(items[:5], 1):
        content += f"{i}. **{title}:** {body} [Impact: {impact}]\n\n"

    content += f"""---
*Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Thresholds: May 2026 empirical calibration*
"""

    with open(path, "w", encoding="utf-8") as f:
        f.write(content)

    log.info(f"Retrospective written: {os.path.basename(path)}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 11: RUN QA CSV
# ─────────────────────────────────────────────────────────────────────────────

def write_run_qa(outputs_dir, run_date, run_qa):
    """Write run_qa dict to CSV in outputs directory."""
    path = os.path.join(outputs_dir, f"{run_date}_ENERGY_run_qa.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["metric", "value"])
        writer.writeheader()
        for k, v in run_qa.items():
            writer.writerow({"metric": k, "value": str(v)})
    return path


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 12: MAIN RUN FUNCTION
# ─────────────────────────────────────────────────────────────────────────────

def run_energy(run_date, outputs_dir, inv_dir):
    """
    Full energy screening run.
    run_date: YYYYMMDD string
    outputs_dir: ephemeral outputs path (cleared between sessions)
    inv_dir: Investment Analysis folder path (permanent deliverables)
    """
    start_time = time.time()
    log.info(f"=== ENERGY SCREENER — {run_date} ===")
    log.info(f"screener_core utilities: {'imported' if _SCREENER_CORE_AVAILABLE else 'inline fallback'}")

    os.makedirs(outputs_dir, exist_ok=True)

    run_qa = {
        "group":      "ENERGY",
        "run_date":   run_date,
        "start_time": datetime.utcnow().isoformat(),
        "screener_core_imported": _SCREENER_CORE_AVAILABLE,
    }

    # ── Step 1: Load watchlist ────────────────────────────────────────────────
    watchlist = load_watchlist(inv_dir)
    tickers = [e["ticker"] for e in watchlist]
    watchlist_map = {e["ticker"]: e for e in watchlist}
    run_qa["watchlist_size"] = len(watchlist)

    # ── Step 2: Fetch all data ────────────────────────────────────────────────
    info_map, stmt_map, fetch_errors = fetch_all_data(tickers)
    run_qa["info_fetched"]     = len(info_map)
    run_qa["stmts_fetched"]    = len(stmt_map)
    run_qa["fetch_errors"]     = len(fetch_errors)

    full_rows = []
    gate_rows = []

    # ── Step 3: Gates → Score → Assemble ─────────────────────────────────────
    for entry in watchlist:
        ticker = entry["ticker"]
        info        = info_map.get(ticker, {})
        stmts       = stmt_map.get(ticker, {})
        income_stmt = stmts.get("income_stmt")
        cashflow    = stmts.get("cashflow")
        balance_sheet = stmts.get("balance_sheet")

        # Handle complete data miss
        if not info and income_stmt is None:
            gate_data = {
                "gate_pass": None,
                "gate_code": "GATE_DATA_UNRESOLVED",
                "gate_reason": "No yfinance data returned",
            }
            row = build_output_row(entry, {}, gate_data, None, None)
            full_rows.append(row)
            gate_rows.append(row)
            continue

        # Apply gates
        gate_data = apply_energy_gates(ticker, info, income_stmt, entry)
        gate_rows.append({**{"ticker": ticker, "company": entry["company"],
                              "sector": entry.get("sector",""), "sub_type": entry.get("sub_type",""),
                              "region": entry.get("region","")}, **gate_data})

        if not gate_data.get("gate_pass"):
            # Gate fail or unresolved — build output row without scores
            row = build_output_row(entry, info, gate_data, None, None)
            full_rows.append(row)
            continue

        # Part A scoring
        try:
            part_a = score_part_a(ticker, info, income_stmt, cashflow, balance_sheet, gate_data)
        except Exception as e:
            log.warning(f"Part A scoring failed for {ticker}: {e}")
            part_a = {"part_a_score": 0, "part_a_grade": "Not Growth"}

        # Part B scoring (only for Strong Growth)
        part_b = None
        if (part_a.get("part_a_score") or 0) >= PART_A_STRONG_GROWTH:
            try:
                part_b = score_part_b(ticker, info, income_stmt, balance_sheet, part_a)
            except Exception as e:
                log.warning(f"Part B scoring failed for {ticker}: {e}")
                part_b = {"part_b_score": 0, "part_b_grade": "Avoid"}
        else:
            # Still score Part B for Watch companies (for full ranking picture)
            try:
                part_b = score_part_b(ticker, info, income_stmt, balance_sheet, part_a)
            except Exception as e:
                part_b = {"part_b_score": 0, "part_b_grade": "Avoid"}

        row = build_output_row(entry, info, gate_data, part_a, part_b)
        if _SCREENER_CORE_AVAILABLE:
            # Forward axis (Part 3 §13) — ENERGY PARITY via the SHARED compute: eps_trend momentum, margin
            # trajectory, forward revenue estimate, price momentum, revision-journey stage. Additive.
            compute_forward_axis(row, {**info, "eps_trend": stmts.get("eps_trend"),
                                       "growth_estimates": stmts.get("growth_estimates")},
                                 stmts.get("quarterly_income_stmt"), stmts.get("history"))
        full_rows.append(row)

    # ── Step 4: Write CSVs ────────────────────────────────────────────────────
    full_csv = os.path.join(outputs_dir, f"{run_date}_ENERGY_full_data.csv")
    gate_csv = os.path.join(outputs_dir, f"{run_date}_ENERGY_gate_results.csv")
    write_csv(full_csv, full_rows, FULL_DATA_COLUMNS)
    write_csv(gate_csv, gate_rows, GATE_COLUMNS)
    log.info(f"CSVs written: {len(full_rows)} full_data, {len(gate_rows)} gate_results")

    # ── Step 5: Coverage stats ────────────────────────────────────────────────
    strong_buys  = len([r for r in full_rows if r.get("final_status") == STATUS_STRONG_BUY])
    watch_count  = len([r for r in full_rows if r.get("final_status") == STATUS_WATCH])
    acceptable   = len([r for r in full_rows if r.get("final_status") == STATUS_ACCEPTABLE])
    not_growth   = len([r for r in full_rows if r.get("final_status") == STATUS_NOT_GROWTH])
    gate_fail    = len([r for r in full_rows if r.get("final_status") == STATUS_GATE_FAIL])
    data_issues  = len([r for r in full_rows if r.get("final_status") == STATUS_DATA_ISSUE])
    scored_count = strong_buys + watch_count + acceptable + not_growth

    run_qa.update({
        "strong_buys": strong_buys, "watch": watch_count, "acceptable": acceptable,
        "not_growth": not_growth, "gate_fail": gate_fail, "data_issues": data_issues,
        "scored": scored_count, "total": len(full_rows),
    })

    log.info(
        f"Results: {strong_buys} Strong Buy | {watch_count} Watch | "
        f"{acceptable} Acceptable | {not_growth} Not Growth | "
        f"{gate_fail} Gate Fail | {data_issues} Data Issues"
    )

    # ── Step 6: Excel ─────────────────────────────────────────────────────────
    excel_path = None
    try:
        excel_path = build_excel(full_rows, gate_rows, run_date, inv_dir)
        run_qa["excel_status"] = "BUILT"
        run_qa["excel_path"]   = excel_path or ""
    except Exception as e:
        log.error(f"Excel build failed: {e}")
        run_qa["excel_status"] = f"FAILED: {e}"

    # ── Step 7: Retrospective ─────────────────────────────────────────────────
    retro_path = None
    try:
        retro_path = write_retrospective(inv_dir, run_date, full_rows, gate_rows, fetch_errors, run_qa)
        run_qa["retrospective_status"] = "WRITTEN"
        run_qa["retrospective_path"]   = retro_path or ""
    except Exception as e:
        log.error(f"Retrospective write failed: {e}")
        run_qa["retrospective_status"] = f"FAILED: {e}"

    # ── Step 8: Write run QA CSV ──────────────────────────────────────────────
    run_qa["elapsed_seconds"] = round(time.time() - start_time, 1)
    run_qa_path = write_run_qa(outputs_dir, run_date, run_qa)

    log.info(f"=== ENERGY run complete in {run_qa['elapsed_seconds']}s ===")

    # Print summary for the scheduled task prompt to pick up
    date_display = datetime.strptime(run_date, "%Y%m%d").strftime("%d-%b-%y") if len(run_date) == 8 else run_date
    print("\n" + "="*60)
    print("ENERGY SCREENER COMPLETE")
    print(f"  Run date:      {date_display}")
    print(f"  Watchlist:     {len(full_rows)} companies")
    print(f"  Strong Buys:   {strong_buys}")
    print(f"  Watch:         {watch_count}")
    print(f"  Gate Fails:    {gate_fail}")
    print(f"  Excel:         {os.path.basename(excel_path) if excel_path else 'FAILED'}")
    print(f"  Retrospective: {os.path.basename(retro_path) if retro_path else 'FAILED'}")
    print("="*60)
    print(f"FULL_DATA_CSV: {full_csv}")
    print(f"GATE_CSV:      {gate_csv}")
    print(f"RETRO_PATH:    {retro_path or ''}")
    print(f"RUN_QA_PATH:   {run_qa_path}")
    print(f"STRONG_BUYS:   {strong_buys}")
    print(f"DATE_DISPLAY:  {date_display}")

    return {
        "full_csv":      full_csv,
        "gate_csv":      gate_csv,
        "retro_path":    retro_path,
        "run_qa_path":   run_qa_path,
        "excel_path":    excel_path,
        "strong_buys":   strong_buys,
        "date_display":  date_display,
        "run_qa":        run_qa,
    }


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 13: CLI ENTRY
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="ISA Energy Growth Stock Screener",
        epilog="Example: python3 energy_screener.py --date 2026-06-22 --outputs /tmp/out/ --inv-dir '/path/Investment Analysis/'"
    )
    parser.add_argument("--date",     required=True,
                        help="Run date YYYY-MM-DD or YYYYMMDD")
    parser.add_argument("--outputs",  required=True,
                        help="Ephemeral outputs directory path")
    parser.add_argument("--inv-dir",  required=True, dest="inv_dir",
                        help="Investment Analysis folder path")
    parser.add_argument("--preflight", action="store_true",
                        help="Local-primary preflight (yfinance/dev-shm/Yahoo). On failure prints "
                             "FALLBACK_TO_COMPOSIO and exits 3. Default off = scheduled run unchanged.")
    args = parser.parse_args()

    # Local-primary guardrail parity (opt-in). Fails over to Composio (exit 3) when the local
    # sandbox can't fetch — the same decision screener_local makes for the growth path.
    if getattr(args, "preflight", False):
        try:
            import isa_env_guard as _guard
            _guard.run_preflight_or_fallback(outputs_dir=args.outputs)
        except SystemExit:
            raise
        except Exception:
            pass

    # Normalise date
    run_date = args.date.replace("-", "")
    if len(run_date) != 8 or not run_date.isdigit():
        parser.error(f"Invalid date format: {args.date}. Use YYYY-MM-DD or YYYYMMDD.")

    run_energy(run_date, args.outputs, args.inv_dir)


if __name__ == "__main__":
    main()
