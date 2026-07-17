#!/usr/bin/env python3
"""
build_excel.py — ISA Growth Stock Analysis | Pre-built Excel builder
Version: 1.0 | 2026-05-24

Single canonical script for all 8 scheduled task groups (SP500, NASDAQ, MIDCAP400,
F250-SPI, STOXX600, OTHER). Reads scored data CSVs from outputs/ directory and builds
a consistent 7-tab workbook. Never rebuild this inline — always call this script.

USAGE (called from within analysis bash script):
    python3 build_excel.py \
        --group     NASDAQ \
        --run_date  "22-May-26" \
        --full_data /path/to/20260522_NASDAQ_full_data.csv \
        --gates     /path/to/20260522_NASDAQ_yf_gate_results.csv \
        --output    "/path/to/Investment Analysis/Growth Stock Analysis Nasdaq W-e 22-May-26.xlsx" \
        [--unresolved  /path/to/20260522_NASDAQ_unresolved_metrics.csv] \
        [--tech_fails  /path/to/20260522_NASDAQ_technical_failures.csv] \
        [--run_qa      /path/to/20260522_NASDAQ_run_qa.csv] \
        [--constituent /path/to/20260522_NASDAQ_constituent_master.csv]

EXPECTED COLUMN NAMES in full_data.csv (see FIELD_MAP below for aliases):
  Identity : ticker | company | sector | industry | index | final_status
  Scores   : part_a_score | part_b_score | total_score | part_a_status | part_b_status
  Part A   : rev_cagr | recent_rev_growth | eps_cagr | share_count_change | fcf_positive_years |
             fcf_cagr | fcf_margin | gross_margin | operating_margin | op_margin_trend |
             roic | net_debt_ebitda | interest_coverage | capex_intensity
  Pt A Scr : score_roic | score_fcf_pos | score_rev_cagr | score_recent_rev |
             score_eps_cagr | score_share_count | score_fcf_cagr | score_fcf_margin |
             score_gross_margin | score_op_margin | score_op_margin_trend |
             score_nd_ebitda | score_int_cov | score_capex
  Part B   : fwd_pe | ev_ebitda | price_fcf | fcf_yield | earnings_yield | position_52wk |
             div_payout_fcf | fwd_eps_growth | target_upside | stress_nd_ebitda | stress_int_cov |
             current_price | target_price_mean | analyst_rating | num_analysts | next_earnings |
             currency
  Pt B Scr : score_b_fwd_pe | score_b_ev_ebitda | score_b_price_fcf | score_b_fcf_yield |
             score_b_earn_yield | score_b_52wk | score_b_div_payout | score_b_fwd_eps |
             score_b_target_upside | score_b_stress |
             score_b_book_to_bill | score_b_backlog_ev  (conditional: equipment sectors only)
  Overlays : est_rev_direction |
             wacc_pct | roic_vs_wacc_spread |
             val_hist_pe_premium_disc | val_hist_pfcf_premium_disc |
             trailing_pe | val_hist_pe_status | overlay_status
  Other    : qualitative_commentary | gate_code | gate_reason | sector_bucket |
             book_to_bill_trailing_2q | backlog_ttm | backlog_ev_ratio |
             b2b_applicable | book_to_bill_status | backlog_ev_status
"""

import argparse
try:
    import isa_env_guard  # noqa  (disk guardrail: forces temp + yfinance cache onto tmpfs /dev/shm)
except Exception:
    pass
import os
import sys

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# COLOUR / STYLE CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

FILL_GREEN       = fill("E2EFDA")
FILL_AMBER       = fill("FFEB9C")
FILL_RED         = fill("FFC7CE")
FILL_BLUE        = fill("DDEBF7")
FILL_GREY        = fill("F5F5F5")
FILL_HDR_DARK    = fill("1A3A6B")
FILL_GRP1        = fill("1A3A6B")   # Identity — navy
FILL_GRP2        = fill("155724")   # Scores — dark green
FILL_GRP3        = fill("0C4A6E")   # Growth Quality — deep blue
FILL_GRP4        = fill("5C3A1E")   # Valuation & Risk — dark brown
FILL_GRP5        = fill("3A1A6B")   # Overlays — dark purple
FILL_WHITE       = fill("FFFFFF")

def font(bold=False, color="000000", size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

FONT_BOLD_WHITE  = font(bold=True,  color="FFFFFF", size=10)
FONT_BOLD        = font(bold=True,  color="000000", size=10)
FONT_NORM        = font(bold=False, color="000000", size=10)
FONT_SMALL       = font(bold=False, color="444444", size=9)

ALIGN_CTR        = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN             = Border(bottom=Side(style="thin", color="D0D0D0"))

STATUS_FILL = {
    "STRONG_BUY":                      FILL_GREEN,
    "CANDIDATE_RANKABLE":              FILL_GREEN,
    "FAIR_MIXED":                      FILL_AMBER,
    "ACCEPTABLE":                      FILL_BLUE,
    "AVOID":                           FILL_RED,
    "HARD_GATE_FAIL":                  FILL_RED,
    "MANDATORY_MINIMUM_FAIL":          FILL_RED,
    "TECHNICAL_SOURCE_FAILURE":        FILL_RED,
    "UNRESOLVED_HARD_GATE_NOT_RANKABLE": FILL_AMBER,
    "LOW_CONFIDENCE_SCORED":           FILL_AMBER,
    "PRE_SCREEN_EXCLUDED":             FILL_GREY,
    "GATE_DATA_UNRESOLVED":            FILL_AMBER,
    "STRUCTURAL_NON_APPLICABLE":       FILL_GREY,
}

# ─────────────────────────────────────────────────────────────────────────────
# FIELD NAME ALIASES  (map alternative CSV column names → canonical name)
# ─────────────────────────────────────────────────────────────────────────────

FIELD_MAP = {
    # Identity
    "symbol": "ticker", "Ticker": "ticker", "Symbol": "ticker",
    "Company": "company", "Name": "company",
    "Sector": "sector", "Industry": "industry", "Index": "index",
    "Final Status": "final_status", "Status": "final_status",
    # Scores
    "Part A": "part_a_score", "Part B": "part_b_score", "Total": "total_score",
    "PartA": "part_a_score", "PartB": "part_b_score",
    "part_a": "part_a_score", "part_b": "part_b_score",
    # Growth metrics
    "rev_cagr_3_5yr": "rev_cagr", "revenue_cagr": "rev_cagr",
    "recent_rev_growth_yoy": "recent_rev_growth",
    "eps_cagr_3_5yr": "eps_cagr", "eps_cagr_3yr": "eps_cagr",
    "share_count_change_3yr": "share_count_change",
    "fcf_positive_yrs": "fcf_positive_years",
    "fcf_cagr_3_5yr": "fcf_cagr",
    "op_margin_trend_3yr": "op_margin_trend",
    "operating_margin_trend": "op_margin_trend",
    "net_debt_ebitda": "net_debt_ebitda", "nd_ebitda": "net_debt_ebitda",
    "interest_cov": "interest_coverage", "int_coverage": "interest_coverage",
    # Valuation
    "forward_pe": "fwd_pe", "forwardPE": "fwd_pe",
    "ev_ebitda_ratio": "ev_ebitda",
    "price_fcf": "price_fcf", "p_fcf": "price_fcf",
    "fcf_yield_pct": "fcf_yield",
    "earnings_yield_pct": "earnings_yield",
    "52wk_position": "position_52wk", "wk52_position": "position_52wk",
    "div_payout_vs_fcf": "div_payout_fcf", "dividend_payout_fcf": "div_payout_fcf",
    "fwd_eps_growth_proxy": "fwd_eps_growth", "forward_eps_growth": "fwd_eps_growth",
    "upside": "target_upside", "implied_upside": "target_upside",
    "stress_net_debt_ebitda": "stress_nd_ebitda",
    "stress_interest_cov": "stress_int_cov",
    "price": "current_price", "currentPrice": "current_price",
    "target_mean": "target_price_mean", "targetMeanPrice": "target_price_mean",
    "analyst_rating_key": "analyst_rating", "recommendationKey": "analyst_rating",
    "number_of_analysts": "num_analysts", "analyst_count": "num_analysts",
    "next_earnings_date": "next_earnings",
    # Overlays
    "est_rev_dir": "est_rev_direction", "est_revision_direction": "est_rev_direction",
    "wacc": "wacc_pct", "wacc_percent": "wacc_pct",
    "roic_wacc_spread": "roic_vs_wacc_spread", "roic_minus_wacc": "roic_vs_wacc_spread",
    "val_hist_pe_prem_disc": "val_hist_pe_premium_disc",
    "val_hist_pfcf_prem_disc": "val_hist_pfcf_premium_disc",
    "trailingPE": "trailing_pe", "trailing_p_e": "trailing_pe",
    "commentary": "qualitative_commentary",
    # Gate
    "Gate": "gate_code", "Gate Code": "gate_code",
    "Reason": "gate_reason", "Exclusion Reason": "gate_reason",
}


def load_csv(path):
    if not path or not os.path.exists(str(path)):
        return pd.DataFrame()
    try:
        df = pd.read_csv(path, low_memory=False)
        df.rename(columns=FIELD_MAP, inplace=True)
        return df
    except Exception as e:
        print(f"[WARN] Could not load {path}: {e}", file=sys.stderr)
        return pd.DataFrame()


# ─────────────────────────────────────────────────────────────────────────────
# FORMATTING HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def pct(v, dec=1):
    try:
        f = float(v)
        if pd.isna(f): return "N/A"
        return f"{f*100:.{dec}f}%"
    except: return "N/A"

def pct_already(v, dec=1):
    """For values ALREADY stored as a percentage number (e.g. -78.8 -> '-78.8%').
    screener_core stores val_hist premium/discount fields x100, so they must NOT be
    multiplied by 100 again (that produced the -7880% 'thousands' display bug)."""
    try:
        f = float(v)
        if pd.isna(f): return "N/A"
        return f"{f:.{dec}f}%"
    except: return "N/A"

def mult(v, dec=1):
    try:
        f = float(v)
        if pd.isna(f): return "N/A"
        return f"{f:.{dec}f}x"
    except: return "N/A"

def num(v, dec=1, suffix=""):
    try:
        f = float(v)
        if pd.isna(f): return "N/A"
        return f"{f:.{dec}f}{suffix}"
    except: return "N/A"

def s(v, default="N/A"):
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)): return default
        return str(v).strip() or default
    except: return default

def score_int(v):
    try:
        f = float(v)
        if pd.isna(f): return "N/A"
        return int(round(f))
    except: return "N/A"

def upside_fmt(v):
    """Format upside with + or - prefix."""
    try:
        f = float(v) * 100
        if pd.isna(f): return "N/A"
        sign = "+" if f >= 0 else ""
        return f"{sign}{f:.1f}%"
    except: return "N/A"

def fmt_price(v, currency_sym=""):
    try:
        f = float(v)
        if pd.isna(f): return "N/A"
        return f"{currency_sym}{f:.2f}"
    except: return s(v)

def fcf_years_fmt(v):
    """e.g. 4.0 → '4/5', 'NET_CASH_NO_MATERIAL_INTEREST_BURDEN' → label"""
    try:
        f = float(v)
        if pd.isna(f): return "N/A"
        return f"{int(round(f))}/5"
    except: return s(v)

def pp_fmt(v):
    """Format percentage-point change e.g. 0.023 → '+2.3pp'"""
    try:
        f = float(v) * 100
        if pd.isna(f): return "N/A"
        sign = "+" if f >= 0 else ""
        return f"{sign}{f:.1f}pp"
    except: return s(v)

def trailing_pe_fmt(v):
    try:
        f = float(v)
        if pd.isna(f): return "N/A"
        return f"{f:.1f}x"
    except: return s(v)


def get_currency_sym(row):
    """Return currency symbol based on currency field or ticker suffix."""
    cur = s(row.get("currency", "N/A")).upper()
    ticker = s(row.get("ticker", ""))
    if cur == "GBP" or cur == "GBX" or cur == "GBP" or ticker.endswith(".L"):
        return "£"
    if cur == "EUR" or ticker.endswith((".DE", ".PA", ".AS", ".MI", ".MC", ".BR")):
        return "€"
    if cur == "CHF" or ticker.endswith(".SW"):
        return "CHF "
    if cur == "CAD" or ticker.endswith(".TO"):
        return "C$"
    if cur == "BRL" or ticker.endswith(".SA"):
        return "R$"
    if cur == "MXN" or ticker.endswith(".MX"):
        return "MX$"
    return "$"


# ─────────────────────────────────────────────────────────────────────────────
# WORKSHEET HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def set_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def hdr_cell(ws, row, col, text, fill_style=None, font_style=None, align_style=None):
    c = ws.cell(row=row, column=col, value=text)
    if fill_style: c.fill  = fill_style
    if font_style: c.font  = font_style
    c.alignment = align_style or ALIGN_CTR
    return c

def data_cell(ws, row, col, value, is_text=False, fill_style=None, row_fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = ALIGN_LEFT if is_text else ALIGN_CTR
    c.border = THIN
    if fill_style:  c.fill = fill_style
    elif row_fill:  c.fill = row_fill
    c.font = FONT_NORM
    return c

def merge_group_header(ws, row, start_col, end_col, text, fill_style):
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row,   end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=text)
    c.fill      = fill_style
    c.font      = FONT_BOLD_WHITE
    c.alignment = ALIGN_CTR


# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY TAB
# ─────────────────────────────────────────────────────────────────────────────

# Column definitions: (group_label, col_header, data_field, format_fn, is_text)
# group_label = "" for continuation columns within a group
SUMMARY_COLS = [
    # Group 1 — Identity (6 cols)
    ("Group 1 — Identity",        "Ticker",             "ticker",                   s,                 True),
    ("",                           "Company",            "company",                  s,                 True),
    ("",                           "Sector",             "sector",                   s,                 True),
    ("",                           "Industry",           "industry",                 s,                 True),
    ("",                           "Sector Bucket",      "sector_bucket",            s,                 True),
    ("",                           "Index",              "index",                    s,                 True),
    ("",                           "Final Status",       "final_status",             s,                 True),
    # Group 2 — Scores (3 cols)
    ("Group 2 — Scores",           "Part A (/28)",       "part_a_score",             score_int,         False),
    ("",                           "Part B (/26-30)",    "part_b_score",             score_int,         False),
    ("",                           "Total (/50-54)",     "total_score",              score_int,         False),
    ("",                           "Fwd Axis (/100)",    "forward_axis_score",       score_int,         False),
    # Fix Pack A6 (12-Jul-26): "screen_source" — ONE unified Source Score, screen = deploy.
    # Breakdown cells show "pre-weight raw → weighted contribution" for the five terms (A1 spec:
    # full anatomy on every SUMMARY row; computed once in source_score, never recomputed here).
    ("",                           "screen_source",      "source_score",             score_int,         False),
    ("",                           "Stage",              "revision_stage",           s,                 True),
    ("",                           "Fwd (raw→wtd)",      "_src_fwd",                 s,                 False),
    ("",                           "Rev (raw→wtd)",      "_src_rev",                 s,                 False),
    ("",                           "Deploy (raw→wtd)",   "_src_deploy",              s,                 False),
    ("",                           "Qual (raw→wtd)",     "_src_qual",                s,                 False),
    ("",                           "Analyst (raw→wtd)",  "_src_analyst",             s,                 False),
    # Fix Pack A2 (12-Jul-26): E[r] 12-24m block — shadow columns at P1, T1 gate input at P2.
    ("Group 2b — E[r] 12-24m",     "E[r] % pa",          "expected_return_12_24m",   lambda v: num(v, 1, "%"), False),
    ("",                           "E[r] growth",        "er_growth",                lambda v: num(v, 1, "%"), False),
    ("",                           "E[r] rerate",        "er_rerate",                lambda v: num(v, 1, "%"), False),
    ("",                           "E[r] yield",         "er_yield",                 lambda v: num(v, 1, "%"), False),
    ("",                           "E[r] conf",          "er_confidence",            lambda v: num(v, 2, ""),  False),
    # Group 3 — Growth Quality (14 cols)
    ("Group 3 — Growth Quality",   "Rev CAGR 3-5yr",     "rev_cagr",                pct,               False),
    ("",                           "Recent Rev Growth",  "recent_rev_growth",        pct,               False),
    ("",                           "EPS CAGR 3-5yr",     "eps_cagr",                pct,               False),
    ("",                           "Share Cnt Chg 3yr",  "share_count_change",       pp_fmt,            False),
    ("",                           "FCF Positive Yrs",   "fcf_positive_years",       fcf_years_fmt,     False),
    ("",                           "FCF CAGR 3-5yr",     "fcf_cagr",                pct,               False),
    ("",                           "FCF Margin",         "fcf_margin",               pct,               False),
    ("",                           "Gross Margin",       "gross_margin",             pct,               False),
    ("",                           "Op Margin",          "operating_margin",         pct,               False),
    ("",                           "Op Margin Trend",    "op_margin_trend",          pp_fmt,            False),
    ("",                           "ROIC",               "roic",                     pct,               False),
    ("",                           "Net Debt/EBITDA",    "net_debt_ebitda",          mult,              False),
    ("",                           "Interest Cov",       "interest_coverage",        mult,              False),
    ("",                           "Capex Intensity",    "capex_intensity",          pct,               False),
    # Group 4 — Valuation & Risk (16 cols: 11 Part B metrics + 5 market data)
    ("Group 4 — Valuation & Risk", "Forward P/E",        "fwd_pe",                   mult,              False),
    ("",                           "EV/EBITDA",          "ev_ebitda",                mult,              False),
    ("",                           "Price/FCF",          "price_fcf",                mult,              False),
    ("",                           "FCF Yield",          "fcf_yield",                pct,               False),
    ("",                           "Earnings Yield",     "earnings_yield",           pct,               False),
    ("",                           "Div Payout/FCF",     "div_payout_fcf",           pct,               False),
    ("",                           "Fwd EPS Growth",     "fwd_eps_growth",           pct,               False),
    # Fix Pack A6/D7: implied_upside_fv (FV-composite basis) is THE upside capital logic reads;
    # the consensus-target gap survives as DISPLAY ONLY (consensus is sentiment data).
    ("",                           "Impl Upside (FV)",   "implied_upside_fv",        upside_fmt,        False),
    ("",                           "Target Gap (display)", "target_upside",          upside_fmt,        False),
    ("",                           "Stress ND/EBITDA",   "stress_nd_ebitda",         mult,              False),
    ("",                           "Stress Int Cov",     "stress_int_cov",           mult,              False),
    ("",                           "Book-to-Bill",       "book_to_bill_trailing_2q", lambda v: num(v, 2, "x") if v not in (None, "N/A") else "N/A", False),
    ("",                           "Backlog/EV",         "backlog_ev_ratio",         lambda v: num(v, 2, "x") if v not in (None, "N/A") else "N/A", False),
    ("",                           "Current Price",      "__price__",                None,              False),  # special
    ("",                           "Target Price",       "__target__",               None,              False),  # special
    ("",                           "Analyst Rating",     "analyst_rating",           s,                 True),
    ("",                           "# Analysts",         "num_analysts",             score_int,         False),
    ("",                           "Next Earnings",      "next_earnings",            s,                 True),
    # Group 5 — High-Score Overlays (6 overlays + commentary)
    # organic_rev_growth, recurring_rev_pct, peg_3yr removed 10-Jun-26 (not yfinance-obtainable).
    # P/E & P/FCF vs 3yr avg use pct_already (values are stored x100 by screener_core).
    ("Group 5 — High-Score Overlays", "Est Rev Direction",   "est_rev_direction",        s,             True),
    ("",                               "WACC %",              "wacc_pct",                 lambda v: num(v, 1, "%"), False),
    ("",                               "ROIC vs WACC",        "roic_vs_wacc_spread",      lambda v: num(v, 1, "pp"), False),
    ("",                               "P/E vs 3yr Avg",      "val_hist_pe_premium_disc", pct_already,   False),
    ("",                               "P/FCF vs 3yr Avg",    "val_hist_pfcf_premium_disc", pct_already, False),
    ("",                               "Trailing P/E",        "trailing_pe",              trailing_pe_fmt, False),
    ("",                               "Valuation Profile",   "__valprofile__",           None,          True),
    ("",                               "Commentary",          "qualitative_commentary",   s,             True),
]

NUM_COLS = len(SUMMARY_COLS)


def build_summary(wb, df_full, run_date, group):
    ws = wb.create_sheet("SUMMARY")
    ws.freeze_panes = "A5"  # header rows 1-4 always visible

    # --- Row 1: Title ---
    try:
        import os as _os0, sys as _sys0
        _sys0.path.insert(0, _os0.path.dirname(_os0.path.abspath(__file__)))
        import scoring_config as _cfg0
        _cb0 = getattr(_cfg0, "SUMMARY_COUNT_BASED", False)
        _paf0 = getattr(_cfg0, "FORWARD_ELIG_PART_A_FLOOR", 10)
        _fl0 = getattr(_cfg0, "SUMMARY_SOURCE_FLOOR", 70.0)
        _mx0 = getattr(_cfg0, "SUMMARY_MAX_COUNT", 40)
    except Exception:
        _cb0 = False
    if _cb0:
        title = f"ISA Growth Stock Analysis — {group} | Best Opportunities (Fix Pack A1 floor-based: screen_source ≥ {_fl0:g} · cap {_mx0} · Part A viability floor {_paf0} · est-rev not deteriorating) | Run: {run_date}"
    else:
        title = f"ISA Growth Stock Analysis — {group} | Best Opportunities (Part A ≥22 · Total ≥43/50 · est-rev not deteriorating) | Run: {run_date}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    c = ws.cell(row=1, column=1, value=title)
    c.fill = FILL_HDR_DARK; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    # --- Row 2: blank ---
    ws.row_dimensions[2].height = 6

    # --- Rows 3 & 4: group headers + column headers ---
    # Track group spans
    grp_start = {}
    current_grp = None
    for col_idx, (grp, hdr, field, fmt_fn, is_text) in enumerate(SUMMARY_COLS, start=1):
        if grp:
            if current_grp and current_grp in grp_start:
                merge_group_header(ws, 3,
                                   grp_start[current_grp]["start"], col_idx - 1,
                                   current_grp, grp_start[current_grp]["fill"])
            current_grp = grp
            fill_map = {
                "Group 1 — Identity":          FILL_GRP1,
                "Group 2 — Scores":            FILL_GRP2,
                "Group 3 — Growth Quality":    FILL_GRP3,
                "Group 4 — Valuation & Risk":  FILL_GRP4,
                "Group 5 — High-Score Overlays": FILL_GRP5,
            }
            grp_start[current_grp] = {"start": col_idx, "fill": fill_map.get(grp, FILL_HDR_DARK)}

    # Close the last group
    if current_grp:
        merge_group_header(ws, 3,
                           grp_start[current_grp]["start"], NUM_COLS,
                           current_grp, grp_start[current_grp]["fill"])

    # Row 4: individual column headers
    for col_idx, (_, hdr, _, _, _) in enumerate(SUMMARY_COLS, start=1):
        c = ws.cell(row=4, column=col_idx, value=hdr)
        c.fill = FILL_HDR_DARK; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    # --- Filter Strong Buys ---
    if df_full.empty:
        ws.cell(row=5, column=1, value="No data available for this run.")
        return ws

    # SUMMARY rule (Fix Pack A1/A6, 12-Jul-26): floor-based selection via THE shared
    # source_score.select_summary (fixed top-30 retired — one quality bar per D4; cap only
    # truncates; thin-tape warning below SUMMARY_MIN_WARN). Score is the UNIFIED screen=deploy
    # source (A6). Breakdown + E[r] cells come from the single compute in source_score /
    # expected_return — deterministic parity with the values screener_core stamps into full_data.
    import os as _os, sys as _sys
    _sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
    import scoring_config as _cfg
    import source_score as _ss
    import expected_return as _er

    _sel, _sqa = _ss.select_summary(df_full.to_dict("records"))
    _rows_out = []
    for _r, _sc in _sel:
        _comp = _ss.source_score_components_for_row(_r)
        _erd = _er.expected_return_for_row(_r)
        _rows_out.append({
            **_r, **_erd,
            "source_score": _sc,
            "implied_upside_fv": _comp.get("implied_upside_fv"),
            "_src_fwd":     f"{_comp['src_fwd_raw']:g} → {_comp['src_fwd_w']:g}",
            "_src_rev":     f"{_comp['src_rev_raw']:g} → {_comp['src_rev_w']:g}",
            "_src_deploy":  f"{_comp['src_deploy_raw']:g} → {_comp['src_deploy_w']:g}",
            "_src_qual":    f"{_comp['src_qual_raw']:g} → {_comp['src_qual_w']:g}",
            "_src_analyst": f"{_comp['src_analyst_raw']:g} → {_comp['src_analyst_w']:g}",
        })
    sb = pd.DataFrame(_rows_out)

    if not sb.empty:
        sb["source_score"] = sb["source_score"].round().astype("Int64")

    if _sqa.get("summary_thin_warning"):
        _warn = (f"SUMMARY_THIN_WARNING: only {_sqa['summary_count']} rows at/above the "
                 f"screen_source floor {_sqa['summary_floor']:g} (warn <{getattr(_cfg, 'SUMMARY_MIN_WARN', 10)})")
        ws.cell(row=2, column=1, value=_warn).font = FONT_BOLD
        ws.row_dimensions[2].height = 14   # spacer row carries the warning — undo the 6px squeeze

    if sb.empty:
        ws.cell(row=5, column=1, value="No summary candidates identified this run.")
        return ws

    # --- Data rows ---
    for r_idx, row in sb.iterrows():
        excel_row = r_idx + 5
        ws.row_dimensions[excel_row].height = 40

        # Determine row fill from final_status or part_b_status
        status_key = s(row.get("final_status", row.get("part_b_status", "")))
        row_fill = STATUS_FILL.get(status_key.upper(), FILL_GREEN)

        for col_idx, (_, _, field, fmt_fn, is_text) in enumerate(SUMMARY_COLS, start=1):
            # Handle special price/target fields
            if field == "__price__":
                cur_sym = get_currency_sym(row)
                raw_v = row.get("current_price", None)
                val = fmt_price(raw_v, cur_sym)
                fmt_fn = None
            elif field == "__target__":
                cur_sym = get_currency_sym(row)
                raw_v = row.get("target_price_mean", None)
                val = fmt_price(raw_v, cur_sym)
                fmt_fn = None
            elif field == "__valprofile__":
                # Premium-growth tag: growth-adjusted valuation cluster (PEG+EVg+PFCFg, /6).
                try:
                    _cl = (float(row.get("score_b_peg") or 0) + float(row.get("score_b_ev_g") or 0)
                           + float(row.get("score_b_pfcf_g") or 0))
                except (TypeError, ValueError):
                    _cl = 0
                val = "GARP" if _cl >= 4 else "Premium Growth"
                fmt_fn = None
            else:
                raw_v = row.get(field, None)
                val = fmt_fn(raw_v) if fmt_fn else s(raw_v)

            c = ws.cell(row=excel_row, column=col_idx, value=val)
            c.alignment = ALIGN_LEFT if is_text else ALIGN_CTR
            c.fill   = row_fill
            c.border = THIN
            c.font   = FONT_NORM

        # Bold the Total Score cell — Group 2 col 3: now col 10 after Sector Bucket insertion in Group 1
        _total_col = next(
            (ci for ci, (_, _, field, _, _) in enumerate(SUMMARY_COLS, start=1) if field == "total_score"),
            10
        )
        ws.cell(row=excel_row, column=_total_col).font = Font(name="Calibri", bold=True, size=10, color="155724")

    # Auto-filter on row 4
    ws.auto_filter.ref = f"A4:{get_column_letter(NUM_COLS)}{4 + len(sb)}"

    # Column widths — field-name driven to be robust to SUMMARY_COLS insertions
    FIELD_WIDTH = {
        "ticker":                    10,
        "company":                   22,
        "sector":                    16,
        "industry":                  18,
        "sector_bucket":             18,  # new
        "index":                     10,
        "final_status":              20,
        "part_a_score":              8,
        "forward_axis_score":        9,
        "source_score":              11,
        "part_b_score":              9,
        "total_score":               9,
        "book_to_bill_trailing_2q":  11,  # new
        "backlog_ev_ratio":          11,  # new
        "qualitative_commentary":    45,
    }
    for col_idx, (_, _, field, _, _) in enumerate(SUMMARY_COLS, start=1):
        width = FIELD_WIDTH.get(field, 11)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# CANDIDATES TAB  (all gate passers ranked by total score)
# ─────────────────────────────────────────────────────────────────────────────

CAND_COLS = [
    ("Rank",            None,                   None),
    ("Ticker",          "ticker",               s),
    ("Company",         "company",              s),
    ("Sector",          "sector",               s),
    ("Industry",        "industry",             s),
    ("Index",           "index",                s),
    ("Final Status",    "final_status",         s),
    ("Part A (/28)",    "part_a_score",         score_int),
    ("Part B (/26-30)", "part_b_score",         score_int),
    ("Total (/50-54)",  "total_score",          score_int),
    ("Rev CAGR",        "rev_cagr",             pct),
    ("Gross Margin",    "gross_margin",         pct),
    ("ROIC",            "roic",                 pct),
    ("FCF Yield",       "fcf_yield",            pct),
    ("Target Upside",   "target_upside",        upside_fmt),
    ("Fwd P/E",         "fwd_pe",               mult),
    ("EV/EBITDA",       "ev_ebitda",            mult),
    ("Net Debt/EBITDA", "net_debt_ebitda",      mult),
    ("Next Earnings",   "next_earnings",        s),
]

TEXT_CAND_COLS = {1, 2, 3, 4, 5, 6, 7, 19}  # 1-indexed positions that are text/left-aligned


def build_candidates(wb, df_full, group, run_date):
    ws = wb.create_sheet("CANDIDATES")
    ws.freeze_panes = "A3"

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CAND_COLS))
    c = ws.cell(row=1, column=1, value=f"{group} — All Gate Passers Ranked by Total Score | {run_date}")
    c.fill = FILL_HDR_DARK; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    # Headers
    for ci, (hdr, _, _) in enumerate(CAND_COLS, start=1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.fill = FILL_HDR_DARK; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    if df_full.empty:
        ws.cell(row=3, column=1, value="No data.")
        return ws

    # Gate passers = stocks that were scored (not pre-screen excluded)
    excluded_statuses = {"PRE_SCREEN_EXCLUDED", "STRUCTURAL_NON_APPLICABLE"}
    gate_pass = df_full[
        ~df_full.get("final_status", pd.Series(dtype=str)).fillna("").str.upper().isin(excluded_statuses)
    ].copy()
    gate_pass = gate_pass.sort_values(
        ["total_score", "part_b_score", "part_a_score", "company"],
        ascending=[False, False, False, True]
    ).reset_index(drop=True)

    for ri, row in gate_pass.iterrows():
        er = ri + 3
        ws.row_dimensions[er].height = 18
        status_key = s(row.get("final_status", "")).upper()
        row_fill = STATUS_FILL.get(status_key, FILL_WHITE)

        for ci, (_, field, fmt_fn) in enumerate(CAND_COLS, start=1):
            if ci == 1:
                val = ri + 1
            elif field is None:
                val = "N/A"
            else:
                raw_v = row.get(field, None)
                val = fmt_fn(raw_v) if fmt_fn else s(raw_v)

            is_text = ci in TEXT_CAND_COLS
            c = ws.cell(row=er, column=ci, value=val)
            c.alignment = ALIGN_LEFT if is_text else ALIGN_CTR
            c.fill = row_fill; c.border = THIN; c.font = FONT_NORM

    ws.auto_filter.ref = f"A2:{get_column_letter(len(CAND_COLS))}{2 + len(gate_pass)}"

    widths = [6, 10, 22, 16, 18, 10, 20, 8, 8, 9, 10, 10, 10, 10, 11, 10, 10, 12, 14]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SCORES TAB  (Part A + Part B metric scores side by side)
# ─────────────────────────────────────────────────────────────────────────────

PART_A_METRIC_COLS = [
    ("ROIC (HG)",        "score_roic"),
    ("FCF Pos (HG)",     "score_fcf_pos"),
    ("Rev CAGR",         "score_rev_cagr"),
    ("Recent Rev",       "score_recent_rev"),
    ("EPS CAGR",         "score_eps_cagr"),
    ("Share Count",      "score_share_count"),
    ("FCF CAGR",         "score_fcf_cagr"),
    ("FCF Margin",       "score_fcf_margin"),
    ("Gross Margin",     "score_gross_margin"),
    ("Op Margin",        "score_op_margin"),
    ("Op Mgn Trend",     "score_op_margin_trend"),
    ("ND/EBITDA",        "score_nd_ebitda"),
    ("Int Cov",          "score_int_cov"),
    ("Capex",            "score_capex"),
    ("Part A Total",     "part_a_score"),
]

PART_B_METRIC_COLS = [
    ("ROIC (MM)",        "score_b_roic",        "score_roic"),
    ("ND/EBITDA (MM)",   "score_b_nd_ebitda",   "score_nd_ebitda"),
    ("Fwd PEG",          "score_b_peg",          None),
    ("EV/EBITDA-g",      "score_b_ev_g",         None),
    ("P/FCF-g",          "score_b_pfcf_g",       None),
    ("Int Cov (r)",      "score_b_int_cov",      "score_int_cov"),  # reuse Part A
    ("Div Payout",       "score_b_div_payout",   None),
    ("Fwd EPS Grwth",    "score_b_fwd_eps",      None),
    ("Target Upside",    "score_b_target_upside", None),
    ("Est Revision",     "score_b_est_rev",      None),
    ("Stress Test",      "score_b_stress",       None),
    ("Book/Bill*",       "score_b_book_to_bill", None),   # * = conditional: equip/hardware only
    ("Backlog/EV*",      "score_b_backlog_ev",   None),   # * = conditional: equip/hardware only
    ("Part B Total",     "part_b_score",         None),
    ("Grand Total",      "total_score",          None),
]

# Part F — Forward Axis decomposition (S5). kind: axis=/100 aggregate, score=0-2 sub-signal,
# raw=underlying driver (%, pp), text=revision stage label.
PART_F_METRIC_COLS = [
    ("Fwd Axis /100",  "forward_axis_score",   "axis"),
    ("EPS Trend",      "score_f_eps_trend",    "score"),
    ("EPS Mom %",      "eps_trend_mom_pct",    "raw"),
    ("Mgn Traj",       "score_f_margin_traj",  "score"),
    ("Mgn d-pp",       "margin_traj_delta_pp", "raw"),
    ("Rev Est",        "score_f_rev_est",      "score"),
    ("Rev Est %",      "rev_est_fwd_pct",      "raw"),
    ("Price Mom",      "score_f_price_mom",    "score"),
    ("Price 12-1m %",  "price_mom_12_1m_pct",     "raw"),
    ("Est Rev (B)",    "score_b_est_rev",      "score"),
    ("Rev Runway",     "revision_runway",      "score"),
    ("Stage",          "revision_stage",       "text"),
]


def build_scores(wb, df_full, group, run_date):
    ws = wb.create_sheet("SCORES")
    ws.freeze_panes = "D3"

    id_cols = [("Ticker", "ticker"), ("Company", "company"), ("Final Status", "final_status")]
    all_a = [(h, f) for h, f in [(h, f) for h, f in PART_A_METRIC_COLS]]
    all_b = [(h, f, alt) for h, f, alt in PART_B_METRIC_COLS]
    show_f = ("forward_axis_score" in df_full.columns) and bool(df_full["forward_axis_score"].notna().any())
    all_f = list(PART_F_METRIC_COLS) if show_f else []

    # Title
    total_cols = len(id_cols) + len(all_a) + len(all_b) + len(all_f)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value=f"{group} — Part A & Part B Metric Scores | {run_date}")
    c.fill = FILL_HDR_DARK; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    # Part A / Part B group labels (row 2)
    id_end = len(id_cols)
    a_start = id_end + 1; a_end = a_start + len(all_a) - 1
    b_start = a_end + 1; b_end = b_start + len(all_b) - 1

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=id_end)
    c = ws.cell(row=2, column=1, value="Identity"); c.fill = FILL_GRP1; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    ws.merge_cells(start_row=2, start_column=a_start, end_row=2, end_column=a_end)
    c = ws.cell(row=2, column=a_start, value="Part A — Growth Stock Classification (14 metrics, max 28)")
    c.fill = FILL_GRP3; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    ws.merge_cells(start_row=2, start_column=b_start, end_row=2, end_column=b_end)
    c = ws.cell(row=2, column=b_start,
                value="Part B — Strong Buy Assessment (13 base metrics max 26; +2 conditional for equipment/hardware = max 30 | * = N/A for non-equipment)")
    c.fill = FILL_GRP4; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    if all_f:
        f_start = b_end + 1; f_end = f_start + len(all_f) - 1
        ws.merge_cells(start_row=2, start_column=f_start, end_row=2, end_column=f_end)
        c = ws.cell(row=2, column=f_start,
                    value="Part F — Forward Axis (S5: aggregate /100 + 5 sub-signal scores 0-2 with raw drivers + revision stage)")
        c.fill = FILL_GRP5; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    # Column headers (row 3)
    all_hdrs = [h for h, _ in id_cols] + [h for h, _ in all_a] + [h for h, _, *_ in all_b]
    all_hdrs += [h for h, _, _ in all_f]
    for ci, hdr in enumerate(all_hdrs, start=1):
        c = ws.cell(row=3, column=ci, value=hdr)
        c.fill = FILL_HDR_DARK; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    if df_full.empty:
        ws.cell(row=4, column=1, value="No data.")
        return ws

    excluded_statuses = {"PRE_SCREEN_EXCLUDED", "STRUCTURAL_NON_APPLICABLE"}
    scored = df_full[
        ~df_full.get("final_status", pd.Series(dtype=str)).fillna("").str.upper().isin(excluded_statuses)
    ].sort_values(["total_score", "company"], ascending=[False, True]).reset_index(drop=True)

    for ri, row in scored.iterrows():
        er = ri + 4
        status_key = s(row.get("final_status", "")).upper()
        row_fill = STATUS_FILL.get(status_key, FILL_WHITE)
        ci = 1

        # Identity cols
        for _, field in id_cols:
            c = ws.cell(row=er, column=ci, value=s(row.get(field)))
            c.alignment = ALIGN_LEFT; c.fill = row_fill; c.border = THIN; c.font = FONT_NORM
            ci += 1

        # Part A scores
        for _, field in all_a:
            raw = row.get(field, None)
            val = score_int(raw) if field not in ("part_a_score", "part_b_score", "total_score") else score_int(raw)
            c = ws.cell(row=er, column=ci, value=val)
            c.alignment = ALIGN_CTR; c.fill = row_fill; c.border = THIN; c.font = FONT_NORM
            ci += 1

        # Part B scores (try primary field, then alt field if missing)
        for _, field, alt_field in all_b:
            raw = row.get(field, None)
            if (raw is None or (isinstance(raw, float) and pd.isna(raw))) and alt_field:
                raw = row.get(alt_field, None)
            val = score_int(raw)
            c = ws.cell(row=er, column=ci, value=val)
            c.alignment = ALIGN_CTR; c.fill = row_fill; c.border = THIN; c.font = FONT_NORM
            ci += 1

        # Part F — Forward Axis decomposition (S5)
        for _fh, _ff, _fk in all_f:
            raw = row.get(_ff, None)
            _missing = raw is None or (isinstance(raw, float) and pd.isna(raw))
            if _fk == "text":
                val = s(raw) if not _missing else "-"
            elif _fk == "raw":
                try:
                    val = round(float(raw), 1) if not _missing else None
                except (TypeError, ValueError):
                    val = None
            else:
                val = score_int(raw)
            c = ws.cell(row=er, column=ci, value=val)
            c.alignment = ALIGN_CTR; c.fill = row_fill; c.border = THIN; c.font = FONT_NORM
            ci += 1

    ws.auto_filter.ref = f"A3:{get_column_letter(total_cols)}{3 + len(scored)}"

    for ci in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 10 if ci > 3 else (10 if ci == 1 else 22 if ci == 2 else 18)
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# EXCLUSIONS TAB  (PRE-SCREEN + STRUCTURAL)
# ─────────────────────────────────────────────────────────────────────────────

def build_exclusions(wb, df_gates, group, run_date):
    ws = wb.create_sheet("EXCLUSIONS")
    ws.freeze_panes = "A3"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    c = ws.cell(row=1, column=1,
                value=f"{group} — Pre-Screen Exclusions (Gate 1 / Gate 2 / Gate 3 / Gate 4 / MktCap / Unresolved) | Gate 2 threshold sector-segmented from Jun-26 | {run_date}")
    c.fill = FILL_HDR_DARK; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    hdrs = ["Ticker", "Company", "Sector", "Industry", "Gate", "Reason", "Source"]
    for ci, h in enumerate(hdrs, start=1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = FILL_HDR_DARK; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    if df_gates.empty:
        ws.cell(row=3, column=1, value="No gate result data available.")
        return ws

    # Filter to excluded rows
    excluded = df_gates[
        df_gates.get("gate_code", pd.Series(dtype=str)).fillna("").str.upper().isin(
            {"GATE 1", "GATE 2", "GATE 3", "GATE 4", "GATE1", "GATE2", "GATE3", "GATE4",
             "MKTCAP", "GATE_DATA_UNRESOLVED", "PRE_SCREEN_EXCLUDED",
             "STRUCTURAL_NON_APPLICABLE", "STRUCTURAL"}
        )
        # Explicitly exclude 5yr override passes — these are gate passers, not exclusions
        & ~df_gates.get("gate_code", pd.Series(dtype=str)).fillna("").str.contains(
            "5yr override", case=False, na=False
        )
        | df_gates.get("final_status", pd.Series(dtype=str)).fillna("").str.upper().isin(
            {"PRE_SCREEN_EXCLUDED", "STRUCTURAL_NON_APPLICABLE", "GATE_DATA_UNRESOLVED"}
        )
    ].copy()

    if excluded.empty:
        # Try using all non-pass rows
        pass_flags = df_gates.get("gate_pass", pd.Series(dtype=str)).fillna("").astype(str).str.upper()
        excluded = df_gates[~pass_flags.isin(["TRUE", "PASS", "1", "YES"])].copy()

    # Normalise gate code display
    def norm_gate(v):
        v2 = s(v).upper().replace("_", " ").strip()
        if v2 in ("GATE 1", "GATE1"): return "Gate 1"
        if v2 in ("GATE 2", "GATE2"): return "Gate 2"
        if v2 in ("GATE 3", "GATE3"): return "Gate 3"
        if v2 in ("GATE 4", "GATE4"): return "Gate 4"
        # Gate 4 (5yr override) — PASS, not a failure.
        # Should not appear in EXCLUSIONS (gate_pass=True), but handle defensively.
        if "5YR OVERRIDE" in v2 or "5YR_OVERRIDE" in v2:
            return "Gate 4 (5yr override — PASS)"
        if v2 == "MKTCAP": return "MktCap"
        if "UNRESOLVED" in v2: return "GATE_DATA_UNRESOLVED"
        if "STRUCTURAL" in v2: return "STRUCTURAL_NON_APPLICABLE"
        return s(v)

    excluded = excluded.copy()
    excluded["_gate_display"] = excluded.get("gate_code", pd.Series(dtype=str)).fillna("").apply(norm_gate)
    excluded = excluded.sort_values(["_gate_display", "company"], na_position="last").reset_index(drop=True)

    for ri, row in excluded.iterrows():
        er = ri + 3
        ws.row_dimensions[er].height = 16
        gate_str = row.get("_gate_display", "N/A")
        rf = FILL_GREY

        data = [
            s(row.get("ticker")), s(row.get("company")), s(row.get("sector")),
            s(row.get("industry")), gate_str, s(row.get("gate_reason")), s(row.get("source", "yfinance"))
        ]
        for ci, val in enumerate(data, start=1):
            c = ws.cell(row=er, column=ci, value=val)
            c.alignment = ALIGN_LEFT; c.fill = rf; c.border = THIN; c.font = FONT_NORM

    ws.auto_filter.ref = f"A2:G{2 + len(excluded)}"
    for ci, w in enumerate([10, 26, 16, 22, 22, 45, 12], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# DATA QUALITY TAB
# ─────────────────────────────────────────────────────────────────────────────

def build_data_quality(wb, df_full, df_unresolved, group, run_date):
    ws = wb.create_sheet("DATA QUALITY")
    ws.freeze_panes = "A3"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c = ws.cell(row=1, column=1, value=f"{group} — Data Quality Issues & Unresolved Metrics | {run_date}")
    c.fill = FILL_HDR_DARK; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    hdrs = ["Ticker", "Company", "Metric", "Status", "Value / Note", "Source Attempted"]
    for ci, h in enumerate(hdrs, start=1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = FILL_HDR_DARK; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    rows_written = 0

    # From full_data: flags (sanity_elevated, LOW_CONFIDENCE etc.)
    if not df_full.empty and "data_quality_flag" in df_full.columns:
        flagged = df_full[df_full["data_quality_flag"].notna() &
                          (df_full["data_quality_flag"] != "")].copy()
        for _, row in flagged.iterrows():
            er = rows_written + 3
            data = [s(row.get("ticker")), s(row.get("company")),
                    s(row.get("dq_metric", "multiple")), s(row.get("data_quality_flag")),
                    s(row.get("dq_value", "")), s(row.get("dq_source", "yfinance"))]
            for ci, val in enumerate(data, start=1):
                c = ws.cell(row=er, column=ci, value=val)
                c.alignment = ALIGN_LEFT; c.fill = FILL_AMBER; c.border = THIN; c.font = FONT_NORM
            rows_written += 1

    # From unresolved metrics CSV
    if not df_unresolved.empty:
        for _, row in df_unresolved.iterrows():
            er = rows_written + 3
            data = [s(row.get("ticker")), s(row.get("company")),
                    s(row.get("metric")), s(row.get("status")),
                    s(row.get("value", "unresolved")), s(row.get("source", ""))]
            for ci, val in enumerate(data, start=1):
                c = ws.cell(row=er, column=ci, value=val)
                c.alignment = ALIGN_LEFT; c.fill = FILL_AMBER; c.border = THIN; c.font = FONT_NORM
            rows_written += 1

    if rows_written == 0:
        ws.cell(row=3, column=1, value="No data quality issues recorded for this run.")

    ws.auto_filter.ref = f"A2:F{max(3, 2 + rows_written)}"
    for ci, w in enumerate([10, 26, 22, 30, 40, 20], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# DIAGNOSTICS TAB
# ─────────────────────────────────────────────────────────────────────────────

def build_diagnostics(wb, df_full, df_constituent, df_run_qa, df_tech_fails, group, run_date):
    ws = wb.create_sheet("DIAGNOSTICS")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    c = ws.cell(row=1, column=1, value=f"{group} — Run Diagnostics, Coverage & Methodology | {run_date}")
    c.fill = FILL_HDR_DARK; c.font = FONT_BOLD_WHITE; c.alignment = ALIGN_CTR

    row_ptr = [3]

    def section(title):
        ws.cell(row=row_ptr[0], column=1, value=title).font = FONT_BOLD
        ws.cell(row=row_ptr[0], column=1).fill = fill("E8EFF7")
        row_ptr[0] += 1

    def kv(k, v):
        ws.cell(row=row_ptr[0], column=1, value=k).font = FONT_BOLD
        ws.cell(row=row_ptr[0], column=2, value=str(v)).font = FONT_NORM
        row_ptr[0] += 1

    # Run QA
    section("Run QA")
    if not df_run_qa.empty:
        for _, row in df_run_qa.iterrows():
            kv(s(row.get("metric", row.iloc[0] if len(row) > 0 else "")),
               s(row.get("value", row.iloc[1] if len(row) > 1 else "")))
    else:
        kv("Run QA", "Not available")
    row_ptr[0] += 1

    # Coverage summary
    section("Coverage Summary")
    if not df_full.empty:
        total = len(df_full)
        scored = df_full[~df_full.get("final_status", pd.Series(dtype=str)).fillna("").str.upper().isin(
            {"PRE_SCREEN_EXCLUDED", "STRUCTURAL_NON_APPLICABLE"}
        )]
        # Jul-26 forward-led count: SUMMARY-eligible AND Source Score >= floor (matches the SUMMARY tab).
        import source_score as _ss_cov, scoring_config as _cfg_cov
        _cov_floor = getattr(_cfg_cov, "SUMMARY_SOURCE_FLOOR", 70.0)
        _elig_cov = df_full.apply(lambda r: _ss_cov.summary_eligible(r.to_dict())
                                  and _ss_cov.source_score_for_row(r.to_dict()) >= _cov_floor, axis=1)
        strong_buys = df_full[_elig_cov]
        kv("Total constituents",   total)
        kv("Gate passers scored",  len(scored))
        kv("Summary candidates (S5: top-N by Source Score / legacy A≥22 · Total≥43)", len(strong_buys))
        kv("Pre-screen excluded",  total - len(scored))
        kv("Coverage %",           f"{len(scored)/total*100:.1f}%" if total > 0 else "N/A")
    row_ptr[0] += 1

    # Technical failures
    section("Technical Failures")
    if not df_tech_fails.empty:
        for _, row in df_tech_fails.iterrows():
            kv(s(row.iloc[0] if len(row) > 0 else ""), s(row.iloc[1] if len(row) > 1 else ""))
    else:
        kv("Technical failures", "None recorded")
    row_ptr[0] += 1

    # Constituent master
    section("Constituent Master")
    if not df_constituent.empty:
        kv("Constituent count", len(df_constituent))
        kv("Sample tickers (first 5)", ", ".join(df_constituent.get("ticker", df_constituent.iloc[:, 0]).head(5).astype(str).tolist()))
    else:
        kv("Constituent data", "Not available")
    row_ptr[0] += 1

    # Methodology
    section("Methodology Summary")
    methodology = [
        ("Scoring", "Part A: 14 metrics, max 28 pts. Strong Growth band at 22/28. Hard gates: ROIC + FCF Pos Years."),
        ("",        "Part B (v27): 11 base metrics, max 22 pts (+2 conditional = max 26 for semiconductor_hardware/equipment)."),
        ("",        "Part B metrics: ROIC, ND/EBITDA (mandatory mins), Fwd PEG, EV/EBITDA-to-growth, P/FCF-to-growth, Int Cov, Div Payout, Fwd EPS Growth, Target Upside, Est Revision, Stress (+ Book-to-Bill/Backlog conditional)."),
        ("",        "Growth-adjusted valuation uses 3-5yr EPS CAGR (fwd fallback) capped at 50%. Total base max 50."),
        ("",        "SUMMARY (forward-led) = Part A >= viability floor (FORWARD_ELIG_PART_A_FLOOR) AND Part B >= SUMMARY_PART_B_FLOOR AND est-rev not deteriorating AND revision stage has forward runway, then top-N by the single Source Score (SOURCE_WEIGHTS: forward/revisions/deployability/quality/analyst) floored at SUMMARY_SOURCE_FLOOR. Valuation Profile tag: GARP if PEG+EVg+PFCFg cluster is cheap else Premium Growth."),
        ("Gates",   "Standard: Gate 1 (sector), Gate 2 (sector-segmented GM threshold — see below), Gate 3 (FCF 3/5yr), Gate 4 (Rev CAGR>=5%)."),
        ("",        "Gate 2 sector thresholds: software_saas>=40%, semiconductor_fabless>=50%, semiconductor_hardware>=25%, semiconductor_equipment>=20%, default>=20%."),
        ("",        "Gate 4 (2C-1): semiconductor_equipment only — 5yr CAGR >=3% override if 3yr CAGR fails. Flagged as 'Gate 4 (5yr override)' in diagnostics."),
        ("",        "Nasdaq: Pre-gate MktCap>=$2bn, Gate 1 (modified), Gate 2 (sector-segmented), Gate 3, Gate 4."),
        ("Scoring v2","CapEx scoring: hardware/equipment INVERTED — 8-25% intensity = 2pts (capacity investment). Default unchanged."),
        ("",         "Op margin scoring: semiconductor_equipment strong>=12%/acceptable>=5%; semiconductor_hardware strong>=12%/acceptable>=6%."),
        ("",         "Sector bucket field in output: classifies each company as software_saas | semiconductor_fabless | semiconductor_hardware | semiconductor_equipment | default."),
        ("Overlays","7 overlays for the SUMMARY-eligible + Source>=floor set: Organic Rev, Recurring Rev, Est Revisions, ROIC vs WACC,"),
        ("",        "PEG 3yr Fwd, Valuation vs Own History (P/E + P/FCF), Trailing P/E. Max 8 min retrieval."),
        ("Sources", "yfinance primary. US fallbacks: Finnhub -> Finviz -> StockAnalysis -> GuruFocus -> Alpha Vantage."),
        ("",        "EU/CA fallbacks: Finnhub -> MarketScreener -> GuruFocus -> TradingView -> Alpha Vantage."),
        ("",        "BR/MX: yfinance -> MarketScreener -> GuruFocus -> TradingView -> Alpha Vantage."),
    ]
    for k, v in methodology:
        kv(k, v)

    for ci, w in enumerate([28, 70, 20, 20], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="ISA Growth Stock Analysis — Excel builder")
    parser.add_argument("--group",       required=True, help="GROUP label e.g. NASDAQ, SP500")
    parser.add_argument("--run_date",    required=True, help="Display date e.g. 22-May-26")
    parser.add_argument("--full_data",   required=True, help="Path to {DATE}_{GROUP}_full_data.csv")
    parser.add_argument("--gates",       required=True, help="Path to {DATE}_{GROUP}_yf_gate_results.csv")
    parser.add_argument("--output",      required=True, help="Output .xlsx path")
    parser.add_argument("--unresolved",  default="",    help="Path to unresolved_metrics.csv (optional)")
    parser.add_argument("--tech_fails",  default="",    help="Path to technical_failures.csv (optional)")
    parser.add_argument("--run_qa",      default="",    help="Path to run_qa.csv (optional)")
    parser.add_argument("--constituent", default="",    help="Path to constituent_master.csv (optional)")
    args = parser.parse_args()

    print(f"[build_excel] Loading data for {args.group} | {args.run_date} ...", flush=True)

    df_full      = load_csv(args.full_data)
    df_gates     = load_csv(args.gates)
    df_unresolved = load_csv(args.unresolved)
    df_tech_fails = load_csv(args.tech_fails)
    df_run_qa    = load_csv(args.run_qa)
    df_const     = load_csv(args.constituent)

    print(f"[build_excel] Loaded: full_data={len(df_full)} rows, gates={len(df_gates)} rows", flush=True)

    # Merge gate data into full_data if final_status not present
    if not df_full.empty and "final_status" not in df_full.columns and not df_gates.empty:
        df_full = df_full.merge(
            df_gates[["ticker", "final_status"]].drop_duplicates("ticker"),
            on="ticker", how="left"
        )

    # Merge gate data into gates for EXCLUSIONS if needed
    if not df_full.empty and not df_gates.empty:
        if "gate_code" not in df_full.columns:
            df_full = df_full.merge(
                df_gates[["ticker", "gate_code", "gate_reason"]].drop_duplicates("ticker"),
                on="ticker", how="left"
            )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    print(f"[build_excel] Building tabs ...", flush=True)
    build_summary(wb, df_full, args.run_date, args.group)
    build_candidates(wb, df_full, args.group, args.run_date)
    build_scores(wb, df_full, args.group, args.run_date)
    build_exclusions(wb, df_gates, args.group, args.run_date)
    build_data_quality(wb, df_full, df_unresolved, args.group, args.run_date)
    build_diagnostics(wb, df_full, df_const, df_run_qa, df_tech_fails, args.group, args.run_date)

    wb.save(args.output)
    print(f"[build_excel] Saved: {args.output}", flush=True)

    # ── SELF-VALIDATION GATE (anti-corruption / anti-stale-schema) ────────────
    # Reopen the file we just wrote and assert it is a complete, valid v27 workbook.
    # Catches: truncated/corrupt .xlsx (today's failure), missing tabs, empty scores,
    # and pre-v27/stale schema. Exit non-zero so a bad deliverable is never emailed.
    import sys as _sys, zipfile as _zip
    _need = {"SUMMARY", "CANDIDATES", "SCORES", "EXCLUSIONS", "DATA QUALITY", "DIAGNOSTICS"}
    try:
        _chk = openpyxl.load_workbook(args.output, read_only=True)
    except (_zip.BadZipFile, OSError) as e:
        print(f"[build_excel] VALIDATION_FAILED: saved file is not a valid .xlsx ({e}). DO NOT SEND.", flush=True); _sys.exit(2)
    _tabs = set(_chk.sheetnames)
    if not _need.issubset(_tabs):
        print(f"[build_excel] VALIDATION_FAILED: missing tabs {_need - _tabs}. DO NOT SEND.", flush=True); _sys.exit(2)
    _diag = " ".join(str(c.value) for r in _chk["DIAGNOSTICS"].iter_rows() for c in r if c.value)
    if "max 22" not in _diag:
        print("[build_excel] VALIDATION_FAILED: DIAGNOSTICS lacks v27 marker 'max 22' - output may be pre-v27/stale. DO NOT SEND.", flush=True); _sys.exit(2)
    _rows = _chk["SCORES"].max_row or 0
    _chk.close()
    if _rows < 3:
        print(f"[build_excel] VALIDATION_FAILED: SCORES has no data rows ({_rows}). DO NOT SEND.", flush=True); _sys.exit(2)
    print(f"[build_excel] VALIDATION_OK: valid v27 workbook, {len(_tabs)} tabs, SCORES rows={_rows}.", flush=True)


if __name__ == "__main__":
    main()
