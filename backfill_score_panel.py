#!/usr/bin/env python3
"""
backfill_score_panel.py — WP-B (29-Jul-2026): recover point-in-time screen scores from the
archived weekly Growth Stock Analysis workbooks into score_panel.csv.

WHY
---
score_panel.csv only began logging 01-Jul-2026, so calibration_report.py has zero matured
observations and the pre-registered rule (scoring_config, frozen 12-Jul-26) cannot be evaluated
until ~Oct-26 at the earliest. But every weekly screen since 25-Jun-2026 wrote its scores to an
xlsx that is still on disk. Those are NOT reconstructions or simulations — they are the SAME
numbers the same pipeline produced on that date. Recovering them moves the earliest observation
back and brings the first honest IC read forward by weeks.

WHAT IT IS NOT
--------------
This does not synthesise, model or re-score anything. Any metric the workbook did not record
stays NULL. It never touches weights. Rows are written through score_panel_logger.log_from_full_data
so the schema, idempotency key (run_date, group, ticker) and source_score derivation are IDENTICAL
to a live screen — there is exactly one writer.

Provenance is recorded in score_panel_backfill_manifest.json (the panel schema is unchanged, so
tests asserting PANEL_COLS keep passing).

CLI
---
  python3 backfill_score_panel.py --dir . --store score_panel.csv [--dry-run] [--shm /tmp/pylibs]
"""
from __future__ import annotations
import argparse, datetime, glob, json, os, re, sys

# SCORES-tab header -> score_panel column. Anything not listed is deliberately left NULL.
COLMAP = {
    "Ticker": "ticker",
    "Part A Total": "part_a_score",
    "Part B Total": "part_b_score",
    "Grand Total": "total_score",
    "Fwd Axis /100": "forward_axis_score",
    "EPS Trend": "score_f_eps_trend",
    "Mgn Traj": "score_f_margin_traj",
    "Rev Est": "score_f_rev_est",
    "Price Mom": "score_f_price_mom",
    "Price 12-1m %": "price_mom_12_1m_pct",
    "Est Rev (B)": "score_b_est_rev",
    "Rev Runway": "revision_runway",
    "Stage": "revision_stage",
}

# filename token -> canonical group label used by the live screens
GROUPS = [
    (re.compile(r"\bnasdaq\b", re.I), "NASDAQ"),
    (re.compile(r"\bsp500\b", re.I), "SP500"),
    (re.compile(r"midcap ?400", re.I), "MIDCAP400"),
    (re.compile(r"stoxx ?600", re.I), "STOXX600"),
    (re.compile(r"f250|spi", re.I), "F250SPI"),
    (re.compile(r"\benergy\b", re.I), "ENERGY"),
]

MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"], 1)}


def parse_filename(path):
    """-> (group, run_date_iso) or (None, None). Filename form:
    'Growth Stock Analysis <GROUP> W-e DD-MMM-YY.xlsx'."""
    base = os.path.basename(path)
    m = re.search(r"W-e\s+(\d{1,2})-([A-Za-z]{3})-(\d{2})", base)
    if not m:
        return None, None
    dd, mon, yy = int(m.group(1)), m.group(2).lower(), int(m.group(3))
    if mon not in MONTHS:
        return None, None
    try:
        d = datetime.date(2000 + yy, MONTHS[mon], dd)
    except ValueError:
        return None, None
    grp = None
    head = base.split("W-e")[0]
    for rx, label in GROUPS:
        if rx.search(head):
            grp = label
            break
    return grp, d.isoformat()


def read_scores_tab(path):
    """-> list[dict] using score_panel column names. Header row is located by content, not index."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "SCORES" not in wb.sheetnames:
        return []
    ws = wb["SCORES"]
    hdr_i, hdr = None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if "Ticker" in vals and ("Part A Total" in vals or "Fwd Axis /100" in vals):
            hdr_i, hdr = i + 1, vals
            break
    if hdr is None:
        return []
    idx = {}
    for j, name in enumerate(hdr):
        if name in COLMAP:
            idx.setdefault(COLMAP[name], j)          # first occurrence wins
    if "ticker" not in idx:
        return []
    out = []
    for row in ws.iter_rows(min_row=hdr_i + 1, values_only=True):
        tk = row[idx["ticker"]] if idx["ticker"] < len(row) else None
        if tk is None or not str(tk).strip():
            continue
        tk = str(tk).strip()
        if tk.lower() in ("ticker", "none", "nan", "total"):
            continue
        rec = {"ticker": tk}
        for col, j in idx.items():
            if col == "ticker" or j >= len(row):
                continue
            v = row[j]
            rec[col] = None if (v is None or str(v).strip() in ("", "-", "N/A", "NA")) else v
        out.append(rec)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", default=".")
    ap.add_argument("--store", default="score_panel.csv")
    ap.add_argument("--manifest", default="score_panel_backfill_manifest.json")
    ap.add_argument("--pattern", default="Growth Stock Analysis*.xlsx")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--shm", default=None)
    a = ap.parse_args()
    if a.shm and os.path.isdir(a.shm):
        sys.path.insert(0, a.shm)
    import pandas as pd
    sys.path.insert(0, os.path.dirname(os.path.abspath(a.store)) or ".")
    import score_panel_logger as spl

    files = sorted(glob.glob(os.path.join(a.dir, a.pattern)))
    manifest = {"generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
                "store": os.path.basename(a.store), "dry_run": bool(a.dry_run), "files": []}
    total_in = 0
    store_total = None
    for f in files:
        grp, rd = parse_filename(f)
        entry = {"file": os.path.basename(f), "group": grp, "run_date": rd}
        if not grp or not rd:
            entry["status"] = "SKIP_UNPARSEABLE_NAME"
            manifest["files"].append(entry); print("SKIP (name) %s" % entry["file"]); continue
        try:
            recs = read_scores_tab(f)
        except Exception as e:
            entry["status"] = "ERROR: %s" % str(e)[:120]
            manifest["files"].append(entry); print("ERROR %s :: %s" % (entry["file"], e)); continue
        if not recs:
            entry["status"] = "SKIP_NO_SCORES_TAB"
            manifest["files"].append(entry); print("SKIP (no SCORES) %s" % entry["file"]); continue
        df = pd.DataFrame(recs)
        entry["rows"] = len(df)
        entry["cols_recovered"] = sorted(c for c in df.columns if c != "ticker")
        total_in += len(df)
        if a.dry_run:
            entry["status"] = "DRY_RUN"
        else:
            n_new, store_total = spl.log_from_full_data(df, grp, rd, a.store)
            entry["status"] = "LOGGED"; entry["rows_in"] = n_new; entry["store_total"] = store_total
        manifest["files"].append(entry)
        print("%-9s %s %-10s rows=%d" % (entry["status"], rd, grp, len(df)))
    manifest["summary"] = {"files": len(files), "rows_in": total_in, "store_total": store_total}
    if not a.dry_run:
        json.dump(manifest, open(a.manifest, "w", encoding="utf-8"), indent=2)
    print("\nBACKFILL %s files=%d rows_in=%d store_total=%s"
          % ("DRY-RUN" if a.dry_run else "COMPLETE", len(files), total_in, store_total))
    return 0


if __name__ == "__main__":
    sys.exit(main())
