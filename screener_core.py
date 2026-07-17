"""
screener_core.py — ISA Growth Stock Screener (canonical build, 2026-05-30)
Covers: SP500 | NASDAQ | MIDCAP400 | F250-SPI | STOXX600 | OTHER
Usage:
  python screener_core.py --group SP500 --date 2026-05-30
  python screener_core.py --group NASDAQ --date 2026-05-30
  python screener_core.py --mode intramonth --ticker AAPL,MSFT --date 2026-05-30
DO NOT MODIFY VCI workflow files.
"""

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1: IMPORTS & CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
import os, sys, re, time, json, math, logging, argparse, traceback
from io import BytesIO
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

# HARD disk guardrail (shared): keep pip temp + the yfinance cache OFF the tiny HOME fs (/sessions, ~12 MB on
# local; harmless on Composio where /dev/shm also exists). Set BEFORE yfinance is used below.
import os as _os, tempfile as _tf
for _d in ("/dev/shm/piptmp", "/dev/shm/yf_cache"):
    try: _os.makedirs(_d, exist_ok=True)
    except Exception: pass
if _os.path.isdir("/dev/shm"):
    _os.environ.setdefault("TMPDIR", "/dev/shm/piptmp"); _tf.tempdir = "/dev/shm/piptmp"
from pathlib import Path

import requests
import pandas as pd
import numpy as np
import yfinance as yf
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import scoring_config as _cfg  # SINGLE SOURCE OF TRUTH for scoring thresholds (path scorers + pre-run adapter import this)
try:
    yf.set_tz_cache_location("/dev/shm/yf_cache") if _os.path.isdir("/dev/shm") else None
except Exception:
    pass
try:
    import openpyxl
except ImportError:
    openpyxl = None

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("screener_core")

# ── Risk-free rates by market suffix (update quarterly) ──────────────────────
RFR = {
    "":    0.043,   # US (no suffix)
    ".L":  0.041,   # UK LSE
    ".DE": 0.038,   # Germany XETRA
    ".PA": 0.038,   # France Euronext
    ".AS": 0.038,   # Netherlands
    ".MI": 0.038,   # Italy
    ".MC": 0.038,   # Spain
    ".SW": 0.009,   # Switzerland SIX
    ".TO": 0.035,   # Canada TSX
    ".SA": 0.125,   # Brazil B3
    ".MX": 0.095,   # Mexico BMV
}
ERP = 0.055  # Equity Risk Premium — Damodaran base rate

# ── STOXX 600 country→exchange mapping for yfinance.Lookup ISIN resolution ───
COUNTRY_EXCHANGES = {
    "United Kingdom": ["LSE"], "Germany": ["GER"], "France": ["PAR", "ENX"],
    "Switzerland": ["EBS"], "Netherlands": ["AMS"], "Sweden": ["STO"],
    "Spain": ["MAD", "MCE"], "Italy": ["MIL"], "Denmark": ["CPH"],
    "Finland": ["HEL"], "Norway": ["OSL"], "Belgium": ["BRU"],
    "Austria": ["VIE"], "Portugal": ["LIS"], "Ireland": ["ISE"],
    "Poland": ["WSE"], "Luxembourg": ["PAR", "AMS"],
}

# ── Sector / industry exclusion lists ────────────────────────────────────────
GATE1_EXCLUDE_SECTORS = {"Real Estate"}
GATE1_EXCLUDE_KEYWORDS = [
    "bank", "insur", "asset manag", "investment trust", "fund", "reit",
    "mining", "metals", "gold", "silver", "copper", "royalt", "streaming vehicle",
    "spac", "blank check", "holding compan",
]
GATE1_FINTECH_KEEP_KEYWORDS = [
    "credit services", "capital markets", "financial data", "stock exchange",
    "payment", "fintech",
]
# Investment holding companies — structural exclusion
STRUCTURAL_EXCLUSIONS = {"BION.SW"}

# ── Index constituent source bounds ──────────────────────────────────────────
CONSTITUENT_BOUNDS = {
    "SP500":      (490, 515),
    "NASDAQ":     (750, 1300),
    "MIDCAP400":  (390, 415),
    "FTSE250":    (240, 270),
    "STOXX600":   (575, 625),
    "SPI":        (175, 220),
    "TSX":        (200, 250),
    "IBRX50":     (48, 52),
    "IPC35":      (33, 37),
}

# ── Group → index list mapping ────────────────────────────────────────────────
GROUP_INDICES = {
    "SP500":     ["SP500"],
    "NASDAQ":    ["NASDAQ"],
    "MIDCAP400": ["MIDCAP400"],
    "F250-SPI":  ["FTSE250", "SPI"],
    "STOXX600":  ["STOXX600"],
    "OTHER":     ["TSX", "IBRX50", "IPC35"],
}

# ── Batch parameters by group ─────────────────────────────────────────────────
BATCH_PARAMS = {
    "SP500":     {"chunk": 75, "workers": 12, "cooldown": 50, "idx_pause": 105},
    "NASDAQ":    {"chunk": 75, "workers": 12, "cooldown": 50, "idx_pause": 105},
    "MIDCAP400": {"chunk": 60, "workers": 10, "cooldown": 60, "idx_pause": 0},
    "F250-SPI":  {"chunk": 55, "workers": 8,  "cooldown": 75, "idx_pause": 150},
    "STOXX600":  {"chunk": 55, "workers": 8,  "cooldown": 75, "idx_pause": 150},
    "OTHER":     {"chunk": 50, "workers": 6,  "cooldown": 90, "idx_pause": 150},
}

# ── Gross margin gate thresholds — sector-segmented (Enhancement 2B) ─────────
# Replaces GATE2_GM_STANDARD and GATE2_GM_NASDAQ. Threshold driven by sector
# classification (classify_sector_bucket), not by exchange or nasdaq_mode flag.
GATE2_GM_THRESHOLD = {
    "software_saas":           0.40,   # SaaS/platform — retains existing Nasdaq gate
    "semiconductor_fabless":   0.50,   # Fabless designers (NVDA, AMD, QCOM) have structurally high margins
    "semiconductor_hardware":  0.25,   # IDMs (MU, TXN) + optical hardware (LITE, AAOI, STM)
    "semiconductor_equipment": 0.20,   # Fluid delivery, probe cards, test systems (ICHR, AEHR, UCTT, FORM)
    "default":                 0.20,   # All other sectors — matches former GATE2_GM_STANDARD exactly
}

# ── Gross margin Part A scoring thresholds — sector-segmented (Enhancement 2B) ─
# Tuple: (strong_threshold, acceptable_threshold)
GROSS_MARGIN_SCORE_THRESHOLDS = {
    "software_saas":           (0.70, 0.55),   # SaaS: strong ≥70%, acceptable ≥55%
    "semiconductor_fabless":   (0.60, 0.50),   # Fabless: NVDA/AVGO-type margins
    "semiconductor_hardware":  (0.40, 0.25),   # IDMs: MU at 39.8% → 1pt (correct)
    "semiconductor_equipment": (0.30, 0.20),   # Equipment: ICHR/UCTT at 15–20% → 1pt
    "default":                 (0.30, 0.20),   # All others — aligns existing non-Nasdaq scoring
}

# ── CapEx scoring mode by sector bucket (Enhancement 2C-2) ───────────────────
# "invert": high capex = capacity investment signal (hardware)
# "standard": low capex = asset-light quality signal (SaaS/default)
CAPEX_SCORE_MODE = {
    "semiconductor_hardware":  "invert",
    "semiconductor_equipment": "invert",
    "software_saas":           "standard",
    "semiconductor_fabless":   "standard",
    "default":                 "standard",
}

# ── Operating margin scoring thresholds by sector bucket (Enhancement 2C-3) ───
# Tuple: (strong_threshold, acceptable_threshold)
OP_MARGIN_SCORE_THRESHOLDS = {
    "semiconductor_equipment": (0.12, 0.05),   # Equipment at cycle trough: trough margins 3–8%
    "semiconductor_hardware":  (0.12, 0.06),   # IDMs at trough: MU/TXN compressed margins
    "software_saas":           (0.15, 0.08),   # Unchanged — existing universal thresholds
    "semiconductor_fabless":   (0.15, 0.08),   # Unchanged
    "default":                 (0.15, 0.08),   # Unchanged
}

# ── Book-to-bill and Backlog/EV scoring thresholds (Enhancement 2) ───────────
# Conditional Part B metrics — only scored for equipment/project-based sectors.
# In automated screening runs, default to unresolved (0pts) — no yfinance source.
# Applicable sector buckets: semiconductor_equipment + semiconductor_hardware
B2B_APPLICABLE_BUCKETS = {"semiconductor_equipment", "semiconductor_hardware"}

BOOK_TO_BILL_SCORE_THRESHOLDS = (1.20, 1.00)   # (strong >= 1.20, acceptable >= 1.00)
BACKLOG_EV_SCORE_THRESHOLDS   = (2.00, 1.00)   # (strong >= 2.00, acceptable >= 1.00)

# ── Scoring thresholds (do not change without checklist update) ───────────────
PART_A_STRONG_THRESHOLD   = _cfg.GROWTH_PART_A_STRONG       # 22  (scoring_config — single source of truth)
PART_A_ACCEPTABLE_MIN     = _cfg.GROWTH_PART_A_ACCEPTABLE   # 14
PART_B_STRONG_THRESHOLD   = _cfg.GROWTH_PART_B_STRONG       # 16  (v27-recalibrated; was hardcoded 19 on the old /26 scale)
PART_B_ACCEPTABLE_MIN     = _cfg.GROWTH_PART_B_ACCEPTABLE   # 11
# OVERLAY_SCORE_TRIGGER retired (Jul-26 Part 5): overlay set is now SUMMARY-eligible + Source>=floor
OVERLAY_TIME_CAP_SECS     = 480  # 8 minutes

# ── Revenue anomaly multiple ──────────────────────────────────────────────────
REV_ANOMALY_MULTIPLE = 3.0

# ── Alpha Vantage daily cap ───────────────────────────────────────────────────
AV_DAILY_CAP = 25
AV_PER_TASK_CAP = 10


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2: UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def safe_float(v, default=None):
    """Convert value to float, returning default on failure."""
    if v is None:
        return default
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f
    except (TypeError, ValueError):
        return default


# Semiconductor fabless-vs-IDM anchors (deterministic override for bellwethers).
# The capex-intensity heuristic below handles the long tail / non-US names.
IDM_ANCHORS     = {"MU", "TXN", "INTC", "STM", "ON", "ADI", "NXPI", "MCHP", "IFX.DE"}
FABLESS_ANCHORS = {"NVDA", "AMD", "QCOM", "AVGO", "MRVL", "MPWR", "LSCC", "QRVO", "SWKS", "ALGM"}


def compute_capex_intensity(income_stmt, cashflow):
    """abs(CapEx) / Revenue from the latest annual period. Structural fabless-vs-IDM
    discriminator (fabless are asset-light ~1-5%; IDMs own fabs ~15-40%). Returns
    None if either input is missing."""
    try:
        cap = get_stmt_value(cashflow, [
            "Capital Expenditure", "Capital Expenditures", "CapitalExpenditure",
            "Purchase Of PPE", "Purchase of Property Plant and Equipment",
            "Investments in Property Plant and Equipment"])
        rev = get_stmt_value(income_stmt, [
            "Total Revenue", "Operating Revenue", "Revenue", "TotalRevenue"])
        if cap is not None and rev and rev > 0:
            return abs(cap) / rev
    except Exception:
        pass
    return None


def classify_sector_bucket(sector: str, industry: str, gross_margin: float = None,
                           capex_intensity: float = None, ticker: str = None) -> str:
    """
    Map yfinance sector/industry strings to a gross margin sector bucket.

    Used by gate2_pass() (hard gate threshold) and score_part_a() (segmented scoring).

    Fabless vs IDM/hardware split (both return industry="Semiconductors" in yfinance):
    resolved by (1) explicit anchor lists, then (2) CAPEX INTENSITY — the structural
    fab-ownership signal — NOT gross margin. The old GM>=0.55 discriminator was cyclical
    and mis-flagged IDMs like Micron as fabless in up-cycles (see _FIX_CHECKPOINT_partb).
    `gross_margin` is retained in the signature for backward compatibility but no longer
    drives the split.

    Returns one of:
      "software_saas" | "semiconductor_fabless" | "semiconductor_hardware"
      | "semiconductor_equipment" | "default"
    """
    s = (sector or "").lower()
    i = (industry or "").lower()
    si = s + " " + i   # combined string for keyword checks

    # ── Software / SaaS / Platform ───────────────────────────────────────────
    if any(kw in si for kw in [
            "software", "internet content", "data processing",
            "application software", "enterprise software",
            "infrastructure software", "system software"]):
        return "software_saas"

    # ── Semiconductor equipment and materials ─────────────────────────────────
    # Must check before generic "semiconductor" to avoid misclassifying AEHR/ICHR
    if "semiconductor" in si and any(kw in i for kw in ["equipment", "materials"]):
        return "semiconductor_equipment"

    # Electronic components, scientific instruments, industrial machinery
    # Catches ICHR (Electronic Components), AEHR (Scientific Instruments),
    # UCTT (Electronic Components), FORM (Scientific Instruments)
    if any(kw in i for kw in [
            "electronic components", "scientific instruments",
            "industrial machinery", "electronic instruments"]):
        return "semiconductor_equipment"

    # ── Semiconductor — fabless vs IDM/hardware split (anchors → capex) ───────
    if "semiconductor" in si:
        t = (ticker or "").upper()
        if t in IDM_ANCHORS:
            return "semiconductor_hardware"
        if t in FABLESS_ANCHORS:
            return "semiconductor_fabless"
        if capex_intensity is not None:
            # Fabless are asset-light (<8% capex/rev); IDMs own fabs (>=8%).
            return "semiconductor_fabless" if capex_intensity < 0.08 else "semiconductor_hardware"
        # No capex available and not anchored → conservative default of hardware
        # (lower 25% Gate 2 bar) so cyclical IDMs are never gated out on a down-cycle margin.
        return "semiconductor_hardware"

    # ── Default — all other sectors ───────────────────────────────────────────
    return "default"


def get_row(stmt, *keys):
    """Try each key in order against a statement dict/Series; return first non-None value."""
    if stmt is None:
        return None
    for k in keys:
        try:
            v = stmt[k] if isinstance(stmt, dict) else stmt.get(k)
            if v is not None:
                f = safe_float(v)
                if f is not None:
                    return f
        except (KeyError, TypeError):
            pass
    return None


def get_stmt_value(stmt_df, row_keys, col_idx=0):
    """
    Extract a value from a yfinance statement DataFrame.
    stmt_df: DataFrame with dates as columns, metrics as index.
    row_keys: list of label strings to try in order.
    col_idx: column index (0 = most recent).
    """
    if stmt_df is None or stmt_df.empty:
        return None
    cols = stmt_df.columns
    if col_idx >= len(cols):
        return None
    col = cols[col_idx]
    for k in row_keys:
        if k in stmt_df.index:
            v = stmt_df.at[k, col]
            f = safe_float(v)
            if f is not None:
                return f
    return None


def get_stmt_series(stmt_df, row_keys, max_periods=5):
    """
    Return a list of (year, value) tuples for up to max_periods annual columns.
    Only includes columns where value is non-null.
    """
    if stmt_df is None or stmt_df.empty:
        return []
    results = []
    for col_idx, col in enumerate(stmt_df.columns[:max_periods]):
        year = col.year if hasattr(col, "year") else int(str(col)[:4])
        for k in row_keys:
            if k in stmt_df.index:
                v = stmt_df.at[k, col]
                f = safe_float(v)
                if f is not None:
                    results.append((year, f))
                    break
    return results  # descending date order (most recent first)


def pence_divisor(info):
    """Return 100 if stock is priced in GBp (pence), else 1."""
    currency = (info.get("currency") or "").strip()
    price = safe_float(info.get("currentPrice") or info.get("regularMarketPrice"))
    if currency == "GBp":
        return 100
    if currency == "GBP" and price is not None and price > 500:
        return 100
    return 1


def apply_pence_correction(info):
    """
    Return a dict of pence-corrected price fields.
    MUST be called at ingestion time before any downstream calculation.
    """
    div = pence_divisor(info)
    return {
        "current_price":    safe_float(info.get("currentPrice") or info.get("regularMarketPrice"), 0) / div,
        "target_mean":      (safe_float(info.get("targetMeanPrice")) or 0) / div,
        "target_high":      (safe_float(info.get("targetHighPrice")) or 0) / div,
        "target_low":       (safe_float(info.get("targetLowPrice")) or 0) / div,
        "low_52wk":         (safe_float(info.get("fiftyTwoWeekLow")) or 0) / div,
        "high_52wk":        (safe_float(info.get("fiftyTwoWeekHigh")) or 0) / div,
        "pence_div":        div,
    }


def compute_cagr(start_val, end_val, years):
    """Compute CAGR. Returns None if inputs invalid or CAGR would be misleading."""
    if start_val is None or end_val is None or years is None or years <= 0:
        return None
    if start_val <= 0 or end_val <= 0:
        return None
    try:
        return (end_val / start_val) ** (1.0 / years) - 1
    except (ZeroDivisionError, ValueError, OverflowError):
        return None


def detect_rev_anomaly(rev_series):
    """
    rev_series: list of (year, revenue) tuples, descending order (newest first).
    Returns (clean_series, anomaly_detected, excluded_year).
    """
    if len(rev_series) < 3:
        return rev_series, False, None
    vals = [v for _, v in rev_series]
    for i in range(1, len(vals) - 1):
        if vals[i] > REV_ANOMALY_MULTIPLE * vals[i - 1] and vals[i] > REV_ANOMALY_MULTIPLE * vals[i + 1]:
            year = rev_series[i][0]
            clean = [r for r in rev_series if r[0] != year]
            return clean, True, year
    # Boundary anomaly: oldest value > 3x next
    if len(vals) >= 2 and vals[-1] > REV_ANOMALY_MULTIPLE * vals[-2]:
        clean = rev_series[:-1]
        return clean, True, rev_series[-1][0]
    return rev_series, False, None


def retry_fetch(fn, ticker_sym, retries=2, delay=5):
    """Retry a fetch function on exception."""
    for attempt in range(retries + 1):
        try:
            return fn()
        except Exception as e:
            if attempt == retries:
                log.warning(f"fetch failed {ticker_sym} after {retries+1} attempts: {e}")
                return None
            time.sleep(delay)


def get_market_suffix(ticker_sym):
    """Return the exchange suffix portion of a ticker (e.g. '.L', '.SW', '' for US)."""
    m = re.search(r'(\.[A-Z]+)$', ticker_sym)
    return m.group(1) if m else ""


def rfr_for_ticker(ticker_sym):
    """Return risk-free rate for ticker's market."""
    suffix = get_market_suffix(ticker_sym)
    return RFR.get(suffix, RFR[""])


def is_us_ticker(ticker_sym):
    return get_market_suffix(ticker_sym) == ""


def geography_group(ticker_sym):
    """Returns 'US', 'EU_CA', or 'OTHER' for fallback waterfall selection."""
    suffix = get_market_suffix(ticker_sym)
    if suffix == "":
        return "US"
    if suffix in (".L", ".DE", ".PA", ".AS", ".MI", ".MC", ".SW", ".TO"):
        return "EU_CA"
    return "OTHER"


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3: CONSTITUENT FETCHERS
# ─────────────────────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json, text/plain, */*",
}

def _assert_count(rows, index_key, source_label, allow_retry=True):
    """Assert row count is within bounds. Returns (rows, warning_str)."""
    lo, hi = CONSTITUENT_BOUNDS[index_key]
    n = len(rows)
    if lo <= n <= hi:
        return rows, None
    warn = f"COUNT_RECONCILIATION_WARNING: {index_key} from {source_label}: {n} rows (expected {lo}–{hi})"
    log.warning(warn)
    return rows, warn


def fetch_sp500():
    """Fetch S&P 500 constituents. Primary: SPY SSGA XLSX. Fallback: S&P DJI XLS."""
    log.info("Fetching S&P 500 constituents...")
    try:
        url = "https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-spy.xlsx"
        r = requests.get(url, headers=HEADERS, timeout=45)
        r.raise_for_status()
        df = pd.read_excel(BytesIO(r.content), header=None)
        # Find header row
        header_row = None
        for i, row in df.iterrows():
            if any(str(c).strip().lower() in ("ticker", "symbol") for c in row):
                header_row = i
                break
        if header_row is None:
            raise ValueError("SPY XLSX: header row not found")
        df.columns = [str(c).strip() for c in df.iloc[header_row]]
        df = df.iloc[header_row + 1:].copy()
        # Find ticker and name columns
        ticker_col = next((c for c in df.columns if c.lower() in ("ticker", "symbol")), None)
        name_col = next((c for c in df.columns if "name" in c.lower()), None)
        weight_col = next((c for c in df.columns if "weight" in c.lower()), None)
        if not ticker_col:
            raise ValueError("SPY XLSX: no ticker column")
        df = df.rename(columns={ticker_col: "ticker"})
        if name_col:
            df = df.rename(columns={name_col: "company"})
        # Filter: valid ticker (alpha + . and -), exclude cash/residual
        df["ticker"] = df["ticker"].astype(str).str.strip()
        df = df[df["ticker"].apply(lambda t: bool(re.match(r'^[A-Za-z][A-Za-z0-9.\-]*$', t)))]
        if weight_col:
            df[weight_col] = pd.to_numeric(df[weight_col], errors="coerce")
            df = df[df[weight_col].notna() & (df[weight_col] > 0)]
        df["index"] = "SP500"
        rows, warn = _assert_count(df, "SP500", "SPY_SSGA_XLSX")
        return rows[["ticker", "company", "index"]].reset_index(drop=True) if "company" in rows.columns else rows[["ticker", "index"]].reset_index(drop=True), warn
    except Exception as e:
        log.warning(f"SPY SSGA XLSX failed: {e} — trying S&P DJI fallback")
    # Fallback: S&P DJI (URL often 403 from sandbox — kept as structural fallback)
    try:
        url = "https://www.spglobal.com/spdji/en/idsexport/file.xls?redesignExport=true&indexId=340"
        r = requests.get(url, headers=HEADERS, timeout=45)
        r.raise_for_status()
        df = pd.read_excel(BytesIO(r.content), skiprows=9)
        df.columns = [str(c).strip() for c in df.columns]
        ticker_col = next((c for c in df.columns if "ticker" in c.lower()), None)
        if not ticker_col:
            raise ValueError("DJI XLS: no ticker column")
        df = df.rename(columns={ticker_col: "ticker"})
        df["ticker"] = df["ticker"].astype(str).str.strip()
        df = df[df["ticker"].apply(lambda t: bool(re.match(r'^[A-Za-z]', t)))]
        df["index"] = "SP500"
        rows, warn = _assert_count(df, "SP500", "SPDJI_XLS_FALLBACK")
        return rows[["ticker", "index"]].reset_index(drop=True), warn
    except Exception as e:
        log.error(f"S&P 500 fallback also failed: {e}")
        return pd.DataFrame(columns=["ticker", "index"]), "CONSTITUENT_SOURCE_FAILURE:SP500"


def fetch_nasdaq():
    """
    Fetch Nasdaq constituents via Nasdaq Screener API.
    $2bn market cap filter applied at sourcing time.
    """
    log.info("Fetching Nasdaq constituents (Screener API, $2bn filter)...")
    url = "https://api.nasdaq.com/api/screener/stocks?tableonly=true&limit=5000&exchange=nasdaq&download=true"
    headers = {**HEADERS, "Referer": "https://www.nasdaq.com/"}
    try:
        r = requests.get(url, headers=headers, timeout=45)
        r.raise_for_status()
        data = r.json()
        rows = data.get("data", {}).get("rows", [])
        if not rows:
            raise ValueError("Nasdaq API: empty rows")
        records = []
        for row in rows:
            sym = str(row.get("symbol", "")).strip()
            if not sym or not re.match(r'^[A-Za-z]', sym):
                continue
            mc_raw = str(row.get("marketCap", "")).replace(",", "").strip()
            try:
                mc = float(mc_raw)
            except (ValueError, TypeError):
                continue
            if mc < 2_000_000_000:
                continue
            records.append({
                "ticker": sym,
                "company": str(row.get("name", "")).strip(),
                "sector": str(row.get("sector", "")).strip(),
                "index": "NASDAQ",
            })
        df = pd.DataFrame(records)
        rows_df, warn = _assert_count(df, "NASDAQ", "NASDAQ_SCREENER_API")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.warning(f"Nasdaq Screener primary failed: {e} — trying fallback URL")
    # Fallback: omit download=true param
    try:
        url2 = "https://api.nasdaq.com/api/screener/stocks?tableonly=true&limit=5000&exchange=nasdaq"
        r = requests.get(url2, headers=headers, timeout=45)
        r.raise_for_status()
        data = r.json()
        rows = data.get("data", {}).get("table", {}).get("rows", [])
        records = []
        for row in rows:
            sym = str(row.get("symbol", "")).strip()
            if not sym or not re.match(r'^[A-Za-z]', sym):
                continue
            mc_raw = str(row.get("marketCap", "")).replace(",", "").strip()
            try:
                mc = float(mc_raw)
            except (ValueError, TypeError):
                continue
            if mc < 2_000_000_000:
                continue
            records.append({"ticker": sym, "company": str(row.get("name", "")).strip(), "sector": str(row.get("sector", "")).strip(), "index": "NASDAQ"})
        df = pd.DataFrame(records)
        rows_df, warn = _assert_count(df, "NASDAQ", "NASDAQ_SCREENER_API_FALLBACK")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.error(f"Nasdaq constituent fetch failed entirely: {e}")
        return pd.DataFrame(columns=["ticker", "company", "sector", "index"]), "CONSTITUENT_SOURCE_FAILURE:NASDAQ"


def fetch_midcap400():
    """Fetch S&P MidCap 400. Primary: IJH iShares CSV. Fallback: S&P DJI XLS."""
    log.info("Fetching S&P MidCap 400 constituents...")
    # PRIMARY (local-friendly): SSGA SPDR MDY xlsx — same SSGA infra as SPY, not IP-blocked on local sandbox.
    # (iShares IJH serves an HTML anti-bot page to the local sandbox; S&P DJI export 403s — both kept as fallbacks.)
    try:
        url = "https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-mdy.xlsx"
        r = requests.get(url, headers=HEADERS, timeout=45)
        r.raise_for_status()
        df = pd.read_excel(BytesIO(r.content), skiprows=4)
        df.columns = [str(c).strip() for c in df.columns]
        tcol = next((c for c in df.columns if c.lower() == "ticker"), None)
        wcol = next((c for c in df.columns if c.lower() == "weight"), None)
        ncol = next((c for c in df.columns if c.lower() == "name"), None)
        if not tcol:
            raise ValueError("SSGA MDY: no Ticker column")
        if wcol:
            df = df[pd.to_numeric(df[wcol], errors="coerce").notna()]
        df = df.rename(columns={tcol: "ticker"})
        df["ticker"] = df["ticker"].astype(str).str.strip()
        df = df[df["ticker"].apply(lambda t: bool(re.match(r'^[A-Za-z]', t)))]
        if ncol:
            df = df.rename(columns={ncol: "company"})
        df = df[~df["ticker"].astype(str).str.upper().str.contains(r"CASH|^USD$|RECEIVABLE|PAYABLE|FUTURE", regex=True, na=False)]
        df["index"] = "MIDCAP400"
        cols = ["ticker", "company", "index"] if "company" in df.columns else ["ticker", "index"]
        rows_df, warn = _assert_count(df[cols], "MIDCAP400", "SSGA_MDY_XLSX")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.warning(f"SSGA MDY failed: {e} — trying iShares IJH")
    try:
        url = "https://www.ishares.com/us/products/239467/ishares-core-sp-mid-cap-etf/1467271812596.ajax?fileType=csv&fileName=IJH_holdings&dataType=fund"
        r = requests.get(url, headers=HEADERS, timeout=45)
        r.raise_for_status()
        # Skip metadata rows until header found
        lines = r.content.decode("utf-8", errors="replace").splitlines()
        header_idx = next((i for i, l in enumerate(lines) if "Ticker" in l or "ticker" in l.lower()), None)
        if header_idx is None:
            raise ValueError("IJH CSV: no header row")
        from io import StringIO
        df = pd.read_csv(StringIO("\n".join(lines[header_idx:])))
        df.columns = [str(c).strip() for c in df.columns]
        ticker_col = next((c for c in df.columns if c.lower() == "ticker"), None)
        asset_col = next((c for c in df.columns if "asset" in c.lower() and "class" in c.lower()), None)
        if not ticker_col:
            raise ValueError("IJH CSV: no Ticker column")
        if asset_col:
            df = df[df[asset_col].astype(str).str.lower() == "equity"]
        df = df.rename(columns={ticker_col: "ticker"})
        df["ticker"] = df["ticker"].astype(str).str.strip()
        df = df[df["ticker"].apply(lambda t: bool(re.match(r'^[A-Za-z]', t)))]
        name_col = next((c for c in df.columns if "name" in c.lower()), None)
        if name_col:
            df = df.rename(columns={name_col: "company"})
        df["index"] = "MIDCAP400"
        cols = ["ticker", "company", "index"] if "company" in df.columns else ["ticker", "index"]
        rows_df, warn = _assert_count(df[cols], "MIDCAP400", "IJH_ISHARES_CSV")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.warning(f"IJH CSV failed: {e} — trying S&P DJI fallback")
    try:
        url = "https://www.spglobal.com/spdji/en/idsexport/file.xls?redesignExport=true&indexId=748"
        r = requests.get(url, headers=HEADERS, timeout=45)
        r.raise_for_status()
        df = pd.read_excel(BytesIO(r.content), skiprows=9)
        df.columns = [str(c).strip() for c in df.columns]
        ticker_col = next((c for c in df.columns if "ticker" in c.lower()), None)
        if not ticker_col:
            raise ValueError("DJI XLS MidCap: no ticker column")
        df = df.rename(columns={ticker_col: "ticker"})
        df["ticker"] = df["ticker"].astype(str).str.strip()
        df = df[df["ticker"].apply(lambda t: bool(re.match(r'^[A-Za-z]', t)))]
        df["index"] = "MIDCAP400"
        rows_df, warn = _assert_count(df[["ticker", "index"]], "MIDCAP400", "SPDJI_XLS_FALLBACK")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.error(f"MidCap400 constituent fetch failed: {e}")
        return pd.DataFrame(columns=["ticker", "index"]), "CONSTITUENT_SOURCE_FAILURE:MIDCAP400"


def fetch_ftse250():
    """Fetch FTSE 250 constituents.

    Primary:  DWS Xtrackers XMCX JSON API — 260 holdings (fully replicated), no
              authentication or browser session required.
              Endpoint: https://etf.dws.com/api/pdp/en-gb/etf/LU0292097317-ftse-250-ucits-etf-1d/holdings
              Returns JSON with all rows in a single response; filter column_5 == "Equities"
              to exclude ~7 money-market / mutual-fund buffer rows.
    Fallback: DWS Xtrackers XMCX Excel download — same underlying data, different endpoint.
              Requires Referer header (same pattern as fetch_stoxx600).
              Endpoint: https://etf.dws.com/etfdata/export/GBR/ENG/excel/product/constituent/LU0292097317/
    All FTSE 250 constituents trade on the London Stock Exchange regardless of ISIN country
    prefix (some are Jersey/Guernsey/Israel-domiciled but LSE-listed). ISINs are resolved
    to yfinance tickers via _resolve_stoxx_ticker targeting the LSE exchange.
    """
    log.info("Fetching FTSE 250 constituents...")

    def _resolve_lse(isin, name):
        """Resolve a FTSE 250 ISIN to an LSE .L ticker; return None if unresolvable."""
        time.sleep(0.15)
        ticker = _resolve_stoxx_ticker(isin, "United Kingdom", name)
        if not ticker:
            log.warning(f"FTSE250: unresolvable ISIN {isin} ({name}) — skipped")
            return None
        clean = ticker.rstrip("*")
        if not clean.endswith(".L"):
            log.warning(f"FTSE250: {isin} resolved to non-LSE ticker {clean} — skipped")
            return None
        return clean

    # ── Primary: DWS Xtrackers XMCX JSON API ────────────────────────────────
    try:
        url = "https://etf.dws.com/api/pdp/en-gb/etf/LU0292097317-ftse-250-ucits-etf-1d/holdings"
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        data = r.json()
        rows = data["tables"][0]["values"]
        if not rows:
            raise ValueError("DWS XMCX API returned empty holdings table")
        records = []
        for row in rows:
            isin = (row.get("header") or {}).get("value", "")
            name = (row.get("column_0") or {}).get("value", "")
            asset_class = (row.get("column_5") or {}).get("value", "")
            # Exclude money-market buffer and mutual fund rows (~7 of 260)
            if asset_class.lower() != "equities":
                continue
            if not isin or len(isin) != 12:
                continue
            ticker = _resolve_lse(isin, name)
            if ticker:
                records.append({
                    "ticker": ticker,
                    "company": name,
                    "isin": isin,
                    "index": "FTSE250",
                })
        df = pd.DataFrame(records).drop_duplicates("ticker")
        rows_df, warn = _assert_count(df, "FTSE250", "DWS_XMCX_API")
        if len(rows_df) < CONSTITUENT_BOUNDS["FTSE250"][0]:
            raise ValueError(
                f"DWS_XMCX_API resolved only {len(rows_df)} valid LSE tickers — falling back"
            )
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.warning(f"DWS XMCX API FTSE 250 fetch failed: {e} — trying DWS Excel download")

    # ── Fallback: DWS Xtrackers XMCX Excel download ─────────────────────────
    try:
        url = "https://etf.dws.com/etfdata/export/GBR/ENG/excel/product/constituent/LU0292097317/"
        dws_headers = {
            **HEADERS,
            "Referer": "https://etf.dws.com/en-gb/LU0292097317-ftse-250-ucits-etf-1d/",
        }
        r = requests.get(url, headers=dws_headers, timeout=60)
        r.raise_for_status()
        if openpyxl is None:
            raise ImportError("openpyxl required for DWS XMCX Excel fallback")
        wb = openpyxl.load_workbook(BytesIO(r.content))
        ws = wb.active
        # Row layout (min_row=5 skips disclaimer + blank + header row):
        # col[0]=index, col[1]=Name, col[2]=ISIN, col[3]=Country, col[4]=Currency,
        # col[5]=Exchange, col[6]=Type of Security, col[7]=Rating,
        # col[8]=Primary Listing, col[9]=Industry Classification, col[10]=Weighting
        records = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or row[6] != "Equities":
                continue
            name = row[1]
            isin = row[2]
            if not name or not isin:
                continue
            ticker = _resolve_lse(str(isin).strip(), str(name).strip())
            if ticker:
                records.append({
                    "ticker": ticker,
                    "company": name,
                    "isin": isin,
                    "index": "FTSE250",
                })
        df = pd.DataFrame(records).drop_duplicates("ticker")
        rows_df, warn = _assert_count(df, "FTSE250", "DWS_XMCX_EXCEL_FALLBACK")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.error(f"FTSE 250 constituent fetch failed completely: {e}")
        return pd.DataFrame(columns=["ticker", "index"]), "CONSTITUENT_SOURCE_FAILURE:FTSE250"


def _resolve_stoxx_ticker(isin, country, name=""):
    """Resolve STOXX600 ISIN to yfinance ticker via yf.Lookup."""
    try:
        lk = yf.Lookup(isin)
        stocks = lk.get_stock(count=8)
        if stocks is None or stocks.empty:
            return None
        target_exchanges = COUNTRY_EXCHANGES.get(country, [])
        for sym, row in stocks.iterrows():
            exch = row.get("exchange", "")
            if exch in target_exchanges:
                return sym
        # Fallback: return first result (flag as unverified)
        return stocks.index[0] + "*"
    except Exception:
        return None


def fetch_stoxx600():
    """Fetch STOXX Europe 600. Primary: Xtrackers XLSX. Fallback: stoxx.com HTML."""
    log.info("Fetching STOXX Europe 600 constituents...")
    try:
        url = "https://etf.dws.com/etfdata/export/LUX/ENG/excel/product/constituent/LU0328475792/"
        r = requests.get(url, headers={**HEADERS, "Referer": "https://etf.dws.com/en-lu/LU0328475792-stoxx-europe-600-ucits-etf-1c/"}, timeout=60)
        r.raise_for_status()
        if openpyxl is None:
            raise ImportError("openpyxl required for STOXX600 fetch")
        wb = openpyxl.load_workbook(BytesIO(r.content))
        ws = wb.active
        records = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or row[6] != "Equities":
                continue
            name = row[1]
            isin = row[2]
            country = row[3] or ""
            currency = row[4] or ""
            exchange = row[5] or ""
            sector = row[9] if len(row) > 9 else ""
            weight = row[10] if len(row) > 10 else None
            if not name or not isin:
                continue
            records.append({
                "name": name, "isin": isin, "country": country,
                "currency": currency, "exchange_raw": exchange,
                "sector": sector, "weight": weight,
            })
        # Resolve tickers
        resolved = []
        for rec in records:
            time.sleep(0.15)
            ticker = _resolve_stoxx_ticker(rec["isin"], rec["country"], rec["name"])
            if ticker:
                resolved.append({
                    "ticker": ticker.rstrip("*"),
                    "company": rec["name"],
                    "sector": rec["sector"],
                    "index": "STOXX600",
                    "isin": rec["isin"],
                    "ticker_verified": not ticker.endswith("*"),
                })
        df = pd.DataFrame(resolved)
        rows_df, warn = _assert_count(df, "STOXX600", "XTRACKERS_XLSX")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.warning(f"Xtrackers STOXX600 XLSX failed: {e} — trying stoxx.com HTML fallback")
    try:
        from bs4 import BeautifulSoup
        url = "https://stoxx.com/index/sxxp/?components=true"
        r = requests.get(url, headers=HEADERS, timeout=45)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        tables = pd.read_html(r.text)
        if not tables:
            raise ValueError("stoxx.com: no tables found")
        df = tables[0]
        df.columns = [str(c).strip() for c in df.columns]
        name_col = next((c for c in df.columns if "company" in c.lower() or "name" in c.lower()), None)
        if not name_col:
            raise ValueError("stoxx.com: no name column")
        # Without tickers this is last resort only; return names for manual review
        df["ticker"] = ""  # no tickers available
        df["company"] = df[name_col]
        df["index"] = "STOXX600"
        rows_df, warn = _assert_count(df, "STOXX600", "STOXX_HTML_FALLBACK")
        return rows_df.reset_index(drop=True), "CONSTITUENT_SOURCE_FAILURE:STOXX600_NO_TICKERS"
    except Exception as e:
        log.error(f"STOXX600 constituent fetch failed entirely: {e}")
        return pd.DataFrame(columns=["ticker", "index"]), "CONSTITUENT_SOURCE_FAILURE:STOXX600"


def fetch_spi():
    """Fetch Swiss Performance Index. Primary: CHSPI iShares CSV."""
    log.info("Fetching SPI constituents...")
    try:
        url = "https://www.ishares.com/ch/individual/en/products/264107/ishares-spi-ch-fund/1495092304805.ajax?fileType=csv&fileName=CHSPI_holdings&dataType=fund"
        r = requests.get(url, headers=HEADERS, timeout=45)
        r.raise_for_status()
        from io import StringIO
        lines = r.content.decode("utf-8", errors="replace").splitlines()
        header_idx = next((i for i, l in enumerate(lines) if "Ticker" in l or "ticker" in l.lower()), None)
        if header_idx is None:
            raise ValueError("CHSPI CSV: no header row")
        df = pd.read_csv(StringIO("\n".join(lines[header_idx:])))
        df.columns = [str(c).strip() for c in df.columns]
        ticker_col = next((c for c in df.columns if c.lower() == "ticker"), None)
        asset_col = next((c for c in df.columns if "asset" in c.lower() and "class" in c.lower()), None)
        if not ticker_col:
            raise ValueError("CHSPI CSV: no Ticker column")
        if asset_col:
            df = df[df[asset_col].astype(str).str.contains("equity", case=False, na=False)]
        df = df.rename(columns={ticker_col: "ticker"})
        df["ticker"] = df["ticker"].astype(str).str.strip()
        df = df[df["ticker"].apply(lambda t: bool(re.match(r'^[A-Za-z]', t)))]
        name_col = next((c for c in df.columns if "name" in c.lower()), None)
        if name_col:
            df = df.rename(columns={name_col: "company"})
        # Yahoo needs the .SW suffix for SIX Swiss stocks; the CHSPI CSV lists BARE local tickers (NESN, NOVN...).
        df["ticker"] = df["ticker"].apply(lambda t: str(t) if "." in str(t) else str(t) + ".SW")
        df["index"] = "SPI"
        cols = ["ticker", "company", "index"] if "company" in df.columns else ["ticker", "index"]
        rows_df, warn = _assert_count(df[cols], "SPI", "CHSPI_ISHARES_CSV")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.error(f"SPI constituent fetch failed: {e}")
        return pd.DataFrame(columns=["ticker", "index"]), "CONSTITUENT_SOURCE_FAILURE:SPI"


def fetch_tsx():
    """Fetch S&P/TSX Composite. Primary: TMX Money HTML. Fallback: XIC BlackRock CSV."""
    log.info("Fetching S&P/TSX Composite constituents...")
    try:
        url = "https://money.tmx.com/en/quote/%5ETSX/constituents"
        tables = pd.read_html(url, attrs=None)
        df = None
        for t in tables:
            cols = [str(c).lower() for c in t.columns]
            if any("ticker" in c or "symbol" in c for c in cols):
                df = t
                break
        if df is None and tables:
            df = tables[0]
        if df is None:
            raise ValueError("TMX: no table found")
        df.columns = [str(c).strip() for c in df.columns]
        ticker_col = next((c for c in df.columns if c.lower() in ("ticker", "symbol")), None)
        if not ticker_col:
            ticker_col = df.columns[0]
        df = df.rename(columns={ticker_col: "ticker"})
        df["ticker"] = df["ticker"].astype(str).str.strip()
        df["ticker"] = df["ticker"].apply(lambda t: t if t.endswith(".TO") else t + ".TO")
        df = df[df["ticker"].apply(lambda t: bool(re.match(r'^[A-Za-z]', t)))]
        name_col = next((c for c in df.columns if "name" in c.lower() or "company" in c.lower()), None)
        if name_col:
            df = df.rename(columns={name_col: "company"})
        df["index"] = "TSX"
        cols = ["ticker", "company", "index"] if "company" in df.columns else ["ticker", "index"]
        rows_df, warn = _assert_count(df[cols], "TSX", "TMX_HTML")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.warning(f"TMX HTML failed: {e} — trying XIC BlackRock CSV")
    try:
        url = "https://www.blackrock.com/ca/investors/en/products/239837/ishares-sptsx-capped-composite-index-etf/1359936950992.ajax?fileType=csv&fileName=XIU_holdings&dataType=fund"
        r = requests.get(url, headers=HEADERS, timeout=45)
        r.raise_for_status()
        from io import StringIO
        lines = r.content.decode("utf-8", errors="replace").splitlines()
        header_idx = next((i for i, l in enumerate(lines) if "Ticker" in l), 0)
        df = pd.read_csv(StringIO("\n".join(lines[header_idx:])))
        df.columns = [str(c).strip() for c in df.columns]
        ticker_col = next((c for c in df.columns if c.lower() == "ticker"), None)
        if not ticker_col:
            raise ValueError("XIC CSV: no ticker col")
        df = df.rename(columns={ticker_col: "ticker"})
        df["ticker"] = df["ticker"].astype(str).str.strip().apply(lambda t: t if t.endswith(".TO") else t + ".TO")
        df = df[df["ticker"].apply(lambda t: bool(re.match(r'^[A-Za-z]', t)))]
        df["index"] = "TSX"
        rows_df, warn = _assert_count(df[["ticker", "index"]], "TSX", "XIC_BLACKROCK_CSV_FALLBACK")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.error(f"TSX constituent fetch failed: {e}")
        return pd.DataFrame(columns=["ticker", "index"]), "CONSTITUENT_SOURCE_FAILURE:TSX"


def fetch_ibrx50():
    """Fetch IBrX 50 via B3 JSON API."""
    log.info("Fetching IBrX 50 constituents...")
    import base64
    try:
        payload = json.dumps({"index": "IBXL", "language": "pt-br"}).encode()
        encoded = base64.b64encode(payload).decode()
        url = f"https://sistemaswebb3-listados.b3.com.br/indexProxy/indexCall/GetPortfolioDay/{encoded}?language=pt-br"
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        data = r.json()
        results = data.get("results", [])
        if not results:
            raise ValueError("B3 API: empty results")
        records = []
        for item in results:
            ticker_raw = str(item.get("cod", "")).strip()
            # Normalize: append .SA suffix
            ticker = ticker_raw + ".SA" if ticker_raw and not ticker_raw.endswith(".SA") else ticker_raw
            name = str(item.get("asset", item.get("name", ""))).strip()
            if ticker:
                records.append({"ticker": ticker, "company": name, "index": "IBRX50"})
        df = pd.DataFrame(records)
        rows_df, warn = _assert_count(df, "IBRX50", "B3_JSON_API")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.warning(f"B3 JSON API failed: {e} — trying B3 composition page")
    try:
        # B3 composition page fallback
        url = "https://sistemaswebb3-listados.b3.com.br/indexPage/day/IBXL?language=pt-br"
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        tables = pd.read_html(r.text)
        if not tables:
            raise ValueError("B3 page: no tables")
        df = tables[0]
        df.columns = [str(c).strip() for c in df.columns]
        ticker_col = next((c for c in df.columns if "cod" in c.lower() or "ticker" in c.lower() or "ativo" in c.lower()), df.columns[0])
        df = df.rename(columns={ticker_col: "ticker"})
        df["ticker"] = df["ticker"].astype(str).str.strip().apply(lambda t: t if t.endswith(".SA") else t + ".SA")
        df["index"] = "IBRX50"
        rows_df, warn = _assert_count(df[["ticker", "index"]], "IBRX50", "B3_PAGE_FALLBACK")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.error(f"IBrX 50 constituent fetch failed: {e}")
        return pd.DataFrame(columns=["ticker", "index"]), "CONSTITUENT_SOURCE_FAILURE:IBRX50"


def fetch_ipc35():
    """Fetch S&P/BMV IPC 35. Primary: S&P DJI XLS."""
    log.info("Fetching S&P/BMV IPC 35 constituents...")
    try:
        url = "https://www.spglobal.com/spdji/en/idsexport/file.xls?redesignExport=true&indexId=92"
        r = requests.get(url, headers=HEADERS, timeout=45)
        r.raise_for_status()
        df = pd.read_excel(BytesIO(r.content), skiprows=9)
        df.columns = [str(c).strip() for c in df.columns]
        ticker_col = next((c for c in df.columns if "ticker" in c.lower()), None)
        name_col = next((c for c in df.columns if "name" in c.lower() or "company" in c.lower() or "security" in c.lower()), None)
        if not ticker_col:
            raise ValueError("IPC XLS: no ticker col")
        df = df.rename(columns={ticker_col: "ticker"})
        if name_col:
            df = df.rename(columns={name_col: "company"})
        df["ticker"] = df["ticker"].astype(str).str.strip()
        # Strip wildcard suffix (e.g. WALMEX*.MX → WALMEXV.MX handled by lookup;
        # strip asterisk, keep the base, append .MX if missing)
        def clean_ipc_ticker(t):
            t = t.replace("*", "").strip()
            if t and not t.endswith(".MX"):
                t = t + ".MX"
            return t
        df["ticker"] = df["ticker"].apply(clean_ipc_ticker)
        df = df[df["ticker"].apply(lambda t: bool(re.match(r'^[A-Za-z]', t)))]
        df["index"] = "IPC35"
        cols = ["ticker", "company", "index"] if "company" in df.columns else ["ticker", "index"]
        rows_df, warn = _assert_count(df[cols], "IPC35", "SPDJI_XLS")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.warning(f"IPC35 DJI XLS failed: {e} — trying BMV fallback")
    try:
        url = "https://www.bmv.com.mx/en/indices/main/"
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        tables = pd.read_html(r.text)
        if not tables:
            raise ValueError("BMV: no tables")
        df = tables[0]
        df.columns = [str(c).strip() for c in df.columns]
        ticker_col = df.columns[0]
        df = df.rename(columns={ticker_col: "ticker"})
        df["ticker"] = df["ticker"].astype(str).str.strip()
        df["ticker"] = df["ticker"].apply(lambda t: t if t.endswith(".MX") else t + ".MX")
        df["index"] = "IPC35"
        rows_df, warn = _assert_count(df[["ticker", "index"]], "IPC35", "BMV_FALLBACK")
        return rows_df.reset_index(drop=True), warn
    except Exception as e:
        log.error(f"IPC35 constituent fetch failed: {e}")
        return pd.DataFrame(columns=["ticker", "index"]), "CONSTITUENT_SOURCE_FAILURE:IPC35"


def fetch_constituents(group):
    """
    Dispatch constituent fetch for all indices in a group.
    Returns (combined_df, warnings_list).
    """
    index_list = GROUP_INDICES.get(group, [])
    all_dfs = []
    warnings = []
    fetchers = {
        "SP500":     fetch_sp500,
        "NASDAQ":    fetch_nasdaq,
        "MIDCAP400": fetch_midcap400,
        "FTSE250":   fetch_ftse250,
        "STOXX600":  fetch_stoxx600,
        "SPI":       fetch_spi,
        "TSX":       fetch_tsx,
        "IBRX50":    fetch_ibrx50,
        "IPC35":     fetch_ipc35,
    }
    for idx in index_list:
        fn = fetchers.get(idx)
        if fn is None:
            log.error(f"No fetcher for index {idx}")
            continue
        df, warn = fn()
        if warn:
            warnings.append(warn)
        if df is not None and not df.empty:
            all_dfs.append(df)
        else:
            warnings.append(f"CONSTITUENT_SOURCE_FAILURE:{idx}")
    if not all_dfs:
        return pd.DataFrame(columns=["ticker", "index"]), warnings
    combined = pd.concat(all_dfs, ignore_index=True)
    # Universal identifier filter: discard rows with no valid ticker
    combined = combined[combined["ticker"].apply(lambda t: bool(t and re.match(r'^[A-Za-z]', str(t))))]
    combined = combined.drop_duplicates("ticker").reset_index(drop=True)
    # Apply structural exclusions
    combined = combined[~combined["ticker"].isin(STRUCTURAL_EXCLUSIONS)].reset_index(drop=True)
    return combined, warnings


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4: GATE FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def _kw_word_start_hit(text, keywords):
    """Return the first keyword occurring at a WORD START in text, else None.
    Word-start (left-boundary) matching fixes the substring-within-word false positive
    (e.g. 'spac' must NOT match the 'spac' inside 'aerospace') while preserving intended
    PREFIX keywords: 'insur'->'insurance', 'royalt'->'royalty', 'asset manag'->'asset
    management', 'holding compan'->'holding company'."""
    for kw in keywords:
        if re.search(r"\b" + re.escape(kw), text):
            return kw
    return None


def _sector_excluded_standard(info):
    """
    Returns (excluded: bool, reason: str) for standard Gate 1 (non-Nasdaq).
    Also handles pre-revenue biotech check identically to Nasdaq-modified Gate 1.
    """
    sector   = str(info.get("sector", "") or "").lower()
    industry = str(info.get("industry", "") or "").lower()

    # Real Estate / REITs
    if "real estate" in sector:
        return True, "Sector exclusion: Real Estate/REIT"

    # Pre-revenue biotech
    if "biotechnology" in industry:
        eps = safe_float(info.get("trailingEps"))
        if eps is None or eps < 0:
            return True, "Sector exclusion: pre-revenue biotech"

    # Financial Services — exclude non-FinTech
    if "financial services" in sector or "financials" in sector:
        keep = any(kw in industry for kw in GATE1_FINTECH_KEEP_KEYWORDS)
        if not keep:
            return True, f"Sector exclusion: financial services ({industry})"

    # Broad keyword exclusions (word-start match — see _kw_word_start_hit)
    combined_text = sector + " " + industry
    _hit = _kw_word_start_hit(combined_text, GATE1_EXCLUDE_KEYWORDS)
    if _hit:
        return True, f"Sector exclusion: {_hit}"

    return False, ""


def _sector_excluded_nasdaq(info):
    """
    Returns (excluded: bool, reason: str) for Nasdaq-modified Gate 1.
    More granular than standard — keeps semiconductors, software, pharma explicitly.
    """
    sector   = str(info.get("sector", "") or "").lower()
    industry = str(info.get("industry", "") or "").lower()

    # Pre-revenue biotech
    if "biotechnology" in industry:
        eps = safe_float(info.get("trailingEps"))
        if eps is None or eps < 0:
            return True, "Sector exclusion (Nasdaq): pre-revenue biotech"

    # Real Estate
    if sector == "real estate":
        return True, "Sector exclusion (Nasdaq): Real Estate"

    # Financial Services — keep FinTech/payments explicitly
    if sector == "financial services":
        keep_kws = ["credit services", "capital markets", "financial data",
                    "stock exchange", "payment", "fintech", "insurance"]
        if not any(kw in industry for kw in keep_kws):
            return True, f"Sector exclusion (Nasdaq): financial services ({industry})"

    # SPACs / investment trusts / royalty (word-start match — 'spac' must not hit 'aerospace')
    spac_kws = ["spac", "blank check", "investment trust", "royalt", "streaming vehicle"]
    if _kw_word_start_hit(industry, spac_kws):
        return True, f"Sector exclusion (Nasdaq): {industry}"

    # Mining (word-start match)
    mining_kws = ["gold", "silver", "copper", "mining", "royalty"]
    if _kw_word_start_hit(industry, mining_kws):
        return True, f"Sector exclusion (Nasdaq): mining/metals ({industry})"

    return False, ""


def gate1_pass(info, nasdaq_mode=False):
    """Returns (pass: bool, reason: str, gate_code: str)."""
    fn = _sector_excluded_nasdaq if nasdaq_mode else _sector_excluded_standard
    excluded, reason = fn(info)
    if excluded:
        return False, reason, "Gate 1"
    return True, "", ""


def gate2_pass(income_stmt, info, nasdaq_mode=False, cashflow=None, ticker=None):
    """
    Returns (pass: bool, reason: str, gate_code: str, gm_value: float|None).
    CRITICAL: None grossMargins → GATE_DATA_UNRESOLVED, NOT a gate fail.
    info["grossMargins"] must NOT be used as Gate 2 source.

    Threshold is now sector-segmented via GATE2_GM_THRESHOLD dict.
    The nasdaq_mode parameter is retained for backward-compatible call sites
    but no longer drives the threshold — sector classification does.
    For stocks where gross_margin is not yet computed at call time,
    classify_sector_bucket is called with gross_margin=None (defaults to
    semiconductor_hardware bucket conservatively).
    """
    sector   = (info.get("sector",   "") or "") if info else ""
    industry = (info.get("industry", "") or "") if info else ""
    tkr      = ticker or (info.get("symbol") if info else None)
    # Classify via capex intensity (+ anchors), NOT gross margin — see classify_sector_bucket.
    capex_int = compute_capex_intensity(income_stmt, cashflow)
    bucket    = classify_sector_bucket(sector, industry, capex_intensity=capex_int, ticker=tkr)
    threshold = GATE2_GM_THRESHOLD.get(bucket, GATE2_GM_THRESHOLD["default"])
    if income_stmt is None or income_stmt.empty:
        return None, "GATE_DATA_UNRESOLVED: income_stmt empty", "GATE_DATA_UNRESOLVED", None

    gross_profit = get_stmt_value(income_stmt, ["Gross Profit", "GrossProfit"])
    revenue = get_stmt_value(income_stmt,
        ["Total Revenue", "Operating Revenue", "Revenue", "TotalRevenue"])

    if gross_profit is None or revenue is None or revenue == 0:
        return None, "GATE_DATA_UNRESOLVED: gross_profit or revenue missing", "GATE_DATA_UNRESOLVED", None

    gm = gross_profit / revenue
    # Redesign Part 3 §8: GM is a SECTOR-SEGMENTED SCORE (GROSS_MARGIN_SCORE_THRESHOLDS),
    # not a hard quality gate. When RELAX_GM_GATE is on, only genuinely non-viable businesses
    # (GM < GM_VIABILITY_FLOOR) are gated; low-GM-for-sector names survive and are differentiated
    # by the sector-segmented Part A GM score. Legacy hard gate retained behind the flag.
    if not getattr(_cfg, "RELAX_GM_GATE", False):
        if gm >= threshold:
            return True, "", "", gm
        return False, f"Gross margin {gm*100:.1f}% below {threshold*100:.0f}% threshold ({bucket})", "Gate 2", gm
    if gm < getattr(_cfg, "GM_VIABILITY_FLOOR", 0.0):
        return False, f"Gross margin {gm*100:.1f}% below viability floor ({bucket})", "Gate 2", gm
    note = "" if gm >= threshold else f"low GM {gm*100:.1f}% < {threshold*100:.0f}% sector floor — scored not gated ({bucket})"
    return True, note, "", gm


def gate3_pass(cashflow):
    """
    Returns (pass: bool|None, reason: str, gate_code: str, fcf_pos_years: int, avail_years: int).
    CRITICAL: threshold is ALWAYS 3 — never reduced for fewer available years.
    """
    if cashflow is None or cashflow.empty:
        return None, "GATE_DATA_UNRESOLVED: cashflow empty", "GATE_DATA_UNRESOLVED", 0, 0

    fcf_series = get_stmt_series(cashflow, ["Free Cash Flow"], max_periods=5)
    # Also try derived FCF
    if not fcf_series:
        ocf_series = get_stmt_series(cashflow,
            ["Operating Cash Flow", "Cash Flow From Continuing Operating Activities",
             "Net Cash Provided By Operating Activities", "Total Cash From Operating Activities"])
        cap_series = get_stmt_series(cashflow,
            ["Capital Expenditure", "Capital Expenditures", "CapitalExpenditure",
             "Purchase Of PPE", "Purchase of Property Plant and Equipment",
             "Investments in Property Plant and Equipment"])
        if ocf_series and cap_series:
            ocf_dict = dict(ocf_series)
            cap_dict = dict(cap_series)
            fcf_series = [(y, ocf_dict[y] + cap_dict[y]) for y in ocf_dict if y in cap_dict]

    avail = len([v for _, v in fcf_series if v is not None])
    if avail < 3:
        return None, f"GATE_DATA_UNRESOLVED: only {avail} FCF years available", "GATE_DATA_UNRESOLVED", 0, avail

    pos = len([v for _, v in fcf_series if v is not None and v > 0])
    if pos >= 3:
        return True, "", "", pos, avail
    if getattr(_cfg, "RELAX_FCF_GATE", False):
        # Redesign Part 3 §8: FCF-negative from strategic capex is acceptable IF operations
        # generate cash (most-recent OCF > 0). Genuine ops cash-burners (OCF<=0) still fail.
        ocf_series = get_stmt_series(cashflow,
            ["Operating Cash Flow", "Cash Flow From Continuing Operating Activities",
             "Net Cash Provided By Operating Activities", "Total Cash From Operating Activities"])
        ocf_recent = next((v for _, v in ocf_series if v is not None), None)
        if ocf_recent is not None and ocf_recent > 0:
            return True, f"FCF positive {pos}/{avail}yr but OCF>0 (capex-driven; scored not gated)", "", pos, avail
    return False, f"FCF positive only {pos}/{avail} years", "Gate 3", pos, avail


def gate4_pass(income_stmt, sector_bucket: str = "default"):
    """
    Returns (pass: bool|None, reason: str, gate_code: str, rev_cagr: float|None).
    Revenue CAGR (3yr) >= 5%.

    Enhancement 2C-1: For semiconductor_equipment bucket companies only, if
    3yr CAGR < 5% (fails the primary gate), compute a 5yr CAGR as a fallback.
    If 5yr CAGR >= 3%, the gate passes with a GATE4_5YR_OVERRIDE flag embedded
    in the reason string. The 5yr CAGR uses whatever is available from the
    rev_series — the existing 5-period fetch (max_periods=5) already provides
    up to 5 years of data, so no additional fetch is required.

    All other sector buckets: unchanged primary 3yr CAGR >= 5% logic.
    """
    if income_stmt is None or income_stmt.empty:
        return None, "GATE_DATA_UNRESOLVED: income_stmt empty", "GATE_DATA_UNRESOLVED", None

    rev_series = get_stmt_series(income_stmt,
        ["Total Revenue", "Operating Revenue", "Revenue", "TotalRevenue"], max_periods=5)

    if len(rev_series) < 2:
        return None, "GATE_DATA_UNRESOLVED: insufficient revenue history", "GATE_DATA_UNRESOLVED", None

    rev_series, anomaly, excluded_yr = detect_rev_anomaly(rev_series)
    if len(rev_series) < 2:
        return None, "GATE_DATA_UNRESOLVED: revenue anomaly left <2 clean periods", "GATE_DATA_UNRESOLVED", None

    # Primary: 3yr CAGR window (same logic as before)
    use = rev_series[:4]  # up to 4 periods (gives 3yr CAGR from index 0 to 3)
    if len(use) >= 4:
        years = 3
        end_val, start_val = use[0][1], use[3][1]
    elif len(use) == 3:
        years = 2
        end_val, start_val = use[0][1], use[2][1]
    else:
        years = 1
        end_val, start_val = use[0][1], use[1][1]

    cagr = compute_cagr(start_val, end_val, years)
    if cagr is None:
        return None, "GATE_DATA_UNRESOLVED: CAGR computation failed", "GATE_DATA_UNRESOLVED", None

    if cagr >= 0.05:
        return True, "", "", cagr

    # Primary gate failed. For semiconductor_equipment only: attempt 5yr CAGR override.
    if sector_bucket == "semiconductor_equipment" and len(rev_series) >= 5:
        end_5   = rev_series[0][1]
        start_5 = rev_series[4][1]
        cagr_5  = compute_cagr(start_5, end_5, 4)   # 5 data points = 4 year span
        if cagr_5 is not None and cagr_5 >= 0.03:
            reason = (
                f"GATE4_5YR_OVERRIDE: 3yr_cagr={cagr*100:.1f}% < 5% threshold; "
                f"5yr_cagr={cagr_5*100:.1f}% >= 3% override — semiconductor_equipment cycle trough"
            )
            return True, reason, "Gate 4 (5yr override)", cagr_5

    # Redesign Part 3 §8: forward-inclusive relaxation — low-trailing-growth names (UNH +2%,
    # turnarounds) survive if not declining; forward growth/estimate momentum is scored downstream.
    if getattr(_cfg, "RELAX_CAGR_GATE", False) and cagr is not None and cagr >= getattr(_cfg, "GATE4_RELAXED_CAGR_MIN", 0.0):
        return True, f"Revenue CAGR {cagr*100:.1f}% < 5% but >= relaxed floor — scored not gated", "", cagr
    return False, f"Revenue CAGR {cagr*100:.1f}% below 5% threshold", "Gate 4", cagr


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 5: DATA FETCHING (batched, chunked, with cooldowns)
# ─────────────────────────────────────────────────────────────────────────────

def _resilient(fn, sym, tries=4, base=2.0, cap=45.0):
    """Retry fn() on exception with exponential backoff + full jitter. Call sites RAISE on an
    empty/throttle response so yfinance's silent empty-after-200 (the dominant rate-limit failure)
    becomes retryable instead of an instant fetch-miss. Strictly more robust; no scoring change."""
    import random as _rand
    for _attempt in range(tries):
        try:
            return fn()
        except Exception as _e:
            if _attempt == tries - 1:
                raise
            # crumb/401 refresh: a stale Yahoo crumb won't recover on a plain retry — force YfData
            # to re-fetch cookie+crumb so the next attempt is authenticated (cuts the 'Invalid Crumb'
            # 401 jitter). Defensive: never let the internal poke break the retry loop.
            if any(_k in str(_e).lower() for _k in ("crumb", "401", "unauthorized")):
                try:
                    import yfinance.data as _yd
                    _ys = _yd.YfData(); _ys._crumb = None; _ys._cookie = None
                except Exception:
                    pass
            time.sleep(min(cap, base * (2 ** _attempt)) * (0.5 + _rand.random()))


def _fetch_ticker_info(ticker_sym):
    """Fetch ticker.info for a single ticker. Returns (ticker_sym, info_dict | None, error)."""
    def _pull():
        info = yf.Ticker(ticker_sym).info or {}
        if not info or len(info) < 5:
            raise RuntimeError("empty_info (throttle?)")
        return info
    try:
        return ticker_sym, _resilient(_pull, ticker_sym), None
    except Exception as e:
        return ticker_sym, None, str(e)


def _fetch_ticker_statements(ticker_sym):
    """Fetch annual statements for a single ticker."""
    def _pull():
        tk = yf.Ticker(ticker_sym)
        d = {"income_stmt": tk.income_stmt, "cashflow": tk.cashflow,
             "balance_sheet": tk.balance_sheet}
        if all((v is None or (hasattr(v, "empty") and v.empty)) for v in d.values()):
            raise RuntimeError("empty_statements (throttle?)")
        return d
    try:
        return ticker_sym, _resilient(_pull, ticker_sym), None
    except Exception as e:
        return ticker_sym, None, str(e)


def _fetch_ticker_scoring_data(ticker_sym, score_gt38=False):
    """Phase-3 fetch with throttle-resilient retry (wraps _once; retries on TOTAL failure only,
    so partial-data tolerance in the body is preserved)."""
    def _pull():
        sym, data, err = _fetch_ticker_scoring_data_once(ticker_sym, score_gt38)
        if err or data is None:
            raise RuntimeError(err or "empty_scoring (throttle?)")
        return data
    try:
        return ticker_sym, _resilient(_pull, ticker_sym), None
    except Exception as e:
        return ticker_sym, None, str(e)


def _fetch_ticker_scoring_data_once(ticker_sym, score_gt38=False):
    """Fetch all incremental data needed for Part A/B scoring and optionally overlays."""
    try:
        tk = yf.Ticker(ticker_sym)
        data = {
            "quarterly_income_stmt": tk.quarterly_income_stmt,
            "quarterly_cashflow":    tk.quarterly_cashflow,
            "quarterly_balance_sheet": tk.quarterly_balance_sheet,
            "earnings_estimate":     tk.earnings_estimate,
            "growth_estimates":      tk.growth_estimates,
            "analyst_price_targets": tk.analyst_price_targets,
            "dividends":             tk.dividends,
        }
        try:
            data["eps_revisions"] = tk.eps_revisions   # now scored in Part B for ALL gate-passers
        except Exception:
            data["eps_revisions"] = None
        try:
            data["eps_trend"] = tk.eps_trend   # forward axis: +1y consensus EPS momentum (Part 3 §13)
        except Exception:
            data["eps_trend"] = None
        # next_earnings via calendar
        try:
            cal = tk.calendar
            if cal is not None and not (isinstance(cal, dict) and not cal):
                if isinstance(cal, dict):
                    date_val = cal.get("Earnings Date") or cal.get("earningsDate")
                    if date_val:
                        if isinstance(date_val, (list, tuple)):
                            date_val = date_val[0]
                        data["next_earnings"] = str(date_val)[:10]
                    else:
                        data["next_earnings"] = "Unknown"
                else:
                    data["next_earnings"] = "Unknown"
            else:
                data["next_earnings"] = "Unknown"
        except Exception:
            # Fallback: earnings_dates
            try:
                ed = tk.earnings_dates
                if ed is not None and not ed.empty:
                    future = ed[ed.index > pd.Timestamp.now()]
                    data["next_earnings"] = str(future.index[0])[:10] if not future.empty else "Unknown"
                else:
                    data["next_earnings"] = "Unknown"
            except Exception:
                data["next_earnings"] = "Unknown"

        # Price history
        period = "5y" if score_gt38 else "2y"   # >=2y so the 12-1m price-momentum window (~273d) is always spannable
        try:
            data["history"] = tk.history(period=period)
        except Exception:
            data["history"] = None

        # High-score overlays (eps_revisions already fetched above for all gate-passers)
        if score_gt38:
            try:
                data["upgrades_downgrades"] = tk.upgrades_downgrades
            except Exception:
                data["upgrades_downgrades"] = None
            try:
                data["recommendations_summary"] = tk.recommendations_summary
            except Exception:
                data["recommendations_summary"] = None

        return ticker_sym, data, None
    except Exception as e:
        return ticker_sym, None, str(e)


def _run_batch(tickers, fetch_fn, chunk_size, max_workers, cooldown, fn_kwargs=None):
    """
    Run fetch_fn over tickers in chunks with ThreadPoolExecutor.
    Returns dict: {ticker: result_or_None}, dict: {ticker: error_str}
    """
    results = {}
    errors  = {}
    chunks  = [tickers[i:i+chunk_size] for i in range(0, len(tickers), chunk_size)]
    fn_kwargs = fn_kwargs or {}

    for ci, chunk in enumerate(chunks):
        log.info(f"  Chunk {ci+1}/{len(chunks)} ({len(chunk)} tickers)...")
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futures = {ex.submit(fetch_fn, t, **fn_kwargs): t for t in chunk}
            for fut in as_completed(futures):
                sym, data, err = fut.result()
                if err:
                    errors[sym] = err
                else:
                    results[sym] = data

        success_rate = len([t for t in chunk if t in results]) / len(chunk)
        if success_rate < 0.70 and ci < len(chunks) - 1:
            log.warning(f"  Chunk {ci+1} success rate {success_rate:.0%} < 70% — pausing 3min then retrying failed tickers")
            time.sleep(180)
            retry_tickers = [t for t in chunk if t not in results]
            with ThreadPoolExecutor(max_workers=max(max_workers // 2, 3)) as ex:
                futures = {ex.submit(fetch_fn, t, **fn_kwargs): t for t in retry_tickers}
                for fut in as_completed(futures):
                    sym, data, err = fut.result()
                    if err:
                        errors[sym] = err
                    else:
                        results[sym] = data
                        errors.pop(sym, None)

        if ci < len(chunks) - 1:
            time.sleep(cooldown)

    return results, errors


def fetch_phase1_info(tickers, group):
    """Fetch ticker.info for all tickers. Returns (info_map, error_map)."""
    params = BATCH_PARAMS.get(group, BATCH_PARAMS["OTHER"])
    log.info(f"Phase 1: fetching info for {len(tickers)} tickers...")
    return _run_batch(tickers, _fetch_ticker_info,
                      params["chunk"], params["workers"], params["cooldown"])


def fetch_phase2_statements(tickers, group):
    """Fetch annual statements for gate-screen tickers."""
    params = BATCH_PARAMS.get(group, BATCH_PARAMS["OTHER"])
    log.info(f"Phase 2: fetching statements for {len(tickers)} tickers...")
    return _run_batch(tickers, _fetch_ticker_statements,
                      params["chunk"], params["workers"], params["cooldown"])


def fetch_phase3_scoring(tickers, group, high_score_tickers=None):
    """
    Fetch incremental scoring data for gate passers.
    high_score_tickers: set of SUMMARY-eligible names at/above the Source floor (fetch 5yr history + overlay objects).
    """
    params  = BATCH_PARAMS.get(group, BATCH_PARAMS["OTHER"])
    high_set = set(high_score_tickers or [])
    log.info(f"Phase 3: fetching scoring data for {len(tickers)} tickers ({len(high_set)} high-score)...")

    results = {}
    errors  = {}
    # Separate high-score (need 5y + overlay) from standard
    standard = [t for t in tickers if t not in high_set]
    high     = [t for t in tickers if t in high_set]

    if standard:
        r, e = _run_batch(standard, _fetch_ticker_scoring_data,
                          params["chunk"], params["workers"], params["cooldown"],
                          fn_kwargs={"score_gt38": False})
        results.update(r); errors.update(e)

    if high:
        r, e = _run_batch(high, _fetch_ticker_scoring_data,
                          params["chunk"], params["workers"], params["cooldown"],
                          fn_kwargs={"score_gt38": True})
        results.update(r); errors.update(e)

    return results, errors


def apply_gates_standard(ticker_sym, info, income_stmt, cashflow):
    """
    Apply Gates 1–4 for non-Nasdaq groups.
    Returns dict with keys: gate_pass, gate_code, gate_reason, gross_margin,
                            fcf_pos_years, avail_fcf_years, rev_cagr_3yr
    """
    # Gate 1
    g1, reason1, code1 = gate1_pass(info, nasdaq_mode=False)
    if not g1:
        return {"gate_pass": False, "gate_code": code1, "gate_reason": reason1}

    # Gate 2
    g2, reason2, code2, gm = gate2_pass(income_stmt, info, nasdaq_mode=False, cashflow=cashflow, ticker=ticker_sym)
    if g2 is None:
        return {"gate_pass": None, "gate_code": code2, "gate_reason": reason2, "gross_margin": None}
    if not g2:
        return {"gate_pass": False, "gate_code": code2, "gate_reason": reason2, "gross_margin": gm}

    # Gate 3
    g3, reason3, code3, fcf_pos, avail = gate3_pass(cashflow)
    if g3 is None:
        return {"gate_pass": None, "gate_code": code3, "gate_reason": reason3,
                "gross_margin": gm, "fcf_pos_years": fcf_pos, "avail_fcf_years": avail}
    if not g3:
        return {"gate_pass": False, "gate_code": code3, "gate_reason": reason3,
                "gross_margin": gm, "fcf_pos_years": fcf_pos, "avail_fcf_years": avail}

    # Gate 4
    # Compute sector_bucket for Gate 4 cycle-trough override (Enhancement 2C-1)
    _sector   = (info.get("sector",   "") or "") if info else ""
    _industry = (info.get("industry", "") or "") if info else ""
    _bucket   = classify_sector_bucket(_sector, _industry)
    g4, reason4, code4, rev_cagr = gate4_pass(income_stmt, sector_bucket=_bucket)
    if g4 is None:
        return {"gate_pass": None, "gate_code": code4, "gate_reason": reason4,
                "gross_margin": gm, "fcf_pos_years": fcf_pos, "avail_fcf_years": avail}
    if not g4:
        return {"gate_pass": False, "gate_code": code4, "gate_reason": reason4,
                "gross_margin": gm, "fcf_pos_years": fcf_pos, "avail_fcf_years": avail,
                "rev_cagr_3yr": rev_cagr}

    return {"gate_pass": True, "gate_code": "", "gate_reason": "",
            "gross_margin": gm, "fcf_pos_years": fcf_pos, "avail_fcf_years": avail,
            "rev_cagr_3yr": rev_cagr}


def screen_group_standard(constituents_df, info_map, stmt_map, save_csv_fn=None):
    """
    Run Gates 1–4 for all tickers in a standard (non-Nasdaq) group.
    Returns (passers_df, exclusions_df, gate_data_dict).
    gate_data_dict: {ticker: gate_result_dict}
    """
    passers    = []
    exclusions = []
    gate_data  = {}

    for _, row in constituents_df.iterrows():
        sym = row["ticker"]
        info = info_map.get(sym)
        if info is None:
            exclusions.append({**row, "gate_code": "TECHNICAL_SOURCE_FAILURE",
                                "gate_reason": "info fetch failed"})
            continue

        stmts = stmt_map.get(sym, {})
        income_stmt   = stmts.get("income_stmt")
        cashflow      = stmts.get("cashflow")

        result = apply_gates_standard(sym, info, income_stmt, cashflow)
        gate_data[sym] = result

        if result["gate_pass"] is True:
            passers.append(row)
        elif result["gate_pass"] is None:
            exclusions.append({**row, "gate_code": result["gate_code"],
                                "gate_reason": result["gate_reason"]})
        else:
            exclusions.append({**row, "gate_code": result["gate_code"],
                                "gate_reason": result["gate_reason"]})

    passers_df    = pd.DataFrame(passers).reset_index(drop=True)
    exclusions_df = pd.DataFrame(exclusions).reset_index(drop=True)
    # Attach yfinance sector/industry to every exclusion row (the bare constituent row
    # carries neither for most indices). info_map holds the values used during gating.
    if not exclusions_df.empty:
        exclusions_df["sector"]   = exclusions_df["ticker"].map(lambda t: (info_map.get(t, {}) or {}).get("sector", "") or "")
        exclusions_df["industry"] = exclusions_df["ticker"].map(lambda t: (info_map.get(t, {}) or {}).get("industry", "") or "")
    return passers_df, exclusions_df, gate_data


def screen_group_nasdaq(constituents_df, outputs_dir, group, run_date):
    """
    3-phase Nasdaq gate execution.
    Phase 1: info fetch → Gate 1 (sector) + MktCap >= $2bn
    Phase 2: statements → Gate 2 (GM 40%) + Gate 3 (FCF) + Gate 4 (RevCAGR)
    Phase 3: scoring data fetch (handled outside this function)
    Returns (passers_df, exclusions_df, gate_data_dict, info_map, stmt_map)
    """
    tickers = constituents_df["ticker"].tolist()

    # Phase 1
    log.info(f"NASDAQ Phase 1: info fetch for {len(tickers)} tickers")
    info_map, info_errors = fetch_phase1_info(tickers, group)
    phase1_passers = []
    phase1_excl    = []
    gate_data      = {}

    for _, row in constituents_df.iterrows():
        sym  = row["ticker"]
        info = info_map.get(sym)
        if info is None:
            phase1_excl.append({**row, "gate_code": "TECHNICAL_SOURCE_FAILURE",
                                 "gate_reason": "info fetch failed", "phase": 1})
            continue
        g1, reason1, code1 = gate1_pass(info, nasdaq_mode=True)
        if not g1:
            phase1_excl.append({**row, "gate_code": code1, "gate_reason": reason1, "phase": 1})
            gate_data[sym] = {"gate_pass": False, "gate_code": code1, "gate_reason": reason1}
        else:
            phase1_passers.append(row)

    survivors1_df = pd.DataFrame(phase1_passers).reset_index(drop=True)
    log.info(f"NASDAQ Phase 1 complete: {len(survivors1_df)} survivors, {len(phase1_excl)} excluded")

    # Save phase 1 survivors
    if outputs_dir:
        phase1_csv = os.path.join(outputs_dir, f"{run_date}_NASDAQ_phase1_survivors.csv")
        survivors1_df.to_csv(phase1_csv, index=False)

    # Phase 2 — statements for survivors
    log.info(f"NASDAQ Phase 2: statements for {len(survivors1_df)} tickers")
    stmt_tickers = survivors1_df["ticker"].tolist()
    stmt_map, stmt_errors = fetch_phase2_statements(stmt_tickers, group)

    phase2_passers = []
    phase2_excl    = []

    for _, row in survivors1_df.iterrows():
        sym   = row["ticker"]
        stmts = stmt_map.get(sym, {})
        income_stmt = stmts.get("income_stmt")
        cashflow    = stmts.get("cashflow")

        # Gate 2 (Nasdaq: 40% GM)
        g2, reason2, code2, gm = gate2_pass(income_stmt, info_map.get(sym, {}), nasdaq_mode=True, cashflow=cashflow, ticker=sym)
        if g2 is None:
            phase2_excl.append({**row, "gate_code": code2, "gate_reason": reason2, "phase": 2})
            gate_data[sym] = {"gate_pass": None, "gate_code": code2, "gate_reason": reason2}
            continue
        if not g2:
            phase2_excl.append({**row, "gate_code": code2, "gate_reason": reason2,
                                 "gross_margin": gm, "phase": 2})
            gate_data[sym] = {"gate_pass": False, "gate_code": code2, "gate_reason": reason2, "gross_margin": gm}
            continue

        # Gate 3
        g3, reason3, code3, fcf_pos, avail = gate3_pass(cashflow)
        if g3 is None:
            phase2_excl.append({**row, "gate_code": code3, "gate_reason": reason3, "phase": 2})
            gate_data[sym] = {"gate_pass": None, "gate_code": code3, "gate_reason": reason3}
            continue
        if not g3:
            phase2_excl.append({**row, "gate_code": code3, "gate_reason": reason3,
                                 "fcf_pos_years": fcf_pos, "phase": 2})
            gate_data[sym] = {"gate_pass": False, "gate_code": code3, "gate_reason": reason3,
                               "fcf_pos_years": fcf_pos}
            continue

        # Gate 4
        _sector_n   = (info_map.get(sym, {}).get("sector",   "") or "")
        _industry_n = (info_map.get(sym, {}).get("industry", "") or "")
        _bucket_n   = classify_sector_bucket(_sector_n, _industry_n)
        g4, reason4, code4, rev_cagr = gate4_pass(income_stmt, sector_bucket=_bucket_n)
        if g4 is None:
            phase2_excl.append({**row, "gate_code": code4, "gate_reason": reason4, "phase": 2})
            gate_data[sym] = {"gate_pass": None, "gate_code": code4, "gate_reason": reason4}
            continue
        if not g4:
            phase2_excl.append({**row, "gate_code": code4, "gate_reason": reason4,
                                 "rev_cagr_3yr": rev_cagr, "phase": 2})
            gate_data[sym] = {"gate_pass": False, "gate_code": code4, "gate_reason": reason4,
                               "rev_cagr_3yr": rev_cagr}
            continue

        phase2_passers.append(row)
        gate_data[sym] = {"gate_pass": True, "gate_code": "", "gate_reason": "",
                          "gross_margin": gm, "fcf_pos_years": fcf_pos, "rev_cagr_3yr": rev_cagr}

    all_excl = phase1_excl + phase2_excl
    passers_df    = pd.DataFrame(phase2_passers).reset_index(drop=True)
    exclusions_df = pd.DataFrame(all_excl).reset_index(drop=True)
    # Attach yfinance sector/industry to every exclusion row (constituent feed has no industry
    # and we prefer the yfinance sector used during gating).
    if not exclusions_df.empty:
        exclusions_df["sector"]   = exclusions_df["ticker"].map(lambda t: (info_map.get(t, {}) or {}).get("sector", "") or "")
        exclusions_df["industry"] = exclusions_df["ticker"].map(lambda t: (info_map.get(t, {}) or {}).get("industry", "") or "")
    log.info(f"NASDAQ Phase 2 complete: {len(passers_df)} gate passers, {len(all_excl)} total excluded")
    return passers_df, exclusions_df, gate_data, info_map, stmt_map


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 6: SCORING ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def compute_roic(income_stmt, balance_sheet, col_idx=0):
    """
    Compute ROIC per Metric_Label_Formula_Reference Section 2.1.
    Returns (roic: float|None, label: str).
    """
    ebit = get_stmt_value(income_stmt, ["EBIT", "Ebit"], col_idx)
    tax_rate_raw = get_stmt_value(income_stmt, ["Tax Rate For Calcs"], col_idx)
    tax_rate = max(0.0, min(safe_float(tax_rate_raw) or 0.25, 0.50))

    invested_cap = get_stmt_value(balance_sheet, ["Invested Capital"], col_idx)
    label = "ROIC_DIRECT"

    if not invested_cap or invested_cap <= 0:
        # Construct from components
        equity = get_stmt_value(balance_sheet, ["Common Stock Equity", "Stockholders Equity"], col_idx)
        debt   = get_stmt_value(balance_sheet, ["Total Debt"], col_idx)
        leases = get_stmt_value(balance_sheet, ["Capital Lease Obligations", "Lease Liabilities"], col_idx) or 0
        cash   = get_stmt_value(balance_sheet, ["Cash And Cash Equivalents",
                                                "Cash Cash Equivalents And Short Term Investments"], col_idx) or 0
        sti    = get_stmt_value(balance_sheet, ["Short Term Investments"], col_idx) or 0
        if equity is not None and debt is not None:
            invested_cap = equity + debt + leases - cash - sti
            label = "ROIC_COMPONENT_CONSTRUCTED"

    if invested_cap and invested_cap > 0 and ebit is not None:
        roic = ebit * (1 - tax_rate) / invested_cap
        return roic, label
    return None, "ROIC_UNRESOLVED"


def compute_fcf_series(cashflow, max_periods=5):
    """
    Return list of (year, fcf) tuples (descending, most recent first).
    Uses direct FCF first; derives from OCF+CapEx if unavailable.
    """
    direct = get_stmt_series(cashflow, ["Free Cash Flow"], max_periods)
    if direct:
        return direct, "FCF_DIRECT"
    # Derive
    ocf_series = get_stmt_series(cashflow,
        ["Operating Cash Flow", "Cash Flow From Continuing Operating Activities",
         "Net Cash Provided By Operating Activities", "Total Cash From Operating Activities"], max_periods)
    cap_series = get_stmt_series(cashflow,
        ["Capital Expenditure", "Capital Expenditures", "CapitalExpenditure",
         "Purchase Of PPE", "Purchase of Property Plant and Equipment",
         "Investments in Property Plant and Equipment"], max_periods)
    if ocf_series and cap_series:
        ocf_d = dict(ocf_series)
        cap_d = dict(cap_series)
        # CapEx stored negative in yfinance → addition gives FCF
        derived = sorted([(y, ocf_d[y] + cap_d[y]) for y in ocf_d if y in cap_d], reverse=True)
        if derived:
            return derived, "FCF_DERIVED"
    return [], "FCF_UNRESOLVED"


def compute_net_debt(balance_sheet, col_idx=0):
    """Net Debt per Section 2.3. Returns (net_debt, label)."""
    nd = get_stmt_value(balance_sheet, ["Net Debt"], col_idx)
    if nd is not None:
        return nd, "NET_DEBT_DIRECT"
    debt = get_stmt_value(balance_sheet, ["Total Debt"], col_idx)
    cash = get_stmt_value(balance_sheet,
        ["Cash And Cash Equivalents", "Cash Cash Equivalents And Short Term Investments"], col_idx) or 0
    sti  = get_stmt_value(balance_sheet, ["Short Term Investments"], col_idx) or 0
    if debt is not None:
        return debt - cash - sti, "NET_DEBT_CONSTRUCTED"
    return None, "NET_DEBT_UNRESOLVED"


def compute_ebitda(income_stmt, info, col_idx=0):
    """EBITDA per Section 2.4. Returns (ebitda, label)."""
    ebitda = get_stmt_value(income_stmt, ["EBITDA", "Normalized EBITDA"], col_idx)
    if ebitda is not None:
        return ebitda, "EBITDA_DIRECT"
    ebitda = safe_float(info.get("ebitda"))
    if ebitda is not None:
        return ebitda, "EBITDA_INFO"
    ebit = get_stmt_value(income_stmt, ["Operating Income", "OperatingIncome", "EBIT", "Ebit"], col_idx)
    da   = get_stmt_value(income_stmt, ["Depreciation And Amortization", "Reconciled Depreciation"], col_idx)
    if ebit is not None and da is not None:
        return ebit + da, "EBITDA_CALCULATED"
    return None, "EBITDA_UNRESOLVED"


def compute_interest_coverage(income_stmt, balance_sheet, info, col_idx=0):
    """
    Interest coverage per Section 2.5.
    Returns (coverage: float|None|str, label: str).
    Net cash → returns string sentinel 'NET_CASH_NO_MATERIAL_INTEREST_BURDEN'.
    """
    net_debt, _ = compute_net_debt(balance_sheet, col_idx)
    total_debt   = get_stmt_value(balance_sheet, ["Total Debt"], col_idx) or 0

    # Net cash check
    if (net_debt is not None and net_debt < 0) or total_debt == 0:
        return "NET_CASH_NO_MATERIAL_INTEREST_BURDEN", "NET_CASH"

    ebit = None
    for lbl in ["Operating Income", "OperatingIncome", "EBIT", "Ebit"]:
        ebit = get_stmt_value(income_stmt, [lbl], col_idx)
        if ebit is not None:
            break

    interest = None
    for lbl in ["Interest Expense", "InterestExpense",
                "Interest Expense Non Operating", "Interest Expense Operating"]:
        interest = get_stmt_value(income_stmt, [lbl], col_idx)
        if interest is not None:
            break

    if ebit is not None and interest is not None and abs(interest) > 0:
        return ebit / abs(interest), "INT_COV_COMPUTED"
    return None, "INT_COV_UNRESOLVED"


def _score2(value, strong_thresh, acceptable_thresh, direction="above"):
    """Generic 2/1/0 scorer. direction='above': strong if value >= strong_thresh."""
    if value is None:
        return 0
    if direction == "above":
        if value >= strong_thresh:
            return 2
        if value >= acceptable_thresh:
            return 1
        return 0
    else:  # 'below' — lower is better
        if value <= strong_thresh:
            return 2
        if value <= acceptable_thresh:
            return 1
        return 0


def score_part_a(ticker_sym, info, income_stmt, cashflow, balance_sheet,
                 quarterly_income_stmt=None, gate_data=None):
    """
    Score Part A (14 metrics, max 28).
    Returns dict with all metric values, scores, hard gate results, and classification.
    """
    out = {"ticker": ticker_sym}
    prices = apply_pence_correction(info)
    div    = prices["pence_div"]

    # ── Sector / industry (from info — used in SUMMARY tab and email) ─────────
    out["sector"]   = info.get("sector",   "") or ""
    out["industry"] = info.get("industry", "") or ""

    # ── ROIC (hard gate) ─────────────────────────────────────────────────────
    roic, roic_label = compute_roic(income_stmt, balance_sheet)
    out["roic"] = roic
    out["roic_label"] = roic_label
    if roic is None:
        out["score_roic"]    = 0
        out["roic_hardgate"] = "UNRESOLVED_HARD_GATE_NOT_RANKABLE"
    elif roic > 0.15:
        out["score_roic"]    = 2
        out["roic_hardgate"] = "pass"
    elif roic >= 0.08:
        out["score_roic"]    = 1
        out["roic_hardgate"] = "pass"
    else:
        out["score_roic"]    = 0
        if getattr(_cfg, "RELAX_PARTA_HARDGATES", False):
            out["roic_hardgate"] = "FLAG_LOW_ROIC"          # H7: rankable with a low-quality flag, not a hard fail
            out.setdefault("quality_flags", []).append("low_roic")
        else:
            out["roic_hardgate"] = "HARD_GATE_FAIL"

    # Sanity flag
    if roic and abs(roic) > 1.0:
        out["roic_sanity_flag"] = f"ROIC_SANITY: {roic*100:.1f}% — review IC denominator"

    # ── FCF series ───────────────────────────────────────────────────────────
    fcf_series, fcf_series_label = compute_fcf_series(cashflow)
    fcf_vals = [v for _, v in fcf_series]
    avail_fcf = len([v for v in fcf_vals if v is not None])
    pos_fcf   = len([v for v in fcf_vals if v is not None and v > 0])
    out["fcf_positive_years"] = pos_fcf
    out["fcf_series_label"]   = fcf_series_label

    if avail_fcf < 3:
        out["score_fcf_pos"]    = 0
        out["fcf_hardgate"]     = "UNRESOLVED_HARD_GATE_NOT_RANKABLE"
    elif pos_fcf >= 4:
        out["score_fcf_pos"]    = 2
        out["fcf_hardgate"]     = "pass"
    elif pos_fcf == 3:
        out["score_fcf_pos"]    = 1
        out["fcf_hardgate"]     = "pass"
    else:
        out["score_fcf_pos"]    = 0
        if getattr(_cfg, "RELAX_PARTA_HARDGATES", False):
            out["fcf_hardgate"] = "FLAG_LOW_FCF"            # H7: capex-driven negative FCF stays rankable (flag)
            out.setdefault("quality_flags", []).append("low_fcf_positive_years")
        else:
            out["fcf_hardgate"]     = "HARD_GATE_FAIL"

    # ── Revenue series ───────────────────────────────────────────────────────
    rev_series = get_stmt_series(income_stmt,
        ["Total Revenue", "Operating Revenue", "Revenue", "TotalRevenue"], 5)
    rev_series, rev_anomaly, rev_excl_yr = detect_rev_anomaly(rev_series)
    out["rev_anomaly_detected"] = rev_anomaly
    out["excluded_revenue_year"] = rev_excl_yr

    # Rev CAGR (use up to 5 periods for 3-4yr window)
    rev_cagr = None
    if len(rev_series) >= 4:
        rev_cagr = compute_cagr(rev_series[3][1], rev_series[0][1], 3)
        out["revenue_cagr_basis"] = "3yr"
    elif len(rev_series) >= 3:
        rev_cagr = compute_cagr(rev_series[2][1], rev_series[0][1], 2)
        out["revenue_cagr_basis"] = "2yr"
    out["rev_cagr"] = rev_cagr
    out["score_rev_cagr"] = _score2(rev_cagr, 0.15, 0.05)

    # Recent revenue growth (QoQ YoY preferred)
    recent_rev_growth = None
    recent_rev_basis  = None
    if quarterly_income_stmt is not None and not quarterly_income_stmt.empty:
        q_rev = get_stmt_series(quarterly_income_stmt,
            ["Total Revenue", "Operating Revenue", "Revenue", "TotalRevenue"], 8)
        if len(q_rev) >= 5:
            latest_q  = q_rev[0][1]
            py_q_same = q_rev[4][1]  # prior year same quarter
            if latest_q and py_q_same and py_q_same > 0:
                recent_rev_growth = (latest_q / py_q_same) - 1
                recent_rev_basis  = "quarterly_yoy"
    if recent_rev_growth is None and len(rev_series) >= 2:
        if rev_series[0][1] and rev_series[1][1] and rev_series[1][1] > 0:
            recent_rev_growth = (rev_series[0][1] / rev_series[1][1]) - 1
            recent_rev_basis  = "annual_yoy"
    if recent_rev_growth is None:
        rv = safe_float(info.get("revenueGrowth"))
        if rv is not None:
            recent_rev_growth = rv
            recent_rev_basis  = "info_revenue_growth"
    out["recent_rev_growth"] = recent_rev_growth
    out["recent_rev_basis"]  = recent_rev_basis
    out["score_recent_rev"]  = _score2(recent_rev_growth, 0.12, 0.03)

    # ── EPS CAGR ─────────────────────────────────────────────────────────────
    eps_series = get_stmt_series(income_stmt,
        ["Diluted EPS", "Diluted EPS From Continuing Operations",
         "Basic EPS", "Basic EPS From Continuing Operations"], 5)
    eps_cagr = None
    eps_cagr_math_unreliable = False
    if len(eps_series) >= 3:
        eps_end   = eps_series[0][1]
        eps_start = eps_series[min(3, len(eps_series)-1)][1]
        years_eps = eps_series[0][0] - eps_series[min(3, len(eps_series)-1)][0]
        if eps_end is not None and eps_start is not None:
            if eps_start <= 0 or eps_end <= 0:
                eps_cagr_math_unreliable = True
            else:
                eps_cagr = compute_cagr(eps_start, eps_end, max(years_eps, 1))
    out["eps_cagr"] = eps_cagr
    out["eps_cagr_math_unreliable"] = eps_cagr_math_unreliable
    if eps_cagr_math_unreliable:
        out["score_eps_cagr"] = 0
    else:
        out["score_eps_cagr"] = _score2(eps_cagr, 0.15, 0.05)

    # ── Share count change ────────────────────────────────────────────────────
    shares_series = get_stmt_series(income_stmt,
        ["Diluted Average Shares", "Weighted Average Shares Diluted", "Diluted Weighted Average Shares",
         "Basic Average Shares", "Weighted Average Shares Basic", "Basic Weighted Average Shares"], 4)
    share_chg_ann = None
    if len(shares_series) >= 3:
        s_end   = shares_series[0][1]
        s_start = shares_series[min(2, len(shares_series)-1)][1]
        years_s = max(shares_series[0][0] - shares_series[min(2, len(shares_series)-1)][0], 1)
        if s_end and s_start and s_start > 0:
            share_chg_ann = (s_end / s_start) ** (1.0 / years_s) - 1
    out["share_count_change"] = share_chg_ann
    if share_chg_ann is None:
        out["score_share_count"] = 0
    elif share_chg_ann < 0:
        out["score_share_count"] = 2  # shrinking
    elif share_chg_ann <= 0.01:
        out["score_share_count"] = 1  # 0–1% dilution
    else:
        out["score_share_count"] = 0  # > 1% dilution (strictly > 2% = weak, but 1-2% is also 0 here)

    # ── FCF CAGR ──────────────────────────────────────────────────────────────
    fcf_cagr = None
    fcf_cagr_math_unreliable = False
    if len(fcf_series) >= 3:
        fcf_end   = fcf_series[0][1]
        fcf_start = fcf_series[min(3, len(fcf_series)-1)][1]
        yrs_fcf   = fcf_series[0][0] - fcf_series[min(3, len(fcf_series)-1)][0]
        if fcf_end is not None and fcf_start is not None:
            if fcf_start <= 0 or fcf_end <= 0:
                fcf_cagr_math_unreliable = True
            else:
                fcf_cagr = compute_cagr(fcf_start, fcf_end, max(yrs_fcf, 1))
    out["fcf_cagr"] = fcf_cagr
    out["fcf_cagr_math_unreliable"] = fcf_cagr_math_unreliable
    if fcf_cagr_math_unreliable:
        out["score_fcf_cagr"] = 0
    else:
        out["score_fcf_cagr"] = _score2(fcf_cagr, 0.12, 0.03)

    # ── FCF Margin ────────────────────────────────────────────────────────────
    latest_fcf = fcf_series[0][1] if fcf_series else None
    latest_rev = rev_series[0][1] if rev_series else None
    fcf_margin = (latest_fcf / latest_rev) if (latest_fcf and latest_rev and latest_rev > 0) else None
    out["fcf_margin"] = fcf_margin
    out["score_fcf_margin"] = _score2(fcf_margin, 0.10, 0.05)

    # ── Gross Margin (preferred: computed from statements) ────────────────────
    gross_profit = get_stmt_value(income_stmt, ["Gross Profit", "GrossProfit"])
    revenue_latest = latest_rev or get_stmt_value(income_stmt,
        ["Total Revenue", "Operating Revenue", "Revenue", "TotalRevenue"])
    if gross_profit and revenue_latest and revenue_latest > 0:
        gross_margin = gross_profit / revenue_latest
        gross_margin_label = "computed"
    else:
        gross_margin = safe_float(info.get("grossMargins"))
        gross_margin_label = "info_fallback" if gross_margin else "unresolved"
    out["gross_margin"] = gross_margin
    out["gross_margin_label"] = gross_margin_label

    # ── Sector bucket classification — used for all sector-segmented scoring ──
    # (Enhancement 2B / 2C) Computed once here; reused for capex and op margin.
    sector_bucket = classify_sector_bucket(
        out.get("sector", ""), out.get("industry", ""), gross_margin,
        capex_intensity=compute_capex_intensity(income_stmt, cashflow),
        ticker=ticker_sym,
    )
    out["sector_bucket"] = sector_bucket   # stored for transparency in Excel output

    # ── Gross margin scoring — sector-segmented thresholds (Enhancement 2B) ───
    gm_strong, gm_acceptable = GROSS_MARGIN_SCORE_THRESHOLDS.get(
        sector_bucket, GROSS_MARGIN_SCORE_THRESHOLDS["default"]
    )
    if gross_margin is None:
        out["score_gross_margin"] = 0
    elif gross_margin > gm_strong:
        out["score_gross_margin"] = 2
    elif gross_margin >= gm_acceptable:
        out["score_gross_margin"] = 1
    else:
        out["score_gross_margin"] = 0

    # ── Operating Margin ──────────────────────────────────────────────────────
    op_income = get_stmt_value(income_stmt, ["Operating Income", "OperatingIncome"])
    op_margin = (op_income / revenue_latest) if (op_income and revenue_latest and revenue_latest > 0) else None
    if op_margin is None:
        op_margin = safe_float(info.get("operatingMargins"))
    out["operating_margin"] = op_margin
    # Operating margin scoring — sector-segmented thresholds (Enhancement 2C-3)
    # Prevents 0pt scoring for semiconductor equipment at cycle trough (3–8% OM)
    _om_strong, _om_acceptable = OP_MARGIN_SCORE_THRESHOLDS.get(
        sector_bucket, OP_MARGIN_SCORE_THRESHOLDS["default"]
    )
    out["score_op_margin"] = _score2(op_margin, _om_strong, _om_acceptable)

    # ── Operating Margin Trend (3yr) ──────────────────────────────────────────
    op_margins_hist = []
    for col_i in range(min(3, len(income_stmt.columns) if income_stmt is not None and not income_stmt.empty else 0)):
        rev_i = get_stmt_value(income_stmt, ["Total Revenue", "Operating Revenue", "Revenue", "TotalRevenue"], col_i)
        op_i  = get_stmt_value(income_stmt, ["Operating Income", "OperatingIncome"], col_i)
        if rev_i and op_i and rev_i > 0:
            op_margins_hist.append(op_i / rev_i)
    op_margin_trend = None
    if len(op_margins_hist) >= 2:
        op_margin_trend = op_margins_hist[0] - op_margins_hist[-1]  # positive = improving
    out["op_margin_trend"] = op_margin_trend
    if op_margin_trend is None:
        out["score_op_margin_trend"] = 0
    elif op_margin_trend > 0.02:
        out["score_op_margin_trend"] = 2
    elif op_margin_trend >= -0.02:
        out["score_op_margin_trend"] = 1
    else:
        out["score_op_margin_trend"] = 0

    # ── Net Debt / EBITDA ─────────────────────────────────────────────────────
    net_debt, nd_label = compute_net_debt(balance_sheet)
    ebitda, ebitda_label = compute_ebitda(income_stmt, info)
    nd_ebitda = None
    if net_debt is not None and ebitda and ebitda > 0:
        nd_ebitda = net_debt / ebitda
    out["net_debt_ebitda"] = nd_ebitda
    out["net_debt_ebitda_label"] = f"{nd_label}/{ebitda_label}"
    out["score_nd_ebitda"] = _score2(nd_ebitda, 1.5, 3.0, direction="below")
    # Note: net cash (net_debt < 0) → ratio negative → scores 2 (correctly)

    # ── Interest Coverage ─────────────────────────────────────────────────────
    int_cov, int_cov_label = compute_interest_coverage(income_stmt, balance_sheet, info)
    out["interest_coverage"] = None if isinstance(int_cov, str) else int_cov
    out["interest_coverage_label"] = int_cov_label if isinstance(int_cov, str) else int_cov_label
    if isinstance(int_cov, str) and int_cov == "NET_CASH_NO_MATERIAL_INTEREST_BURDEN":
        out["score_int_cov"] = 2
        out["interest_coverage_label"] = "NET_CASH_NO_MATERIAL_INTEREST_BURDEN"
    elif int_cov is None:
        out["score_int_cov"] = 0
    else:
        out["score_int_cov"] = _score2(int_cov, 8.0, 3.0)

    # ── CapEx Intensity ───────────────────────────────────────────────────────
    capex = get_stmt_value(cashflow,
        ["Capital Expenditure", "Capital Expenditures", "CapitalExpenditure",
         "Purchase Of PPE", "Purchase of Property Plant and Equipment",
         "Investments in Property Plant and Equipment"])
    capex_label = "direct"
    if capex is None:
        ocf = get_stmt_value(cashflow,
            ["Operating Cash Flow", "Cash Flow From Continuing Operating Activities",
             "Net Cash Provided By Operating Activities"])
        if ocf is not None and latest_fcf is not None:
            capex = ocf - latest_fcf  # derived: CapEx = OCF - FCF (both in £/$ units)
            capex_label = "capex_derived"
    capex_intensity = (abs(capex) / revenue_latest) if (capex is not None and revenue_latest and revenue_latest > 0) else None
    out["capex_intensity"] = capex_intensity
    out["capex_label"] = capex_label
    # CapEx intensity scoring — sector-segmented mode (Enhancement 2C-2)
    # For semiconductor_hardware and semiconductor_equipment: invert the scoring.
    # High capex (8–25%) = capacity investment signal = moat-building = 2pts.
    # For software_saas and default: unchanged — low capex = asset-light = 2pts.
    if capex_intensity is None:
        out["score_capex"] = 0
    elif CAPEX_SCORE_MODE.get(sector_bucket, "standard") == "invert":
        # Hardware/equipment inversion: productive investment range scores best
        if 0.08 <= capex_intensity <= 0.25:
            out["score_capex"] = 2   # active capacity investment — moat building
        elif capex_intensity < 0.08:
            out["score_capex"] = 1   # light capex — less differentiation
        else:
            out["score_capex"] = 0   # >25% — potentially distressed or excessive
    else:
        # Standard scoring: asset-light = best
        if capex_intensity < 0.08:
            out["score_capex"] = 2
        elif capex_intensity <= 0.15:
            out["score_capex"] = 1
        else:
            out["score_capex"] = 0

    # ── Part A total ─────────────────────────────────────────────────────────
    scored_metrics = [
        out["score_rev_cagr"], out["score_recent_rev"], out["score_eps_cagr"],
        out["score_share_count"], out["score_fcf_cagr"], out["score_fcf_margin"],
        out["score_gross_margin"], out["score_op_margin"], out["score_op_margin_trend"],
        out["score_nd_ebitda"], out["score_int_cov"], out["score_capex"],
    ]
    hard_gate_scores = [out["score_roic"], out["score_fcf_pos"]]
    part_a_score = sum(scored_metrics) + sum(hard_gate_scores)
    out["part_a_score"] = part_a_score

    # Hard gate failure overrides
    if out["roic_hardgate"] == "HARD_GATE_FAIL" or out["fcf_hardgate"] == "HARD_GATE_FAIL":
        out["part_a_status"] = "Not Growth"
        out["final_status"] = "HARD_GATE_FAIL"
    elif out["roic_hardgate"] == "UNRESOLVED_HARD_GATE_NOT_RANKABLE" or out["fcf_hardgate"] == "UNRESOLVED_HARD_GATE_NOT_RANKABLE":
        out["final_status"] = "UNRESOLVED_HARD_GATE_NOT_RANKABLE"
        out["part_a_status"] = "Unresolved"
    elif part_a_score >= PART_A_STRONG_THRESHOLD:
        out["part_a_status"] = "Strong Growth"
    elif part_a_score >= PART_A_ACCEPTABLE_MIN:
        out["part_a_status"] = "Acceptable"
    else:
        out["part_a_status"] = "Not Growth"

    # Store values needed for Part B reuse
    out["_net_debt"] = net_debt
    out["_ebitda"]   = ebitda
    out["_int_cov_raw"] = int_cov
    out["_latest_fcf"]  = latest_fcf
    out["_latest_rev"]  = latest_rev
    out["_prices"]      = prices

    return out


def score_part_b(ticker_sym, info, income_stmt, cashflow, balance_sheet,
                 quarterly_income_stmt=None):
    """
    Score Part B (13 base metrics max 26 pts; +2 conditional metrics for equipment = max 30 pts).

    Base metrics (all companies):
      2 mandatory minimums: ROIC + Net Debt/EBITDA (reused from Part A compute functions)
      1 reused scored metric: Interest Coverage
      10 new scored metrics: Forward P/E, EV/EBITDA, Price/FCF, FCF Yield, Earnings Yield,
        Price vs 52wk, Dividend Payout vs FCF, Forward EPS Growth, Target Upside, Stress Test

    Conditional metrics (semiconductor_hardware + semiconductor_equipment buckets only):
      +2 stub metrics: Book-to-Bill Trailing 2Q, Backlog/EV Ratio
      These default to 0/unresolved in automated screening runs (no yfinance source).
      Data injected via fetch_watchlist_metrics.py for watchlist-level analysis.

    Returns partial dict; _score_ticker finalises part_b_score using pa scores.
    sector_bucket is read from info via classify_sector_bucket() to determine
    whether conditional metrics are applicable.
    """
    out = {}
    prices = apply_pence_correction(info)
    div    = prices["pence_div"]

    mktcap        = safe_float(info.get("marketCap"))
    current_price = prices["current_price"]
    low_52wk      = prices["low_52wk"]
    high_52wk     = prices["high_52wk"]
    target_price  = prices["target_mean"]

    # Reused computations (same logic as Part A; cheap recompute)
    net_debt, net_debt_label = compute_net_debt(balance_sheet)
    ebitda, ebitda_label   = compute_ebitda(income_stmt, info)
    int_cov, int_cov_label = compute_interest_coverage(income_stmt, balance_sheet, info)
    fcf_series, _          = compute_fcf_series(cashflow)
    fcf_vals               = [v for _, v in fcf_series if v is not None]
    latest_fcf             = fcf_vals[0] if fcf_vals else None

    # Net Debt/EBITDA mandatory minimum
    nd_ebitda = None
    if net_debt is not None and ebitda and ebitda > 0:
        nd_ebitda = net_debt / ebitda
    out["net_debt_ebitda"] = nd_ebitda

    nd_mand_fail = (nd_ebitda is not None and nd_ebitda > 3.0)
    if nd_mand_fail and getattr(_cfg, "RELAX_ND_MANDATORY", False):
        # H8: ND/EBITDA>3 is NOT a hard fail when the leverage is comfortably SERVICEABLE (net cash, or
        # interest coverage >= threshold) -> flag instead, name stays rankable (leverage still scored, so
        # penalised). Genuinely distressed over-leverage (weak/zero coverage) still fails.
        _serviceable = ((isinstance(int_cov, str) and "NET_CASH" in int_cov)
                        or (isinstance(int_cov, (int, float))
                            and int_cov >= getattr(_cfg, "ND_SERVICEABLE_INT_COV", 4.0)))
        if _serviceable:
            nd_mand_fail = False
            out.setdefault("quality_flags", []).append("high_leverage_serviceable")

    # ── Metric 1: Forward P/E ─────────────────────────────────────────────
    fwd_pe = safe_float(info.get("forwardPE"))
    out["fwd_pe"] = fwd_pe
    if fwd_pe is None:
        out["score_b_fwd_pe"] = 0
    elif fwd_pe < 20:
        out["score_b_fwd_pe"] = 2
    elif fwd_pe <= 30:
        out["score_b_fwd_pe"] = 1
    else:
        out["score_b_fwd_pe"] = 0

    # ── Metric 2: EV/EBITDA ───────────────────────────────────────────────
    ev_ebitda = safe_float(info.get("enterpriseToEbitda"))
    out["ev_ebitda"] = ev_ebitda
    if ev_ebitda is None:
        out["score_b_ev_ebitda"] = 0
    elif ev_ebitda < 12:
        out["score_b_ev_ebitda"] = 2
    elif ev_ebitda <= 20:
        out["score_b_ev_ebitda"] = 1
    else:
        out["score_b_ev_ebitda"] = 0

    # ── Metric 3: Price/FCF ───────────────────────────────────────────────
    price_fcf = None
    if mktcap and latest_fcf and latest_fcf > 0:
        price_fcf = mktcap / latest_fcf
    out["price_fcf"] = price_fcf
    if price_fcf is None:
        out["score_b_price_fcf"] = 0
    elif price_fcf < 20:
        out["score_b_price_fcf"] = 2
    elif price_fcf <= 35:
        out["score_b_price_fcf"] = 1
    else:
        out["score_b_price_fcf"] = 0

    # ── Metric 4: FCF Yield ───────────────────────────────────────────────
    fcf_yield = None
    if mktcap and mktcap > 0 and latest_fcf is not None:
        fcf_yield = latest_fcf / mktcap
    out["fcf_yield"] = fcf_yield
    if fcf_yield is None:
        out["score_b_fcf_yield"] = 0
    elif fcf_yield > 0.05:
        out["score_b_fcf_yield"] = 2
    elif fcf_yield >= 0.03:
        out["score_b_fcf_yield"] = 1
    else:
        out["score_b_fcf_yield"] = 0

    # ── Metric 5: Earnings Yield ──────────────────────────────────────────
    earnings_yield = None
    if fwd_pe and fwd_pe > 0:
        earnings_yield = 1 / fwd_pe
    elif current_price and current_price > 0:
        trail_eps = safe_float(info.get("trailingEps"))
        if trail_eps:
            earnings_yield = trail_eps / current_price
    out["earnings_yield"] = earnings_yield
    if earnings_yield is None:
        out["score_b_earn_yield"] = 0
    elif earnings_yield > 0.06:
        out["score_b_earn_yield"] = 2
    elif earnings_yield >= 0.03:
        out["score_b_earn_yield"] = 1
    else:
        out["score_b_earn_yield"] = 0

    # ── Metric 6: Price vs 52-week range ──────────────────────────────────
    position_52wk = None
    pct_above_low = None
    if current_price and low_52wk and high_52wk and (high_52wk - low_52wk) > 0:
        position_52wk = (current_price - low_52wk) / (high_52wk - low_52wk)
    if current_price and low_52wk and low_52wk > 0:
        pct_above_low = (current_price - low_52wk) / low_52wk
    out["position_52wk"] = position_52wk
    if pct_above_low is None:
        out["score_b_52wk"] = 0
    elif pct_above_low < 0.15:
        out["score_b_52wk"] = 2
    elif pct_above_low < 0.30:
        out["score_b_52wk"] = 1
    else:
        out["score_b_52wk"] = 0

    # ── Metric 7: Interest Coverage (reused from Part A — value only) ─────
    # score_int_cov is set in score_part_a; we don't override it here.
    out["interest_coverage"] = int_cov if not isinstance(int_cov, str) else None

    # ── Metric 8: Dividend Payout vs FCF ─────────────────────────────────
    div_paid = get_stmt_value(cashflow, ["Cash Dividends Paid", "Common Stock Dividends Paid",
                                         "Dividends Paid", "Payment Of Dividends"])
    if div_paid is not None:
        div_paid = abs(div_paid)

    div_payout_fcf = None
    if div_paid is None or div_paid == 0:
        # No dividend → score Strong
        div_payout_fcf = 0.0
        out["score_b_div_payout"] = 2
        out["div_payout_label"] = "No dividend / no FCF payout burden"
    elif latest_fcf and latest_fcf > 0:
        div_payout_fcf = div_paid / latest_fcf
        if div_payout_fcf < 0.60:
            out["score_b_div_payout"] = 2
        elif div_payout_fcf <= 0.85:
            out["score_b_div_payout"] = 1
        else:
            out["score_b_div_payout"] = 0
    else:
        # Dividend paid on negative/zero FCF → score 0
        div_payout_fcf = None
        out["score_b_div_payout"] = 0
        out["div_payout_label"] = "Dividend on negative/zero FCF"
    out["div_payout_fcf"] = div_payout_fcf

    # ── Metric 9: Forward EPS Growth Proxy ───────────────────────────────
    fwd_eps_growth = None
    ge = info.get("growth_estimates")
    try:
        if ge is not None and not (hasattr(ge, "empty") and ge.empty):
            df = ge if isinstance(ge, pd.DataFrame) else pd.DataFrame(ge)
            for idx in df.index:
                lbl = str(idx).lower()
                if "next year" in lbl or "+1 year" in lbl or "1y" == lbl:
                    col = df.columns[0] if len(df.columns) > 0 else None
                    if col is not None:
                        val = safe_float(df.at[idx, col])
                        if val is not None:
                            fwd_eps_growth = val
                            break
    except Exception:
        pass
    if fwd_eps_growth is None:
        fwd_eps_growth = safe_float(info.get("earningsGrowth"))
    out["fwd_eps_growth"] = fwd_eps_growth
    if fwd_eps_growth is None:
        out["score_b_fwd_eps"] = 0
    elif fwd_eps_growth > 0.12:
        out["score_b_fwd_eps"] = 2
    elif fwd_eps_growth >= 0.05:
        out["score_b_fwd_eps"] = 1
    else:
        out["score_b_fwd_eps"] = 0

    # ── Metric 10: Price vs Fair Value / Target ───────────────────────────
    target_upside = None
    if target_price and current_price and current_price > 0:
        target_upside = (target_price - current_price) / current_price
    out["target_upside"]     = target_upside
    out["current_price"]     = current_price
    out["target_price_mean"] = target_price
    # Recalibrated (v27): don't zero a fast grower that has merely caught up to a lagging
    # target; only penalise a genuine premium to target. >=15% upside=2, -10%..15%=1, >10% above target=0.
    if target_upside is not None and target_upside >= 0.15:
        out["score_b_target_upside"] = 2
    elif target_upside is not None and target_upside >= -0.10:
        out["score_b_target_upside"] = 1
    else:
        out["score_b_target_upside"] = 0

    # ── Metric 11: Downside Stress Test ──────────────────────────────────
    stress_nd_ebitda = None
    stress_int_cov   = None
    if ebitda and ebitda > 0:
        ebitda_str = ebitda * 0.75
        if net_debt is not None and ebitda_str > 0:
            stress_nd_ebitda = net_debt / ebitda_str
        int_exp = get_stmt_value(
            income_stmt,
            ["Interest Expense", "InterestExpense",
             "Interest Expense Non Operating", "Interest Expense Operating"],
        )
        if int_exp and ebitda_str > 0:
            stress_int_cov = ebitda_str / abs(int_exp)
        elif int_cov == "NET_CASH_NO_MATERIAL_INTEREST_BURDEN":
            stress_int_cov = 99.0
    out["stress_nd_ebitda"] = stress_nd_ebitda
    out["stress_int_cov"]   = stress_int_cov

    def _ic_strong(ic):
        return ic == "NET_CASH_NO_MATERIAL_INTEREST_BURDEN" or (isinstance(ic, (int, float)) and ic > 5)

    if stress_nd_ebitda is None and stress_int_cov is None:
        out["score_b_stress"] = 0
    else:
        nd_str_val = stress_nd_ebitda if stress_nd_ebitda is not None else 0.0
        ic_str_ok  = (stress_int_cov is not None and stress_int_cov > 5.0) \
                     or (stress_int_cov is None and _ic_strong(int_cov))
        ic_acc_ok  = (stress_int_cov is not None and stress_int_cov > 3.0) \
                     or (stress_int_cov is None and _ic_strong(int_cov))
        if nd_str_val < 2.0 and ic_str_ok:
            out["score_b_stress"] = 2
        elif nd_str_val < 3.0 and ic_acc_ok:
            out["score_b_stress"] = 1
        else:
            out["score_b_stress"] = 0

    # ── Conditional Metrics 12–13: Book-to-Bill + Backlog/EV (Enhancement 2) ──
    # Only applicable for semiconductor_equipment and semiconductor_hardware.
    # In automated screening runs: always unresolved (no yfinance source).
    # Data is injected from earnings call disclosures in fetch_watchlist_metrics.py.
    # sector_bucket is sourced from info via classify_sector_bucket.
    _sector_pb   = (info.get("sector",   "") or "") if info else ""
    _industry_pb = (info.get("industry", "") or "") if info else ""
    # gross_margin not available here; use info["grossMargins"] for classification only
    _gm_pb       = safe_float(info.get("grossMargins")) if info else None
    _bucket_pb   = classify_sector_bucket(_sector_pb, _industry_pb, _gm_pb, capex_intensity=compute_capex_intensity(income_stmt, cashflow), ticker=ticker_sym)
    out["sector_bucket_pb"] = _bucket_pb
    out["b2b_applicable"]   = (_bucket_pb in B2B_APPLICABLE_BUCKETS)
    # Per-stock Part B / Total max: 22/50 base; 26/54 for semi-hardware/equipment with the
    # book-to-bill + backlog/EV conditional metrics. Read by the pre-run adapter for max-aware
    # /50-vs-/54 display + conviction brackets (single source of truth = scoring_config).
    out["part_b_max"] = _cfg.GROWTH_PART_B_MAX_EXTENDED if out["b2b_applicable"] else _cfg.GROWTH_PART_B_MAX
    out["total_max"]  = _cfg.GROWTH_PART_A_MAX + out["part_b_max"]

    # Metric 12: Book-to-Bill Trailing 2 Quarters
    # Source: earnings call disclosures — not available in automated run.
    # To inject: set info["_book_to_bill_trailing_2q"] before calling score_part_b.
    btb_val = safe_float(info.get("_book_to_bill_trailing_2q")) if info else None
    out["book_to_bill_trailing_2q"] = btb_val
    if not out["b2b_applicable"]:
        out["score_b_book_to_bill"] = None   # not applicable — excluded from scoring
        out["book_to_bill_status"]  = "not_applicable"
    elif btb_val is None:
        out["score_b_book_to_bill"] = 0
        out["book_to_bill_status"]  = "unresolved"
    elif btb_val >= BOOK_TO_BILL_SCORE_THRESHOLDS[0]:    # strong >= 1.20
        out["score_b_book_to_bill"] = 2
        out["book_to_bill_status"]  = "scored"
    elif btb_val >= BOOK_TO_BILL_SCORE_THRESHOLDS[1]:    # acceptable >= 1.00
        out["score_b_book_to_bill"] = 1
        out["book_to_bill_status"]  = "scored"
    else:
        out["score_b_book_to_bill"] = 0
        out["book_to_bill_status"]  = "scored"

    # Metric 13: Backlog / EV Ratio
    # Source: earnings disclosure (backlog) + yfinance info["enterpriseValue"] (EV).
    # To inject: set info["_backlog_ttm"] before calling score_part_b.
    backlog_val = safe_float(info.get("_backlog_ttm")) if info else None
    ev_val      = safe_float(info.get("enterpriseValue")) if info else None
    backlog_ev  = (backlog_val / ev_val) if (backlog_val and ev_val and ev_val > 0) else None
    out["backlog_ttm"]      = backlog_val
    out["backlog_ev_ratio"] = backlog_ev
    if not out["b2b_applicable"]:
        out["score_b_backlog_ev"] = None   # not applicable — excluded from scoring
        out["backlog_ev_status"]  = "not_applicable"
    elif backlog_ev is None:
        out["score_b_backlog_ev"] = 0
        out["backlog_ev_status"]  = "unresolved"
    elif backlog_ev >= BACKLOG_EV_SCORE_THRESHOLDS[0]:    # strong >= 2.00
        out["score_b_backlog_ev"] = 2
        out["backlog_ev_status"]  = "scored"
    elif backlog_ev >= BACKLOG_EV_SCORE_THRESHOLDS[1]:    # acceptable >= 1.00
        out["score_b_backlog_ev"] = 1
        out["backlog_ev_status"]  = "scored"
    else:
        out["score_b_backlog_ev"] = 0
        out["backlog_ev_status"]  = "scored"

    # ── Additional info fields ────────────────────────────────────────────
    out["analyst_rating"] = info.get("recommendationKey", "")
    out["num_analysts"]   = safe_float(info.get("numberOfAnalystOpinions"))
    # next_earnings: phase3 fetch merges it into `info` (run loop: info = {**info_map, **d}),
    # so read it directly here. (Previously read info["calendar"], which is never present on
    # the .info dict, leaving next_earnings blank for every stock.)
    out["next_earnings"] = info.get("next_earnings", "") or "Unknown"
    out["currency"] = info.get("currency", "")

    # ── Partial Part B score (10 base + up to 2 conditional) ─────────────
    # score_roic + score_nd_ebitda + score_int_cov (from Part A) added in _score_ticker
    # Conditional metrics (book_to_bill, backlog_ev) contribute 0 when unresolved,
    # None when not applicable (excluded from sum by the or 0 handling).
    _b_new = sum([
        out.get("score_b_fwd_pe", 0) or 0,
        out.get("score_b_ev_ebitda", 0) or 0,
        out.get("score_b_price_fcf", 0) or 0,
        out.get("score_b_fcf_yield", 0) or 0,
        out.get("score_b_earn_yield", 0) or 0,
        out.get("score_b_52wk", 0) or 0,
        out.get("score_b_div_payout", 0) or 0,
        out.get("score_b_fwd_eps", 0) or 0,
        out.get("score_b_target_upside", 0) or 0,
        out.get("score_b_stress", 0) or 0,
        # Conditional: only counted when applicable and scored (not None)
        out.get("score_b_book_to_bill") or 0,   # None → 0 (not applicable)
        out.get("score_b_backlog_ev")   or 0,   # None → 0 (not applicable)
    ])
    out["_part_b_new_scores_sum"] = _b_new
    out["_nd_mand_fail"]          = nd_mand_fail

    return out


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 7: OVERLAY RETRIEVAL (SUMMARY-eligible + Source Score >= floor only)
# ─────────────────────────────────────────────────────────────────────────────

def _parse_eps_revisions_robust(eps_rev_df):
    """
    Robust eps_revisions parser. Row label format varies across yfinance versions.
    Returns dict: {period: {"up": int, "down": int}} for 7d/30d/60d/90d windows.
    """
    if eps_rev_df is None or (hasattr(eps_rev_df, "empty") and eps_rev_df.empty):
        return {}

    result = {}
    UP_VARIANTS   = ["upLast7days", "up_last7days", "Up Last 7 days",
                     "upLast30days", "up_last30days", "Up Last 30 days",
                     "upLast60days", "up_last60days", "Up Last 60 days",
                     "upLast90days", "up_last90days", "Up Last 90 days"]
    DOWN_VARIANTS = ["downLast7days", "down_last7days", "Down Last 7 days",
                     "downLast30days", "down_last30days", "Down Last 30 days",
                     "downLast60days", "down_last60days", "Down Last 60 days",
                     "downLast90days", "down_last90days", "Down Last 90 days"]

    try:
        df = eps_rev_df
        index_vals = list(df.index) if hasattr(df, "index") else []

        def _find_row(variants):
            for v in variants:
                if v in index_vals:
                    return df.loc[v]
            for idx in index_vals:
                for v in variants:
                    if v.lower() in str(idx).lower():
                        return df.loc[idx]
            return None

        cols = list(df.columns) if hasattr(df, "columns") else []
        for i, period in enumerate(["7d", "30d", "60d", "90d"]):
            up_row_idx   = i * 2
            down_row_idx = i * 2 + 1
            up_val, down_val = 0, 0
            up_series   = _find_row(UP_VARIANTS[i*3:(i+1)*3])
            down_series = _find_row(DOWN_VARIANTS[i*3:(i+1)*3])
            if up_series is not None:
                up_val = int(safe_float(up_series.iloc[0] if hasattr(up_series, "iloc") else up_series) or 0)
            elif len(df) > up_row_idx:
                up_val = int(safe_float(df.iloc[up_row_idx, 0] if df.shape[1] > 0 else 0) or 0)
            if down_series is not None:
                down_val = int(safe_float(down_series.iloc[0] if hasattr(down_series, "iloc") else down_series) or 0)
            elif len(df) > down_row_idx:
                down_val = int(safe_float(df.iloc[down_row_idx, 0] if df.shape[1] > 0 else 0) or 0)
            result[period] = {"up": up_val, "down": down_val}
    except Exception as e:
        log.debug(f"eps_revisions parse error: {e}")
    return result


def overlay_estimate_revisions(ticker_sym, scoring_data, info):
    """Estimate Revisions overlay. Returns dict with est_rev_* fields."""
    out = {"est_rev_source": "yfinance", "est_rev_status": "retrieved"}

    rev_parsed = _parse_eps_revisions_robust(scoring_data.get("eps_revisions"))
    up_30   = rev_parsed.get("30d", {}).get("up", 0)
    down_30 = rev_parsed.get("30d", {}).get("down", 0)
    out["est_rev_eps_up_30d"]   = up_30
    out["est_rev_eps_down_30d"] = down_30

    consensus_trend = "neutral"
    rec_sum = scoring_data.get("recommendations_summary")
    if rec_sum is not None and not (hasattr(rec_sum, "empty") and rec_sum.empty):
        try:
            df = rec_sum
            if hasattr(df, "tail"):
                recent    = df.tail(2)
                buy_cols  = [c for c in df.columns if "buy" in str(c).lower()]
                sbuy_cols = [c for c in df.columns if "strong" in str(c).lower() and "buy" in str(c).lower()]
                if len(recent) >= 2 and (buy_cols or sbuy_cols):
                    curr_buy  = sum(recent.iloc[-1][c] for c in buy_cols + sbuy_cols if c in recent.columns and not pd.isna(recent.iloc[-1][c]))
                    prior_buy = sum(recent.iloc[-2][c] for c in buy_cols + sbuy_cols if c in recent.columns and not pd.isna(recent.iloc[-2][c]))
                    consensus_trend = "improving" if curr_buy > prior_buy else ("deteriorating" if curr_buy < prior_buy else "neutral")
        except Exception:
            pass

    out["est_rev_consensus_trend"] = consensus_trend
    net_revision = up_30 - down_30
    total_rev  = up_30 + down_30
    down_share = (down_30 / total_rev) if total_rev > 0 else 0.0
    up_share   = (up_30 / total_rev) if total_rev > 0 else 0.0
    # Jul-2026 (Raj): SIGNIFICANCE-GATED direction. A single-window net-count is only a signal when
    # it is distinguishable from noise — a margin (>=3 net) OR a clear breadth (>=65% of >=3 revisions).
    # A net of -1 on 13 estimates (MU 6/7) is a coin flip (binomial p~1.0) and must NOT force a
    # 'deteriorating' -> mandatory SELL. Symmetric bar both sides; a genuine consensus-trend cut still
    # counts. (yfinance only populates the 30d window; 60d/90d come back empty, so multi-window is n/a.)
    _det_sig = (down_30 >= up_30 + 3) or (total_rev >= 3 and down_share >= 0.65)
    _imp_sig = (up_30 >= down_30 + 3) or (total_rev >= 3 and up_share >= 0.65)
    if _det_sig or consensus_trend == "deteriorating":
        out["est_rev_direction_raw"] = "deteriorating"
    elif _imp_sig or (net_revision > 0 and consensus_trend == "improving"):
        out["est_rev_direction_raw"] = "improving"
    else:
        out["est_rev_direction_raw"] = "neutral"

    if not rev_parsed:
        out["est_rev_status"]    = "unresolved"
        out["est_rev_direction_raw"] = "neutral"
    return out


def compute_wacc(ticker_sym, info, income_stmt, balance_sheet, part_a_out):
    """WACC per Section 2.10. Beta capped at 2.5, floored at 0.5."""
    out = {}
    beta_raw = safe_float(info.get("beta"))
    mktcap   = safe_float(info.get("marketCap"))
    rfr      = rfr_for_ticker(ticker_sym)
    total_debt = get_stmt_value(balance_sheet, ["Total Debt"])
    cash       = get_stmt_value(balance_sheet, ["Cash And Cash Equivalents", "Cash Cash Equivalents And Short Term Investments"]) or 0
    sti        = get_stmt_value(balance_sheet, ["Short Term Investments"]) or 0
    int_exp    = get_stmt_value(income_stmt, ["Interest Expense", "InterestExpense", "Interest Expense Non Operating", "Interest Expense Operating"])
    tax_raw    = get_stmt_value(income_stmt, ["Tax Rate For Calcs"])
    tax_rate   = max(0.0, min(safe_float(tax_raw) or 0.25, 0.50))

    if beta_raw is None:
        out.update({"wacc_status": "beta_unresolved", "wacc_computation_basis": "unresolved",
                    "wacc_pct": None, "roic_vs_wacc_spread": None,
                    "wacc_beta_used": None, "wacc_riskfree_used": rfr,
                    "wacc_beta_floored": False, "wacc_beta_capped": False})
        return out

    beta = max(0.5, min(beta_raw, 2.5))
    out["wacc_beta_floored"] = beta_raw < 0.5
    out["wacc_beta_capped"]  = beta_raw > 2.5
    out["wacc_beta_used"]    = beta
    out["wacc_riskfree_used"] = rfr

    cost_of_equity = rfr + beta * ERP
    ev_approx = (mktcap or 0) + (total_debt or 0) - cash - sti
    eq_weight   = (mktcap or 0) / ev_approx if ev_approx > 0 else 1.0
    debt_weight = (total_debt or 0) / ev_approx if (ev_approx > 0 and total_debt) else 0.0
    cost_debt_post = (abs(int_exp) / total_debt * (1 - tax_rate)) if (total_debt and total_debt > 0 and int_exp) else 0.0

    wacc = eq_weight * cost_of_equity + debt_weight * cost_debt_post
    out["wacc_pct"]              = round(wacc * 100, 2)
    out["wacc_computation_basis"] = "yfinance_computed"
    out["wacc_status"]           = "computed"
    roic = part_a_out.get("roic")
    out["roic_vs_wacc_spread"] = round((roic - wacc) * 100, 2) if roic and wacc else None
    return out


def compute_val_hist(ticker_sym, info, scoring_data, income_stmt, cashflow):
    """
    Valuation vs Own History per Section 2.11.
    Critical: tz-naive normalisation + GBp/100 fix + elevated_base_period.
    """
    out = {"val_hist_status": "unresolved", "val_hist_basis": "unresolved",
           "val_hist_pe_status": "unresolved", "val_hist_periods_used": 0}

    price_history = scoring_data.get("history")
    if price_history is None or (hasattr(price_history, "empty") and price_history.empty):
        out["val_hist_status"] = "insufficient_history"
        return out

    gbp_div = 100 if get_market_suffix(ticker_sym) == ".L" else 1
    try:
        price_index_naive = price_history.index.tz_localize(None) if price_history.index.tz is not None else price_history.index
    except Exception:
        price_index_naive = price_history.index

    eps_series = get_stmt_series(income_stmt,
        ["Diluted EPS", "Diluted EPS From Continuing Operations",
         "Basic EPS", "Basic EPS From Continuing Operations"], 5)
    fcf_series, _ = compute_fcf_series(cashflow)
    shares_series  = get_stmt_series(income_stmt,
        ["Diluted Average Shares", "Weighted Average Shares Diluted",
         "Basic Average Shares", "Weighted Average Shares Basic"], 5)

    annual_eps = dict(eps_series); annual_fcf = dict(fcf_series); annual_shares = dict(shares_series)
    hist_pe = {}; hist_pfcf = {}
    mktcap  = safe_float(info.get("marketCap"))

    for year, eps in annual_eps.items():
        try:
            fy_end = pd.Timestamp(f"{year}-12-31").tz_localize(None)
            idx    = min(price_index_naive.searchsorted(fy_end), len(price_history) - 1)
            close  = safe_float(price_history.iloc[idx]["Close"])
            if close and eps and eps > 0:
                hist_pe[year] = (close / gbp_div) / eps
        except Exception:
            continue

    for year, fcf in annual_fcf.items():
        try:
            shares = annual_shares.get(year)
            fy_end = pd.Timestamp(f"{year}-12-31").tz_localize(None)
            idx    = min(price_index_naive.searchsorted(fy_end), len(price_history) - 1)
            close  = safe_float(price_history.iloc[idx]["Close"])
            if close and fcf and fcf > 0 and shares:
                hist_pfcf[year] = ((close / gbp_div) * shares) / fcf
        except Exception:
            continue

    valid_pe   = [v for v in list(hist_pe.values())[-3:]   if v and not math.isnan(v)]
    valid_pfcf = [v for v in list(hist_pfcf.values())[-3:] if v and not math.isnan(v)]

    if len(valid_pe) >= 2:
        pe_3yr_avg  = sum(valid_pe) / len(valid_pe)
        current_pe  = safe_float(info.get("trailingPE"))
        pe_premium  = ((current_pe / pe_3yr_avg) - 1) if (current_pe and pe_3yr_avg) else None
        out.update({"val_hist_pe_3yr_avg": round(pe_3yr_avg, 1),
                    "val_hist_current_pe": current_pe,
                    "val_hist_pe_premium_disc": round(pe_premium * 100, 1) if pe_premium is not None else None,
                    "val_hist_periods_used": len(valid_pe),
                    "val_hist_status": "computed",
                    "val_hist_basis": "yfinance_computed",
                    "val_hist_pe_status": "elevated_base_period" if pe_3yr_avg > 80 else "normal"})
    if len(valid_pfcf) >= 2:
        pfcf_3yr_avg = sum(valid_pfcf) / len(valid_pfcf)
        latest_fcf   = annual_fcf.get(max(annual_fcf.keys())) if annual_fcf else None
        curr_pfcf    = (mktcap / latest_fcf) if (mktcap and latest_fcf and latest_fcf > 0) else None
        pfcf_prem    = ((curr_pfcf / pfcf_3yr_avg) - 1) if (curr_pfcf and pfcf_3yr_avg) else None
        out.update({"val_hist_pfcf_3yr_avg": round(pfcf_3yr_avg, 1),
                    "val_hist_current_pfcf": round(curr_pfcf, 1) if curr_pfcf else None,
                    "val_hist_pfcf_premium_disc": round(pfcf_prem * 100, 1) if pfcf_prem is not None else None})
    return out


def overlay_trailing_pe(ticker_sym, info):
    """Trailing P/E per Section 6.7. No additional fetch required."""
    out = {"trailing_pe_status": "unresolved", "trailing_pe_source": "unresolved"}
    trailing_pe = safe_float(info.get("trailingPE"))
    out["trailing_pe"] = trailing_pe
    prices = apply_pence_correction(info)
    trail_eps = safe_float(info.get("trailingEps"))
    trailing_pe_computed = (prices["current_price"] / trail_eps) if (prices["current_price"] and trail_eps and trail_eps != 0) else None
    out["trailing_pe_computed"] = trailing_pe_computed

    if trailing_pe and trailing_pe_computed:
        if abs(trailing_pe - trailing_pe_computed) / abs(trailing_pe) > 0.20:
            out["trailing_pe_status"] = "pe_crosscheck_mismatch"
            out["trailing_pe_source"] = "yfinance_direct"
            return out

    if trailing_pe is not None:
        out["trailing_pe_source"] = "yfinance_direct"
        out["trailing_pe_status"] = "pe_sanity_elevated" if trailing_pe > 100 else "retrieved"
    elif trailing_pe_computed is not None:
        out["trailing_pe"] = trailing_pe_computed
        out["trailing_pe_source"] = "yfinance_computed"
        out["trailing_pe_status"] = "retrieved"
    else:
        trail_eps_v = safe_float(info.get("trailingEps"))
        out["trailing_pe_status"] = "negative_earnings" if (trail_eps_v and trail_eps_v <= 0) else "unresolved"
    return out


def run_overlays(ticker_sym, info, income_stmt, cashflow, balance_sheet,
                 scoring_data, part_a_out, geography):
    """Run all 7 overlays. Time cap enforced by caller."""
    out = {}

    # organic_rev_growth, recurring_rev_pct and peg_3yr removed 10-Jun-26 — none are
    # obtainable from yfinance (organic growth & recurring-rev % are disclosure-only;
    # a true 3yr-forward EPS CAGR needs +2y/+3y estimates yfinance does not provide).
    # See _FIX_CHECKPOINT_overlays_exclusions.md.
    out.update(overlay_estimate_revisions(ticker_sym, scoring_data, info))
    out.update(compute_wacc(ticker_sym, info, income_stmt, balance_sheet, part_a_out))
    out.update(compute_val_hist(ticker_sym, info, scoring_data, income_stmt, cashflow))
    out.update(overlay_trailing_pe(ticker_sym, info))

    unresolved = [n for n, k in [
        ("est_rev", "est_rev_status"), ("wacc", "wacc_status"),
        ("val_hist", "val_hist_status"), ("trailing_pe", "trailing_pe_status"),
    ] if out.get(k) in ("unresolved", "beta_unresolved", "insufficient_history", "insufficient_analysts")]
    out["overlay_status"]      = "partial" if unresolved else "complete"
    out["overlays_unresolved"] = ",".join(unresolved)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 8: OUTPUT FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

FIELD_MAP = [
    "ticker", "company", "sector", "industry", "index", "final_status",
    "part_a_score", "part_b_score", "total_score", "part_a_status", "part_b_status",
    "rev_cagr", "recent_rev_growth", "eps_cagr", "share_count_change",
    "fcf_positive_years", "fcf_cagr", "fcf_margin", "gross_margin",
    "gross_margin_label", "sector_bucket",    # sector classification used for segmented scoring
    "operating_margin", "op_margin_trend", "roic", "net_debt_ebitda",
    "interest_coverage", "capex_intensity",
    "score_roic", "score_fcf_pos", "score_rev_cagr", "score_recent_rev",
    "score_eps_cagr", "score_share_count", "score_fcf_cagr", "score_fcf_margin",
    "score_gross_margin", "score_op_margin", "score_op_margin_trend",
    "score_nd_ebitda", "score_int_cov", "score_capex",
    "book_to_bill_trailing_2q", "backlog_ttm", "backlog_ev_ratio",
    "b2b_applicable", "book_to_bill_status", "backlog_ev_status",
    "score_b_book_to_bill", "score_b_backlog_ev",
    "fwd_pe", "ev_ebitda", "price_fcf", "fcf_yield", "earnings_yield",
    "position_52wk", "div_payout_fcf", "fwd_eps_growth", "target_upside",
    "stress_nd_ebitda", "stress_int_cov", "current_price", "target_price_mean",
    "analyst_rating", "num_analysts", "next_earnings", "currency",
    "score_b_fwd_pe", "score_b_ev_ebitda", "score_b_price_fcf", "score_b_fcf_yield",
    "score_b_earn_yield", "score_b_52wk", "score_b_div_payout", "score_b_fwd_eps",
    "score_b_target_upside", "score_b_stress",
    # Forward axis (Part 3 §13) + per-stock max — carried into the screen output for SUMMARY + shadow
    "forward_axis_score", "revisions_score", "score_f_eps_trend", "score_f_margin_traj", "score_f_rev_est", "score_f_price_mom",
    "eps_trend_mom_pct", "margin_traj_delta_pp", "rev_est_fwd_pct", "price_mom_12_1m_pct", "total_max", "part_b_max",
    "revision_stage", "revision_runway",
    "score_b_peg", "score_b_ev_g", "score_b_pfcf_g", "score_b_est_rev",
    "est_rev_direction",
    "wacc_pct", "roic_vs_wacc_spread",
    "val_hist_pe_premium_disc", "val_hist_pfcf_premium_disc",
    "trailing_pe", "val_hist_pe_status", "overlay_status",
    # ── Fix Pack Jul-26 (A1/A2/A6) — raw val-hist anchors (FV-composite inputs, previously
    # dropped here), unified Source-Score anatomy, and the E[r] block. Stamped post-overlay
    # by the run flow via source_score.source_score_components_for_row + expected_return.
    "val_hist_pe_3yr_avg", "val_hist_current_pe", "val_hist_pfcf_3yr_avg", "val_hist_current_pfcf",
    "screen_source", "src_fwd_raw", "src_fwd_w", "src_rev_raw", "src_rev_w",
    "src_deploy_raw", "src_deploy_w", "src_qual_raw", "src_qual_w", "src_analyst_raw", "src_analyst_w",
    "implied_upside_fv", "display_target_gap", "fv_basis", "fv_conf", "source_input_missing",
    "expected_return_12_24m", "er_growth", "er_rerate", "er_yield", "er_confidence", "er_basis",
    "qualitative_commentary", "gate_code", "gate_reason",
]


def _to_row(scored, gate_d, constituent_row, overlay_d=None):
    """Convert scored dict + overlay → FIELD_MAP-aligned row."""
    row = {col: None for col in FIELD_MAP}
    row["ticker"]       = scored.get("ticker", str(constituent_row.get("ticker", "")))
    row["company"]      = str(constituent_row.get("company", ""))
    row["sector"]       = scored.get("sector", str(constituent_row.get("sector", "")))
    row["industry"]     = scored.get("industry", str(constituent_row.get("industry", "")))
    row["index"]        = str(constituent_row.get("index", ""))
    row["final_status"] = scored.get("final_status", "")
    row["part_a_score"] = scored.get("part_a_score"); row["part_b_score"] = scored.get("part_b_score")
    row["total_score"]  = scored.get("total_score")
    row["part_a_status"] = scored.get("part_a_status", ""); row["part_b_status"] = scored.get("part_b_status", "")
    for col in FIELD_MAP:
        if col in scored and row[col] is None:
            row[col] = scored[col]
    row["gate_code"]   = gate_d.get("gate_code", "")
    row["gate_reason"] = gate_d.get("gate_reason", "")
    if overlay_d:
        for oc in ["wacc_pct", "roic_vs_wacc_spread",
                   "val_hist_pe_premium_disc", "val_hist_pfcf_premium_disc",
                   "trailing_pe", "val_hist_pe_status", "overlay_status"]:
            row[oc] = overlay_d.get(oc)
    return row


def save_csv(df, path):
    df.to_csv(path, index=False)
    log.info(f"Saved: {path} ({len(df)} rows)")


def save_full_data(rows, outputs_dir, run_date, group):
    df   = pd.DataFrame(rows, columns=FIELD_MAP)
    path = os.path.join(outputs_dir, f"{run_date}_{group}_full_data.csv")
    save_csv(df, path)
    # Jul-26 Part 9a: append the point-in-time score panel (learning module). Best-effort — a logging
    # failure must never break a screen. Persistent store lives beside the scripts (synced with the repo).
    try:
        import score_panel_logger as _spl
        _store = os.path.join(os.path.dirname(os.path.abspath(__file__)), "score_panel.csv")
        _spl.log_from_full_data(df, group=group, run_date=run_date, store=_store)
    except Exception as _e:
        try:
            log.warning(f"score-panel log skipped: {_e}")
        except Exception:
            pass
    return path


def save_gate_results(passers_df, exclusions_df, outputs_dir, run_date, group):
    path = os.path.join(outputs_dir, f"{run_date}_{group}_yf_gate_results.csv")
    combined = pd.concat([passers_df.assign(gate_outcome="pass"), exclusions_df], ignore_index=True)
    save_csv(combined, path)
    return path


def save_run_qa(qa_dict, inv_analysis_dir, run_date, group, outputs_dir=None,
                tech_failures=None):
    """
    Save run QA only in outputs_dir (session-temp, auto-cleared — never written to OneDrive).
    CSV version for build_excel.py --run_qa.
    Tech failures CSV also written to outputs_dir.
    No files written to inv_analysis_dir — the Excel DIAGNOSTICS tab is the permanent record.
    """
    csv_path = None
    tf_path  = None
    if outputs_dir:
        # Key-value CSV for build_excel.py
        flat = [{"key": k, "value": str(v)} for k, v in qa_dict.items()]
        csv_path = os.path.join(outputs_dir, f"{run_date}_{group}_run_qa.csv")
        pd.DataFrame(flat).to_csv(csv_path, index=False)
        log.info(f"Saved run QA CSV: {csv_path}")

        # Tech failures CSV
        if tech_failures is not None:
            tf_path = os.path.join(outputs_dir, f"{run_date}_{group}_technical_failures.csv")
            pd.DataFrame(tech_failures or []).to_csv(tf_path, index=False)
            log.info(f"Saved tech failures: {tf_path}")

    return csv_path, tf_path


def build_gate4_sector_summary(exclusions_df):
    if exclusions_df is None or exclusions_df.empty or "gate_code" not in exclusions_df.columns:
        return {}
    gate4 = exclusions_df[exclusions_df["gate_code"] == "Gate 4"]
    if gate4.empty or "sector" not in gate4.columns:
        return {}
    counts = gate4["sector"].value_counts().to_dict()
    total  = sum(counts.values())
    counts["_total"]                  = total
    counts["_concentration_warning"]  = (max(counts.values()) / total > 0.35) if total > 0 else False
    return counts


# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9: MAIN — run_scheduled / run_intramonth / CLI
# ─────────────────────────────────────────────────────────────────────────────

def _est_rev_score(eps_rev_df):
    """Estimate-revision score (v27): 2 improving / 0 deteriorating / 1 neutral or no-data.
    Reads the current-fiscal-year ('0y') row: upLast30days vs downLast30days."""
    try:
        if eps_rev_df is None or (hasattr(eps_rev_df, "empty") and eps_rev_df.empty):
            return 1
        idx = list(getattr(eps_rev_df, "index", []))
        row = eps_rev_df.loc["0y"] if "0y" in idx else eps_rev_df.iloc[0]
        up = safe_float(row.get("upLast30days")) or 0
        dn = safe_float(row.get("downLast30days")) or 0
        return 2 if up > dn else (0 if dn > up else 1)
    except Exception:
        return 1


def _growth_for_peg(eps_cagr, fwd_eps_growth):
    """Growth denominator for growth-adjusted valuation: 3-5yr EPS CAGR primary,
    forward EPS growth fallback, capped at 50% (so hyper-growth can't excuse any multiple)."""
    g = eps_cagr if (eps_cagr is not None and eps_cagr > 0) else fwd_eps_growth
    if g is None or g <= 0:
        return None
    return min(g, 0.50)


def _band(v, lo, hi):
    """2 if v<lo, 1 if v<hi, else 0. None passes through."""
    return None if v is None else (2 if v < lo else (1 if v < hi else 0))


# ── Forward axis (redesign Part 3 §13) — ADDITIVE: separate from Part A/B; carried for the Source Score ──
def _eps_trend_momentum(eps_trend_df):
    """(+1y consensus EPS now vs 90d ago) % magnitude + 0/1/2 score. Annual row for non-US robustness."""
    try:
        if eps_trend_df is None or (hasattr(eps_trend_df, "empty") and eps_trend_df.empty):
            return None, None
        idx = list(getattr(eps_trend_df, "index", []))
        key = "+1y" if "+1y" in idx else ("0y" if "0y" in idx else (idx[-1] if idx else None))
        if key is None:
            return None, None
        row = eps_trend_df.loc[key]
        cur = safe_float(row.get("current")); old = safe_float(row.get("90daysAgo"))
        if cur is None or old is None or old == 0:
            return None, None
        mom = (cur - old) / abs(old) * 100.0
        strong, accept = _cfg.EPS_TREND_MOM_THRESHOLDS
        return round(mom, 1), (2 if mom >= strong else (1 if mom >= accept else 0))
    except Exception:
        return None, None


def _margin_trajectory_score(quarterly_income_stmt):
    """Revenue up AND operating margin up over the last 2 quarters (pricing power). 2/1/0; None if no data."""
    try:
        if quarterly_income_stmt is None or quarterly_income_stmt.empty:
            return None, None
        rev = get_stmt_series(quarterly_income_stmt, ["Total Revenue", "Operating Revenue", "Revenue", "TotalRevenue"], max_periods=3)
        oi  = get_stmt_series(quarterly_income_stmt, ["Operating Income", "OperatingIncome", "Total Operating Income As Reported"], max_periods=3)
        if len(rev) < 2 or len(oi) < 2:
            return None, None
        rn, rp = rev[0][1], rev[1][1]; on, op = oi[0][1], oi[1][1]
        if None in (rn, rp, on, op) or rn <= 0 or rp <= 0:
            return None, None
        rev_up = rn > rp
        m_delta = (on / rn) - (op / rp)
        margin_up = m_delta > 0
        return round(m_delta * 100, 1), (2 if (rev_up and margin_up) else (1 if (rev_up or margin_up) else 0))
    except Exception:
        return None, None


def _rev_estimate_score(growth_estimates, info):
    """Forward revenue growth consensus % + 0/1/2. growth_estimates (already fetched) or info fallback."""
    try:
        g = None
        if growth_estimates is not None and hasattr(growth_estimates, "index") and "+1y" in list(growth_estimates.index):
            r = growth_estimates.loc["+1y"]
            for col in ("stockTrend", "growth", "revenueGrowth"):
                try:
                    v = safe_float(r.get(col))
                except Exception:
                    v = None
                if v is not None:
                    g = v; break
        if g is None and info is not None:
            g = safe_float(info.get("revenueGrowth"))
        if g is None:
            return None, None
        gp = g * 100.0 if abs(g) < 3 else g
        strong, accept = _cfg.REV_EST_FWD_THRESHOLDS
        return round(gp, 1), (2 if gp >= strong else (1 if gp >= accept else 0))
    except Exception:
        return None, None


def _price_momentum_score(history, lookback=None, skip=None):
    """Price-momentum return % + 0/1/2. DEFAULT = 12-1 month (a 12-month window ending ~1 month ago).
    Jun-26 backtest (280 names / 42 dates): trailing 63-day (one-quarter) momentum had ZERO / mildly NEGATIVE forward
    edge (reversal-prone); 12-1m carries the real edge (top-quintile fwd return beats 3m at 3/6/12m).
    The ~1-month skip removes short-term reversal noise. Window + thresholds are config-driven
    (PRICE_MOM_LOOKBACK, PRICE_MOM_SKIP, PRICE_MOM_THRESHOLDS)."""
    try:
        if history is None or (hasattr(history, "empty") and history.empty):
            return None, None
        cols = getattr(history, "columns", [])
        close = history["Close"].dropna() if "Close" in cols else None
        if close is None or len(close) < 40:
            return None, None
        lb = int(getattr(_cfg, "PRICE_MOM_LOOKBACK", 252)) if lookback is None else lookback
        sk = int(getattr(_cfg, "PRICE_MOM_SKIP", 21)) if skip is None else skip
        n = len(close)
        end = float(close.iloc[-(sk + 1)]) if n > sk else float(close.iloc[-1])
        ref_pos = n - 1 - sk - lb
        ref = float(close.iloc[ref_pos]) if ref_pos >= 0 else float(close.iloc[0])  # short history -> degrade
        if ref <= 0:
            return None, None
        mom = (end / ref - 1) * 100.0
        strong, accept = _cfg.PRICE_MOM_THRESHOLDS
        return round(mom, 1), (2 if mom >= strong else (1 if mom >= accept else 0))
    except Exception:
        return None, None


def _revision_journey_stage(eps_trend_df):
    """Classify WHERE a rising +1y EPS estimate sits in its UPGRADE JOURNEY (revision drift decays late).
    Returns (stage_label, runway 0-2 | None) from the eps_trend 90d trajectory; degrades gracefully.
    Igniting / Accelerating = most drift ahead (runway 2); Sustained = 1; Maturing / Rolling-over = 0."""
    try:
        if eps_trend_df is None or (hasattr(eps_trend_df, "empty") and eps_trend_df.empty):
            return None, None
        idx = list(getattr(eps_trend_df, "index", []))
        key = "+1y" if "+1y" in idx else ("0y" if "0y" in idx else (idx[-1] if idx else None))
        if key is None:
            return None, None
        row = eps_trend_df.loc[key]
        cur = safe_float(row.get("current")); v30 = safe_float(row.get("30daysAgo")); v90 = safe_float(row.get("90daysAgo"))
        if cur is None or v90 is None or v90 == 0:
            return None, None
        tot = (cur - v90) / abs(v90)
        if tot <= 0.005:                       # not a meaningful upgrade journey
            return ("Flat/Down" if tot <= 0 else "Marginal"), None
        if v30 is not None and v30 != 0:
            recent_rate = (cur - v30) / abs(v30) / 30.0   # per-day, last 30d
            early_rate  = (v30 - v90) / abs(v90) / 60.0   # per-day, prior 60d
            if early_rate <= 0 and recent_rate > 0:
                return "Igniting", 2                       # just turned up
            if recent_rate >= early_rate * 1.25:
                return "Accelerating", 2                   # upgrades speeding up — runway ahead
            if recent_rate <= max(0.0, early_rate * 0.5):
                return ("Rolling over" if recent_rate <= 0 else "Maturing"), 0  # rising but late
            return "Sustained", 1
        return ("Sustained" if tot >= 0.03 else "Early/unconfirmed"), 1          # endpoints-only fallback
    except Exception:
        return None, None


def compute_forward_axis(scored, info, quarterly_income_stmt, history=None):
    """Compute forward-axis sub-signals + a 0-100 Forward score (F). Sets fields on `scored` ADDITIVELY
    (does NOT change Part A/B/total). Shared by the weekly screen and the pre-run fetch.
    NOTE: price-momentum is the ABSOLUTE 12-1m window (252d skip 21d; sector-relative is a later refinement); the
    weekly screen passes history=None until the two-pass fetch supplies it (watchlist path has 5y history)."""
    info = info or {}
    mom, scored["score_f_eps_trend"] = _eps_trend_momentum(info.get("eps_trend"))
    scored["eps_trend_mom_pct"] = mom
    mtj, scored["score_f_margin_traj"] = _margin_trajectory_score(quarterly_income_stmt)
    scored["margin_traj_delta_pp"] = mtj
    rev, scored["score_f_rev_est"] = _rev_estimate_score(info.get("growth_estimates"), info)
    scored["rev_est_fwd_pct"] = rev
    # Price momentum: thread the screen's already-fetched history (info["history"]) when not passed
    # explicitly. FIX (Jun-26): the weekly screen previously passed history=None, so price momentum
    # was silently absent for EVERY name. It is now wired in here.
    if history is None:
        history = info.get("history")
    pm, scored["score_f_price_mom"] = _price_momentum_score(history)
    scored["price_mom_12_1m_pct"] = pm
    scored["revision_stage"], scored["revision_runway"] = _revision_journey_stage(info.get("eps_trend"))
    # Stage-label gate (Jul-26): a high-conviction price/estimate label ("Accelerating"/"Igniting") is only
    # credible when est-rev direction confirms; otherwise downgrade to "Sustained" so the label can't
    # over-read (e.g. ADBE "Accelerating" while direction is only "neutral").
    if str(scored.get("est_rev_direction") or "").lower() != "improving" and scored.get("revision_stage") in ("Accelerating", "Igniting"):
        scored["revision_stage"] = "Sustained"
    # Runway cap: a rising-estimate "runway" of 2 is only credible when estimate revisions are
    # actually improving; cap at 1 when est-rev direction is neutral/deteriorating (avoids the
    # over-crediting that lifted e.g. ADBE 50->60 while its est-rev was only "neutral").
    if getattr(_cfg, "REVISION_RUNWAY_CAP", True) and scored.get("revision_runway") == 2:
        if str(scored.get("est_rev_direction") or "").lower() != "improving":
            scored["revision_runway"] = 1
    _runway_on = getattr(_cfg, "REVISION_RUNWAY_IN_F", False) and scored.get("revision_runway") is not None

    # ── Jul-26 Part 2: STRUCTURAL SPLIT ──────────────────────────────────────────────────────────
    # The forward AXIS is now PRICE + MARGIN only (the estimate-revision signals are pulled OUT into a
    # separate `revisions_score` so SOURCE_WEIGHTS can weight "forward" (0.60) and "revisions" (0.15)
    # independently without double-counting). Bucket weights: margin 0.30 / price 0.70 (price-dominant).
    bw = getattr(_cfg, "FORWARD_AXIS_BUCKET_WEIGHTS", {"margin": 0.30, "price": 0.70})
    # Price bucket: impute 0 when momentum is unmeasurable (penalise, do NOT exclude the stock).
    pm_score = scored.get("score_f_price_mom")
    pm_bucket = pm_score if pm_score is not None else 0
    buckets = [(bw.get("price", 0.70), pm_bucket / 2.0)]
    if scored.get("score_f_margin_traj") is not None:
        buckets.append((bw.get("margin", 0.30), scored["score_f_margin_traj"] / 2.0))
    wsum = sum(w for w, _ in buckets)
    scored["forward_axis_score"] = round(100.0 * sum(w * f for w, f in buckets) / wsum, 1) if wsum > 0 else None

    # Separate revisions_score (0-100): the estimate-revision JOURNEY (eps_trend, fwd rev-est,
    # scored est-rev, and the runway stage) — averaged over available sub-signals, each 0-2.
    rev_subs = [scored.get("score_f_eps_trend"), scored.get("score_f_rev_est"),
                scored.get("score_b_est_rev"), scored.get("revision_runway")]
    rev_av = [v for v in rev_subs if v is not None]
    scored["revisions_score"] = round(100.0 * sum(rev_av) / (len(rev_av) * 2.0), 1) if rev_av else None
    return scored


def _score_ticker(ticker, info, inc, cf, bal, inc_q, constituents_df):
    """Helper: score one ticker and return a FIELD_MAP-aligned row dict."""
    pa = score_part_a(ticker, info, inc, cf, bal, inc_q)
    pb = score_part_b(ticker, info, inc, cf, bal, inc_q)
    scored = {**pa, **pb}  # pb overrides pa on key conflicts (e.g. net_debt_ebitda, interest_coverage)

    # ── Part B v27 redesign: growth-adjusted valuation + estimate-revision ───────
    # Replaces the absolute Fwd P/E / EV/EBITDA / Price/FCF scores with growth-adjusted
    # (PEG-style) ones; drops the duplicate FCF Yield / Earnings Yield and the 52wk metric;
    # adds a scored estimate-revision metric. Part B = 11 metrics, max 22; Total max 50.
    _g  = _growth_for_peg(safe_float(scored.get("eps_cagr")), safe_float(scored.get("fwd_eps_growth")))
    _fp = safe_float(scored.get("fwd_pe")); _ev = safe_float(scored.get("ev_ebitda")); _pf = safe_float(scored.get("price_fcf"))
    if _g:
        _gp = _g * 100.0
        scored["score_b_peg"]    = _band(_fp / _gp, 1.0, 2.0) if (_fp and _fp > 0) else 0
        scored["score_b_ev_g"]   = _band(_ev / _gp, 1.0, 2.0) if (_ev and _ev > 0) else 0
        scored["score_b_pfcf_g"] = _band(_pf / _gp, 1.5, 3.0) if (_pf and _pf > 0) else 0
    else:
        # No positive growth measure → fall back to the original absolute-multiple scores
        scored["score_b_peg"]    = scored.get("score_b_fwd_pe", 0) or 0
        scored["score_b_ev_g"]   = scored.get("score_b_ev_ebitda", 0) or 0
        scored["score_b_pfcf_g"] = scored.get("score_b_price_fcf", 0) or 0
    scored["score_b_est_rev"]   = _est_rev_score(info.get("eps_revisions"))
    # Bug#4 fix: SINGLE canonical est_rev_direction via CONSERVATIVE MERGE of the two sources
    # (revision-count "_raw" from the fetch layer + the scored Est-Rev metric). Deteriorating if EITHER
    # flags it (capital protection); improving only when neither deteriorates and at least one improves.
    _raw = str(info.get("est_rev_direction_raw") or scored.get("est_rev_direction_raw") or "").lower()
    _scb = scored.get("score_b_est_rev")
    if _raw == "deteriorating" or _scb == 0:
        scored["est_rev_direction"] = "deteriorating"
    elif _raw == "improving" or _scb == 2:
        scored["est_rev_direction"] = "improving"
    else:
        scored["est_rev_direction"] = "neutral"
    # Deprecated metrics dropped from the Part B sum (values may remain for display)
    for _dead in ("score_b_fcf_yield", "score_b_earn_yield", "score_b_52wk"):
        scored[_dead] = None
    scored["_part_b_new_scores_sum"] = sum([
        scored["score_b_peg"], scored["score_b_ev_g"], scored["score_b_pfcf_g"],
        scored.get("score_b_div_payout", 0) or 0,
        scored.get("score_b_fwd_eps", 0) or 0,
        scored.get("score_b_target_upside", 0) or 0,
        scored.get("score_b_stress", 0) or 0,
        scored["score_b_est_rev"],
        scored.get("score_b_book_to_bill") or 0,   # conditional (semis only); None→0
        scored.get("score_b_backlog_ev") or 0,
    ])

    # ── Finalise Part B score ─────────────────────────────────────────────
    # Part B = score_roic (mandatory) + score_nd_ebitda (mandatory) +
    #          score_int_cov (reused) + 10 new score_b_* metrics
    roic_s   = scored.get("score_roic", 0) or 0
    nd_s     = scored.get("score_nd_ebitda", 0) or 0
    ic_s     = scored.get("score_int_cov", 0) or 0
    new_b    = scored.get("_part_b_new_scores_sum", 0) or 0
    part_b   = roic_s + nd_s + ic_s + new_b
    scored["part_b_score"] = part_b

    # Mandatory minimum fail (ND/EBITDA > 3x) overrides Part B status
    if scored.get("_nd_mand_fail") or scored.get("final_status") in (
            "HARD_GATE_FAIL", "MANDATORY_MINIMUM_FAIL", "UNRESOLVED_HARD_GATE_NOT_RANKABLE"):
        if scored.get("_nd_mand_fail"):
            scored["final_status"]  = "MANDATORY_MINIMUM_FAIL"
        scored["part_b_status"] = "Avoid"
    elif part_b >= PART_B_STRONG_THRESHOLD:
        scored["part_b_status"] = "Strong Buy"
    elif part_b >= PART_B_ACCEPTABLE_MIN:
        scored["part_b_status"] = "Fair / Mixed"
    else:
        scored["part_b_status"] = "Avoid"

    scored["total_score"] = (scored.get("part_a_score") or 0) + part_b

    # Forward axis (Part 3 §13) — additive, separate from Part A/B; consumed by the Source Score (rerank)
    compute_forward_axis(scored, info, inc_q)

    # ── Set final_status for ranked stocks (only failures were set above) ─────
    if not scored.get("final_status"):
        scored["final_status"] = "CANDIDATE_RANKABLE"

    # ── Fallback commentary when overlays unavailable ─────────────────────────
    # Written now from scored metrics; will be overwritten if overlays succeed.
    if not scored.get("qualitative_commentary"):
        pa_s = scored.get("part_a_score", 0) or 0
        pb_s = part_b
        tot  = scored.get("total_score", 0) or 0
        roic_v = scored.get("roic")
        rev_v  = scored.get("rev_cagr")
        gm_v   = scored.get("gross_margin")
        fcf_v  = scored.get("fcf_margin")
        nd_v   = scored.get("net_debt_ebitda")
        parts  = []
        if roic_v:  parts.append(f"ROIC {roic_v*100:.1f}%")
        if rev_v:   parts.append(f"Rev CAGR {rev_v*100:.1f}%")
        if gm_v:    parts.append(f"Gross margin {gm_v*100:.1f}%")
        if fcf_v:   parts.append(f"FCF margin {fcf_v*100:.1f}%")
        if nd_v is not None: parts.append(f"ND/EBITDA {nd_v:.1f}x")
        metrics_str = "; ".join(parts) if parts else "see scored metrics"
        sector_str  = scored.get("sector", "") or scored.get("industry", "") or "unknown sector"
        status_str  = scored.get("part_b_status", "")
        scored["qualitative_commentary"] = (
            f"Scores {pa_s}/28 Part A and {pb_s}/22 Part B (total {tot}/50). "
            f"{sector_str}. Key metrics: {metrics_str}. "
            f"Part B status: {status_str}. Overlays not available — verify valuation and estimate trends before acting."
        )


    _mask     = constituents_df["ticker"] == ticker
    const_row = (constituents_df[_mask].iloc[0].to_dict()
                 if _mask.any() else {"ticker": ticker})
    return _to_row(scored, {}, const_row)


def run_scheduled(group: str, run_date: str, outputs_dir: str, inv_analysis_dir: str):
    """
    Full scheduled run for one group.
    NASDAQ: 3-phase via screen_group_nasdaq.
    All others: standard via screen_group_standard (fetch info+stmts first).
    Returns (scored_rows: list[dict], run_qa: dict)
    """
    start_time = time.time()
    is_nasdaq  = (group == "NASDAQ")
    run_qa = {
        "group": group, "run_date": run_date,
        "start_time": datetime.utcnow().isoformat(),
        "constituent_warnings": [], "phases_completed": [],
        "gate4_sector_summary": {}, "gate4_sector_concentration_warning": False,
    }
    os.makedirs(outputs_dir, exist_ok=True)
    log.info(f"=== {group} run starting — {run_date} ===")

    # ── CONSTITUENTS ──────────────────────────────────────────────────────
    constituents_df, const_warnings = fetch_constituents(group)
    run_qa["constituent_warnings"] = const_warnings
    run_qa["clean_equity_count"]   = len(constituents_df)
    log.info(f"Constituents: {len(constituents_df)} clean equities for {group}")

    scored_rows   = []
    tech_failures = []

    if is_nasdaq:
        # ── NASDAQ 3-PHASE ────────────────────────────────────────────────
        passers_df, exclusions_df, gate_data, info_map, stmt_map = \
            screen_group_nasdaq(constituents_df, outputs_dir, group, run_date)
        run_qa["phases_completed"].append("gates_nasdaq")

        # Gate4 sector concentration check
        gate4_summary = build_gate4_sector_summary(exclusions_df)
        run_qa["gate4_sector_summary"] = gate4_summary
        if gate4_summary.get("_concentration_warning"):
            run_qa["gate4_sector_concentration_warning"] = True
            log.warning("Gate4 sector concentration >35% detected")

        # Save gate results
        save_gate_results(
            passers_df,
            exclusions_df,
            outputs_dir, run_date, group,
        )

        # Phase 3 — scoring data for gate passers
        passers = passers_df["ticker"].tolist() if not passers_df.empty else []
        phase3_results, phase3_errors = fetch_phase3_scoring(passers, group)
        run_qa["phases_completed"].append("phase3")
        run_qa["phase3_errors"]    = len(phase3_errors)

        for ticker in passers:
            d = phase3_results.get(ticker)
            if d is None:
                tech_failures.append({"ticker": ticker, "reason": "phase3_fetch_failed"})
                continue
            try:
                info  = {**(info_map.get(ticker) or {}), **d}   # merge phase1 info + phase3 extras
                inc   = stmt_map.get(ticker, {}).get("income_stmt")
                cf    = stmt_map.get(ticker, {}).get("cashflow")
                bal   = stmt_map.get(ticker, {}).get("balance_sheet")
                _inc_q_d = d.get("quarterly_income_stmt")
                inc_q = _inc_q_d if (_inc_q_d is not None and not (hasattr(_inc_q_d, "empty") and _inc_q_d.empty)) \
                        else stmt_map.get(ticker, {}).get("income_stmt_quarterly")
                row   = _score_ticker(ticker, info, inc, cf, bal, inc_q, constituents_df)
                scored_rows.append(row)
            except Exception as e:
                tech_failures.append({"ticker": ticker, "reason": f"scoring_exception:{e}"})
                log.warning(f"Scoring failed {ticker}: {e}")

    else:
        # ── STANDARD GROUPS ───────────────────────────────────────────────
        all_tickers = constituents_df["ticker"].tolist()

        # Phase 1 — info
        info_map, info_errors = fetch_phase1_info(all_tickers, group)
        run_qa["phases_completed"].append("phase1")
        run_qa["info_errors"] = len(info_errors)

        # Phase 2 — statements
        stmt_map, stmt_errors = fetch_phase2_statements(all_tickers, group)
        run_qa["phases_completed"].append("phase2")
        run_qa["stmt_errors"] = len(stmt_errors)

        # Apply gates 1-4
        passers_df, exclusions_df, gate_data = \
            screen_group_standard(constituents_df, info_map, stmt_map)
        run_qa["phases_completed"].append("gates_standard")

        # Gate4 sector concentration check
        gate4_summary = build_gate4_sector_summary(exclusions_df)
        run_qa["gate4_sector_summary"] = gate4_summary
        if gate4_summary.get("_concentration_warning"):
            run_qa["gate4_sector_concentration_warning"] = True
            log.warning("Gate4 sector concentration >35% detected")

        # Save gate results
        save_gate_results(passers_df, exclusions_df, outputs_dir, run_date, group)

        # Phase 3 — incremental scoring data
        passers = passers_df["ticker"].tolist() if not passers_df.empty else []
        phase3_results, phase3_errors = fetch_phase3_scoring(passers, group)
        run_qa["phases_completed"].append("phase3")
        run_qa["phase3_errors"] = len(phase3_errors)

        for ticker in passers:
            d = phase3_results.get(ticker)
            if d is None:
                tech_failures.append({"ticker": ticker, "reason": "phase3_fetch_failed"})
                continue
            try:
                base_info = info_map.get(ticker) or {}
                info  = {**base_info, **d}
                inc   = stmt_map.get(ticker, {}).get("income_stmt")
                cf    = stmt_map.get(ticker, {}).get("cashflow")
                bal   = stmt_map.get(ticker, {}).get("balance_sheet")
                _inc_q_d = d.get("quarterly_income_stmt")
                inc_q = _inc_q_d if (_inc_q_d is not None and not (hasattr(_inc_q_d, "empty") and _inc_q_d.empty)) \
                        else stmt_map.get(ticker, {}).get("income_stmt_quarterly")
                row   = _score_ticker(ticker, info, inc, cf, bal, inc_q, constituents_df)
                scored_rows.append(row)
            except Exception as e:
                tech_failures.append({"ticker": ticker, "reason": f"scoring_exception:{e}"})
                log.warning(f"Scoring failed {ticker}: {e}")

    run_qa["phases_completed"].append("scoring")
    run_qa["scored_count"]       = len(scored_rows)
    run_qa["tech_failure_count"] = len(tech_failures)
    log.info(f"Scored: {len(scored_rows)} | Tech failures: {len(tech_failures)}")

    # ── OVERLAYS (Jul-26 Part 5: SUMMARY-eligible AND Source Score >= floor, 8-min cap) ──
    # Overlays are only shown for SUMMARY names, so gate on the SUMMARY-eligibility set + the single
    # Source Score (>= SUMMARY_SOURCE_FLOOR) instead of the old fixed total-score cut, which missed
    # high-forward / low-total names (e.g. MU: total 32 / source ~80).
    # Fix Pack A6 (12-Jul-26): the score here is the UNIFIED source (deployability = FV-composite
    # upside x conf; analyst real). At this pre-overlay point the FV composite runs on the
    # analyst-target basis (val_hist arrives WITH the overlay fetch); the FINAL screen_source is
    # re-stamped post-overlay below — "Source final before overlays" no longer holds by design.
    import source_score as _ss
    _ov_floor = getattr(_cfg, "SUMMARY_SOURCE_FLOOR", 70.0)
    high_score_tickers = [
        r["ticker"] for r in scored_rows
        if _ss.summary_eligible(r) and _ss.source_score_for_row(r) >= _ov_floor
    ]
    log.info(f"High-score overlay candidates: {len(high_score_tickers)}")

    if high_score_tickers:
        hs_results, _ = fetch_phase3_scoring(high_score_tickers, group,
                                             high_score_tickers=high_score_tickers)
        overlay_start = time.time()
        for row in scored_rows:
            ticker = row["ticker"]
            if ticker not in high_score_tickers:
                continue
            if time.time() - overlay_start > 480:
                log.warning("Overlay 8-min cap reached — remaining skipped")
                for r2 in scored_rows:
                    if r2["ticker"] in high_score_tickers and not r2.get("overlay_status"):
                        r2["overlay_status"] = "session_limit_reached"
                break
            d      = hs_results.get(ticker) or {}
            b_info = info_map.get(ticker) or {} if not is_nasdaq else {}
            info   = {**b_info, **d}
            inc    = stmt_map.get(ticker, {}).get("income_stmt")
            cf     = stmt_map.get(ticker, {}).get("cashflow")
            bal    = stmt_map.get(ticker, {}).get("balance_sheet")
            pa_out = {k: v for k, v in row.items() if k.startswith("score_") or k == "roic"}
            geo    = geography_group(ticker)
            try:
                ovl = run_overlays(ticker, info, inc, cf, bal, d, pa_out, geo)
                row.update(ovl)
            except Exception as e:
                log.warning(f"Overlay failed {ticker}: {e}")
                row["overlay_status"] = f"error:{e}"

    run_qa["phases_completed"].append("overlays")

    # ── Fix Pack Jul-26 (A1/A2/A6): stamp unified Source anatomy + E[r] on EVERY scored row ──
    # AFTER overlays so the FV-composite inputs (val_hist_*) are present for the gated set. ONE
    # compute — source_score.source_score_components_for_row — build_excel / build_email /
    # score_panel read the stamped fields (or recompute via the same function: deterministic
    # parity). summary_count / SUMMARY_THIN_WARNING via THE shared select_summary (A1/D4).
    try:
        import expected_return as _er
        for _srow in scored_rows:
            try:
                _srow.update(_ss.source_score_components_for_row(_srow))
                _srow.update(_er.expected_return_for_row(_srow))
            except Exception as _se:
                _srow.setdefault("source_input_missing", f"stamp_error:{_se}")
        _sel, _sqa = _ss.select_summary(scored_rows)
        run_qa.update(_sqa)
        if _sqa.get("summary_thin_warning"):
            log.warning(f"SUMMARY_THIN_WARNING: only {_sqa['summary_count']} SUMMARY rows "
                        f"(floor {_sqa['summary_floor']})")
        log.info(f"SUMMARY (floor-based, A1): {_sqa['summary_count']} rows "
                 f"(eligible {_sqa['summary_eligible_count']}, cap {_sqa['summary_cap']})")
        run_qa["phases_completed"].append("fixpack_stamp")
    except Exception as _e:
        log.warning(f"Fix-Pack stamping failed (non-fatal, rows keep pre-stamp fields): {_e}")

    run_qa["end_time"]  = datetime.utcnow().isoformat()
    run_qa["elapsed_s"] = round(time.time() - start_time, 1)

    # ── SAVE OUTPUTS ──────────────────────────────────────────────────────
    save_full_data(scored_rows, outputs_dir, run_date, group)
    save_run_qa(run_qa, inv_analysis_dir, run_date, group,
                outputs_dir=outputs_dir, tech_failures=tech_failures)

    log.info(f"=== {group} run complete in {run_qa['elapsed_s']}s ===")
    return scored_rows, run_qa


# ─────────────────────────────────────────────────────────────────────────────
# 9b. INTRAMONTH RUN
# ─────────────────────────────────────────────────────────────────────────────

def run_intramonth(tickers: list, run_date: str, outputs_dir: str, inv_analysis_dir: str):
    """
    On-demand intramonth run for a specific list of tickers.
    Runs the full pipeline (info → statements → scoring → overlays) for given tickers.
    """
    log.info(f"=== Intramonth run: {len(tickers)} tickers, {run_date} ===")
    os.makedirs(outputs_dir, exist_ok=True)
    start_time = time.time()
    group = "INTRAMONTH"

    # Build a minimal constituents_df
    constituents_df = pd.DataFrame([{"ticker": t, "company": t, "sector": "",
                                      "industry": "", "index": "INTRAMONTH"} for t in tickers])

    scored_rows   = []
    tech_failures = []

    # Phase 1
    info_map, _ = fetch_phase1_info(tickers, group)
    # Phase 2
    stmt_map, _ = fetch_phase2_statements(tickers, group)
    # Phase 3
    phase3_results, _ = fetch_phase3_scoring(tickers, group)

    for ticker in tickers:
        d = phase3_results.get(ticker)
        if d is None:
            tech_failures.append({"ticker": ticker, "reason": "phase3_fetch_failed"})
            continue
        try:
            base_info = info_map.get(ticker) or {}
            info  = {**base_info, **d}
            inc   = stmt_map.get(ticker, {}).get("income_stmt")
            cf    = stmt_map.get(ticker, {}).get("cashflow")
            bal   = stmt_map.get(ticker, {}).get("balance_sheet")
            _inc_q_d = d.get("quarterly_income_stmt")
            inc_q = _inc_q_d if (_inc_q_d is not None and not (hasattr(_inc_q_d, "empty") and _inc_q_d.empty)) \
                    else stmt_map.get(ticker, {}).get("income_stmt_quarterly")
            row   = _score_ticker(ticker, info, inc, cf, bal, inc_q, constituents_df)
            scored_rows.append(row)
        except Exception as e:
            tech_failures.append({"ticker": ticker, "reason": f"scoring_exception:{e}"})
            log.warning(f"Intramonth scoring failed {ticker}: {e}")

    # Overlays for all tickers (intramonth runs are small — skip 8-min cap)
    if scored_rows:
        hs_results, _ = fetch_phase3_scoring(tickers, group, high_score_tickers=tickers)
        for row in scored_rows:
            ticker = row["ticker"]
            d      = hs_results.get(ticker) or {}
            base_info = info_map.get(ticker) or {}
            info   = {**base_info, **d}
            inc    = stmt_map.get(ticker, {}).get("income_stmt")
            cf     = stmt_map.get(ticker, {}).get("cashflow")
            bal    = stmt_map.get(ticker, {}).get("balance_sheet")
            pa_out = {k: v for k, v in row.items() if k.startswith("score_") or k == "roic"}
            geo    = geography_group(ticker)
            try:
                ovl = run_overlays(ticker, info, inc, cf, bal, d, pa_out, geo)
                row.update(ovl)
            except Exception as e:
                log.warning(f"Intramonth overlay failed {ticker}: {e}")

    run_qa = {
        "group": group, "run_date": run_date,
        "tickers": tickers,
        "scored_count": len(scored_rows),
        "tech_failure_count": len(tech_failures),
        "elapsed_s": round(time.time() - start_time, 1),
    }
    save_full_data(scored_rows, outputs_dir, run_date, group)
    save_run_qa(run_qa, inv_analysis_dir, run_date, group,
                outputs_dir=outputs_dir, tech_failures=tech_failures)
    log.info(f"=== Intramonth run complete in {run_qa['elapsed_s']}s ===")
    return scored_rows, run_qa


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 9c. CLI ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="ISA Growth Stock Screener")
    parser.add_argument("--group",    required=False, help="Index group (SP500, NASDAQ, MIDCAP400, STOXX600, F250-SPI, OTHER)")
    parser.add_argument("--date",     required=True,  help="Run date YYYY-MM-DD")
    parser.add_argument("--outputs",  required=True,  help="Session temp outputs directory path")
    parser.add_argument("--inv-dir",  required=True,  dest="inv_dir", help="Investment Analysis folder path")
    parser.add_argument("--ticker",   nargs="+",      help="Ticker list for intramonth mode")
    parser.add_argument("--mode",     default="scheduled", choices=["scheduled", "intramonth"])
    args = parser.parse_args()

    # Normalise date to YYYYMMDD for filenames
    run_date = args.date.replace("-", "")

    if args.mode == "intramonth" or args.ticker:
        if not args.ticker:
            parser.error("--ticker required for intramonth mode")
        run_intramonth(args.ticker, run_date, args.outputs, args.inv_dir)
    else:
        if not args.group:
            parser.error("--group required for scheduled mode")
        run_scheduled(args.group, run_date, args.outputs, args.inv_dir)


if __name__ == "__main__":
    main()
