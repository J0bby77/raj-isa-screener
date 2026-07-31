#!/usr/bin/env python3
"""
restamp_write.py — WP-M7 (29-Jul-2026). PHASE 4 of the restamp pipeline: WRITE.

THE GAP THIS CLOSES
-------------------
restamp_screener_outputs.py (phase 1) / restamp_fetch.py (phase 2) / restamp_apply.py (phase 3)
recompute the axis, fetch the missing FV inputs and re-select SUMMARY under the CURRENT config —
but phase 3 only PRINTS the new selection. The step that actually rewrites the workbook SUMMARY
tab was done ad hoc on 29-Jul and never committed, so the pipeline could not be re-run. This is
that step, made reproducible.

DESIGN — two paths, both loss-free:
  Path 1 (ticker already present in the workbook's CURRENT or _PRE_WPM SUMMARY tab): COPY the
      rendered row verbatim and overwrite ONLY the columns the reweight actually changes
      (RECOMPUTED below). Nothing else can drift, because nothing else is touched.
  Path 2 (ticker newly promoted, never in any SUMMARY tab): render from build_excel.SUMMARY_COLS
      using field values from the screen's *_full_data.csv. If no full_data exists for that group
      the row is written partial and LOGGED — never silently blank.

The header row is READ, never rewritten: each workbook has its own column count (55-71) and rows
are aligned to that workbook's own headers by NAME, not by position.

Usage:
  python3 restamp_write.py --selection /tmp/wk/dec_rows.json --outdir /tmp/wb [--inplace]
"""
from __future__ import annotations
import argparse, json, os, sys, shutil, math

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

# Columns whose value is a FUNCTION OF THE CALIBRATION and must be re-rendered every restamp.
# Anything not in this set is name-static fundamental data and is copied untouched (Path 1).
# NOTE the LEGACY ALIASES. Workbook SUMMARY headers are not uniform across screeners: F250SPI
# (55 cols, older build) says "Source Score" where the newer 67/71-col sheets say "screen_source",
# and "Target Upside" where they say "Target Gap (display)". The 29-Jul restamp missed both, so
# F250SPI kept a pre-WP-M Source Score column (CFR.SW showed 51 against a real 72.6) while its row
# ORDER came from the new scores — the worst kind of stale: internally inconsistent, not obviously
# wrong. Any new header spelling MUST be added here or the column silently keeps its old value.
RECOMPUTED = {
    "Fwd Axis (/100)", "screen_source", "Source Score", "Stage",
    "Door (B7 shadow)", "Door admit (shadow)",
    "Fwd (raw→wtd)", "Rev (raw→wtd)", "Deploy (raw→wtd)",
    "Qual (raw→wtd)", "Analyst (raw→wtd)",
    "Impl Upside (FV)", "Target Gap (display)", "Target Upside",
    "Current Price", "Target Price", "Analyst Rating", "# Analysts",
    "Momentum State", "Price Mom 1m %", "Timing Gate", "Calibration",
}


def _f(v):
    try:
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _header_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "Ticker" in cells:
            return i, cells
        if i > 8:
            break
    return None, None


def read_summary_bank(paths):
    """{ticker: {header: value}} from every SUMMARY tab given (later paths do not overwrite)."""
    import openpyxl
    bank = {}
    for p in paths:
        if not os.path.exists(p):
            continue
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        if "SUMMARY" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["SUMMARY"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        hi = None
        for i, r in enumerate(rows[:8]):
            if r and "Ticker" in [str(c).strip() if c else "" for c in r]:
                hi = i
                break
        if hi is None:
            continue
        hdr = [str(c).strip() if c is not None else "" for c in rows[hi]]
        for r in rows[hi + 1:]:
            if not r or not r[0] or not str(r[0]).strip():
                continue
            t = str(r[0]).strip()
            if t.lower().startswith(("note", "total", "—", "glossary")):
                continue
            rec = {h: (r[i] if i < len(r) else None) for i, h in enumerate(hdr) if h}
            bank.setdefault(t, rec)
    return bank


def recomputed_cells(row, fingerprint):
    """The calibration-dependent cells, rendered exactly as build_excel renders them."""
    def upside(v):
        """build_excel.upside_fmt semantics: value is a FRACTION, rendered x100 with a sign."""
        v = _f(v)
        if v is None:
            return "N/A"
        f_ = v * 100.0
        return f"{'+' if f_ >= 0 else ''}{f_:.1f}%"

    def rw(raw, wtd, nd=3):
        r_, w_ = _f(raw), _f(wtd)
        if r_ is None and w_ is None:
            return "N/A"
        rs = f"{r_:g}" if r_ is not None else "N/A"
        return f"{rs} → {w_:g}" if w_ is not None else rs

    fwd = _f(row.get("forward_axis_score"))
    src = _f(row.get("screen_source")) or _f(row.get("source_score"))
    out = {
        "Fwd Axis (/100)":      (round(fwd) if fwd is not None else None),
        "screen_source":        src,
        "Source Score":         src,            # legacy header (F250SPI)
        "Stage":                row.get("revision_stage") or "N/A",
        "Door (B7 shadow)":     row.get("door") or "",
        "Door admit (shadow)":  row.get("door_admit_shadow") or "",
        "Fwd (raw→wtd)":   rw(row.get("src_fwd_raw"), row.get("src_fwd_w")),
        "Rev (raw→wtd)":   rw(row.get("src_rev_raw"), row.get("src_rev_w")),
        "Deploy (raw→wtd)": rw(row.get("src_deploy_raw"), row.get("src_deploy_w")),
        "Qual (raw→wtd)":  rw(row.get("src_qual_raw"), row.get("src_qual_w")),
        "Analyst (raw→wtd)": rw(row.get("src_analyst_raw"), row.get("src_analyst_w")),
        "Impl Upside (FV)":     upside(row.get("implied_upside_fv")),
        "Target Gap (display)": upside(row.get("display_target_gap")),
        "Target Upside":        upside(row.get("display_target_gap")),   # legacy header (F250SPI)
        "Current Price":        _f(row.get("current_price")),
        "Target Price":         _f(row.get("target_price_mean")),
        "Analyst Rating":       row.get("analyst_rating") or "N/A",
        "# Analysts":           _f(row.get("num_analysts")),
        "Momentum State":       row.get("momentum_state") or "N/A",
        "Price Mom 1m %":       _f(row.get("price_mom_1m_pct")),
        "Timing Gate":          row.get("timing_gate_shadow") or "N/A",
        "Calibration":          fingerprint,
    }
    return out


def render_from_full_data(ticker, hdr, fd_row, row):
    """Path 2 — build a full rendered row for a newly promoted name via build_excel's own spec."""
    import build_excel as be
    fields = dict(fd_row or {})
    fields.update({k: v for k, v in row.items() if v not in (None, "")})
    spec = {c[1]: (c[2], c[3]) for c in be.SUMMARY_COLS}
    out = {}
    for h in hdr:
        if h not in spec:
            continue
        field, fmt = spec[h]
        v = fields.get(field)
        try:
            out[h] = fmt(v)
        except Exception:
            out[h] = v if v not in (None, "") else "N/A"
    out["Ticker"] = ticker
    return out


def write_workbook(src, dst, group, sel_rows, bank, full_data, fingerprint, log):
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(src)
    ws = wb["SUMMARY"]
    hi, hdr = _header_row(ws)
    if hi is None:
        raise RuntimeError(f"{group}: no Ticker header in SUMMARY")
    hdr_row = hi + 1                      # 1-based
    first_data = hdr_row + 1

    # style template from the existing first data row (so new rows look identical)
    tmpl = {c: ws.cell(first_data, c)._style for c in range(1, len(hdr) + 1)} \
        if ws.max_row >= first_data else {}

    # clear existing data rows
    for r in range(ws.max_row, hdr_row, -1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    n_bank = n_fd = n_partial = n_filled = 0
    for i, row in enumerate(sel_rows):
        t = row["ticker"]
        rec = dict(bank.get(t) or {})
        fd = (full_data or {}).get(t)
        if rec:
            n_bank += 1
            # A bank row can itself be sparse: names promoted by an EARLIER restamp were written
            # partial (AGYS 27%, ATCO-A.ST 27%). Backfill blanks from full_data — never overwrite
            # a value the workbook already has, so this can only add information.
            if fd:
                rendered = render_from_full_data(t, hdr, fd, row)
                gaps = [h for h in hdr if h and rec.get(h) in (None, "", "N/A")
                        and rendered.get(h) not in (None, "", "N/A")]
                for h in gaps:
                    rec[h] = rendered[h]
                if gaps:
                    n_filled += 1
        else:
            rec = render_from_full_data(t, hdr, fd, row)
            if fd:
                n_fd += 1
            else:
                n_partial += 1
                log.append(f"{group}: {t} written PARTIAL (no bank row, no full_data)")
        rec.update(recomputed_cells(row, fingerprint))
        rec["Ticker"] = t
        rn = first_data + i
        for ci, h in enumerate(hdr, start=1):
            if not h:
                continue
            cell = ws.cell(rn, ci)
            if ci in tmpl:
                cell._style = tmpl[ci]
            v = rec.get(h)
            cell.value = None if v in ("", None) else v

    last = first_data + len(sel_rows) - 1
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(hdr))}{max(last, first_data)}"
    wb.save(dst)
    log.append(f"{group}: wrote {len(sel_rows)} rows (bank {n_bank}, full_data {n_fd}, "
               f"partial {n_partial}, gap-filled {n_filled})")
    return {"group": group, "rows": len(sel_rows), "from_bank": n_bank,
            "from_full_data": n_fd, "partial": n_partial, "gap_filled": n_filled}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--selection", required=True, help="JSON {group: {ticker: row}} in rank order")
    ap.add_argument("--files", required=True, help="JSON {group: workbook filename}")
    ap.add_argument("--outdir", required=True)
    ap.add_argument("--full-data", default="{}", help="JSON {group: csv path}")
    a = ap.parse_args()

    os.makedirs(a.outdir, exist_ok=True)
    sel = json.load(open(a.selection))
    files = json.load(open(a.files))
    fdmap = json.loads(a.full_data)
    try:
        import calibration_guard as cg
        fingerprint = cg.config_fingerprint()["hash"]
    except Exception as e:
        fingerprint = "UNSTAMPED"
        print(f"WARNING: fingerprint unavailable ({e})")

    log, report = [], []
    for g, fname in files.items():
        src = os.path.join(SCRIPT_DIR, fname)
        dst = os.path.join(a.outdir, fname)
        bank = read_summary_bank([src, src.replace(".xlsx", "_PRE_WPM.xlsx")])
        fd = None
        if fdmap.get(g) and os.path.exists(fdmap[g]):
            import pandas as pd
            d = pd.read_csv(fdmap[g], low_memory=False)
            key = "ticker" if "ticker" in d.columns else d.columns[0]
            fd = {str(r[key]): {k: (None if (isinstance(v, float) and math.isnan(v)) else v)
                                for k, v in r.items()} for r in d.to_dict("records")}
        rows = list(sel[g].values()) if isinstance(sel[g], dict) else sel[g]
        report.append(write_workbook(src, dst, g, rows, bank, fd, fingerprint, log))
    for line in log:
        print("  " + line)
    print(json.dumps({"fingerprint": fingerprint, "report": report}))


if __name__ == "__main__":
    main()
