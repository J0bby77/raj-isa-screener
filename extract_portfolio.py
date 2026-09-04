#!/usr/bin/env python3
"""
extract_portfolio.py  --  AJ Bell ISA Portfolio xlsx parser
Version: 1.1  |  2026-06-03

Reads the latest AJ Bell ISA Portfolio xlsx from the ISA folder and produces
a clean portfolio_data_mmm_yyyy.json. Called by monthly_isa_prerun.py.

Usage (standalone):
    python3 extract_portfolio.py [--xlsx /path/to/file.xlsx] [--out /path/to/output.json]

Output: portfolio_data_mmm_yyyy.json with full holdings, sleeve breakdown,
effective cash (with standing order adjustment), and ISA allowance info.

Stock classification rules (from AJ Bell format):
  - Stock  = ticker format "EXCHANGE:SYMBOL" e.g. NASDAQ:ADBE, NYSE:X, LSE:BGFD
             BUT NOT LSE:VUAG or LSE:SMT — see fund_tickers override below
  - Fund   = ticker format "FUND:ISIN" e.g. FUND:B2PLJM6
             OR known LSE-listed ETFs/trusts that are classified as funds for
             sleeve purposes (e.g. SMT is a closed-end trust counted as fund)
  - Cash   = "Cash GBP" row

SMT (Scottish Mortgage) is an exception: it is LSE:SMT but classified as FUND
for sleeve purposes per standing rule in Run_Context.
"""

import argparse
import glob
import json
import os
import re
import sys
from datetime import datetime

import openpyxl


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
ISA_FOLDER = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), ".."
)

# ONE HOME: extract_cash_statement (R4.4). Restating these here is what let the same rule
# live under two names in three modules.
from extract_cash_statement import (  # noqa: E402
    CASH_BUFFER_MAX, CASH_BUFFER_MIN, SO_CLEAR_WORKING_DAYS,
    STANDING_ORDER as STANDING_ORDER_MONTHLY, STANDING_ORDER_ACTIVE,
)

# Tickers classified as FUND for sleeve purposes even if listed on an exchange.
# Add any new ETF/trust that should NOT count toward the stock sleeve.
# THE single home for this set (R4.4). extract_transactions imports it; it must never
# be restated. Until 12-Aug-2026 extract_transactions held its own copy carrying IWMO and
# SGLN under a comment claiming it "mirrors" this one — it did not, and the divergence
# would have booked both into the STOCK sleeve on first purchase.
FUND_OVERRIDE_TICKERS = {
    "SMT",    # Scottish Mortgage — closed-end investment trust
    "VUAG",   # Vanguard S&P 500 ETF — UCITS ETF, fund sleeve
    "IWMO",   # iShares MSCI World Momentum — UCITS ETF, B2 satellite (target_weights WP-6)
    "SGLN",   # iShares Physical Gold — ETC, hedge bucket (target_weights WP-5)
}

# Exchanges that indicate a direct stock holding (not a fund)
STOCK_EXCHANGES = {"NASDAQ", "NYSE", "LSE", "XTRA", "EPA", "TSX"}

# Exchanges where a listing could be either a stock or a fund — check FUND_OVERRIDE
AMBIGUOUS_EXCHANGES = {"LSE"}

# Non-exchange identifier schemes the broker sometimes uses for DIRECT equity
# holdings, e.g. "(SEDOL:BKM4N88)". Funds always carry a "FUND:" prefix, so a
# SEDOL/ISIN/CUSIP-tagged line is treated as a stock unless the mapped ticker is
# in FUND_OVERRIDE_TICKERS.
IDENTIFIER_SCHEMES = {"SEDOL", "ISIN", "CUSIP"}

# Map a broker identifier (SEDOL/ISIN/symbol) -> canonical exchange ticker, so
# the stock sleeve and watchlist reconcile on the same symbol. Extend as needed.
IDENTIFIER_TICKER_MAP = {
    "BKM4N88": "PCTY",   # Paylocity Holding Corp (held under SEDOL, not NASDAQ:PCTY)
}


def _canonical_ticker(ticker) -> str:
    """Map a broker identifier to its canonical exchange ticker (upper-cased)."""
    t = str(ticker or "").strip().upper()
    return IDENTIFIER_TICKER_MAP.get(t, t)


# ---------------------------------------------------------------------------
# Standing order helpers
# ---------------------------------------------------------------------------
def _working_days_after_1st(file_dt: datetime) -> int:
    """
    Count Mon-Fri working days elapsed strictly after the 1st of the file's
    month up to and including file_dt. Returns 0 if file_dt is on or before
    the 1st.

    Examples (assuming no bank holidays):
      file_dt = 1 Jun (Mon) → 0  (S/O just paid — not cleared)
      file_dt = 2 Jun (Tue) → 1  (1 working day — not cleared)
      file_dt = 3 Jun (Wed) → 2  (2 working days — not cleared)
      file_dt = 4 Jun (Thu) → 3  (3 working days — cleared)
      file_dt = 6 Jun (Sat) → 4  (counts Mon 2, Tue 3, Wed 4, Thu 5 = 4 — cleared)
    """
    from datetime import timedelta
    the_first = file_dt.replace(day=1).date()
    target    = file_dt.date() if hasattr(file_dt, "date") else file_dt
    if target <= the_first:
        return 0
    count   = 0
    current = the_first
    while current < target:
        current += timedelta(days=1)
        if current.weekday() < 5:   # 0=Mon … 4=Fri
            count += 1
    return count


def _standing_order_adjustment(file_dt: datetime | None) -> tuple[float, bool, int | None]:
    """
    Returns (adjustment_amount, applied, working_days_elapsed).

    Logic:
      - If we cannot determine the file date, apply the adjustment
        conservatively (fail-safe: never under-count cash).
      - If fewer than SO_CLEAR_WORKING_DAYS working days have elapsed since
        the 1st, the S/O has not cleared → add STANDING_ORDER_MONTHLY.
      - Otherwise the S/O is already reflected in the stated cash → add 0.
    """
    # ⚑ 12-Aug-2026 (register ISA-0011). The standing order is PAUSED. There is no
    # unprocessed payment to credit, so the adjustment is 0 in every branch - including the
    # unknown-date branch below, whose "fail-safe: never under-count cash" is only fail-safe
    # while a payment is actually coming. With the S/O paused, adding it OVER-counts
    # deployable cash by £1,250, and it did so on any file dated within SO_CLEAR_WORKING_DAYS
    # of the 1st - which is precisely when the monthly review runs.
    if not STANDING_ORDER_ACTIVE:
        return 0.0, False, (None if file_dt is None else _working_days_after_1st(file_dt))

    if file_dt is None:
        return STANDING_ORDER_MONTHLY, True, None

    elapsed = _working_days_after_1st(file_dt)
    if elapsed < SO_CLEAR_WORKING_DAYS:
        return STANDING_ORDER_MONTHLY, True, elapsed
    else:
        return 0.0, False, elapsed


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------
def find_latest_xlsx(folder: str) -> str:
    """
    Finds the most recently dated AJ Bell ISA Portfolio xlsx in the given folder.
    Filename pattern: 'AJ Bell ISA Portfolio DD-Mmm-YY.xlsx'
    Ignores 'Example' files and the old 'portfolio-ACB8G2I-ISA May 2025.xlsx'.
    Returns the full path of the latest file.
    """
    pattern = os.path.join(folder, "AJ Bell ISA Portfolio *.xlsx")
    candidates = [
        f for f in glob.glob(pattern)
        if "example" not in os.path.basename(f).lower()
        and "x-ray" not in os.path.basename(f).lower()
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No AJ Bell ISA Portfolio xlsx found in: {folder}\n"
            f"Expected filename: 'AJ Bell ISA Portfolio DD-Mmm-YY.xlsx'"
        )

    def _parse_date(path):
        name = os.path.basename(path)
        # e.g. "AJ Bell ISA Portfolio 07-May-26.xlsx"
        m = re.search(r"(\d{2}-\w{3}-\d{2,4})\.xlsx$", name, re.IGNORECASE)
        if m:
            try:
                return datetime.strptime(m.group(1), "%d-%b-%y")
            except ValueError:
                try:
                    return datetime.strptime(m.group(1), "%d-%b-%Y")
                except ValueError:
                    pass
        return datetime.min

    return max(candidates, key=_parse_date)


# ---------------------------------------------------------------------------
# Classification helpers
# ---------------------------------------------------------------------------
def _iso_or(d):
    """'31-Aug-2026' | '2026-08-31' -> ISO. Returns the input unchanged if unparseable, so a
    comparison against it fails CLOSED (the statement is treated as not-newer) rather than open."""
    import datetime as _dt
    t = str(d or "").strip()
    for f in ("%Y-%m-%d", "%d-%b-%Y", "%d-%b-%y", "%d/%m/%Y"):
        try:
            return _dt.datetime.strptime(t[:11], f).date().isoformat()
        except ValueError:
            continue
    return t


def classify_holding(investment_name: str, ticker) -> str:
    # Doc B B2 (P2): MMF/ultra-short UCITS tickers are CASH EQUIVALENTS — counted as cash in
    # deployable-cash, the B1 reserve and drift denominators; excluded from equity buckets.
    # ⚑ MATCH ON THE LISTING BASE, NOT THE EXACT STRING (ISA-0564, 02-Sep-2026). The declared
    # list carries `CSH2.L`; the AJ Bell export names the holding
    # `Amundi Smart Overnight Ret GBP H ETF Acc (LSE:CSH2)`, which parses to `CSH2`. Exact
    # comparison therefore never matched, the money-market ETF was filed as a STOCK, and the
    # rule three lines above — "counted as cash in deployable-cash, the B1 reserve and drift
    # denominators; excluded from equity buckets" — never fired ONCE. Measured on the September
    # book: GBP 1,249.40 invisible as deployable capital, the stock sleeve reported 11.12% when
    # the equity sleeve is 10.27%, and the B2 MMF sweep could not see its own instrument.
    # ⚑ This is the THIRD instance of ISA-0549's class in this framework: one holding, two
    # spellings, the alias resolved on the way IN and not on the way OUT. The normalisation is
    # applied to BOTH sides here so the declared list can be written either way and neither
    # spelling can go unmatched again.
    try:
        import scoring_config as _sc_b2

        def _ceq_key(t):
            """Uppercased ticker with a single trailing exchange suffix removed. `.L` and `CSH2`
            are the same instrument; `BRK.B` is a share CLASS, so only a suffix that looks like
            an exchange code (1-3 letters after the final dot, on a base of 2+ chars) is cut."""
            t = str(t or "").strip().upper()
            if "." in t:
                base, _, suf = t.rpartition(".")
                if len(base) >= 2 and 1 <= len(suf) <= 3 and suf.isalpha():
                    return base
            return t

        _declared = getattr(_sc_b2, "CASH_EQUIVALENT_TICKERS", []) or []
        _keys = {_ceq_key(t) for t in _declared} | {str(t or "").strip().upper() for t in _declared}
        _t = str(ticker or "").strip().upper()
        if _t and (_t in _keys or _ceq_key(_t) in _keys):
            return "cash_equivalent"
        # the list also carries a NAME entry for the OEIC alternative, matched on the name
        if any(str(d or "").strip().upper() in (investment_name or "").upper()
               for d in _declared if " " in str(d or "")):
            return "cash_equivalent"
    except Exception:
        pass
    """
    Returns 'stock', 'fund', or 'cash'.

    investment_name: full string e.g. "Broadcom Inc (NASDAQ:AVGO)"
    ticker: the Ticker column value e.g. "AVGO" or "B2PLJM6" or None
    """
    name_upper = (investment_name or "").upper()

    # Cash row
    if "CASH GBP" in name_upper or "CASH" in name_upper:
        return "cash"

    # FUND: prefix in parentheses → always a fund
    if "FUND:" in name_upper:
        return "fund"

    # Extract exchange from parenthetical "(EXCHANGE:SYMBOL)"
    m = re.search(r"\(([A-Z]+):([A-Z0-9]+)\)", investment_name or "")
    if m:
        exchange = m.group(1).upper()
        symbol   = m.group(2).upper()

        # Check fund overrides first
        if symbol in FUND_OVERRIDE_TICKERS:
            return "fund"

        # Exchange known to list stocks
        if exchange in STOCK_EXCHANGES:
            return "stock"

        # Identifier-scheme listing (SEDOL/ISIN/CUSIP) of a direct equity.
        # Funds carry a "FUND:" prefix (handled above), so map to canonical
        # ticker and treat as a stock unless that ticker is a fund override.
        if exchange in IDENTIFIER_SCHEMES:
            mapped = IDENTIFIER_TICKER_MAP.get(symbol, symbol)
            if mapped in FUND_OVERRIDE_TICKERS:
                return "fund"
            return "stock"

    # Fallback: if ticker looks like a fund ISIN (7–9 alphanumerics, starts letter)
    t = str(ticker or "").strip()
    if re.match(r"^[A-Z][A-Z0-9]{6,8}$", t) and not re.match(r"^[A-Z]{1,5}$", t):
        return "fund"

    # Short alphabetic ticker without exchange → assume stock
    if re.match(r"^[A-Z]{1,5}$", t):
        return "stock"

    return "unknown"


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Fix Pack A22 (P2, from ISA_Monthly_Review_Framework_Fixes P3): allowance from BROKER
# transaction reconciliation — never assume the S/O schedule. A £5,000 lump sum was invisible
# in July (used = £8,750 not £3,750). Contributions are summed from actual deposit rows in
# "ISA Transaction History*.xlsx"; the S/O schedule survives only as a cross-check. When the
# files do not cover the whole tax year to the data date, allowance_reconciled=False and the
# email must print "UNRECONCILED — verify AJ Bell", never a confident figure.
# ---------------------------------------------------------------------------
from extract_cash_statement import ISA_ALLOWANCE_GBP  # ONE HOME (R4.4)  # noqa: E402
_DEPOSIT_TYPES = ("deposit", "subscription", "regular investment", "direct debit",
                  "cash in", "transfer in", "credit")


# ⚑ A22 GLOB DEFECT — fixed 05-Aug-2026, and it cost £5,000 of stated allowance.
# The original pattern was the literal `ISA Transaction History*.xlsx`. Of the three transaction
# files Raj actually keeps, it matched exactly ONE:
#     "ISA Transaction History 1st Week May 2026.xlsx"      matched  — ONE WEEK of May
#     "Transaction History 07-2026.xlsx"                    missed   — no "ISA " prefix
#     "transactionhistory 1st April 2025 to 26th July 2026.xlsx"  missed — lowercase, no spaces
# So the allowance was "broker-reconciled" against a single week of one month, and reported
# ~£3,750 used against a true £8,750 — a £5,000 Faster Payment In on 06-Apr-2026 that no input
# ever saw. The failure was silent because a non-matching glob returns [] and one matching file
# returns a plausible number; nothing anywhere asserted the files COVERED the tax year.
#
# Two changes, and the second matters more than the first:
#   1. match all three conventions, case-insensitively;
#   2. `parse_contributions` already computes coverage — it is now compared against the tax-year
#      window and `allowance_reconciled` is False whenever coverage falls short, so a partial
#      file set can no longer masquerade as a reconciliation.
_TXN_FILE_PATTERNS = ("ISA Transaction History*.xlsx", "Transaction History*.xlsx",
                      "transactionhistory*.xlsx", "*ransaction*istory*.xlsx")


def _find_transaction_files(folder):
    """Every transaction-history export in `folder`, under any of Raj's naming conventions.
    Case-insensitive, de-duplicated by real path. Temporary Office lock files (~$...) excluded."""
    import glob as _glob
    found = {}
    for pat in _TXN_FILE_PATTERNS:
        for fp in _glob.glob(os.path.join(folder, pat)):
            base = os.path.basename(fp)
            if base.startswith("~$"):
                continue
            found[os.path.realpath(fp)] = fp
    # a case-insensitive sweep catches conventions not yet invented
    try:
        for base in os.listdir(folder):
            b = base.lower()
            if b.endswith(".xlsx") and not base.startswith("~$") \
                    and "ransaction" in b and ("history" in b or "histry" in b):
                fp = os.path.join(folder, base)
                found[os.path.realpath(fp)] = fp
    except OSError:
        pass
    return list(found.values())


def _tax_year_start(ref: datetime) -> datetime:
    yr = ref.year if (ref.month, ref.day) >= (4, 6) else ref.year - 1
    return datetime(yr, 4, 6)


def parse_contributions(folder: str, data_dt: datetime | None) -> dict:
    """Sum external contributions (S/O + lump sums) for the current tax year from every
    'ISA Transaction History*.xlsx' in `folder`. Returns {allowance_used_gbp,
    contributions_detail, allowance_reconciled, coverage_note}."""
    import glob as _glob
    ref = data_dt or datetime.now()
    ty_start = _tax_year_start(ref)
    files = sorted(_find_transaction_files(folder))
    detail, seen_refs = [], set()
    cover_min = cover_max = None
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            for ws in wb.worksheets:
                headers = None
                for row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(c or "").strip().lower() for c in row]
                        continue
                    r = dict(zip(headers, row))
                    dt = r.get("date")
                    if isinstance(dt, str):
                        try:
                            dt = datetime.fromisoformat(dt[:10])
                        except ValueError:
                            dt = None
                    if not isinstance(dt, datetime):
                        continue
                    cover_min = dt if (cover_min is None or dt < cover_min) else cover_min
                    cover_max = dt if (cover_max is None or dt > cover_max) else cover_max
                    ttype = str(r.get("transaction") or "").strip().lower()
                    if not any(k in ttype for k in _DEPOSIT_TYPES):
                        continue
                    if dt < ty_start:
                        continue
                    ref_id = str(r.get("reference") or f"{fp}:{dt}:{r.get('amount (gbp)')}")
                    if ref_id in seen_refs:
                        continue        # files may overlap — dedupe on broker reference
                    seen_refs.add(ref_id)
                    amt = r.get("amount (gbp)")
                    try:
                        amt = abs(float(amt))
                    except (TypeError, ValueError):
                        continue
                    detail.append({"date": dt.strftime("%Y-%m-%d"),
                                   "type": ("S/O" if "regular" in ttype or "direct debit" in ttype
                                            else "lump_sum"),
                                   "transaction": str(r.get("transaction") or ""),
                                   "amount_gbp": round(amt, 2)})
        except Exception as e:
            detail.append({"date": None, "type": "PARSE_ERROR", "transaction": os.path.basename(fp),
                           "amount_gbp": None, "error": str(e)})
    used = round(sum(d["amount_gbp"] or 0 for d in detail if d["type"] != "PARSE_ERROR"), 2)
    # Coverage proxy: earliest row within 30d of the tax-year start (an export from the TY
    # start may simply have no activity on 6-Apr itself) AND latest row within 7d of the data
    # date AND no parse errors. Anything less -> UNRECONCILED (never a confident figure).
    covered = (bool(files) and cover_min is not None
               and (cover_min - ty_start).days <= 30 and cover_min >= ty_start - __import__("datetime").timedelta(days=3)
               and cover_max is not None
               and (ref - cover_max).days <= 7
               and not any(d["type"] == "PARSE_ERROR" for d in detail))
    note = ("reconciled from broker transactions" if covered else
            ("no transaction files found" if not files else
             f"files cover {cover_min:%d-%b-%y} to {cover_max:%d-%b-%y}" if cover_min else
             "transaction files unparseable"))
    return {"allowance_used_gbp": used if covered else None,
            "allowance_used_partial_gbp": used,
            "allowance_remaining_gbp": round(ISA_ALLOWANCE_GBP - used, 2) if covered else None,
            "contributions_detail": sorted(detail, key=lambda d: d["date"] or ""),
            "allowance_reconciled": covered,
            "coverage_note": f"UNRECONCILED — verify AJ Bell ({note})" if not covered else note}


def _contributions_golden(folder, data_dt):
    """ISA contributions, from the document that actually contains them.

    ⚑ THE STRUCTURAL DEFECT THIS REPLACES (register H13, root-caused 05-Aug-2026).
    `parse_contributions()` reads AJ Bell's **Transaction History**, which is a DEALING record:
    Purchase, Sale, Transfer In, Equalisation, Fund Class Conversion. It has **no cash-deposit
    rows of any kind**. Contributions appear only in the **Cash Statement**, a separate export.
    So the allowance was being derived from a document that structurally cannot answer the
    question — and the framework could not notice, because "found no deposits" and "there were
    no deposits" produce the identical number. The £5,000 `Faster Payment In` of 06-Apr-2026
    was invisible for four months for exactly that reason.

    The cash statement is now the GOLDEN SOURCE. The dealing record is retained as an
    independent cross-check and its disagreement is recorded rather than resolved silently —
    two derivations, both reported, per the engineering standard.

    Cash statement absent -> fall back to the dealing record and say so. Never a silent
    substitution, and never a confident figure from a source that cannot hold one.
    """
    dealing = parse_contributions(folder, data_dt)
    try:
        import extract_cash_statement as _ecs
        cs = _ecs.parse(folder=folder,
                        as_of=(data_dt.date() if hasattr(data_dt, "date") else data_dt) or None)
    except Exception as e:
        dealing["golden_source"] = "dealing_record_only"
        dealing["golden_source_note"] = (
            f"cash statement unavailable ({type(e).__name__}: {e}) — the DEALING record cannot "
            f"contain cash deposits, so this figure is a floor, not the allowance")
        return dealing
    al = cs.get("allowance")
    if not al or not cs.get("reconciled"):
        dealing["golden_source"] = "dealing_record_only"
        dealing["golden_source_note"] = (
            "cash statement present but not reconciled (" +
            "; ".join(cs.get("warnings") or ["reason unstated"]) + ")")
        dealing["cash_statement_warnings"] = cs.get("warnings")
        return dealing
    return {
        "allowance_used_gbp": al["used_gbp"],
        "allowance_remaining_gbp": al["remaining_gbp"],
        "allowance_used_partial_gbp": al["used_gbp"],
        "allowance_reconciled": True,
        "golden_source": "cash_statement",
        "coverage_note": (f"reconciled from the AJ Bell cash statement "
                          f"({', '.join(cs['source_files'])}); tax year from "
                          f"{al['tax_year_start']}"),
        "contributions_detail": [
            {"date": r["date"], "transaction": r["description"],
             "amount_gbp": r["receipt_gbp"], "type": "CONTRIBUTION",
             "reference": r["reference"]} for r in al["contribution_rows"]],
        "opening_balance_gbp": cs.get("opening_balance_gbp"),
        "closing_balance_gbp": cs.get("closing_balance_gbp"),
        "fx_rate_pct": cs.get("fx_rate_pct"),
        "invariants": cs.get("invariants"),
        # the dealing record's own view, retained so a future disagreement is visible rather
        # than lost — it will normally be LOWER, because it cannot see deposits at all
        "cross_check_dealing_record": {
            "allowance_used_gbp": dealing.get("allowance_used_gbp"),
            "allowance_reconciled": dealing.get("allowance_reconciled"),
            "coverage_note": dealing.get("coverage_note"),
            "note": "a dealing record contains no cash deposits; a LOWER figure here is "
                    "expected and is not a break"},
    }


def parse_portfolio(xlsx_path: str) -> dict:
    """
    Parses the AJ Bell ISA Portfolio xlsx and returns a structured dict.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Worksheet is empty: {xlsx_path}")

    # Identify header row (first row with 'Investment' in first cell)
    header_row = None
    data_start = 0
    for i, row in enumerate(rows):
        if row and str(row[0]).strip().lower() == "investment":
            header_row = [str(c).strip() if c else "" for c in row]
            data_start = i + 1
            break

    if header_row is None:
        raise ValueError(
            f"Could not find header row with 'Investment' column in: {xlsx_path}"
        )

    # Map column names to indices
    col = {name: idx for idx, name in enumerate(header_row)}
    required = ["Investment", "Quantity", "Price", "Value (£)", "Cost (£)", "Ticker"]
    for r in required:
        if r not in col:
            raise ValueError(f"Missing required column '{r}' in: {xlsx_path}")

    # Extract date from file
    file_date_str = None
    m = re.search(r"(\d{2}-\w{3}-\d{2,4})\.xlsx$",
                  os.path.basename(xlsx_path), re.IGNORECASE)
    if m:
        file_date_str = m.group(1)

    # Parse holdings
    stocks = []
    funds  = []
    cash_value = 0.0
    mmf_value = 0.0
    mmf_holdings = []
    data_date = None

    for row in rows[data_start:]:
        if not row or row[0] is None:
            continue

        investment = str(row[col["Investment"]]).strip()
        if not investment or investment.lower() == "investment":
            continue

        qty     = row[col["Quantity"]]
        price   = row[col["Price"]]
        value   = row[col["Value (£)"]]
        cost    = row[col["Cost (£)"]]
        ticker  = row[col["Ticker"]]

        # Change and gain
        change_gbp = row[col.get("Change (£)", -1)] if "Change (£)" in col else None
        change_pct = row[col.get("Change (%)", -1)] if "Change (%)" in col else None

        # Valuation currency
        val_ccy = row[col.get("Valuation currency", -1)] if "Valuation currency" in col else None
        fx_rate = row[col.get("Exchange rate", -1)] if "Exchange rate" in col else None

        # Date
        date_val = row[col.get("Date", -1)] if "Date" in col else None
        if date_val and hasattr(date_val, "strftime"):
            data_date = date_val.strftime("%d-%b-%Y")

        kind = classify_holding(investment, ticker)

        # Safe numeric conversion
        def safe_float(v):
            try:
                return float(v) if v is not None else None
            except (TypeError, ValueError):
                return None

        v = safe_float(value) or 0.0
        c = safe_float(cost)  or v   # if cost missing, use value
        g = safe_float(change_gbp)
        gp = safe_float(change_pct)

        if kind == "cash":
            cash_value += v
            continue
        if kind == "cash_equivalent":
            # B2: MMF value counts as CASH everywhere (deployable, B1 reserve, drift
            # denominators); tracked separately in the record list for the sweep line.
            cash_value += v
            mmf_value += v
            mmf_holdings.append({"name": investment, "ticker": _canonical_ticker(ticker),
                                 "value_gbp": v})
            continue

        # Extract clean name (remove exchange/ISIN suffix)
        clean_name = re.sub(r"\s*\([A-Z]+:[A-Z0-9]+\)\s*$", "", investment).strip()

        record = {
            "name":         clean_name,
            "full_name":    investment,
            "ticker":       _canonical_ticker(ticker),
            "quantity":     safe_float(qty),
            "price":        safe_float(price),
            "value_gbp":    v,
            "cost_gbp":     c,
            "gain_gbp":     round(v - c, 2),
            "gain_pct":     round((v - c) / c * 100, 2) if c and c != 0 else 0.0,
            "currency":     str(val_ccy).strip() if val_ccy else "GBP",
            "fx_rate":      safe_float(fx_rate) or 1.0,
            "kind":         kind,
        }

        if kind == "stock":
            stocks.append(record)
        elif kind == "fund":
            funds.append(record)

    # Portfolio totals
    total_stock = round(sum(h["value_gbp"] for h in stocks), 2)
    total_funds = round(sum(h["value_gbp"] for h in funds), 2)
    total_value = round(total_stock + total_funds + cash_value, 2)

    if total_value == 0:
        raise ValueError("Total portfolio value is zero — check xlsx format.")

    # Resolve data_dt from data_date or file_date_str (needed for S/O and month label)
    data_dt = None
    if data_date:
        try:
            data_dt = datetime.strptime(data_date, "%d-%b-%Y")
        except ValueError:
            pass
    if data_dt is None and file_date_str:
        try:
            data_dt = datetime.strptime(file_date_str, "%d-%b-%y")
        except ValueError:
            pass

    # Standing order adjustment — conditional on file date
    # The S/O is paid on the 1st of each month and typically clears within
    # SO_CLEAR_WORKING_DAYS (3) working days. If the AJ Bell file was saved
    # before the S/O cleared, we add £1,250 to stated cash. If it was saved
    # after clearing, the S/O is already reflected in the stated balance.
    so_adj, so_applied, so_elapsed = _standing_order_adjustment(data_dt)
    effective_cash  = round(cash_value + so_adj, 2)
    # ⚑ THE GBP 150 BUFFER IS RETIRED HERE (ISA-0565, 02-Sep-2026). D23 (ISA-0544, Raj,
    # 02-Sep-2026) declares the GBP 250 ISA cash reserve "the SINGLE control on residual
    # deployment". This line was a SECOND control on the same quantity, applied earlier and by
    # a different module, so the framework was actually retaining GBP 400 while reporting a
    # GBP 250 reserve. That is the FC-D class (R4.4) that `consistency_check.
    # pair_cash_reserve_is_single_topup_control` exists to prevent — and it missed this one,
    # because it enforces the retirement of the constant NAMED in D23 (`min_topup_gbp`) rather
    # than the RULE D23 states. The pair has been widened accordingly.
    # ⚑ Deployable is now the effective cash in full. The reserve is applied ONCE, downstream,
    # by position_sizing.allocate() reading target_weights.stock_sleeve.cash_reserve_gbp.
    deployable_cash = effective_cash
    # ══════════════════════════════════════════════════════════════════════════════════════
    # DEPLOYABLE CASH COMES FROM THE CASH STATEMENT (Raj, 02-Sep-2026 — ISA-0574)
    # ══════════════════════════════════════════════════════════════════════════════════════
    # ⚑ TWO BASES, ON PURPOSE, AND THEY ARE NOT IN CONFLICT. Raj's convention: the portfolio and
    # X-Ray are the PREVIOUS MONTH-END, because they are the EXPOSURE basis — co-dating them is
    # what keeps XR1, the 6.05 two-derivation fund reconciliation and the look-through comparing
    # the same book, and what makes month-on-month comparison mean anything. Cash is the CAPITAL
    # basis and must be current, because it is the only input that decides how much money is
    # actually on the table. A fund reports NAV at month-end and knows its cash for dealing;
    # this is the same split, declared.
    # ⚑ THIS IS NOT A NEW RULE — IT IS AN UNIMPLEMENTED ONE. Run_Context's Step 1b-2 row already
    # names the cash statement "THE GOLDEN SOURCE FOR THE ISA ALLOWANCE, platform fees, FX
    # charges, dividends, interest AND THE DAILY CASH BALANCE". The allowance was wired to it;
    # the balance never was. Two homes for one quantity, with the non-authoritative one winning
    # (R4.4). Measured 02-Sep-2026: the 31-Aug snapshot says GBP 1,065.39, the statement says
    # GBP 12,315.39, and the router was sizing on the first.
    # ⚑ EXPOSURE FIGURES ARE UNTOUCHED. `cash_value`, `total_value_gbp`, every weight, the sleeve
    # percentages and the band positions all stay on the month-end book. Only `deployable` moves.
    # ⚑ THE MMF IS ADDED BACK because the statement's balance is broker CASH and the money-market
    # ETF is a HOLDING there — it is cash to this framework (ISA-0564) but not to AJ Bell.
    # ⚑ AND IT DEGRADES BY NAMING, NEVER BY GUESSING: no statement, an unreadable one, or one
    # OLDER than the broker file, and deployable falls back to the month-end figure with the
    # basis stated on the artefact. "We could not read it" and "there is no more cash" must
    # never produce the same number (R2.10).
    class _MMFHandled(Exception):
        """Sentinel: the MMF branch already set the basis; skip the default assignment."""
    deployable_basis = "month_end_broker_file"
    deployable_as_of = str(data_date)
    deployable_note = ("cash statement not consulted")
    try:
        import extract_cash_statement as _ecs_dep
        # the ISA folder is the parent of the Investment Analysis dir the xlsx was found in;
        # parse_portfolio only receives the workbook path, so derive it from that (one home:
        # the same folder every other reader of the cash statement uses).
        _isa_dir = os.path.dirname(os.path.abspath(xlsx_path))
        _cs_dep = _ecs_dep.parse(folder=_isa_dir)
        _close = _cs_dep.get("closing_balance_gbp")
        _rows_dep = [r for r in (_cs_dep.get("rows") or []) if r.get("date")]
        _cs_last = max((str(r["date"])[:10] for r in _rows_dep), default=None)
        if not _cs_dep.get("source_files"):
            deployable_note = ("no Cash Statement export on file — deployable falls back to the "
                               "month-end broker cash line. This is a COVERAGE GAP, not a "
                               "statement that no further cash exists.")
        elif _close is None or _cs_last is None:
            deployable_note = ("the cash statement carries no closing balance or no dated rows — "
                               "deployable falls back to the month-end broker cash line.")
        elif _cs_last < _iso_or(data_date):
            deployable_note = ("the cash statement's last row (%s) PREDATES the broker file (%s), "
                               "so the statement is the stale side here — deployable stays on the "
                               "month-end broker cash line." % (_cs_last, _iso_or(data_date)))
        else:
            # ⚑ THE MMF IS ADDED BACK, AND THAT IS ONLY SAFE WHILE THE TWO PARTS ARE CONSISTENT
            # (Raj, 02-Sep-2026). The statement's balance is broker CASH and does NOT read the
            # money-market holding — correct, because to AJ Bell the MMF is a position. So
            # deployable = current cash (statement) + the MMF (snapshot). Those come from
            # DIFFERENT DATES, which is fine for an instrument that moves ~GBP 0.06/day — but
            # NOT if the MMF itself was traded in between: a sale after the snapshot date raises
            # the statement's cash AND leaves the holding still showing in the month-end file,
            # so the sum would count the same money twice. A purchase loses it twice over.
            # The one condition that breaks the arithmetic is therefore checked, not assumed.
            _mmf_names = []
            try:
                import scoring_config as _sc_mmf
                _mmf_names = [str(t).upper() for t in
                              (getattr(_sc_mmf, "CASH_EQUIVALENT_TICKERS", []) or [])]
            except Exception:                                           # noqa: BLE001
                pass
            _mmf_trades = [r for r in _rows_dep
                           if str(r.get("date"))[:10] > _iso_or(data_date)
                           and any(n.split(".")[0] in str(r.get("description") or "").upper()
                                   or n in str(r.get("description") or "").upper()
                                   or "OVERNIGHT" in str(r.get("description") or "").upper()
                                   for n in _mmf_names)]
            if _mmf_trades:
                deployable_cash = round(float(_close), 2)
                deployable_basis = "cash_statement_closing_balance_mmf_excluded"
                deployable_as_of = _cs_last
                deployable_note = (
                    "Closing GBP %.2f as at %s. ⚑ The money-market holding is NOT added: %d "
                    "MMF trade(s) settled after the %s snapshot, so the month-end holding figure "
                    "and the statement's cash are no longer consistent and summing them would "
                    "double-count. The MMF is EXCLUDED rather than estimated — deployable is "
                    "understated by at most the stale holding, which is the safe direction. "
                    "Rows: %s" % (float(_close), _cs_last, len(_mmf_trades), data_date,
                                  "; ".join(str(r.get("description"))[:40] for r in _mmf_trades[:3])))
                raise _MMFHandled
            deployable_cash = round(float(_close) + float(mmf_value or 0.0), 2)
            deployable_basis = "cash_statement_closing_balance_plus_mmf"
            deployable_as_of = _cs_last
            deployable_note = ("Run_Context Step 1b-2 declares the cash statement the golden "
                               "source for the daily cash balance. Closing GBP %.2f as at %s, "
                               "plus GBP %.2f held in the money-market ETF (cash to this "
                               "framework, a holding to the broker). Exposure figures remain on "
                               "the %s month-end book."
                               % (float(_close), _cs_last, float(mmf_value or 0.0), data_date))
    except _MMFHandled:
        pass
    except Exception as _dep_e:                                         # noqa: BLE001
        deployable_note = ("the cash statement could not be read (%s: %s) — deployable falls back "
                           "to the month-end broker cash line, UNMEASURED rather than confirmed."
                           % (type(_dep_e).__name__, _dep_e))

    # Weights
    for h in stocks + funds:
        h["weight_pct"] = round(h["value_gbp"] / total_value * 100, 2)

    cash_pct         = round(cash_value / total_value * 100, 2)
    stock_sleeve_pct = round(total_stock / total_value * 100, 2)
    fund_sleeve_pct  = round(total_funds / total_value * 100, 2)

    # Validation flag: stock position > 12.5%
    concentration_flags = [
        h["ticker"] for h in stocks + funds
        if h["weight_pct"] > 12.5
    ]

    # VUAG + B5B71Q7 combined check (must not exceed 12.5%)
    vuag_combined = sum(
        h["value_gbp"] for h in funds
        if h["ticker"] in ("VUAG", "B5B71Q7")
    )
    vuag_combined_pct = round(vuag_combined / total_value * 100, 2)
    vuag_combined_flag = vuag_combined_pct > 12.5

    # Month label
    month_label = data_dt.strftime("%b_%Y").lower() if data_dt else "unknown"
    run_month   = data_dt.strftime("%b %Y") if data_dt else "Unknown"

    # Build standing order notes
    if so_applied:
        so_note = (
            f"Standing order NOT YET cleared ({so_elapsed} working day(s) after 1st — "
            f"threshold {SO_CLEAR_WORKING_DAYS}). "
            f"£{so_adj:,.0f} added: £{cash_value:,.2f} → £{effective_cash:,.2f} effective."
        ) if so_elapsed is not None else (
            f"Standing order adjustment applied (file date unknown — conservative default). "
            f"£{so_adj:,.0f} added: £{cash_value:,.2f} → £{effective_cash:,.2f} effective."
        )
    else:
        so_note = (
            f"Standing order ALREADY CLEARED ({so_elapsed} working day(s) after 1st — "
            f"threshold {SO_CLEAR_WORKING_DAYS}). "
            f"No adjustment: stated cash £{cash_value:,.2f} = effective cash."
        )

    return {
        "_meta": {
            "source_file":                    os.path.basename(xlsx_path),
            "file_date":                      file_date_str or "unknown",
            "data_date":                      data_date or "unknown",
            "month_label":                    month_label,
            "run_month":                      run_month,
            "extracted_at":                   datetime.now().strftime("%Y-%m-%d %H:%M"),
            "standing_order_applied":         so_applied,
            "standing_order_working_days":    so_elapsed,
            "standing_order_note":            so_note,
        },
        "summary": {
            "total_value_gbp":        total_value,
            "total_cost_gbp":         round(sum(h["cost_gbp"] for h in stocks + funds), 2),
            "total_gain_gbp":         round(sum(h["gain_gbp"] for h in stocks + funds), 2),
            "cash_stated_gbp":        round(cash_value, 2),
            "mmf_value_gbp":          round(mmf_value, 2),      # B2 — included in cash figures
            "mmf_holdings":           mmf_holdings,
            "cash_effective_gbp":     effective_cash,
            "cash_deployable_gbp":    deployable_cash,
            "cash_deployable_basis":  deployable_basis,
            "cash_deployable_as_of":  deployable_as_of,
            "standing_order_applied": so_applied,
            "standing_order_adj":     so_adj,
            "cash_pct":               cash_pct,
            "stock_sleeve_value_gbp": total_stock,
            "stock_sleeve_pct":       stock_sleeve_pct,
            "fund_sleeve_value_gbp":  total_funds,
            "fund_sleeve_pct":        fund_sleeve_pct,
            "num_stock_positions":    len(stocks),
            "num_fund_positions":     len(funds),
        },
        "flags": {
            "concentration_over_12_5pct": concentration_flags,
            "vuag_plus_vanguard_us_combined_pct": vuag_combined_pct,
            "vuag_plus_vanguard_us_exceeds_12_5pct": vuag_combined_flag,
        },
        # Fix Pack A22: broker-reconciled allowance (S/O + lump sums) — feeds email §10 and
        # the A19 contribution-history cross-check; UNRECONCILED prints as such, never assumed.
        "contributions": _contributions_golden(os.path.dirname(os.path.abspath(xlsx_path)), data_dt),
        "stocks": sorted(stocks, key=lambda h: h["value_gbp"], reverse=True),
        "funds":  sorted(funds,  key=lambda h: h["value_gbp"], reverse=True),
        "cash": {
            "value_gbp":      round(cash_value, 2),
            "effective_gbp":  effective_cash,
            "deployable_gbp": deployable_cash,
            # ⚑ R4.2 — the figure states WHEN it was true and WHERE it came from, because the
            # exposure basis and the capital basis are deliberately different dates (ISA-0574).
            "deployable_basis": deployable_basis,
            "deployable_as_of": deployable_as_of,
            "deployable_note": deployable_note,
        },
        "notes": (
            f"Cash per file: £{cash_value:,.2f}. {so_note} "
            f"Deployable is effective cash in full: £{deployable_cash:,.2f} — the GBP 250 "
            f"D23 reserve is the single control and is applied downstream, once (ISA-0565)."
        ),
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Extract AJ Bell ISA Portfolio xlsx to structured JSON."
    )
    parser.add_argument(
        "--xlsx", default=None,
        help="Path to xlsx file. If omitted, auto-detects latest in ISA folder."
    )
    parser.add_argument(
        "--out", default=None,
        help="Output JSON path. If omitted, writes portfolio_data_mmm_yyyy.json to Investment Analysis folder."
    )
    parser.add_argument(
        "--isa-folder", default=ISA_FOLDER,
        help="ISA root folder (default: parent of this script's directory)."
    )
    args = parser.parse_args()

    # Resolve xlsx
    if args.xlsx:
        xlsx_path = args.xlsx
        if not os.path.exists(xlsx_path):
            print(f"ERROR: xlsx not found: {xlsx_path}")
            sys.exit(1)
    else:
        try:
            xlsx_path = find_latest_xlsx(args.isa_folder)
        except FileNotFoundError as e:
            print(f"ERROR: {e}")
            sys.exit(1)
        print(f"Auto-detected: {os.path.basename(xlsx_path)}")

    # Parse
    print(f"Parsing: {xlsx_path}")
    try:
        data = parse_portfolio(xlsx_path)
    except Exception as e:
        print(f"ERROR parsing portfolio: {e}")
        sys.exit(1)

    # Resolve output path
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if args.out:
        out_path = args.out
    else:
        month_label = data["_meta"]["month_label"]
        out_path = os.path.join(script_dir, f"portfolio_data_{month_label}.json")

    # Write
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)

    # Summary
    s = data["summary"]
    print(f"\nPortfolio extracted: {data['_meta']['run_month']}")
    print(f"  Total value:       £{s['total_value_gbp']:>12,.2f}")
    print(f"  Fund sleeve:       £{s['fund_sleeve_value_gbp']:>12,.2f}  ({s['fund_sleeve_pct']:.1f}%)")
    print(f"  Stock sleeve:      £{s['stock_sleeve_value_gbp']:>12,.2f}  ({s['stock_sleeve_pct']:.1f}%)")
    print(f"  Cash (stated):     £{s['cash_stated_gbp']:>12,.2f}  ({s['cash_pct']:.1f}%)")
    if s.get("standing_order_applied"):
        so_days = data["_meta"].get("standing_order_working_days")
        days_str = f"{so_days} working day(s) after 1st" if so_days is not None else "date unknown"
        print(f"  Cash (effective):  £{s['cash_effective_gbp']:>12,.2f}  (+£{s['standing_order_adj']:,.0f} S/O not yet cleared — {days_str})")
    else:
        so_days = data["_meta"].get("standing_order_working_days")
        print(f"  Cash (effective):  £{s['cash_effective_gbp']:>12,.2f}  (S/O already cleared — {so_days} working day(s) after 1st)")
    print(f"  Deployable cash:   £{s['cash_deployable_gbp']:>12,.2f}")
    print(f"\n  Stock positions:   {s['num_stock_positions']}")
    print(f"  Fund positions:    {s['num_fund_positions']}")

    if data["flags"]["concentration_over_12_5pct"]:
        print(f"\n  FLAG: Position(s) over 12.5%: {data['flags']['concentration_over_12_5pct']}")
    if data["flags"]["vuag_plus_vanguard_us_exceeds_12_5pct"]:
        print(f"\n  FLAG: VUAG + Vanguard US combined {data['flags']['vuag_plus_vanguard_us_combined_pct']:.1f}% > 12.5%")

    print(f"\nOutput written: {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Extract AJ Bell ISA Portfolio xlsx to structured JSON."
    )
    parser.add_argument(
        "--xlsx", default=None,
        help="Path to xlsx file. If omitted, auto-detects latest in ISA folder."
    )
    parser.add_argument(
        "--out", default=None,
        help="Output JSON path. If omitted, writes portfolio_data_mmm_yyyy.json to Investment Analysis folder."
    )
    parser.add_argument(
        "--isa-folder", default=ISA_FOLDER,
        help="ISA root folder (default: parent of this script's directory)."
    )
    args = parser.parse_args()

    # Resolve xlsx
    if args.xlsx:
        xlsx_path = args.xlsx
        if not os.path.exists(xlsx_path):
            print(f"ERROR: xlsx not found: {xlsx_path}")
            sys.exit(1)
    else:
        try:
            xlsx_path = find_latest_xlsx(args.isa_folder)
        except FileNotFoundError as e:
            print(f"ERROR: {e}")
            sys.exit(1)
        print(f"Auto-detected: {os.path.basename(xlsx_path)}")

    # Parse
    print(f"Parsing: {xlsx_path}")
    try:
        data = parse_portfolio(xlsx_path)
    except Exception as e:
        print(f"ERROR parsing portfolio: {e}")
        sys.exit(1)

    # Resolve output path
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if args.out:
        out_path = args.out
    else:
        month_label = data["_meta"]["month_label"]
        out_path = os.path.join(script_dir, f"portfolio_data_{month_label}.json")

    # Write
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)

    # Summary
    s = data["summary"]
    print(f"\nPortfolio extracted: {data['_meta']['run_month']}")
    print(f"  Total value:       £{s['total_value_gbp']:>12,.2f}")
    print(f"  Fund sleeve:       £{s['fund_sleeve_value_gbp']:>12,.2f}  ({s['fund_sleeve_pct']:.1f}%)")
    print(f"  Stock sleeve:      £{s['stock_sleeve_value_gbp']:>12,.2f}  ({s['stock_sleeve_pct']:.1f}%)")
    print(f"  Cash (stated):     £{s['cash_stated_gbp']:>12,.2f}  ({s['cash_pct']:.1f}%)")
    if s.get("standing_order_applied"):
        so_days = data["_meta"].get("standing_order_working_days")
        days_str = f"{so_days} working day(s) after 1st" if so_days is not None else "date unknown"
        print(f"  Cash (effective):  £{s['cash_effective_gbp']:>12,.2f}  (+£{s['standing_order_adj']:,.0f} S/O not yet cleared -- {days_str})")
    else:
        so_days = data["_meta"].get("standing_order_working_days")
        print(f"  Cash (effective):  £{s['cash_effective_gbp']:>12,.2f}  (S/O already cleared -- {so_days} working day(s) after 1st)")
    print(f"  Deployable cash:   £{s['cash_deployable_gbp']:>12,.2f}")
    print(f"\n  Stock positions:   {s['num_stock_positions']}")
    print(f"  Fund positions:    {s['num_fund_positions']}")

    if data["flags"]["concentration_over_12_5pct"]:
        print(f"\n  FLAG: Position(s) over 12.5%: {data['flags']['concentration_over_12_5pct']}")
    if data["flags"]["vuag_plus_vanguard_us_exceeds_12_5pct"]:
        print(f"\n  FLAG: VUAG + Vanguard US combined {data['flags']['vuag_plus_vanguard_us_combined_pct']:.1f}% > 12.5%")

    print(f"\nOutput written: {out_path}")
    return out_path


if __name__ == "__main__":
    main()
