#!/usr/bin/env python3
"""
calibration_universe.py — Capture Layer Item 5. 02-Aug-2026.

THE PROBLEM
-----------
`calibration_prices.csv` covers only the names that reached SCORING — i.e. the gate PASSERS
that got logged to `score_panel.csv`. For STOXX600 that is 273 of ~600 constituents; for
F250SPI, 138 of ~250. Every name a gate BLOCKED is absent.

That was tolerable while the only question was "how did our signals do on the names we
scored?". It is not tolerable now. Capture Layer Item 1 records gate variables for every
constituent precisely so §7.2's `rule_frictions` can ask **"did the names our gates blocked
subsequently perform?"** — and that question cannot be answered without prices for the blocked
names, which are exactly the ones missing.

So Items 1 and 5 are one mechanism: Item 1 supplies the variables, Item 5 supplies the
outcomes. Either alone answers nothing.

THE RESOLUTION BUG
------------------
The build order notes it exactly: "the same tickers resolve during screening but not during
calibration, which is itself the bug." `screener_core.fetch_constituents()` already produces
correctly-suffixed symbols (`.L`, `.DE`, `.SW`, `.ST`, `.MI` ...) because the screen would
fetch nothing otherwise. The calibration path never used it — it took whatever strings happened
to be in the panel. This module reuses `fetch_constituents()` rather than re-deriving suffixes,
so there is one resolution rule and it is the one already proven by every weekly screen.

RESUMABLE BY CONSTRUCTION
-------------------------
~2,600 tickers cannot be fetched inside a 45-second bash ceiling. Every call fetches one chunk,
writes the cache, and reports what remains. Interrupting it loses nothing.

CLI:
  python3 calibration_universe.py --group STOXX600 --chunk 150     # one resumable pass
  python3 calibration_universe.py --group STOXX600 --loop 4        # several passes
  python3 calibration_universe.py --coverage                       # the acceptance report
  python3 calibration_universe.py --selftest
"""
from __future__ import annotations
import argparse, json, os, sys, time

HERE = os.path.dirname(os.path.abspath(__file__))
PRICE_CACHE = os.path.join(HERE, "calibration_prices.csv")
UNIVERSE_CACHE = os.path.join(HERE, "calibration_universe.json")

# The acceptance criterion names these two because they are the worst cases: both are
# non-US, both carry exchange suffixes, and both are where resolution silently failed.
ACCEPTANCE_GROUPS = ("STOXX600", "F250-SPI")
# Renamed from PRICE_RESOLUTION_FLOOR 12-Aug-2026 (register ISA-0012). gate_variables held a
# DIFFERENT 0.95 under the SAME name: two numeric coverage floors, one flat namespace, and
# an import of the wrong one would have produced a plausible number - FC-B by construction.
# The emitted JSON key stays "acceptance_floor": that is an artefact contract with the
# dashboard and the pre-run, and a rename must not move it.
PRICE_RESOLUTION_FLOOR = 0.85
MIN_OBS_1M = 15          # trading days that must be present for "1m resolved"


# Group -> the token that appears in an emitted screen filename. F250-SPI is written
# "F250SPI" by the screen, which is exactly the kind of one-character mismatch that makes a
# lookup silently return nothing.
_SCREEN_TOKEN = {"F250-SPI": "F250SPI", "STOXX600": "STOXX600", "SP500": "SP500",
                 "NASDAQ": "NASDAQ", "MIDCAP400": "MIDCAP400"}


# Group -> the string that appears in an emitted workbook filename.
_WORKBOOK_TOKEN = {"F250-SPI": "F250", "STOXX600": "Stoxx", "SP500": "SP500",
                   "NASDAQ": "Nasdaq", "MIDCAP400": "MIDCAP400"}


def _universe_from_workbook(group, here=None):
    """Constituents recovered from the emitted weekly workbook (SCORES + EXCLUSIONS)."""
    import glob as _glob
    here = here or HERE
    token = _WORKBOOK_TOKEN.get(group, group)
    files = [f for f in _glob.glob(os.path.join(here, "Growth Stock Analysis*.xlsx"))
             if token.lower() in os.path.basename(f).lower()]
    files += [f for f in _glob.glob(os.path.join(here, "archive", "*",
                                                 "Growth Stock Analysis*.xlsx"))
              if token.lower() in os.path.basename(f).lower()]
    if not files:
        return [], None
    latest = sorted(files, key=os.path.getmtime)[-1]
    try:
        import openpyxl
        wb_ = openpyxl.load_workbook(latest, read_only=True, data_only=True)
    except Exception:
        return [], None
    out = set()
    for sheet, header_row in (("SCORES", 3), ("EXCLUSIONS", 2)):
        if sheet not in wb_.sheetnames:
            continue
        ws = wb_[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i <= header_row or not row:
                continue
            v = row[0]
            if v and isinstance(v, str) and 1 <= len(v.strip()) <= 14 and " " not in v.strip():
                out.add(v.strip())
    return sorted(out), f"last_workbook:{os.path.basename(latest)}"


def _universe_from_last_screen(group, here=None):
    """The full, ALREADY-RESOLVED constituent list from the most recent weekly screen.

    `{date}_{GROUP}_yf_gate_results.csv` contains every constituent the screen fetched —
    passers AND every excluded name — written with the exact symbols the screen resolved and
    successfully fetched against. That makes it a strictly better universe source than
    re-deriving one:

      * it is the same resolution rule, already applied and already proven to work;
      * it carries no live-network dependency, so a calibration pass cannot be blocked by an
        index-page source being slow or unreachable (which is precisely what happens in a
        sandbox, and would have made this whole item untestable);
      * it is point-in-time correct — the universe as it was when the signals were logged,
        not as it is today.

    Returns (tickers, warnings, source_label).
    """
    import pandas as pd
    import glob as _glob
    here = here or HERE
    token = _SCREEN_TOKEN.get(group, group)
    pats = [os.path.join(here, f"*_{token}_yf_gate_results.csv"),
            os.path.join(here, "archive", "*", f"*_{token}_yf_gate_results.csv"),
            os.path.join(here, f"*_{token}_full_data.csv")]
    files = []
    for p in pats:
        files.extend(_glob.glob(p))
    if not files:
        # Third source: the emitted WORKBOOK. SCORES holds the scored names and EXCLUSIONS
        # holds every rejected one, so together they are the same full constituent list --
        # just in the artefact that survived rather than the CSV that did not. F250-SPI has
        # no gate-results CSV on disk at all, so without this the group would have no
        # universe and the acceptance criterion could never be evaluated for it.
        tk, warn = _universe_from_workbook(group, here)
        if tk:
            return tk, [], warn
        return [], [f"no emitted screen file or workbook found for {group} (token {token})"], None
    latest = sorted(files, key=lambda f: os.path.basename(f))[-1]
    try:
        d = pd.read_csv(latest, usecols=["ticker"])
    except Exception as e:
        return [], [f"could not read {os.path.basename(latest)}: {e}"], None
    tickers = sorted({str(t) for t in d["ticker"].dropna() if str(t).strip()})
    return tickers, [], f"last_screen:{os.path.basename(latest)}"


def universe_for(group, refresh=False, cache=None):
    """The FULL constituent list for a group, with the screen's own suffix resolution.

    Cached to disk: the constituent fetch hits several external sources, and re-fetching it on
    every resumable price pass would make the resumability itself expensive.
    """
    # Resolved at CALL time, not bound as a default: a default argument would capture the
    # module constant at import and silently ignore any later override (including the
    # self-test's).
    cache = cache or UNIVERSE_CACHE
    store = {}
    if os.path.exists(cache) and not refresh:
        try:
            with open(cache, encoding="utf-8") as f:
                store = json.load(f)
        except Exception:
            store = {}
    if not refresh and group in store and store[group].get("tickers"):
        return store[group]["tickers"], store[group].get("warnings", [])

    tickers, warns, src = _universe_from_last_screen(group)
    if not tickers:
        # Fallback: re-fetch live. Slower, network-dependent, and unnecessary whenever a
        # screen has run — which is why it is the fallback and not the primary.
        sys.path.insert(0, HERE)
        import screener_core as SC
        df, warns = SC.fetch_constituents(group)      # ← the SAME resolution rule
        tickers = (sorted({str(t) for t in df["ticker"].dropna()})
                   if df is not None and not df.empty else [])
        src = "live_fetch_constituents"
    store[group] = {"tickers": tickers, "warnings": warns, "source": src,
                    "fetched_at": time.strftime("%Y-%m-%d %H:%M:%S")}
    with open(cache, "w", encoding="utf-8") as f:
        json.dump(store, f, indent=1)
    return tickers, warns


def _load_cache(path=PRICE_CACHE):
    import pandas as pd
    if not os.path.exists(path):
        return pd.DataFrame()
    try:
        return pd.read_csv(path, index_col=0, parse_dates=True)
    except Exception:
        return pd.DataFrame()


def extend(group, chunk=150, period="1y", price_cache=PRICE_CACHE, refresh_universe=False):
    """One resumable pass. Returns a dict describing what happened and what remains."""
    import pandas as pd
    import yfinance as yf
    try:
        from fetch_guard import with_backoff
    except Exception:
        def with_backoff(fn, *a, **k):
            return fn(*a, **k)

    tickers, warns = universe_for(group, refresh=refresh_universe)
    cache = _load_cache(price_cache)
    have = set(cache.columns)
    todo = [t for t in tickers if t not in have]
    if not todo:
        return {"group": group, "universe": len(tickers), "fetched": 0, "remaining": 0,
                "done": True, "warnings": warns}

    batch = todo[:chunk]
    px = with_backoff(yf.download, batch, period=period, progress=False,
                      auto_adjust=True, threads=True)
    px = px["Close"] if "Close" in getattr(px, "columns", []) or (
        hasattr(px, "columns") and "Close" in px.columns.get_level_values(0)) else px
    if isinstance(px, pd.Series):
        px = px.to_frame(batch[0])
    if getattr(px.index, "tz", None) is not None:
        px.index = px.index.tz_localize(None)

    # A ticker that resolved to nothing must still be RECORDED as attempted, or every pass
    # will retry it forever and the cache will never report itself complete.
    for t in batch:
        if t not in px.columns:
            px[t] = float("nan")

    cache = px if cache.empty else pd.concat([cache, px[[c for c in px.columns
                                                         if c not in cache.columns]]], axis=1)
    cache = cache.loc[:, ~cache.columns.duplicated()].sort_index()
    cache.to_csv(price_cache)

    remaining = len(todo) - len(batch)
    nn = px[batch].notna().sum()
    resolved = int((nn >= MIN_OBS_1M).sum())
    return {"group": group, "universe": len(tickers), "fetched": len(batch),
            "resolved_in_batch": resolved,
            "batch_resolution": round(resolved / len(batch), 4),
            "remaining": remaining, "done": remaining == 0,
            "unresolved_sample": [t for t in batch if nn.get(t, 0) < MIN_OBS_1M][:10],
            "warnings": warns}


def coverage(groups=None, price_cache=PRICE_CACHE, universe_cache=None):
    """The Item 5 acceptance report: 1m price resolution over the FULL constituent universe."""
    import pandas as pd
    groups = groups or list(ACCEPTANCE_GROUPS)
    cache = _load_cache(price_cache)
    if cache.empty:
        return {"error": "no price cache"}
    recent = cache.tail(30)
    nn = recent.notna().sum()
    out = {"price_cache": price_cache, "cache_columns": int(len(cache.columns)), "groups": {}}
    for g in groups:
        try:
            tickers, _ = universe_for(g, cache=universe_cache)
        except Exception as e:
            out["groups"][g] = {"error": f"universe unavailable: {e}"}
            continue
        if not tickers:
            out["groups"][g] = {"error": "empty universe"}
            continue
        attempted = [t for t in tickers if t in cache.columns]
        resolved = [t for t in attempted if nn.get(t, 0) >= MIN_OBS_1M]
        out["groups"][g] = {
            "universe": len(tickers),
            "attempted": len(attempted),
            "resolved_1m": len(resolved),
            # over the FULL universe — the number the acceptance criterion is about
            "resolution_vs_universe": round(len(resolved) / len(tickers), 4),
            # over what has actually been fetched so far — tells you whether a shortfall is
            # "not finished yet" or "finished and genuinely unresolvable"
            "resolution_vs_attempted": (round(len(resolved) / len(attempted), 4)
                                        if attempted else 0.0),
            "not_yet_attempted": len(tickers) - len(attempted),
            "unresolved_sample": [t for t in attempted if nn.get(t, 0) < MIN_OBS_1M][:10],
        }
    vals = [v["resolution_vs_universe"] for v in out["groups"].values() if "error" not in v]
    out["worst"] = min(vals) if vals else 0.0
    out["acceptance_floor"] = PRICE_RESOLUTION_FLOOR
    out["acceptance"] = "PASS" if out["worst"] >= PRICE_RESOLUTION_FLOOR else "FAIL"
    return out


def _selftest():
    fails = []

    def ok(label, cond, detail=""):
        print(("  PASS  " if cond else "  FAIL  ") + label + (f"  [{detail}]" if detail else ""))
        if not cond:
            fails.append(label)

    import pandas as pd, tempfile
    with tempfile.TemporaryDirectory() as td:
        pc = os.path.join(td, "px.csv")
        idx = pd.date_range("2026-07-01", periods=25, freq="B")
        df = pd.DataFrame({"AAA.L": range(25), "BBB.DE": [None] * 25}, index=idx)
        df.to_csv(pc)
        uc = os.path.join(td, "uni.json")
        json.dump({"TEST": {"tickers": ["AAA.L", "BBB.DE", "CCC.SW"]}}, open(uc, "w"))
        global UNIVERSE_CACHE
        old = UNIVERSE_CACHE
        UNIVERSE_CACHE = uc
        try:
            # patch universe_for's default cache by calling with the module-level constant
            rep = coverage(groups=["TEST"], price_cache=pc, universe_cache=uc)
            g = rep["groups"]["TEST"]
            ok("U-CU1 resolution measured against the FULL universe, not the fetched subset",
               g["universe"] == 3 and g["attempted"] == 2)
            ok("U-CU2 an all-NaN column counts as ATTEMPTED but NOT resolved",
               g["resolved_1m"] == 1 and "BBB.DE" in g["unresolved_sample"])
            ok("U-CU3 a never-fetched ticker is reported separately from an unresolved one",
               g["not_yet_attempted"] == 1)
            ok("U-CU4 vs-universe (0.33) and vs-attempted (0.50) reported separately",
               abs(g["resolution_vs_universe"] - 0.3333) < 0.01
               and abs(g["resolution_vs_attempted"] - 0.5) < 0.01)
            ok("U-CU5 acceptance FAILS below the floor", rep["acceptance"] == "FAIL")
        finally:
            UNIVERSE_CACHE = old

    ok("U-CU6 acceptance groups are the two the build order names",
       set(ACCEPTANCE_GROUPS) == {"STOXX600", "F250-SPI"})
    print("SELFTEST PASS" if not fails else f"SELFTEST FAIL ({len(fails)}) {fails}")
    return 0 if not fails else 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--group")
    ap.add_argument("--chunk", type=int, default=150)
    ap.add_argument("--period", default="1y")
    ap.add_argument("--loop", type=int, default=1, help="consecutive resumable passes")
    ap.add_argument("--coverage", action="store_true")
    ap.add_argument("--groups", nargs="*")
    ap.add_argument("--refresh-universe", action="store_true")
    ap.add_argument("--selftest", action="store_true")
    a = ap.parse_args()
    if a.selftest:
        return _selftest()
    if a.coverage:
        rep = coverage(groups=a.groups)
        print(json.dumps(rep, indent=2))
        return 0 if rep.get("acceptance") == "PASS" else 1
    if not a.group:
        ap.error("--group or --coverage required")
    for i in range(a.loop):
        r = extend(a.group, chunk=a.chunk, period=a.period,
                   refresh_universe=(a.refresh_universe and i == 0))
        print(f"UNIVERSE_PRICES group={r['group']} universe={r['universe']} "
              f"fetched={r.get('fetched')} batch_resolution={r.get('batch_resolution')} "
              f"remaining={r['remaining']} {'DONE' if r['done'] else 'NOT_DONE (resumable)'}")
        if r.get("unresolved_sample"):
            print("  unresolved in this batch: " + ", ".join(r["unresolved_sample"]))
        if r["done"]:
            break
    return 0


if __name__ == "__main__":
    sys.exit(main())
