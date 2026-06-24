#!/usr/bin/env python3
"""
update_watchlist.py  --  Watchlist Promotion, Score Refresh, and Ranking
Version: 1.0  |  2026-06-02

Runs as Step 4 of the monthly_isa_prerun.py pipeline (after portfolio_analytics.py,
before fetch_watchlist_metrics.py).

Purpose:
  Mechanically apply the Step 3 Watchlist Promotion Framework to produce an updated
  watchlist_tickers.json before the metrics fetch runs, so newly promoted stocks are
  in-window-checked in the same monthly cycle they are identified.

Inputs:
  --portfolio-data   portfolio_data_mmm_yyyy.json  (for portfolio duplicate check)
  --watchlist-json   watchlist_tickers.json         (current watchlist to update)
  --inv-dir          Investment Analysis folder path  (to locate growth stock xlsx)
  --out-json         watchlist_tickers.json          (writes back in place)

Outputs:
  - Updated watchlist_tickers.json (in place)
  - watchlist_promotion_log dict printed as JSON on stdout (embedded in run_context by orchestrator)
  - Deletes all Growth Stock Analysis*.xlsx files after processing

Phase logic:
  Phase 1 — Path classification (energy vs growth stock)
  Phase 2 — Hard exclusions (portfolio duplicates, unchanged watchlist dups, China VIE)
  Phase 3 — Candidate pool assembly (deduplicate by ticker, keep highest score)
  Phase 4 — Existing watchlist score refresh (delta_score, part_b_driver)
  Phase 5 — Provisional entry levels for new Tier 1 candidates
  Phase 6 — Combined ranking and top-10 selection
  Phase 7 — Write outputs and delete xlsx files
"""

import argparse
import glob
import json
import re
import math
import os
import shutil
import sys
from datetime import date, datetime

# Single source of truth for the fluid-pool decay flags (Raj: pool+watchlist must turn over).
try:
    import scoring_config as _cfg
except Exception:
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        import scoring_config as _cfg
    except Exception:
        _cfg = None

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Max watchlist size
MAX_WATCHLIST = 10

# Score history: keep at most this many months
MAX_SCORE_HISTORY = 3

# China VIE hard exclusion list
CHINA_VIE_EXCLUSIONS = {"BZ", "PDD", "NTES", "ATAT", "BEKE", "DIDI", "TUYA", "LAIX"}

# Path A scorecard max (Growth_Stock_Checklist.pdf: 14+13 metrics, 2pts each)
PATH_A_MAX = 54

# Path C scorecard max (energy_screener.py: 10+8 metrics, 2pts each)
PATH_C_MAX = 36

# Score gate thresholds
NORMALISED_SCORE_PROMOTION_GATE = 70.0   # new: was effectively ~50%
NORMALISED_SCORE_PROBATION_FLOOR = 60.0  # existing names 60-69 = probation
NORMALISED_SCORE_HARD_REMOVE_BELOW = 60.0  # existing names <60 = remove unless override

# Archive (replaces delete)
XLSX_ARCHIVE_SUBDIR = "archive"  # created inside inv_dir if not present


# ---------------------------------------------------------------------------
# SUMMARY tab reader
# ---------------------------------------------------------------------------

def read_xlsx_summary_tab(xlsx_path: str) -> list[dict]:
    """
    Read the SUMMARY tab from a growth stock analysis xlsx file.
    Returns a list of row dicts. Tolerates missing columns gracefully.
    Requires openpyxl.
    """
    try:
        import openpyxl
    except ImportError:
        print("  [update_watchlist] openpyxl not installed — attempting install...")
        import subprocess
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"],
            capture_output=True
        )
        import openpyxl

    rows = []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"  [update_watchlist] Cannot open {os.path.basename(xlsx_path)}: {e}")
        return rows

    # Find SUMMARY sheet (case-insensitive)
    sheet = None
    for name in wb.sheetnames:
        if name.strip().upper() == "SUMMARY":
            sheet = wb[name]
            break

    if sheet is None:
        print(f"  [update_watchlist] No SUMMARY tab in {os.path.basename(xlsx_path)} — skipped")
        wb.close()
        return rows

    # Locate the header row dynamically. Analysis SUMMARY tabs carry a title banner
    # in row 1 and group labels in row 3; the real column headers (Ticker, Part A
    # (/28), Total (/54), ...) sit several rows down. Scan for the first row with a
    # Ticker/Symbol/Stock header. Header names are normalised to strip newline and
    # parenthetical suffixes, e.g. "Part A\n(/28)" -> "part a".
    def _norm(val):
        if val is None:
            return ""
        txt = str(val).split("\n")[0]
        txt = re.sub(r"\(.*?\)", "", txt)
        return txt.strip().lower()

    all_rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    header_idx = None
    headers = []
    for i, r in enumerate(all_rows):
        norm = [_norm(c) for c in r]
        if any(h in ("ticker", "symbol", "stock") for h in norm):
            header_idx = i
            headers = norm
            break

    if header_idx is None:
        print(f"  [update_watchlist] No Ticker header found in SUMMARY of {os.path.basename(xlsx_path)} — skipped")
        wb.close()
        return rows

    def col(row_vals: list, *names: str):
        """Return first matching column value (headers already normalised)."""
        for name in names:
            target = name.strip().lower()
            for i, h in enumerate(headers):
                if h == target and i < len(row_vals):
                    return row_vals[i]
        return None

    # Read data rows (everything after the located header row)
    for row in all_rows[header_idx + 1:]:
        row_vals = list(row)
        if not any(v for v in row_vals):
            continue  # skip blank rows

        ticker = col(row_vals, "Ticker", "Symbol", "Stock")
        if not ticker or str(ticker).strip() == "":
            continue
        ticker = str(ticker).strip().upper()

        company = col(row_vals, "Company", "Name", "Company Name")
        sector  = col(row_vals, "Sector", "Industry", "GICS Sector")
        exchange = col(row_vals, "Exchange", "Market")

        # Scores
        def safe_int(v):
            try:
                return int(float(v)) if v is not None else None
            except (TypeError, ValueError):
                return None

        def safe_float(v):
            try:
                return float(v) if v is not None else None
            except (TypeError, ValueError):
                return None

        part_a = safe_int(col(row_vals, "Part A", "Part A Score", "Part A Total"))
        part_b = safe_int(col(row_vals, "Part B", "Part B Score", "Part B Total"))
        total  = safe_int(col(row_vals, "Total", "Total Score", "Combined Score"))

        # Derive total if missing
        if total is None and part_a is not None and part_b is not None:
            total = part_a + part_b

        # Consensus target / current price (for provisional entry level)
        target_price   = safe_float(col(row_vals, "Target Price", "Consensus Target", "Price Target", "Fair Value"))
        current_price  = safe_float(col(row_vals, "Price", "Current Price", "Last Price"))
        currency       = col(row_vals, "Currency") or "USD"

        rows.append({
            "ticker":       ticker,
            "company":      str(company).strip() if company else ticker,
            "sector":       str(sector).strip() if sector else "",
            "exchange":     str(exchange).strip() if exchange else "NASDAQ",
            "currency":     str(currency).strip() if currency else "USD",
            "part_a":       part_a,
            "part_b":       part_b,
            "total":        total,
            "target_price": target_price,
            "current_price": current_price,
            "source_file":  os.path.basename(xlsx_path),
        })

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Phase 1: Path classification
# ---------------------------------------------------------------------------

def classify_path(row: dict, filename: str) -> str | None:
    """
    Returns 'A', 'C', or None (misrouted energy — exclude from promotion).
    'C' for energy/clean-tech/renewables; 'A' for everything else.
    Misrouted energy in a non-ENERGY file returns None (log and exclude).
    """
    fname_upper = filename.upper()
    is_energy_file = "ENERGY" in fname_upper

    sector = (row.get("sector") or "").lower()
    energy_sectors = {
        "energy", "utilities", "oil", "gas", "renewables", "clean-tech",
        "cleantech", "nuclear", "power generation", "upstream", "midstream",
        "downstream", "coal", "lng", "solar", "wind", "hydro",
    }
    is_energy_sector = any(kw in sector for kw in energy_sectors)

    if is_energy_file:
        return "C"
    elif is_energy_sector:
        # Misrouted — energy stock in a non-energy file
        return None  # caller logs as path_c_misrouted
    else:
        return "A"


# ---------------------------------------------------------------------------
# Phase 4: Part B driver classification
# ---------------------------------------------------------------------------

VALUATION_METRICS_PART_B_INDICES = {
    # Metrics 4-8 in Part B are typically valuation metrics:
    # fwd P/E, EV/EBITDA, P/FCF, FCF Yield, Earnings Yield, 52wk position
    # ROIC (index 0), ND/EBITDA (index 1), Int Coverage (index 2) = quality metrics
    "valuation_indices": {3, 4, 5, 6, 7, 8},  # fwd metrics, 52wk range
    "quality_indices":   {0, 1, 2},             # ROIC, ND/EBITDA, Int Coverage
}


def classify_part_b_driver(prior_total: int, new_total: int,
                            prior_part_a: int | None, new_part_a: int | None,
                            prior_part_b: int | None, new_part_b: int | None) -> str | None:
    """
    Classify what drove a score change.
    Returns one of: 'valuation_led', 'quality_concern', 'broad_deterioration',
                    'broad_improvement', 'valuation_improvement', None (no material change).
    Uses Part A/B delta to attribute: Part A = quality metrics; Part B = valuation/resilience.
    This is a heuristic from SUMMARY tab data; normalise_adapter.py provides metric-level detail.
    """
    if new_total is None or prior_total is None:
        return None

    delta = new_total - prior_total
    if abs(delta) < 2:
        return None  # immaterial change — no attribution needed

    part_a_delta = (new_part_a - prior_part_a) if (new_part_a is not None and prior_part_a is not None) else 0
    part_b_delta = (new_part_b - prior_part_b) if (new_part_b is not None and prior_part_b is not None) else 0

    if delta < 0:
        # Score declined
        if part_b_delta < -1 and part_a_delta >= -1:
            return "valuation_led"        # Part B down, Part A stable → valuation repricing
        elif part_a_delta < -1 and part_b_delta >= -1:
            return "quality_concern"      # Part A down, Part B stable → quality erosion
        elif part_a_delta < -1 and part_b_delta < -1:
            return "broad_deterioration"  # Both down → systemic decline
        else:
            return "valuation_led"        # default for decline
    else:
        # Score improved
        if part_a_delta > 1 and part_b_delta <= 1:
            return "quality_improvement"   # Part A up → quality improving
        elif part_b_delta > 1 and part_a_delta <= 1:
            return "valuation_improvement" # Part B up → valuation more attractive
        else:
            return "broad_improvement"    # Both up


# ---------------------------------------------------------------------------
# Normalised score
# ---------------------------------------------------------------------------

def _base_ticker(t: str) -> str:
    """Symbol without exchange suffix, upper-cased (ONT.L -> ONT, CSU.TO -> CSU)."""
    t = str(t or "").strip().upper()
    return t.split(".")[0] if "." in t else t


def normalised_score(total: int | None, path: str) -> float | None:
    if total is None:
        return None
    max_score = PATH_A_MAX if path == "A" else PATH_C_MAX
    return round(total / max_score * 100, 1)


def _pool_admit(ns, part_a, path, override) -> bool:
    """Candidate-pool admission (H2 fix). Default = ns >= NORMALISED_SCORE_PROMOTION_GATE (quality-total).
    When FORWARD_ELIGIBILITY is on, a Part A VIABILITY floor (path-aware) replaces the total gate so
    forward-confirmed lower-total names are NOT dropped here — rerank then applies forward eligibility +
    the Source Score. NOTE: SUMMARY Part A is on the path scale (growth /28, energy /20)."""
    if override:
        return True
    if getattr(_cfg, "FORWARD_ELIGIBILITY", False):
        floor = (getattr(_cfg, "FORWARD_ELIG_PART_A_FLOOR_ENERGY", 14) if path == "C"
                 else getattr(_cfg, "FORWARD_ELIG_PART_A_FLOOR", 21))
        return (part_a or 0) >= floor
    return (ns or 0) >= NORMALISED_SCORE_PROMOTION_GATE


# ---------------------------------------------------------------------------
# Fluid-pool decay helpers (redesign Part3 §4 Layer4 / §13)
# ---------------------------------------------------------------------------

def _fluid_on() -> bool:
    return bool(getattr(_cfg, "FLUID_POOL_DECAY", False))


def _ageout_months() -> int:
    return int(getattr(_cfg, "POOL_AGEOUT_MONTHS", 3))


def _decay_penalty() -> float:
    return float(getattr(_cfg, "POOL_DECAY_PENALTY", 5.0))


def _months_since(label, now_label):
    """Whole months between two '%b-%Y' month labels ('Mar-2026'->'Jun-2026' = 3).
    None if either is unparseable."""
    try:
        a = datetime.strptime(str(label), "%b-%Y")
        b = datetime.strptime(str(now_label), "%b-%Y")
        return (b.year - a.year) * 12 + (b.month - a.month)
    except Exception:
        return None


def _pool_decay_decision(last_confirmed_label, now_label, ageout_months=None):
    """Age-out decision for a name ABSENT from this cycle's screens. Returns
    (months_since_confirmed, age_out:bool). age_out True -> drop (not re-screened for
    > ageout_months ~= 90d). This is the time-based replacement for score_history-length
    staleness, which could freeze a <3-history name on a stale score indefinitely."""
    ageout = _ageout_months() if ageout_months is None else ageout_months
    ms = _months_since(last_confirmed_label, now_label)
    return ms, (ms is not None and ms > ageout)


# ---------------------------------------------------------------------------
# Main logic
# ---------------------------------------------------------------------------

def run(portfolio_path: str, watchlist_path: str, inv_dir: str, out_path: str,
        dry_run: bool = False) -> dict:
    """
    Execute the 7-phase watchlist update. Returns the watchlist_promotion_log dict.
    """
    run_date = date.today().strftime("%Y-%m-%d")
    month_label = date.today().strftime("%b-%Y")

    # Load watchlist_tickers.json
    with open(watchlist_path, encoding="utf-8") as f:
        wl_data = json.load(f)

    # Snapshot the prior pool so carry-forward incumbents survive the ephemeral wipe.
    prior_pool = list(wl_data.get("candidate_pool", []))
    # Fluid pool: preserve first_seen across runs (durable-pool memory, CONTRACTS candidate-pool).
    _prior_first_seen = {e.get("ticker"): e.get("first_seen") for e in prior_pool if e.get("ticker")}

    # Load portfolio_data for duplicate check
    portfolio_tickers = set()
    port_stocks = []
    port_data_date = run_date
    if portfolio_path and os.path.exists(portfolio_path):
        try:
            with open(portfolio_path, encoding="utf-8") as f:
                port = json.load(f)
            port_stocks = port.get("stocks", [])
            port_data_date = port.get("_meta", {}).get("data_date", run_date)
            for s in port_stocks:
                t = s.get("ticker", "")
                if t:
                    portfolio_tickers.add(t.upper())
            # Also from watchlist stock_sleeve
        except Exception as e:
            print(f"  [update_watchlist] Warning: could not read portfolio_data: {e}")

    # Also add stock_sleeve tickers from watchlist_tickers.json itself
    for s in wl_data.get("stock_sleeve", []):
        t = s.get("ticker", "")
        if t:
            portfolio_tickers.add(t.upper())

    # Current watchlist entries (dict by ticker for easy lookup)
    current_watchlist = {e["ticker"].upper(): e for e in wl_data.get("watchlist", [])}

    promotion_log = {
        "run_date":             run_date,
        "additions":            [],
        "removals":             [],
        "score_updates":        [],
        "misrouted_energy":     [],
        "stale_scores":         [],
        "rejected_candidates":  [],   # names failing the 70% gate
        "duplicate_ticker_log": [],   # cross-file duplicate appearances
        "xlsx_files_deleted":   [],   # kept for backward compatibility — always empty after archiving
        "xlsx_files_archived":  [],   # replaces xlsx_files_deleted
        "xlsx_files_read":      [],
        "path_c_misrouted_log": [],
        "rows_parsed":          0,
        "held_removed_from_watchlist": [],
        "held_removed_from_vci":       [],
        "sleeve_phantom_removed":      [],
        "sleeve_added":                [],
    }

    # -----------------------------------------------------------------------
    # Reconcile against broker truth (portfolio_data.stocks)
    # -----------------------------------------------------------------------
    held_base = {_base_ticker(s.get("ticker", "")) for s in port_stocks if s.get("ticker")}
    portfolio_tickers |= held_base

    # (C) Stock-sleeve sync: drop phantom holds, add un-recorded buys.
    held_by_base = {_base_ticker(s.get("ticker", "")): s for s in port_stocks if s.get("ticker")}
    kept_sleeve, sleeve_bases = [], set()
    for s in wl_data.get("stock_sleeve", []):
        b = _base_ticker(s.get("ticker", ""))
        if b in held_by_base:
            kept_sleeve.append(s); sleeve_bases.add(b)
        else:
            promotion_log["sleeve_phantom_removed"].append({
                "ticker": s.get("ticker"), "name": s.get("name"),
                "reason": "in stock_sleeve but absent from broker portfolio file",
            })
    for b, hs in held_by_base.items():
        if b not in sleeve_bases:
            qty = hs.get("quantity") or 0
            cost = hs.get("cost_gbp")
            cps = round(cost / qty, 4) if (cost and qty) else None
            exch = "LSE" if (hs.get("currency") == "GBP") else "NASDAQ"
            kept_sleeve.append({
                "ticker": b, "name": hs.get("name", ""), "exchange": exch,
                "purchase_date": None, "cost_per_share_usd": None,
                "cost_total_gbp": cost, "shares": qty,
                "include_in_metrics_pull": True,
                "note": (f"Auto-added from broker file {port_data_date} — "
                         "verify purchase_date / cost basis / thesis-break conditions"),
                "_auto_added": True,
            })
            promotion_log["sleeve_added"].append({"ticker": b, "shares": qty})
    if port_stocks:
        wl_data["stock_sleeve"] = kept_sleeve

    # (B) Remove held names from the VCI watchlist (suffix-insensitive).
    if held_base:
        new_vci = []
        for e in wl_data.get("vci_watchlist", []):
            if _base_ticker(e.get("ticker", "")) in held_base:
                promotion_log["held_removed_from_vci"].append({
                    "ticker": e.get("ticker"), "reason": "now held in portfolio"})
            else:
                new_vci.append(e)
        wl_data["vci_watchlist"] = new_vci

    # -----------------------------------------------------------------------
    # Find all Growth Stock Analysis xlsx files
    # -----------------------------------------------------------------------
    # Ingest analysis xlsx from BOTH the working dir and THIS MONTH's archive, so
    # the (ephemeral) candidate pool rebuilds completely and idempotently on every
    # run — even after a prior run already archived the source files.
    archive_month_dir   = os.path.join(inv_dir, XLSX_ARCHIVE_SUBDIR, run_date[:7])
    working_xlsx_files  = sorted(glob.glob(os.path.join(inv_dir, "Growth Stock Analysis*.xlsx")))
    archived_xlsx_files = sorted(glob.glob(os.path.join(archive_month_dir, "Growth Stock Analysis*.xlsx")))
    _seen = {os.path.basename(p) for p in working_xlsx_files}
    xlsx_files = sorted(working_xlsx_files + [p for p in archived_xlsx_files
                                             if os.path.basename(p) not in _seen])

    # Screen-freshness guard: warn if the newest growth screen is stale (>35 days).
    SCREEN_STALE_DAYS = 35
    import re as _re
    from datetime import datetime as _DT
    _newest = None
    for _p in xlsx_files:
        _m = _re.search(r"W-e\s+(\d{2})-([A-Za-z]{3})-(\d{2})", os.path.basename(_p))
        if _m:
            try:
                _d = _DT.strptime(f"{_m.group(1)}-{_m.group(2)}-{_m.group(3)}", "%d-%b-%y").date()
                if _newest is None or _d > _newest:
                    _newest = _d
            except Exception:
                pass
    if _newest is not None:
        _age = (date.today() - _newest).days
        promotion_log["screen_freshness"] = {"newest_screen": _newest.isoformat(), "age_days": _age}
        if _age > SCREEN_STALE_DAYS:
            _w = (f"STALE SCREEN: newest growth file is {_age} days old "
                  f"({_newest.isoformat()}). Refresh Growth Stock Analysis xlsx so the "
                  f"top-10 reflects the current universe.")
            promotion_log.setdefault("warnings", []).append(_w)
            print(f"  [update_watchlist] WARNING: {_w}")

    if not xlsx_files:
        print("  [update_watchlist] No Growth Stock Analysis xlsx files found — nothing to promote.")

    # -----------------------------------------------------------------------
    # Phase 1 & 2 & 3: Read files, classify paths, apply exclusions
    # -----------------------------------------------------------------------
    candidate_pool: dict[str, dict] = {}  # ticker → best row

    for xlsx_path in xlsx_files:
        fname = os.path.basename(xlsx_path)
        print(f"  [update_watchlist] Reading: {fname}")
        promotion_log["xlsx_files_read"].append(fname)
        rows = read_xlsx_summary_tab(xlsx_path)
        promotion_log["rows_parsed"] += len(rows)

        for row in rows:
            ticker = row["ticker"]

            # Phase 1: classify path
            path = classify_path(row, fname)
            if path is None:
                # Misrouted energy stock
                promotion_log["misrouted_energy"].append({
                    "ticker": ticker,
                    "file":   fname,
                    "note":   "Energy sector in non-ENERGY file — scored on Path A; advisory only",
                })
                promotion_log["path_c_misrouted_log"].append(ticker)
                continue

            row["path"] = path

            # Phase 2: hard exclusions
            if ticker in portfolio_tickers:
                continue  # already held

            if ticker in CHINA_VIE_EXCLUSIONS:
                continue  # China VIE

            total = row.get("total")
            if total is None:
                continue  # no score — can't rank

            # Watchlist duplicate check: only exclude if not a material score improvement
            if ticker in current_watchlist:
                prior_total = current_watchlist[ticker].get("total") or 0
                # Allow into candidate pool — Phase 4 will handle score refresh
                # (existing names always refreshed if they appear in SUMMARY)

            # Phase 3: deduplicate — keep highest normalised score per ticker
            # Normalised score comparison handles cross-path duplicates correctly
            # (a Path A score of 40/54 vs Path C score of 28/36 both normalise to ~74%)
            new_ns = normalised_score(total, row.get("path", "A")) or 0.0
            existing_ns = (
                normalised_score(
                    candidate_pool[ticker].get("total", 0),
                    candidate_pool[ticker].get("path", "A")
                ) or 0.0
            ) if ticker in candidate_pool else 0.0

            if ticker not in candidate_pool or new_ns > existing_ns:
                # If duplicate appeared, log it
                if ticker in candidate_pool and new_ns != existing_ns:
                    promotion_log.setdefault("duplicate_ticker_log", []).append({
                        "ticker": ticker,
                        "kept_file": row.get("source_file", ""),
                        "kept_score": new_ns,
                        "replaced_score": existing_ns,
                    })
                candidate_pool[ticker] = row

    # -----------------------------------------------------------------------
    # Phase 4: Existing watchlist score refresh
    # -----------------------------------------------------------------------
    updated_existing: dict[str, dict] = {}

    for ticker, entry in current_watchlist.items():
        override_reason = entry.get("eligibility_override_reason")  # None or string

        if ticker in candidate_pool:
            # This month's SUMMARY has a score for this ticker
            new_row   = candidate_pool[ticker]
            new_total = new_row.get("total")
            new_a     = new_row.get("part_a")
            new_b     = new_row.get("part_b")
            new_path  = new_row.get("path", entry.get("path", "A"))

            prior_total = entry.get("total")
            prior_a     = entry.get("part_a")
            prior_b     = entry.get("part_b")

            # Push current score to history
            history = list(entry.get("score_history", []))
            if prior_total is not None:
                if not history or history[0].get("month") != month_label:
                    history.insert(0, {
                        "month":      month_label,
                        "part_a":     prior_a,
                        "part_b":     prior_b,
                        "total":      prior_total,
                        "normalised": entry.get("normalised_score"),
                    })
                    history = history[:MAX_SCORE_HISTORY]

            delta = (new_total - prior_total) if (new_total is not None and prior_total is not None) else None
            # Updated call: pass part_a deltas for better attribution
            pb_driver = classify_part_b_driver(
                prior_total, new_total, prior_a, new_a, prior_b, new_b
            )
            ns = normalised_score(new_total, new_path)

            # --- Probation / removal logic for existing names ---
            probation_flag = False
            removal_flag   = False
            if ns is not None:
                if ns < NORMALISED_SCORE_HARD_REMOVE_BELOW and not override_reason:
                    removal_flag = True
                elif ns < NORMALISED_SCORE_PROMOTION_GATE and not override_reason:
                    probation_flag = True
            # (_removal_flag entries are excluded in Phase 6 ranking)

            updated_entry = dict(entry)
            updated_entry.update({
                "path":             new_path,
                "total":            new_total,
                "part_a":           new_a,
                "part_b":           new_b,
                "normalised_score": ns,
                "score_history":    history,
                "delta_score":      delta,
                "probation_flag":   probation_flag,
                "_removal_flag":    removal_flag,  # underscore prefix = internal, not persisted to JSON
            })
            if pb_driver:
                updated_entry["part_b_driver"] = pb_driver
            elif "part_b_driver" in updated_entry:
                del updated_entry["part_b_driver"]

            if _fluid_on():
                # Confirmed in this cycle's fresh screens -> reset decay. Score-driven removal
                # (removal_flag above, ns < hard floor) still applies — turnover stays score-driven.
                updated_entry["first_seen"]         = entry.get("first_seen") or month_label
                updated_entry["last_confirmed"]     = month_label
                updated_entry["decay_state"]        = "below_floor" if removal_flag else "active"
                updated_entry["reconfirm_required"] = False
                updated_entry["in_pool"]            = not removal_flag

            updated_existing[ticker] = updated_entry

            if delta is not None:
                promotion_log["score_updates"].append({
                    "ticker":       ticker,
                    "delta_score":  delta,
                    "part_b_driver": pb_driver,
                    "probation":    probation_flag,
                    "removed":      removal_flag,
                })
        else:
            # Not in this month's SUMMARY tabs — carry forward
            updated_entry = dict(entry)
            ns = updated_entry.get("normalised_score")
            if _fluid_on():
                # Absence is NOT thesis-break (gated screens are noisy) -> FLAG for re-confirmation,
                # don't auto-eject. Turnover is driven by score (hard floor) + a long-stop age-out;
                # names with a confirmed catalyst / override are protected from the age-out.
                # M4: first fluid run for this entry (no last_confirmed) -> seed to THIS month (grace),
                # NOT the old score_history month, so activating fluidity doesn't mass-age-out incumbents.
                lc = entry.get("last_confirmed") or month_label
                updated_entry["first_seen"]     = entry.get("first_seen") or lc
                updated_entry["last_confirmed"] = lc
                months_since, age_out = _pool_decay_decision(lc, month_label)
                protected = bool(override_reason or entry.get("catalyst_protected")
                                 or entry.get("confirmed_catalyst"))
                removal_flag = False
                probation_flag = False
                if protected:
                    updated_entry["decay_state"] = "protected_catalyst"
                    updated_entry["reconfirm_required"] = True
                elif age_out:
                    removal_flag = True            # LONG-STOP backstop (no catalyst/override)
                    updated_entry["decay_state"] = "aged_out"
                    promotion_log["stale_scores"].append({
                        "ticker": ticker, "last_confirmed": lc, "months_since": months_since,
                        "reason": (f"long-stop age-out — not re-screened {months_since}m "
                                   f"(> {_ageout_months()}) and no catalyst/override"),
                    })
                else:
                    updated_entry["decay_state"] = "stale_unconfirmed"
                    updated_entry["reconfirm_required"] = True
                    probation_flag = True          # surfaces for review re-confirmation, NOT removed
                    promotion_log.setdefault("reconfirm_required", []).append({
                        "ticker": ticker, "last_confirmed": lc, "months_since": months_since,
                    })
                    if ns is not None and months_since:
                        updated_entry["_decayed_score"] = max(0.0, ns - _decay_penalty() * months_since)
                # Score-driven removal still applies to a carried sub-floor score (unless protected).
                if (not protected) and ns is not None and ns < NORMALISED_SCORE_HARD_REMOVE_BELOW and not override_reason:
                    removal_flag = True
                    updated_entry["decay_state"] = "below_floor"
                updated_entry["in_pool"]        = not removal_flag
                updated_entry["probation_flag"] = probation_flag
                updated_entry["_removal_flag"]  = removal_flag
                updated_existing[ticker] = updated_entry
            else:
                # --- legacy (score_history-length) staleness — pre-fluid behaviour ---
                history = list(entry.get("score_history", []))
                stale = False
                if history and len(history) >= MAX_SCORE_HISTORY:
                    stale = True
                    updated_entry["stale_score_flag"] = True
                    promotion_log["stale_scores"].append({
                        "ticker":       ticker,
                        "last_screened": history[0].get("month", "unknown"),
                    })
                # Existing names absent from SUMMARY are not removed automatically.
                # They carry forward at their last known score unless stale + below floor.
                removal_flag = False
                probation_flag = False
                if stale and ns is not None and ns < NORMALISED_SCORE_HARD_REMOVE_BELOW and not override_reason:
                    removal_flag = True
                elif ns is not None and ns < NORMALISED_SCORE_PROMOTION_GATE and not override_reason:
                    probation_flag = True
                updated_entry["probation_flag"]  = probation_flag
                updated_entry["_removal_flag"]   = removal_flag
                updated_existing[ticker] = updated_entry

    # -----------------------------------------------------------------------
    # Phase 5: Provisional entry levels for new candidates
    # -----------------------------------------------------------------------
    new_candidates: dict[str, dict] = {}

    for ticker, row in candidate_pool.items():
        if ticker in current_watchlist:
            continue  # already handled in Phase 4

        path  = row.get("path", "A")
        total = row.get("total", 0)
        ns    = normalised_score(total, path) or 0.0
        override = row.get("eligibility_override_reason")

        # GATE: ns >= 70 for promotion (default) OR Part A viability floor (FORWARD_ELIGIBILITY).
        if not _pool_admit(ns, row.get("part_a"), path, override):
            promotion_log["rejected_candidates"].append({
                "ticker":           ticker,
                "normalised_score": ns,
                "path":             path,
                "source_file":      row.get("source_file", ""),
                "reason":           f"normalised_score {ns:.1f} < {NORMALISED_SCORE_PROMOTION_GATE} threshold",
            })
            continue

        # Provisional entry level
        target  = row.get("target_price")
        current = row.get("current_price")
        if target and target > 0:
            entry_level = round(target * 0.75, 2)
        elif current and current > 0:
            entry_level = round(current * 0.88, 2)
        else:
            entry_level = None

        new_candidates[ticker] = {
            "ticker":                  ticker,
            "name":                    row.get("company", ticker),
            "exchange":                row.get("exchange", "NASDAQ"),
            "entry_level":             entry_level,
            "entry_currency":          row.get("currency", "USD"),
            "entry_level_provisional": True,
            "sector":                  row.get("sector", ""),
            "source_pipeline":         "growth_stock" if path == "A" else "energy",
            "path":                    path,
            "total":                   total,
            "part_a":                  row.get("part_a"),
            "part_b":                  row.get("part_b"),
            "normalised_score":        ns,
            "delta_score":             None,
            "score_history":           [],
            "status":                  "Watchlist — newly promoted",
            "thesis_break_summary":    "[Claude fills at Step 10 — thesis-break conditions]",
            "probation_flag":          False,
            "eligibility_override_reason": override,
        }
        if _fluid_on():
            new_candidates[ticker].update({
                "first_seen": month_label, "last_confirmed": month_label,
                "decay_state": "active", "in_pool": True, "reconfirm_required": False,
            })

    # -----------------------------------------------------------------------
    # Phase 6: Combined ranking and top-10 selection
    # -----------------------------------------------------------------------
    combined: dict[str, dict] = {}
    combined.update(updated_existing)
    for ticker, entry in new_candidates.items():
        if ticker not in combined:
            combined[ticker] = entry

    # Remove currently-held names from the growth watchlist (positions, not candidates).
    if held_base:
        for _t in list(combined.keys()):
            if _base_ticker(_t) in held_base:
                promotion_log["held_removed_from_watchlist"].append({
                    "ticker": _t, "reason": "now held in portfolio — moved to stock sleeve"})
                combined.pop(_t, None)

    def rank_key(item):
        # Fluid pool: stale-but-carried names rank on their decayed score so freshly-confirmed
        # names displace them; falls back to normalised_score when no decay applied.
        s = item.get("_decayed_score")
        if s is None:
            s = item.get("normalised_score") or 0
        return -s

    # Separate removal candidates before ranking
    removal_candidates = {t: e for t, e in combined.items() if e.get("_removal_flag")}
    rankable = {t: e for t, e in combined.items() if not e.get("_removal_flag")}

    ranked = sorted(rankable.values(), key=rank_key)

    # Top MAX_WATCHLIST
    new_wl  = ranked[:MAX_WATCHLIST]
    removed = ranked[MAX_WATCHLIST:]

    # Log names removed by score displacement or removal_flag
    new_wl_tickers  = {e["ticker"] for e in new_wl}
    old_wl_tickers  = set(current_watchlist.keys())

    added_tickers   = new_wl_tickers - old_wl_tickers
    removed_tickers = old_wl_tickers - new_wl_tickers

    for ticker in added_tickers:
        entry = combined[ticker]
        promotion_log["additions"].append({
            "ticker":           ticker,
            "normalised_score": entry.get("normalised_score"),
            "path":             entry.get("path"),
            "source_file":      candidate_pool.get(ticker, {}).get("source_file", ""),
        })

    for e in removed:
        if e["ticker"] in old_wl_tickers:
            new_top10_ns = (new_wl[-1].get("normalised_score") or 0) if new_wl else 0
            promotion_log["removals"].append({
                "ticker": e["ticker"],
                "reason": (
                    f"ranked out — normalised score {e.get('normalised_score', 'N/A')}, "
                    f"below new #10 at {new_top10_ns}"
                ),
            })

    # Explicitly removed by floor / stale+floor combination
    for ticker, e in removal_candidates.items():
        if ticker in old_wl_tickers:
            promotion_log["removals"].append({
                "ticker": ticker,
                "reason": (
                    f"removed — normalised_score {e.get('normalised_score', 'N/A')} "
                    f"below hard floor ({NORMALISED_SCORE_HARD_REMOVE_BELOW}) "
                    + ("and no override" if not e.get("eligibility_override_reason") else "")
                    + (" and stale score" if e.get("stale_score_flag") else "")
                ),
            })

    # Assign ranks to new_wl
    for i, entry in enumerate(new_wl, 1):
        entry["rank"] = i

    # Strip internal (_-prefixed) fields before writing to JSON
    INTERNAL_FIELDS = {"_removal_flag", "_decayed_score"}
    for entry in new_wl:
        for f in INTERNAL_FIELDS:
            entry.pop(f, None)

    # --- Build candidate_pool list: all >=70 that are NOT in new_wl ---
    new_wl_set = {e["ticker"] for e in new_wl}
    pool_entries = []
    for ticker, row in candidate_pool.items():
        ns = normalised_score(row.get("total", 0), row.get("path", "A")) or 0.0
        override = row.get("eligibility_override_reason")
        if not _pool_admit(ns, row.get("part_a"), row.get("path", "A"), override):
            continue  # below gate — already in rejected_candidates log
        if ticker in new_wl_set:
            continue  # already in top-10 — not duplicated in candidate_pool
        if ticker in {s.get("ticker", "") for s in wl_data.get("stock_sleeve", [])}:
            continue  # already held
        if ticker in {e.get("ticker", "") for e in wl_data.get("vci_watchlist", [])}:
            continue  # already in VCI watchlist

        target  = row.get("target_price")
        current_p = row.get("current_price")
        if target and target > 0:
            entry_level = round(target * 0.75, 2)
        elif current_p and current_p > 0:
            entry_level = round(current_p * 0.88, 2)
        else:
            entry_level = None

        pool_entries.append({
            "ticker":                  ticker,
            "name":                    row.get("company", ticker),
            "exchange":                row.get("exchange", "NASDAQ"),
            "path":                    row.get("path", "A"),
            "source_pipeline":         "growth_stock" if row.get("path", "A") == "A" else "energy",
            "normalised_score":        round(ns, 1),
            "total":                   row.get("total"),
            "part_a":                  row.get("part_a"),
            "part_b":                  row.get("part_b"),
            "entry_level":             entry_level,
            "entry_currency":          row.get("currency", "USD"),
            "entry_level_provisional": True,
            "sector":                  row.get("sector", ""),
            "candidate_pool_month":    month_label,
            "eligibility_override_reason": override,
            **({"first_seen": _prior_first_seen.get(ticker) or month_label,
                "last_confirmed": month_label, "decay_state": "active",
                "in_pool": True, "reconfirm_required": False} if _fluid_on() else {}),
        })

    # Sort pool by normalised_score descending
    pool_entries.sort(key=lambda x: -(x.get("normalised_score") or 0))

    # -----------------------------------------------------------------------
    # Phase 7: Write outputs
    # -----------------------------------------------------------------------
    if not dry_run:
        wl_data["watchlist"] = new_wl

        # --- Membership is fresh-screen-driven (anti-stickiness, design decision 2026-06-07) ---
        # The candidate pool membership = ONLY names freshly screened this cycle (>=70); names
        # displaced from the top-10 or absent from this cycle's screen are NOT retained as pool
        # members. The pool represents the CURRENT opportunity set, not an accumulation of prior
        # months. WHEN FLUID_POOL_DECAY is on, prior_pool is used ONLY to carry each name's
        # first_seen forward (durable-pool memory) — membership stays fresh-screen-driven.

        # Write candidate_pool (ephemeral — overwrites any prior month's pool)
        wl_data["candidate_pool"] = pool_entries
        wl_data["_candidate_pool_meta"] = {
            "last_updated":       run_date,
            "pool_month":         month_label,
            "gate_threshold_pct": NORMALISED_SCORE_PROMOTION_GATE,
            "pool_size":          len(pool_entries),
            "note": "EPHEMERAL — wiped and rewritten each pre-run by update_watchlist.py. Do not manually edit.",
        }

        wl_data["_meta"]["last_updated"]   = run_date
        wl_data["_meta"]["updated_by_run"] = f"{month_label} pre-run — update_watchlist.py"

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(wl_data, f, indent=2, ensure_ascii=False)

        print(f"  [update_watchlist] watchlist_tickers.json written: {len(new_wl)} watchlist + {len(pool_entries)} candidate_pool entries")

        # Archive working-dir xlsx files (move to archive/<month>/). Files already in
        # the archive are left in place; re-runs re-read them via the archive glob.
        archive_dir = os.path.join(inv_dir, XLSX_ARCHIVE_SUBDIR, run_date[:7])  # e.g. archive/2026-06/
        if working_xlsx_files:
            os.makedirs(archive_dir, exist_ok=True)

        for xlsx_path in working_xlsx_files:
            try:
                dest = os.path.join(archive_dir, os.path.basename(xlsx_path))
                if os.path.abspath(xlsx_path) == os.path.abspath(dest):
                    continue
                if os.path.exists(dest):
                    os.remove(dest)
                shutil.move(xlsx_path, dest)
                fname = os.path.basename(xlsx_path)
                promotion_log["xlsx_files_archived"].append({
                    "file":        fname,
                    "archived_to": dest,
                })
                print(f"  [update_watchlist] Archived: {fname} → {archive_dir}")
            except Exception as e:
                print(f"  [update_watchlist] Could not archive {os.path.basename(xlsx_path)}: {e}")
                # Fall back to delete if move fails (network drive unavailable etc.)
                try:
                    os.remove(xlsx_path)
                    promotion_log["xlsx_files_deleted"].append(os.path.basename(xlsx_path))
                    print(f"  [update_watchlist] Fallback deleted: {os.path.basename(xlsx_path)}")
                except Exception as e2:
                    print(f"  [update_watchlist] Could not delete either: {e2}")
    else:
        print(f"  [update_watchlist] [DRY RUN] Would write {len(new_wl)} watchlist + {len(pool_entries)} pool entries")
        print(f"  [update_watchlist] [DRY RUN] Would archive {len(xlsx_files)} xlsx file(s) to archive/{run_date[:7]}/")

    # Print summary
    print(f"  [update_watchlist] Additions: {[e['ticker'] for e in promotion_log['additions']]}")
    print(f"  [update_watchlist] Removals:  {[e['ticker'] for e in promotion_log['removals']]}")
    print(f"  [update_watchlist] Score updates: {len(promotion_log['score_updates'])}")
    print(f"  [update_watchlist] Candidate pool: {len(pool_entries)} entries")
    print(f"  [update_watchlist] Rejected (below 70% gate): {len(promotion_log['rejected_candidates'])}")
    if promotion_log["misrouted_energy"]:
        print(f"  [update_watchlist] Misrouted energy: {promotion_log['path_c_misrouted_log']}")
    if promotion_log["stale_scores"]:
        print(f"  [update_watchlist] Stale scores: {[e['ticker'] for e in promotion_log['stale_scores']]}")
    if promotion_log["duplicate_ticker_log"]:
        print(f"  [update_watchlist] Cross-file duplicates resolved: {len(promotion_log['duplicate_ticker_log'])}")

    # Emit the promotion log as JSON on stdout for the orchestrator to capture
    print(json.dumps(promotion_log))

    return promotion_log


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Watchlist promotion, score refresh, and ranking. "
            "Runs as Step 4 of monthly_isa_prerun.py pipeline."
        )
    )
    parser.add_argument("--portfolio-data", required=False, default=None,
                        help="Path to portfolio_data_mmm_yyyy.json (for portfolio duplicate check)")
    parser.add_argument("--watchlist-json", required=True,
                        help="Path to watchlist_tickers.json (current watchlist)")
    parser.add_argument("--inv-dir", required=True,
                        help="Investment Analysis folder path (to locate Growth Stock Analysis xlsx files)")
    parser.add_argument("--out-json", required=True,
                        help="Output path for updated watchlist_tickers.json (usually same as input)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print what would be done without writing or deleting files")
    parser.add_argument("--fluid-pool", action="store_true",
                        help="Activate fluid-pool decay for this run (overrides scoring_config.FLUID_POOL_DECAY)")
    args = parser.parse_args()

    if args.fluid_pool and _cfg is not None:
        _cfg.FLUID_POOL_DECAY = True

    run(
        portfolio_path = args.portfolio_data or "",
        watchlist_path = args.watchlist_json,
        inv_dir        = args.inv_dir,
        out_path       = args.out_json,
        dry_run        = args.dry_run,
    )


if __name__ == "__main__":
    main()
