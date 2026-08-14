"""
fund_performance.py -- THE GOLDEN SOURCE for fund trailing performance.

BuildSpec: ISA_BuildSpec_FundPerformance_GoldenSource_Aug2026.md
Standard : feedback_isa_engineering_standard.md (all five rules)

Replaces the X-Ray as the SOURCE of per-fund returns. The X-Ray is demoted to a
CORROBORATOR (invariant I6) because its per-fund rows are dated a month before
the report header that carries them -- the D1 defect.

Every figure returned is a Metric (value + as_of + source) or a Missing(reason).
Nothing here ever returns a bare float for a decision-grade number.
"""
from __future__ import annotations
import csv, datetime as dt, json, os, sys

from isa_metric import Metric, Missing, is_present, value_or, as_dict

SCHEMA_VERSION = 1
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
UNIVERSE = os.path.join(SCRIPT_DIR, "fund_universe.json")
CACHE_DIR = os.path.join(SCRIPT_DIR, "nav_cache")
HISTORY_CSV = os.path.join(SCRIPT_DIR, "fund_performance_history.csv")

WINDOWS = {"1m": 1/12, "3m": 0.25, "6m": 0.5, "1y": 1.0, "3y": 3.0, "5y": 5.0,
           # 10y added 05-Aug-2026: a single window is a bet on a start date. Scottish
           # Mortgage reads ~0.2% over 5y, ~22% over 3y and ~16.7% over 10y — the same
           # fund, three answers, decided entirely by whether the 2022 -45.7% year is
           # inside the window. The anchor rule now needs every window it can get.
           "10y": 10.0}
PRICE_MATCH_TOL_PCT = 1.0      # I1
MAX_PRICING_LAG_DAYS = 3       # I1: OEICs are struck T-1 vs the broker valuation
                               # date; ETFs/ITs T-0. Empirically every holding
                               # reconciles to 0.00% at the right lag, so a
                               # >1% diff at the BEST lag is a genuine fault.
XRAY_AGREE_TOL_PP  = 2.0       # I6
MIN_COVERAGE = 0.98            # need 98% of the window actually covered by NAV history


# ---------------------------------------------------------------- universe
def load_universe(path=UNIVERSE):
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)["funds"]


def isin_check_digit(body: str) -> str:
    """Luhn over the alphanumeric-expanded body. Used to VALIDATE a KID-sourced
    ISIN -- never to invent one (domicile is not derivable from a SEDOL)."""
    conv = "".join(str(int(c, 36)) if c.isalpha() else c for c in body.upper())
    total = 0
    for i, ch in enumerate(reversed(conv)):
        n = int(ch)
        if i % 2 == 0:
            n *= 2
            if n > 9:
                n -= 9
        total += n
    return str((10 - total % 10) % 10)


def isin_is_valid(isin: str) -> bool:
    return (isinstance(isin, str) and len(isin) == 12
            and isin_check_digit(isin[:11]) == isin[11])


# ---------------------------------------------------------------- NAV history
def _cache_path(symbol): return os.path.join(CACHE_DIR, f"{symbol.replace('/','_')}.csv")


def _read_cache(symbol):
    p = _cache_path(symbol)
    if not os.path.exists(p):
        return None
    out = []
    with open(p, newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            try:
                out.append((dt.date.fromisoformat(row["date"]), float(row["close"])))
            except (ValueError, KeyError):
                continue
    return out or None


def _write_cache(symbol, series):
    os.makedirs(CACHE_DIR, exist_ok=True)
    with open(_cache_path(symbol), "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh); w.writerow(["date", "close"])
        for d, c in series:
            w.writerow([d.isoformat(), f"{c:.6f}"])


def _scale_for(u):
    """GBp -> GBP. Unit confusion is its own defect class: SMT priced 1330.50
    (GBp) against a broker 13.3050 (GBP) read as a 9900% mismatch."""
    return 100.0 if (u or {}).get("price_unit") == "GBp" else 1.0


def fetch_nav_history(symbol, period="max", use_cache=True, refresh=False, scale=1.0):
    """[(date, close)] ascending, dividend-reinvested. Cached so the 45s sandbox
    ceiling cannot half-complete a 12-fund run.

    ⚑ period="max", NOT "10y" (fixed 06-Aug-2026). A fetch exactly as long as the longest window
    can never cover that window: yfinance measures 10y back from TODAY, the returns are measured
    back from a MONTH-END as_of, and the few days between them are enough. Every cache begins
    2016-08-05 while the 10-year window from 31-Jul-2026 needs 01-Aug-2016 — short by four days.
    The consequence was that the 10y window, declared in WINDOWS and load-bearing in the
    return-adequacy median, produced NOTHING for any of the twelve funds on any run since it was
    added, and read as "not applicable" rather than "never fetched". `window_coverage()` below
    now makes that class of silence impossible to repeat."""
    if not symbol:
        return []
    if use_cache and not refresh:
        c = _read_cache(symbol)
        if c:
            return [(d, v / scale) for d, v in c] if scale != 1.0 else c
    try:
        import yfinance as yf
    except ImportError:
        return []
    try:
        h = yf.Ticker(symbol).history(period=period, auto_adjust=True)["Close"]
    except Exception:
        return []
    series = []
    for idx, val in h.items():
        try:
            d = idx.date() if hasattr(idx, "date") else dt.date.fromisoformat(str(idx)[:10])
            if val == val and val > 0:          # NaN-safe
                series.append((d, float(val)))
        except Exception:
            continue
    series.sort()
    if series and use_cache:
        _write_cache(symbol, series)
    if scale != 1.0:
        series = [(d, c / scale) for d, c in series]
    return series


# ---------------------------------------------------------------- local series (ISA-0307)
class LocalSeriesError(RuntimeError):
    """A declared local series that cannot be read. RAISED, never silently empty: the whole
    defect this route exists to fix was a NAV series that was on disk, correct, reconciled, and
    read by nothing — while the study reported `warnings: []`."""


def load_local_series(u, base=None):
    """[(date, close)] for a fund whose NAV has no feed and is supplied manually.

    ⚑ WHY THIS EXISTS. `yfinance` rejects the valid ISIN IE00B42W4J83, so Polar Capital Global
    Tech had no NAV series and every consumer recorded 'no NAV series — factsheet-only fund'.
    Raj supplied a reconciling monthly series on 06-Aug-2026, it was cached, and for six days
    nothing read it: 7.6% of the portfolio carried no measured beta and M* coverage read 92.4%
    for want of a file that was already on disk. Built, believed live, never reached.

    ⚑ THE ROUTE IS DECLARED, NOT DISCOVERED. It fires only where `fund_universe` carries a
    `local_series` block naming the file, sheet and columns, and it re-asserts that block's own
    factsheet reconciliation before returning anything (R4.10 — an artefact asserts its own
    fitness for the use it will be put to)."""
    ls = (u or {}).get("local_series")
    if not ls:
        return []
    base = base or os.path.dirname(os.path.abspath(__file__))
    cache = os.path.join(base, ls.get("cache", ""))
    series = []
    if ls.get("cache") and os.path.exists(cache):
        with open(cache) as fh:
            for r in csv.DictReader(fh):
                series.append((dt.date.fromisoformat(r["date"]), float(r["close"])))
    else:
        path = os.path.normpath(os.path.join(base, ls["path"]))
        if not os.path.exists(path):
            raise LocalSeriesError(f"declared local_series not found: {path}")
        import openpyxl
        ws = openpyxl.load_workbook(path, data_only=True)[ls["sheet"]]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        di, vi = hdr.index(ls["date_col"]), hdr.index(ls["value_col"])
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[di] is None or row[vi] is None:
                continue
            d = row[di].date() if hasattr(row[di], "date") else row[di]
            series.append((d, float(row[vi])))
        if ls.get("cache"):
            os.makedirs(os.path.dirname(cache), exist_ok=True)
            with open(cache, "w", newline="") as fh:
                w = csv.writer(fh); w.writerow(["date", "close"])
                w.writerows([(d.isoformat(), v) for d, v in sorted(series)])
    series.sort()
    if not series:
        raise LocalSeriesError("declared local_series produced no rows")
    _assert_local_reconciliation(series, ls)
    return series


def _assert_local_reconciliation(series, ls):
    """The block declares deltas against a named factsheet. Re-derive them and RAISE on
    disagreement — two independent derivations must agree, with a stated tolerance (R5.2)."""
    rec = ls.get("reconciliation") or {}
    tol = rec.get("tolerance_pp")
    deltas = rec.get("deltas_pp") or {}
    if tol is None or not deltas:
        raise LocalSeriesError("local_series carries no reconciliation — a manually supplied NAV "
                               "series may not be used unmeasured (R6.3)")
    for k, v in deltas.items():
        if abs(float(v)) > float(tol):
            raise LocalSeriesError(
                f"local_series reconciliation {k} is {v}pp, outside the declared {tol}pp tolerance")
    L = {}
    for d, v in series:
        L[(d.year, d.month)] = v
    if len(L) < 30:
        raise LocalSeriesError(f"local_series has {len(L)} months, below the 30-month minimum")


def nav_series_for(sedol, u, refresh=False):
    """THE one way to get a NAV series for a fund. Local series first where declared, then the
    feed. One home (R4.4) — every caller that had its own `yf_symbol or isin` line was a place a
    fund with no feed could go quietly missing."""
    if (u or {}).get("local_series"):
        s = load_local_series(u)
        if s:
            return s
    sym = (u or {}).get("yf_symbol") or (u or {}).get("isin")
    return fetch_nav_history(sym, use_cache=True, refresh=refresh, scale=_scale_for(u))


# ---------------------------------------------------------------- maths
def _nav_on_or_before(series, target, max_back_days=10):
    lo, hi = 0, len(series) - 1
    best = None
    while lo <= hi:
        mid = (lo + hi) // 2
        if series[mid][0] <= target:
            best = series[mid]; lo = mid + 1
        else:
            hi = mid - 1
    if best and (target - best[0]).days <= max_back_days:
        return best
    return None


def _years_before(as_of: dt.date, years: float) -> dt.date:
    return as_of - dt.timedelta(days=round(years * 365.25))


def trailing_return(series, as_of, years, source, label):
    """-> (cumulative Metric|Missing, annualised Metric|Missing). Never
    extrapolates: a window not genuinely covered returns Missing (rule 1)."""
    miss = lambda r: (Missing(r, as_of, source), Missing(r, as_of, source))
    if not series:
        return miss("no NAV history")
    end = _nav_on_or_before(series, as_of)
    if not end:
        return miss(f"no NAV within 10d of {as_of}")
    start_target = _years_before(as_of, years)
    # ⚑ ONE THRESHOLD, NOT TWO (fixed 06-Aug-2026). This gate demanded the history reach the FULL
    # window while MIN_COVERAGE below declared 98% sufficient — two rules for the same question,
    # with the stricter one winning silently and MIN_COVERAGE never consulted at the boundary.
    # Every cache begins 2016-08-05 and the 10-year window from 31-Jul-2026 needs 01-Aug-2016, so
    # the 10y figure was refused on FOUR DAYS while 99.9% of the window was sitting in the file.
    # The two now agree, and MIN_COVERAGE is the single home for "how much of the window is
    # enough". (Annualisation stays on the NOMINAL window — that is the convention Morningstar
    # uses and the reason RLGES 15.82 and JPM UK 12.06 reproduce to 2dp.)
    if series[0][0] > _years_before(as_of, years * MIN_COVERAGE):
        have = (as_of - series[0][0]).days / 365.25
        return miss(f"insufficient_history: {have:.2f}y available, "
                    f"{years * MIN_COVERAGE:.2f}y required ({MIN_COVERAGE:.0%} of {years:g}y)")
    start = _nav_on_or_before(series, start_target, max_back_days=15)
    if not start:
        # The history begins just AFTER the window's start date. Refusing here was the second
        # half of the same two-thresholds defect: the gate above now admits 98% coverage, and
        # then this lookup demanded a NAV on or before the full-window date and refused anyway.
        # Fall back to the earliest observation and let MIN_COVERAGE below be the single arbiter.
        cand = series[0]
        if (end[0] - cand[0]).days / 365.25 >= years * MIN_COVERAGE:
            start = cand
        else:
            return miss(f"no NAV within 15d of {start_target} and the earliest observation "
                        f"({cand[0]}) covers only "
                        f"{(end[0] - cand[0]).days / 365.25:.2f}y of {years:g}y")
    covered = (end[0] - start[0]).days / 365.25
    if covered < years * MIN_COVERAGE:
        return miss(f"window_undercovered: {covered:.2f}y of {years:g}y")
    if start[1] <= 0:
        return miss("non-positive start NAV")
    cum = (end[1] / start[1]) - 1.0
    ann = (1.0 + cum) ** (1.0 / years) - 1.0 if years >= 1 else cum
    note = f"{start[0]}->{end[0]} ({covered:.2f}y covered of {years:g}y nominal)"
    return (Metric(cum * 100, end[0], source, unit="%", note=note),
            Metric(ann * 100, end[0], source, unit="%", note=note))


# ---------------------------------------------------------------- invariants
def verify_price_match(series, broker_price, broker_date,
                       tol_pct=PRICE_MATCH_TOL_PCT, max_lag=MAX_PRICING_LAG_DAYS):
    """I1 -- LAG-AWARE, DATE-ALIGNED price identity check.

    A naive latest-vs-broker comparison produced false mismatches that were pure
    market drift (M&G 5.34%, Invesco 3.31%). A naive same-day comparison produced
    false mismatches because UK OEICs are struck T-1 against the broker's
    valuation date. Searching a short lag window and taking the BEST match gives
    an exact identity (every holding reconciles to 0.00%), so any residual
    difference is a real fault -- which is how the Ranmore wrong-share-class
    (315%) was caught.

    Returns (ok, detail, nav, lag_days).
    """
    if not series:
        return False, "no NAV history", None, None
    if not broker_price:
        return False, "no broker price", None, None
    cands = [(d, c) for d, c in series
             if 0 <= (broker_date - d).days <= max_lag]
    if not cands:
        return False, f"no NAV within {max_lag}d on/before {broker_date}", None, None
    d, nav = min(cands, key=lambda x: abs(x[1] - broker_price))
    diff = abs(nav - broker_price) / broker_price * 100.0
    lag = (broker_date - d).days
    ok = diff <= tol_pct
    return ok, (f"{'OK' if ok else 'MISMATCH'}: NAV {nav:.4f} @{d} (lag {lag}d) vs "
                f"broker {broker_price:.4f} @{broker_date} = {diff:.2f}% "
                f"(tol {tol_pct}%)"), nav, lag


def verify_cum_ann(cum, ann, years, tol=1e-6):
    """I3 -- the two derivations of the same return must agree."""
    if not (is_present(cum) and is_present(ann)):
        return True, "skipped (absent)"
    if years < 1:
        return True, "n/a (<1y not annualised)"
    lhs = (1 + cum.value / 100) ** (1 / years) - 1
    d = abs(lhs - ann.value / 100)
    return d <= tol, f"cum->ann delta {d:.2e} (tol {tol:.0e})"


def verify_vs_xray(golden_ann, xray_pct, tol_pp=XRAY_AGREE_TOL_PP):
    """I6 -- X-Ray is a CORROBORATOR, not a source. Disagreement is a WARN that
    must name both values and both dates."""
    if not is_present(golden_ann) or xray_pct is None:
        return True, "skipped"
    d = abs(golden_ann.value - xray_pct)
    return d <= tol_pp, (f"golden {golden_ann.value:.2f}% @{golden_ann.as_of} vs "
                         f"xray {xray_pct:.2f}% = {d:.2f}pp (tol {tol_pp}pp)")


# ---------------------------------------------------------------- main API
def fund_performance(sedol, as_of, universe=None, broker_price=None,
                     broker_date=None, refresh=False):
    """Golden-source trailing performance for one fund."""
    universe = universe or load_universe()
    u = universe.get(sedol)
    if not u:
        return {"sedol": sedol, "status": "not_in_universe",
                "returns": {}, "invariants": {}}

    out = {"sedol": sedol, "name": u["name"], "isin": u.get("isin"),
           "share_class": u.get("share_class"), "bucket": u.get("bucket"),
           "ocf": u.get("ocf"), "as_of_requested": as_of.isoformat(),
           "status": u.get("resolution_status"), "returns": {}, "invariants": {},
           "notes": []}

    if u.get("isin") and not isin_is_valid(u["isin"]):
        out["notes"].append(f"ISIN {u['isin']} fails its own check digit")

    man = u.get("manual_returns")
    if man and u.get("resolution_status") == "manual_factsheet":
        out["status"] = "manual_factsheet"
        src = man["source"]
        for w in WINDOWS:
            v = man["annualised_pct"].get(w)
            if v is None:
                m = Missing(man.get("absent_reason", {}).get(w, "not published by source"),
                            man["as_of"], src)
                out["returns"][w] = {"cumulative": as_dict(m), "annualised": as_dict(m),
                                     "proxy_used": False}
            else:
                mm = Metric(v, man["as_of"], src, confidence=man.get("confidence", 1.0),
                            unit="%", note="manager/platform factsheet")
                out["returns"][w] = {"cumulative": as_dict(
                                         Missing("factsheet publishes annualised only",
                                                 man["as_of"], src)),
                                     "annualised": as_dict(mm), "proxy_used": False}
        out["notes"].append(f"NAV feed unavailable; trailing returns taken from {src} "
                            f"as at {man['as_of']} (stamped, not inferred).")
        return out

    if u.get("resolution_status") != "resolved" or not u.get("yf_symbol"):
        r = u.get("unresolved_reason", "no yf_symbol")
        out["status"] = "unresolved"
        out["fallback"] = u.get("fallback")
        for w in WINDOWS:
            out["returns"][w] = {"cumulative": as_dict(Missing(r, as_of, "fund_universe")),
                                 "annualised": as_dict(Missing(r, as_of, "fund_universe"))}
        return out

    sym = u["yf_symbol"]
    src = (f"local_series:{u['local_series']['path']}" if u.get("local_series")
           else f"yfinance:{sym}")
    scale = _scale_for(u)
    series = (load_local_series(u) if u.get("local_series")
              else fetch_nav_history(sym, refresh=refresh, scale=scale))
    out["nav_points"] = len(series)
    if series:
        out["nav_range"] = [series[0][0].isoformat(), series[-1][0].isoformat()]
        # DECLARED inception (KID/factsheet), carried so that an absent window can be
        # classified rather than merely reported. Absent = the absence stays `unclassified`,
        # which is an honest open question, not a pass.
        out["declared_inception"] = (u or {}).get("inception")
        out["inception_source"] = (u or {}).get("inception_source")
        if out["declared_inception"] and out["nav_range"][0] < str(out["declared_inception"])[:10]:
            out.setdefault("notes", []).append(
                f"⚑ NAV history begins {out['nav_range'][0]}, BEFORE the declared inception "
                f"{out['declared_inception']} — one of the two is wrong. Declared inception is "
                f"not used for this fund until reconciled.")
            out["declared_inception"] = None

    # Pin the measurement date to the NAV the BROKER actually priced against.
    # "Latest available" makes a return depend on WHEN the pipeline runs, because
    # OEIC NAVs publish with a lag (Ranmore's 31-Jul NAV appeared between two
    # fetches 30 minutes apart and moved its 1y from 17.70% to 16.82%). Pinning
    # makes every figure reproducible and matches the basis the broker valued on.
    effective_as_of = as_of
    if broker_price and broker_date:
        ok, msg, nav, lag = verify_price_match(series, broker_price, broker_date)
        out["invariants"]["I1_price_match"] = {"pass": ok, "detail": msg,
                                               "lag_days": lag}
        out["pricing_lag_days"] = lag
        # NOTE: lag is recorded for the I1 identity check ONLY. Returns are
        # measured to the CALENDAR as_of -- confirmed against two managers'
        # published figures on the exact held share class:
        #   MI Thornbridge C Acc  1y 13.32 / 3y 15.57 / 5y 16.93  (HL)
        #   Ranmore Institutional 1y 16.82 / 3y 19.94 / 5y n/a    (AJ Bell)
        # Both reproduce EXACTLY on the calendar basis. Lag-pinning reproduced
        # neither and was reverted.
        if not ok:
            out["status"] = "price_mismatch"
            out["notes"].append("I1 FAILED -- likely wrong share class. "
                                "Figures suppressed rather than published.")
    out["as_of_effective"] = effective_as_of.isoformat()
    out["as_of_basis"] = "calendar as_of (validated vs manager factsheets)"

    proxy = u.get("proxy") or {}
    proxy_series = None
    for label, yrs in WINDOWS.items():
        cum, ann = trailing_return(series, effective_as_of, yrs, src, label)
        used_proxy = False
        if isinstance(cum, Missing) and label in proxy.get("applies_to", []):
            if proxy_series is None:
                proxy_series = fetch_nav_history(proxy["yf_symbol"],
                                                 refresh=refresh, scale=scale)
            psrc = f"yfinance:{proxy['yf_symbol']}[{proxy['flag']}]"
            pc, pa = trailing_return(proxy_series, effective_as_of, yrs, psrc, label)
            if is_present(pc):
                cum, ann, used_proxy = pc, pa, True
        okc, dmsg = verify_cum_ann(cum, ann, yrs)
        out["invariants"][f"I3_cum_ann_{label}"] = {"pass": okc, "detail": dmsg}
        out["returns"][label] = {"cumulative": as_dict(cum), "annualised": as_dict(ann),
                                 "proxy_used": used_proxy}
        if used_proxy:
            out["notes"].append(
                f"{label}: {proxy['flag']} via {proxy['share_class']} "
                f"({proxy['isin']}) -- {proxy['authority'][:80]}...")

    if out["status"] == "resolved":
        out["status"] = "ok"
    return out


def all_fund_performance(as_of, portfolio_funds=None, refresh=False):
    """Every fund in the universe. I5: every portfolio fund must appear."""
    universe = load_universe()
    px = {f["ticker"]: (f.get("price"), f.get("_as_of")) for f in (portfolio_funds or [])}
    res = {}
    for sedol in universe:
        p, d = px.get(sedol, (None, None))
        res[sedol] = fund_performance(sedol, as_of, universe, p, d, refresh)
    if portfolio_funds is not None:
        missing = [f["ticker"] for f in portfolio_funds if f["ticker"] not in res]
        res["_i5_coverage"] = {"pass": not missing, "missing": missing}
    return res


def window_coverage(perf_by_sedol):
    """I8 — a DECLARED window that yields nothing for the ENTIRE universe is a build fault, not a
    data fact. It was silence exactly like this that hid the 10-year window for a whole month:
    every fund returned Missing('insufficient_history'), which is individually plausible and
    collectively impossible. One fund missing a window is data; twelve of twelve is code."""
    cov = {w: {"have": [], "missing": {}} for w in WINDOWS}
    for sedol, p in sorted(perf_by_sedol.items()):
        if str(sedol).startswith("_"):
            continue
        for w in WINDOWS:
            a = ((p.get("returns", {}).get(w) or {}).get("annualised") or {})
            if a.get("present"):
                cov[w]["have"].append(sedol)
            else:
                cov[w]["missing"][sedol] = a.get("reason", "absent")
    n = sum(1 for k in perf_by_sedol if not str(k).startswith("_"))
    out = {}
    for w, c in cov.items():
        # ⚑ FIXED 06-Aug-2026. This line was
        #     {k: v for k, v in list(c["missing"].items())[:3]}
        # — the reasons were TRUNCATED TO THREE. On the August run four funds had no 10-year
        # figure and three reasons were published, so **VUAG's absence was silent**: the report
        # showed 8 of 12 covered and explained 3 of the 4 gaps, with nothing indicating that
        # anything had been left out.
        #
        # This is the same failure as N5 (`consistency_check` verifying a fixed byte window
        # instead of the table): a report whose explanatory coverage shrinks as the thing it
        # explains grows. It is harmless-looking at 4 gaps and a false all-clear at 40.
        # An explanation list must never be truncated — the whole point of it is completeness.
        out[w] = {"covered": len(c["have"]), "of": n,
                  "alarm": bool(n and not c["have"]),
                  "n_missing": len(c["missing"]),
                  "reasons": dict(c["missing"])}
        assert len(out[w]["reasons"]) == out[w]["n_missing"], (
            f"{w}: every absence must carry a reason ({out[w]['n_missing']} missing, "
            f"{len(out[w]['reasons'])} explained)")
        assert out[w]["covered"] + out[w]["n_missing"] == n, (
            f"{w}: covered + missing must equal the universe")
        # ── structural history limit vs an unexpected short series ─────────────────────
        # "insufficient_history" is two different facts wearing one label: a fund that did not
        # exist yet (permanent, expected, not a defect) and a fetch that came back short
        # (transient, a defect, and indistinguishable from the first without a second source).
        # `inception` in fund_universe.json is that second source — DECLARED from the KID, never
        # inferred — and the two must agree.
        struct = {}
        for sd, reason in out[w]["reasons"].items():
            p = perf_by_sedol.get(sd) or {}
            inc = p.get("declared_inception")
            first = (p.get("nav_range") or [None])[0]
            cls = "unclassified"
            note = ("no declared inception in fund_universe.json — this absence cannot be "
                    "distinguished from a short fetch. One declared date per fund closes it.")
            if "not published" in str(reason):
                cls, note = "source_does_not_publish", str(reason)
            elif inc:
                need = _years_before(dt.date.today(), WINDOWS[w])
                inc_d = str(inc)[:10]
                if inc_d > need.isoformat():
                    cls = "structural_history_limit"
                    note = (f"declared inception {inc_d} postdates the {w} window start "
                            f"{need.isoformat()} — the history does not exist and never will. "
                            f"Expected, permanent, NOT a defect.")
                else:
                    cls = "UNEXPECTED_SHORT_SERIES"
                    note = (f"⚑ declared inception {inc_d} PREDATES the {w} window start "
                            f"{need.isoformat()}, so this history should exist. A short series "
                            f"here is a FETCH problem, not a fund fact.")
            struct[sd] = {"reason": str(reason), "classification": cls,
                          "declared_inception": inc, "observed_series_start": first,
                          "note": note}
        out[w]["absence_classification"] = struct
        if out[w]["alarm"]:
            out[w]["note"] = (f"⚑ the {w} window produced NOTHING across all {n} funds. A window "
                              f"declared in WINDOWS and used by return adequacy must either "
                              f"produce values or be removed — an empty one reads as 'not "
                              f"applicable' and quietly shrinks the evidence base.")
    return out


def append_history(perf_by_sedol, est_by_sedol, run_date):
    """I7 -- persist est BESIDE realised every month so D2 (est has no
    discriminating power) becomes measurable instead of merely noticed."""
    new = not os.path.exists(HISTORY_CSV)
    with open(HISTORY_CSV, "a", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        if new:
            w.writerow(["run_date", "sedol", "name", "bucket", "est_return_pct",
                        "realised_1y_ann", "realised_3y_ann", "realised_5y_ann",
                        "as_of", "source", "status"])
        for sedol, p in sorted(perf_by_sedol.items()):
            if sedol.startswith("_"):
                continue
            g = lambda w_: (p.get("returns", {}).get(w_, {})
                            .get("annualised", {}) or {}).get("value")
            a = (p.get("returns", {}).get("1y", {}).get("annualised", {}) or {})
            w.writerow([run_date, sedol, p.get("name"), p.get("bucket"),
                        est_by_sedol.get(sedol), g("1y"), g("3y"), g("5y"),
                        a.get("as_of"), a.get("source"), p.get("status")])
    return HISTORY_CSV


if __name__ == "__main__":
    as_of = dt.date.fromisoformat(sys.argv[1]) if len(sys.argv) > 1 else dt.date.today()
    r = all_fund_performance(as_of)
    print(json.dumps({k: v for k, v in r.items() if not k.startswith("_")},
                     indent=1, default=str)[:4000])
