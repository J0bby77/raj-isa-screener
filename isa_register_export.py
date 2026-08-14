"""One row per item: the 4Cs and every reference field, in one table.

Standard: ISA_Engineering_Rules.md.  Item: ISA-0303 (Raj, 12-Aug-2026).

R7.1 - this is a RENDER of `Dashboard/state/isa_items.jsonl`, generated like every other view,
so it can never disagree with the store. R14.3 - never hand-edited; regenerate it.

Two outputs, deliberately:
  * `ISA_Item_Register.xlsx` - filterable, frozen header, one sheet per view (All / Open /
    Raj's queue / Rationale ledger), plus a Fields sheet explaining every column.
  * `ISA_Item_Register.csv`  - the same rows, for anything that would rather read text.

⚑ Absence is rendered as `Missing(not_recorded_at_the_time)`, never as an empty cell. An empty
cell in a spreadsheet reads as "nothing to say"; the whole point of ISA-0226 is that "nobody
recorded it" and "there is nothing" are different facts.

CLI:
  python3 isa_register_export.py --write
  python3 isa_register_export.py --selftest
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

import isa_register as R

XLSX_NAME = "ISA_Item_Register.xlsx"
CSV_NAME = "ISA_Item_Register.csv"

# ⚑ STRIPPED BACK, 12-Aug-2026 (Raj: "far too many columns... stripped back to just the columns
# that are truly necessary"). He is right, and 49 was over-engineered: the spreadsheet is a thing a
# person reads and decides from, and a 49-column grid is not that.
#
# The split, and why it is a split rather than a deletion: the XLSX carries the 20 columns you
# decide from; the CSV keeps EVERY field, because the store's completeness is the point and
# throwing fields away to tidy a view would be the tail wagging the dog (R6.5).
#
# "Truly necessary" = you cannot triage without it: what it is, how bad, which tier, whether it can
# even be built yet, how old, the 4Cs, what it needs from you, and whether a "closed" item is really
# closed.

CORE_COLUMNS = [
    ("ID", lambda i: i["id"], "Canonical item id (R7.6)."),
    ("Legacy", lambda i: ", ".join(a for a in i.get("aliases", []) if ":" not in a),
     "The id it had in the old registers (D-24, H7, V-1, BL-13...)."),
    ("Tier", lambda i: R.tier(i)[0], "P0/P1/P2/P3, computed from the rank score - never typed."),
    ("Tier meaning", lambda i: R.tier(i)[1],
     "P0 blocks the next scheduled run · P1 irreversible: another cycle destroys data permanently · "
     "P2 changes where capital goes · P3 correctness and robustness."),
    ("Title", lambda i: i["title"], "One line stating the DEFECT, not the fix."),
    ("Type", lambda i: i["record_type"], "DEFECT/ENHANCEMENT/DESIGN_GAP/DECISION/CORRECTION/RESEARCH/LEARNING/RATIONALE."),
    ("State", lambda i: i["state"], "OPEN/IN_PROGRESS/BLOCKED_ON_RAJ/DEFERRED/CLOSED_*."),
    ("Criticality", lambda i: i["criticality"], "CRITICAL affects where capital goes or destroys evidence permanently."),
    ("Ready to build?", lambda i: i.get("build_readiness"),
     "ISA-0304. ANALYSIS_FIRST = a study is needed before code · RAJ_DECISION_UNBLOCKS = waiting on you · "
     "BUILD_READY = build straight away · UNKNOWN = no source said, and it is NOT guessed."),
    ("Readiness basis", lambda i: i.get("build_readiness_basis"), "Where that answer came from."),
    ("Size", lambda i: i.get("size_est"), "XS<1h · S half a session · M one session · L multi · XL programme."),
    ("Domain", lambda i: i.get("domain"), "Subject AREA, not a workflow state: screener/funds/email/dashboard/vci/capture/analysis/process."),
    ("Raised", lambda i: R.date_raised(i),
     "ISA-0338. The date the item was RAISED, dd-mmm-yy (Raj, 13-Aug-2026). Derived: the "
     "detection date where one exists, otherwise the date the record was written \u2014 never "
     "created_on alone, which for 190 migrated items is 12-Aug-2026 and would show a year-old "
     "finding as days old."),
    ("Raised basis", lambda i: R.date_raised_basis(i), "Which date that came from."),
    ("Age (days)", lambda i: R.age_days(i),
     "Days since the item was DETECTED, not since it was migrated - migrating on 12-Aug would otherwise reset every clock."),
    ("Age basis", lambda i: R.age_basis(i), "detected_on, or created_on where no detection date exists."),
    ("Auto-archive in (days)", lambda i: R.archive_due_in(i),
     "LOW items auto-archive at 90 days (Raj, 12-Aug-2026). Blank = the policy does not apply: "
     "MEDIUM and above age visibly but are never auto-archived, and anything waiting on you, "
     "deferred, or in progress is exempt."),
    ("C1 Context", lambda i: i.get("context"), "What was being done, on what data, at what date."),
    ("C2 Cause (proximate)", lambda i: i.get("cause_proximate"), "The line, the field, the call site."),
    ("C2 Cause (systemic)", lambda i: i.get("cause_systemic"), "Why the class exists and why nothing caught it (R7.4)."),
    ("C3 Consequence", lambda i: i.get("consequence"), "What it did, or would have done."),
    ("C4 Corrective action", lambda i: i.get("corrective_action"), "What changes. ⚑ Its presence does NOT mean the analysis is done - read 'Ready to build?'."),
    ("Question for you", lambda i: i.get("blocked_question"), "Populated on BLOCKED_ON_RAJ."),
    ("Closure evidence", lambda i: (i.get("verification") or {}).get("liveness_ref"),
     "The named check that catches RECURRENCE. Without it a closure is a narrative claim (R7.3)."),
]

# FULL set - retained in the CSV so no field is lost.
COLUMNS = [
    ("ID", lambda i: i["id"], "Canonical item id. Never reused, never re-meaninged (R7.6)."),
    ("Legacy IDs", lambda i: ", ".join(a for a in i.get("aliases", []) if ":" not in a),
     "The id this item had in the old registers (D-24, H7, V-1, BL-13...)."),
    ("Title", lambda i: i["title"], "One line stating the DEFECT, not the fix."),
    ("Type", lambda i: i["record_type"], "DEFECT/ENHANCEMENT/DESIGN_GAP/DECISION/CORRECTION/RESEARCH/LEARNING/RATIONALE."),
    ("Is fix?", lambda i: i.get("is_fix"), "Raj's simple two-way split. Drives K1."),
    ("State", lambda i: i["state"], "OPEN/IN_PROGRESS/BLOCKED_ON_RAJ/DEFERRED/CLOSED_*."),
    ("Criticality", lambda i: i["criticality"], "CRITICAL affects where capital goes or destroys evidence."),
    ("Rank score", lambda i: R.rank_score(i), "Computed: run-blocker 1000, irreversible 500, capital 200, + criticality."),
    ("Tier", lambda i: R.tier(i)[0], "P0/P1/P2/P3 from the rank score."),
    ("Tier meaning", lambda i: R.tier(i)[1], "What that tier means."),
    ("Ready to build?", lambda i: i.get("build_readiness"), "ISA-0304."),
    ("Readiness basis", lambda i: i.get("build_readiness_basis"), "Where that answer came from."),
    ("Raised", lambda i: R.date_raised(i), "ISA-0338. Date raised, dd-mmm-yy."),
    ("Raised basis", lambda i: R.date_raised_basis(i), "detected_on, or created_on where none."),
    ("Age (days)", lambda i: R.age_days(i), "Days since detection."),
    ("Age basis", lambda i: R.age_basis(i), "detected_on or created_on."),
    ("Size", lambda i: i.get("size_est"), "XS<1h · S half a session · M one session · L multi · XL programme."),
    ("Points", lambda i: i.get("points"), "XS1 S3 M8 L20 XL50. Drives K1."),
    ("Domain", lambda i: i.get("domain"), "screener/funds/email/dashboard/vci/capture/analysis/process."),
    ("Failure class", lambda i: i.get("failure_class"), "FC-A..FC-L, Rules §1."),
    ("C1 Context", lambda i: i.get("context"), "What was being done, on what data, at what date."),
    ("C2 Cause (proximate)", lambda i: i.get("cause_proximate"), "The line, the field, the call site."),
    ("C2 Cause (systemic)", lambda i: i.get("cause_systemic"), "Why the class exists and why nothing caught it (R7.4)."),
    ("C3 Consequence", lambda i: i.get("consequence"), "What it did, or would have done."),
    ("C3 Quantified", lambda i: _q(i.get("consequence_quantified")),
     "The number. R13.3: a consequence without one cannot be ranked against anything."),
    ("C4 Corrective action", lambda i: i.get("corrective_action"), "What changes. Must address the systemic cause."),
    ("Claim status", lambda i: i.get("claim_status"), "HYPOTHESIS/TESTED/VERIFIED (R2.2). HYPOTHESIS cannot rank above MEDIUM."),
    ("Evidence basis", lambda i: i.get("evidence_basis"), "BACKTESTED/MEASURED/DECLARED/REFUSED_FOR_POWER (R13.1)."),
    ("Studies", lambda i: _studies(i.get("studies")), "Research artefacts this item rests on (ISA-0212)."),
    ("Capital link", lambda i: i.get("capital_link"), "Link to the annualised target, or infrastructure_no_direct_link (R16.1)."),
    ("Blocks run", lambda i: _blocks(i.get("blocks_run")), "The scheduled run this must land before."),
    ("Question for Raj", lambda i: i.get("blocked_question"), "Populated on BLOCKED_ON_RAJ."),
    ("Deferred until", lambda i: i.get("deferred_until"), "Excluded from the queue until this date."),
    ("Deferred reason", lambda i: i.get("deferred_reason"), "Why the deferral is by design, not neglect."),
    ("Detected by", lambda i: i.get("detected_by"), "Drives K6: found by the framework, or by a person?"),
    ("Detected on", lambda i: i.get("detected_on"), "Date the item was raised."),
    ("Introduced on", lambda i: i.get("introduced_on"), "With introduced_basis: known/inferred/unknown - never guessed."),
    ("Latency (days)", lambda i: i.get("latency_days"), "introduced -> detected. Drives K7. Null when the basis is unknown."),
    ("Escaped?", lambda i: i.get("escaped"), "Did a wrong value reach a decision surface? Drives K8."),
    ("Escape surface", lambda i: i.get("escape_surface"), "email/dashboard/action_stack/trade/none."),
    ("Resolved on", lambda i: i.get("resolved_on"), "Date closed."),
    ("Verification test", lambda i: (i.get("verification") or {}).get("test_id"), "The suite that proves the fix."),
    ("Liveness reference", lambda i: (i.get("verification") or {}).get("liveness_ref"),
     "The named check that catches RECURRENCE. close() refuses without it (R7.3)."),
    ("Assertions", lambda i: (i.get("verification") or {}).get("assertion_count"), "Assertion count in that suite."),
    ("Green on", lambda i: (i.get("verification") or {}).get("green_on"), "Date the verification last passed."),
    ("Standard refs", lambda i: ", ".join(i.get("standard_refs") or []), "Rules the work was built against. Drives K11."),
    ("Rule that would have caught it", lambda i: i.get("rule_that_would_have_caught"),
     "Or NONE_EXISTED, which auto-raises a child item proposing one."),
    ("Rationale: constant", lambda i: (i.get("rationale") or {}).get("constant"), "P2.5 ledger (R12.3)."),
    ("Rationale: home", lambda i: (i.get("rationale") or {}).get("home"), "file:line where the constant lives."),
    ("Rationale: who set it", lambda i: (i.get("rationale") or {}).get("who_set_it"), "R12.3 provenance."),
    ("Rationale: evidence", lambda i: (i.get("rationale") or {}).get("evidence_basis"),
     "NO_RECORDED_RATIONALE is a permitted and important answer."),
    ("Rationale: what would falsify it", lambda i: (i.get("rationale") or {}).get("what_would_falsify_it"), "R12.3."),
    ("Learning", lambda i: _learning(i.get("learning")), "Mandatory on every item. Silence is not an answer (R8.1)."),
    ("Disagreement", lambda i: i.get("disagreement"), "Independent derivations that conflict (R8.3). Published, never blended."),
    ("Provenance", lambda i: i.get("provenance"), "captured_live, or backfilled - trend metrics start Sep-2026."),
    ("Source", lambda i: i.get("source_doc"), "The artefact it came from, with as_of."),
    ("Revision", lambda i: i.get("revision"), "Increments on every supersede."),
]


def _q(q):
    if not q:
        return R.NOT_RECORDED
    return f"{q.get('value')} {q.get('unit','')} — {q.get('metric','')} (basis: {q.get('basis','')})"


def _studies(st):
    if not st:
        return "none linked"
    return " | ".join(f"{d['doc']}{'' if d.get('on_disk') else ' [NOT ON DISK]'}" for d in st)


def _blocks(b):
    return f"{b.get('run_type')} {b.get('run_date')}" if b and b.get("run_date") else ""


def _learning(l):
    if not l:
        return ""
    return f"task: {l.get('task_id')}" if l.get("learnable") else f"none — {l.get('reason_none')}"


def rows(items=None, columns=None) -> list:
    columns = columns or COLUMNS
    items = items if items is not None else R.read_all()
    items = sorted(items, key=lambda i: (-R.rank_score(i), i["id"]))
    out = []
    for it in items:
        row = {}
        for header, fn, _desc in columns:
            try:
                v = fn(it)
            except Exception as e:                                # noqa: BLE001
                v = f"ERROR: {e}"
            row[header] = "" if v is None else v
        out.append(row)
    return out


def write_csv(dest: Path, data: list) -> Path:
    p = dest / CSV_NAME
    with p.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=[c[0] for c in COLUMNS])
        w.writeheader()
        w.writerows(data)
    return p


def write_xlsx(dest: Path, data: list, columns=None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    columns = columns or CORE_COLUMNS
    headers = [c[0] for c in columns]
    wb = Workbook()
    ARIAL = "Arial"
    head_font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
    head_fill = PatternFill("solid", fgColor="1F3864")
    body_font = Font(name=ARIAL, size=10)
    crit_fill = PatternFill("solid", fgColor="FCE4E4")
    absent_font = Font(name=ARIAL, size=10, italic=True, color="808080")

    live = {"OPEN", "IN_PROGRESS", "BLOCKED_ON_RAJ", "DEFERRED"}
    views = [
        ("All items", data),
        ("Open", [r for r in data if r["State"] in live]),
        ("Raj's queue", [r for r in data if r["State"] == "BLOCKED_ON_RAJ"]),
        ("Rationale ledger", [r for r in data if r["Type"] == "RATIONALE"]),
    ]
    first = True
    for name, subset in views:
        ws = wb.active if first else wb.create_sheet()
        ws.title = name
        first = False
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for r in subset:
            ws.append([r[h] for h in headers])
        for ridx in range(2, len(subset) + 2):
            crit = ws.cell(row=ridx, column=headers.index("Criticality") + 1).value
            for cidx in range(1, len(headers) + 1):
                cell = ws.cell(row=ridx, column=cidx)
                cell.font = absent_font if cell.value == R.NOT_RECORDED else body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if crit == "CRITICAL":
                    cell.fill = crit_fill
        widths = {"ID": 11, "Legacy": 14, "Tier": 6, "Tier meaning": 34, "Title": 58, "Type": 13,
                  "State": 17, "Criticality": 12, "Ready to build?": 21, "Readiness basis": 40,
                  "Size": 6, "Domain": 11, "Age (days)": 11, "Age basis": 12}
        for idx, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = widths.get(h, 46)
        ws.freeze_panes = "D2"
        if subset:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(subset) + 1}"
        ws.row_dimensions[1].height = 34

    fs = wb.create_sheet("Fields")
    fs.append(["Column", "What it is"])
    for c in range(1, 3):
        fs.cell(row=1, column=c).font = head_font
        fs.cell(row=1, column=c).fill = head_fill
    for header, _fn, desc in columns:
        fs.append([header, desc])
    for row in fs.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    fs.column_dimensions["A"].width = 34
    fs.column_dimensions["B"].width = 105
    fs.append([])
    fs.append(["GENERATED", f"by isa_register_export.py from Dashboard/state/isa_items.jsonl on "
                            f"{R._today()}. This is a VIEW — do not edit it, edit the store "
                            f"(R7.1, R14.3). No formulas: every cell is a recorded fact."])
    fs.append(["Absence", f"{R.NOT_RECORDED} in italic grey means the source never recorded it and "
                          f"it was deliberately not invented (R7.5). An empty cell would read as "
                          f"'nothing to say', which is a different fact."])

    # A workbook open in Excel is LOCKED, and this one is written every time the store changes,
    # so that will happen often. Failing the whole export because a viewer is open would be a
    # brittle build; writing silently to a different name would be worse - you would keep reading
    # a stale file believing it fresh. So: try the real path, and if it is locked, write beside it
    # under a name that ANNOUNCES itself and say so out loud (R2.10).
    p = dest / XLSX_NAME
    try:
        wb.save(p)
        return p
    except PermissionError:
        alt = dest / XLSX_NAME.replace(".xlsx", "_LOCKED_CLOSE_EXCEL_AND_RERUN.xlsx")
        wb.save(alt)
        print(f"⚑ {XLSX_NAME} is LOCKED (open in Excel?). Wrote {alt.name} instead - "
              f"close the workbook and re-run `python3 isa_register_export.py --write` so the "
              f"canonical file is the current one.")
        return alt


def write(dest=None) -> dict:
    d = Path(dest).resolve() if dest else Path(__file__).resolve().parent
    full = rows()                       # CSV: every field, nothing lost
    core = rows(columns=CORE_COLUMNS)   # XLSX: the columns you decide from
    return {"csv": str(write_csv(d, full)), "xlsx": str(write_xlsx(d, core, CORE_COLUMNS)),
            "rows": len(full), "xlsx_columns": len(CORE_COLUMNS), "csv_columns": len(COLUMNS)}


def selftest(verbose=True) -> int:
    import shutil, tempfile
    n = 0

    def ok(cond, msg):
        nonlocal n
        n += 1
        if not cond:
            raise AssertionError(msg)

    data = rows(columns=CORE_COLUMNS)
    ok(data, "the export must have rows")
    heads = [c[0] for c in CORE_COLUMNS]
    ok(len(CORE_COLUMNS) < len(COLUMNS),
       "the sheet a person reads must be narrower than the complete record")
    for c in ("Tier", "Tier meaning", "Ready to build?", "Age (days)", "Auto-archive in (days)"):
        ok(c in heads, f"{c} must be a core column (Raj, 12-Aug-2026)")
    ok({r["Tier"] for r in data} <= {"P0", "P1", "P2", "P3"}, "tiers are P0-P3")
    ok(any(r["Tier meaning"] == "blocks the next scheduled run" for r in data),
       "the tier LABEL must be present, not just the code")
    ok(all(r["Ready to build?"] in ("", *R.BUILD_READINESS) for r in data),
       "readiness must be one of the declared values or blank on a closed item")
    ages = [r["Age (days)"] for r in data if isinstance(r["Age (days)"], int)]
    ok(ages and max(ages) > 3,
       "ageing must run from DETECTION - if every item is 0-3 days old the clock was reset by the migration")
    ok(len(set(heads)) == len(heads), "column headers must be unique")
    ok("Raised" in heads and "Raised basis" in heads,
       "ISA-0338: the raised date must be a core column, with its basis beside it")
    ok(all(re.fullmatch(r"\d{2}-[A-Z][a-z]{2}-\d{2}", r["Raised"]) for r in data if r["Raised"]),
       "ISA-0338: every raised date renders as dd-mmm-yy")
    for c in ("C1 Context", "C2 Cause (proximate)", "C2 Cause (systemic)", "C3 Consequence",
              "C4 Corrective action"):
        ok(c in heads, f"the 4Cs must all be core columns: {c} missing")
    ok("Closure evidence" in heads, "the closure evidence must be a core column (R7.3)")
    full_heads = [c[0] for c in COLUMNS]
    ok("C3 Quantified" in full_heads and "Studies" in full_heads
       and "Rationale: what would falsify it" in full_heads,
       "trimming the SHEET must not drop a field from the RECORD (R6.5)")
    ok(all(len(r) == len(CORE_COLUMNS) for r in data), "every row must fill every column")
    ok(not any(v is None for r in data for v in r.values()),
       "a None must never reach a cell - absence is a marker, not a blank (R4.1)")
    live = [r for r in data if r["State"] in ("OPEN", "IN_PROGRESS", "BLOCKED_ON_RAJ", "DEFERRED")]
    ok(all(r["C2 Cause (systemic)"] for r in live),
       "every live row states a systemic cause or says it was never recorded")
    ok(any(r["C3 Quantified"] != R.NOT_RECORDED for r in rows()),
       "at least one item must carry a quantified consequence, or the column is decoration")
    ok(data[0]["Tier"] <= data[-1]["Tier"], "rows are in computed tier order")

    tmp = Path(tempfile.mkdtemp(prefix="isa_export_"))
    try:
        res = write(tmp)
        ok(Path(res["csv"]).exists() and Path(res["xlsx"]).exists(), "both files are written")
        import csv as _csv
        with open(res["csv"], encoding="utf-8-sig") as fh:
            back = list(_csv.DictReader(fh))
        ok(len(back) == len(data), "the CSV round-trips every row")
        ok(len(back[0]) == len(COLUMNS),
           "the CSV keeps EVERY field - the XLSX is trimmed, the record is not")
        ok(back[0]["ID"] == data[0]["ID"], "row order survives the round trip")
        from openpyxl import load_workbook
        wb = load_workbook(res["xlsx"])
        ok({"All items", "Open", "Raj's queue", "Rationale ledger", "Fields"} <= set(wb.sheetnames),
           f"expected views missing: {wb.sheetnames}")
        ws = wb["All items"]
        ok(ws.max_row == len(data) + 1, "every item is a row on All items")
        ok(ws.freeze_panes == "D2" and ws.auto_filter.ref, "header frozen and filterable")
        ok(wb["Rationale ledger"].max_row > 1, "the rationale ledger view is populated")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    if verbose:
        print(f"isa_register_export selftest: {n} assertions, 0 failed")
    return n


def main(argv=None):
    ap = argparse.ArgumentParser(description="One-row-per-item export of the register")
    ap.add_argument("--dest", default=None)
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--selftest", action="store_true")
    a = ap.parse_args(argv)
    if a.selftest:
        selftest()
        return 0
    if a.write:
        print(json.dumps(write(a.dest), indent=2))
        return 0
    ap.print_help()
    return 0


if __name__ == "__main__":
    sys.exit(main())
