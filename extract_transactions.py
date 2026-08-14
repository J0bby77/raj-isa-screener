#!/usr/bin/env python3
"""
extract_transactions.py  --  AJ Bell ISA transaction history parser
Version: 1.0  |  2026-07-26

Reads a monthly AJ Bell transaction export and folds it into a PERSISTENT
transaction ledger, so the monthly review reconciles executions from BROKER
TRUTH (what was actually traded, when, at what price and cost) rather than
inferring them from month-to-month holdings deltas.

Expected monthly file (ISA root folder):
    Transaction History MM-YYYY.xlsx        e.g. "Transaction History 07-2026.xlsx"

Columns (AJ Bell standard export):
    Date | Transaction | Description | Quantity | Price | Amount (GBP) | Reference

Outputs
-------
  transaction_ledger.json            PERSISTENT, append-and-dedupe, never
                                     deleted at post-run cleanup. Full history.
  transactions_data_[mmm_yyyy].json  Per-month view consumed by the pre-run
                                     (deleted at post-run cleanup like the
                                     other *_data_[mmm_yyyy].json files).

What it derives
---------------
  * normalised transaction type (buy / sell / transfer_in / distribution /
    equalisation / conversion / unclassified)
  * ticker, resolved from the free-text broker description
  * true dealing cost per trade:  |Amount - (Quantity x Price)|  and cost as a
    % of gross consideration  -> calibrates the Step 10.6 cost estimate
  * cash impact per row (buy negative, sell positive, in-specie/acc rows zero)
  * FIFO round-trips per ticker: realised P&L, holding period in days
    -> feeds the trades log, min-hold rule (audit C-1) and the learning loop
  * an explicit unmapped-description report. Rows are NEVER silently dropped.

Degradation
-----------
A missing monthly file is a WARNING, not an error (exit code 3): the caller
falls back to the legacy holdings-delta inference in
decision_ledger.reconcile_executions.

Usage
-----
  python3 extract_transactions.py --isa-folder /path/to/ISA
  python3 extract_transactions.py --xlsx "/path/Transaction History 07-2026.xlsx"
  python3 extract_transactions.py --seed "/path/transactionhistory ... .xlsx"
  python3 extract_transactions.py --report
"""

import argparse
import glob
import hashlib
import json
import os
import re
import sys
from datetime import datetime

import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_ISA_FOLDER = os.path.abspath(os.path.join(SCRIPT_DIR, ".."))

LEDGER_SCHEMA_VERSION = 1

# Monthly export naming convention agreed with Raj (26-Jul-2026).
MONTHLY_GLOB = "Transaction History [0-9][0-9]-[0-9][0-9][0-9][0-9].xlsx"
MONTHLY_RE = re.compile(r"Transaction History (\d{2})-(\d{4})\.xlsx$", re.I)

EXPECTED_HEADER = ["date", "transaction", "description", "quantity",
                   "price", "amount (gbp)", "reference"]

# Broker transaction label -> normalised type.
TYPE_MAP = {
    "purchase":                  "buy",
    "sale":                      "sell",
    "transfer in":               "transfer_in",
    "transfer out":              "transfer_out",
    "equalisation acc units":    "equalisation",
    "accumulation distribution": "distribution",
    "dividend":                  "distribution",
    "income distribution":       "distribution",
    "fund class conversion":     "conversion",
    "subscription":              "cash_in",
    "contribution":              "cash_in",
    "withdrawal":                "cash_out",
}

# Cash direction by normalised type. Rows that move units but not cash are 0.
CASH_SIGN = {
    "buy": -1, "sell": +1, "cash_in": +1, "cash_out": -1,
    "transfer_in": 0, "transfer_out": 0, "equalisation": 0,
    "distribution": 0, "conversion": 0, "unclassified": 0,
}

# Types that represent an actual dealing execution (what the ledger reconciles).
TRADE_TYPES = {"buy", "sell"}

# Descriptions that will never appear in a saved portfolio file because the
# position has been fully closed, or was bought after the last snapshot.
# Extend as needed -- the unmapped report tells you when.
STATIC_TICKER_OVERRIDES = {
    "adobe inc":                               "ADBE",
    "ptc inc":                                 "PTC",
    "paylocity hldg cor com usd0.001":         "PCTY",
    "paylocity hlding cor com usd0.001":       "PCTY",
    "paylocity hlding corp com usd0.001":      "PCTY",
    "marks & spencer group plc":               "MKS",
    "abcellera biologics inc ordinary shares": "ABCL",
    "micron technology inc":                   "MU",
    "broadcom inc":                            "AVGO",
    "oxford nanopore technologies plc":        "ONT",
    # Fund share classes that no longer appear in any saved portfolio file
    # because they were converted away. Suffixed so they never collide with
    # the surviving class's real SEDOL -- see CLASS_CONVERSIONS.
    "ranmore global equity investor gbp":      "BR2Q8G6-INV",
    "ranmore global equity fund (investor)":   "BR2Q8G6-INV",
}

# Retired share class -> surviving class. A "Fund Class Conversion" pair is a
# continuation of the SAME economic holding, not a sale plus a purchase, so
# open_positions treats the retired leg as transfer_out and the surviving leg
# as transfer_in rather than ignoring both.
CLASS_CONVERSIONS = {
    "BR2Q8G6-INV": "BR2Q8G6",   # Ranmore Investor -> Institutional, 24-Feb-2026
}

# Tickers classified as FUND for sleeve purposes even if exchange-listed.
# ONE HOME: extract_portfolio (R4.4). This module previously kept its own copy under a
# comment claiming it mirrored that one — {SMT, VUAG, IWMO, SGLN} against {SMT, VUAG}.
# A comment asserting agreement is not agreement (FC-B); the two must be one object.
try:
    from extract_portfolio import FUND_OVERRIDE_TICKERS
except ImportError as _exc:  # pragma: no cover - environment, not logic
    raise ImportError(
        "extract_transactions needs extract_portfolio.FUND_OVERRIDE_TICKERS. Refusing to fall "
        "back to a local copy: the local copy is exactly the defect this import removed "
        "(R4.4, R4.3)") from _exc


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _norm(s) -> str:
    """Normalise a broker description for matching: lowercase, collapse space."""
    if s is None:
        return ""
    s = str(s).replace("£", "GBP").replace("–", "-").replace("’", "'")
    return re.sub(r"\s+", " ", s).strip().lower()


def _num(v):
    """Coerce to float, tolerating None / blank / comma-formatted strings."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("£", "").strip())
    except ValueError:
        return None


def _as_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if hasattr(v, "isoformat"):
        return v.isoformat()
    if not v:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%b-%y", "%d %b %Y"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _uid(row):
    """Stable row identity.

    The broker Reference is NOT unique on its own -- equalisation and
    accumulation-distribution pairs share one reference (e.g. '44626##0013').
    So the key is composite.
    """
    key = "|".join(str(row.get(k) or "") for k in
                   ("date", "raw_type", "description", "quantity",
                    "amount_gbp", "reference"))
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:16]


# ---------------------------------------------------------------------------
# Ticker resolution
# ---------------------------------------------------------------------------
def build_ticker_map(isa_folder):
    """Derive description -> ticker from EVERY saved AJ Bell portfolio xlsx.

    The portfolio 'Investment' cell is exactly the transaction 'Description'
    plus ' (EXCHANGE:TICKER)', so the mapping is exact, not fuzzy. Scanning all
    saved snapshots (not just the latest) picks up positions since closed.
    Static overrides fill the remaining gaps.
    """
    mapping = {}
    paths = sorted(set(glob.glob(os.path.join(isa_folder, "*ISA*.xlsx"))
                       + glob.glob(os.path.join(isa_folder, "portfolio-*.xlsx"))))
    for path in paths:
        base = os.path.basename(path)
        if "X-Ray" in base or "Transaction" in base or "transaction" in base:
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        try:
            for sheet in wb.sheetnames:
                for row in wb[sheet].iter_rows(values_only=True):
                    if not row or not row[0]:
                        continue
                    m = re.match(r"^(.*?)\s*\(([A-Z]+):([A-Z0-9\.]+)\)\s*$",
                                 str(row[0]))
                    if not m:
                        continue
                    desc, _exch, ticker = m.groups()
                    mapping.setdefault(_norm(desc), ticker.upper())
        finally:
            try:
                wb.close()
            except Exception:
                pass
    for desc, ticker in STATIC_TICKER_OVERRIDES.items():
        mapping.setdefault(_norm(desc), ticker.upper())
    return mapping


def resolve_ticker(description, mapping):
    """Exact normalised match, then a conservative containment match."""
    key = _norm(description)
    if not key:
        return None, "unmapped"
    if key in mapping:
        return mapping[key], "exact"
    for known, ticker in mapping.items():
        if not known:
            continue
        if key.startswith(known) or known.startswith(key):
            return ticker, "prefix"
    return None, "unmapped"


def asset_class_for(ticker):
    if not ticker:
        return "unknown"
    if ticker in FUND_OVERRIDE_TICKERS:
        return "fund"
    base = ticker.split("-")[0]
    if re.match(r"^[0-9B][A-Z0-9]{6}$", base):  # FUND:SEDOL style, e.g. B2PLJM6
        return "fund"
    return "stock"


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
def parse_xlsx(path, mapping):
    """Return (rows, warnings). Never drops a row silently."""
    warnings = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        raw = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    finally:
        try:
            wb.close()
        except Exception:
            pass
    if not raw:
        return [], ["Empty workbook: %s" % os.path.basename(path)]

    header = [_norm(c) for c in raw[0]]
    if header[:len(EXPECTED_HEADER)] != EXPECTED_HEADER:
        warnings.append(
            "Unexpected header in %s: %s (expected %s) -- parsed positionally"
            % (os.path.basename(path), header, EXPECTED_HEADER))

    rows = []
    for i, r in enumerate(raw[1:], start=2):
        if r is None or all(c is None or c == "" for c in r):
            continue
        date = _as_date(r[0] if len(r) > 0 else None)
        if not date:
            warnings.append("Row %d skipped: unparseable date %r"
                            % (i, r[0] if r else None))
            continue
        raw_type = (str(r[1]).strip()
                    if len(r) > 1 and r[1] is not None else "")
        ttype = TYPE_MAP.get(_norm(raw_type), "unclassified")
        description = (str(r[2]).strip()
                       if len(r) > 2 and r[2] is not None else "")
        qty = _num(r[3]) if len(r) > 3 else None
        price = _num(r[4]) if len(r) > 4 else None
        amount = _num(r[5]) if len(r) > 5 else None
        reference = (str(r[6]).strip()
                     if len(r) > 6 and r[6] is not None else "")

        if ttype == "unclassified":
            if not qty and amount:
                # AJ Bell occasionally exports an income row with a blank
                # Transaction cell. Zero units + a cash amount is an
                # accumulation distribution; classify it so the cost base
                # reconciles, but say so.
                ttype = "distribution"
                warnings.append(
                    "Row %d (%s, %s): blank transaction type, zero units with "
                    "amount GBP %.2f -- inferred as 'distribution'"
                    % (i, date, description[:40], amount))
            else:
                warnings.append(
                    "Row %d (%s, %s): unrecognised transaction type %r -- "
                    "recorded as unclassified, excluded from reconciliation"
                    % (i, date, description[:40], raw_type))

        ticker, match = resolve_ticker(description, mapping)
        if ticker is None:
            warnings.append(
                "Row %d (%s, %s): description not mapped to a ticker -- add it "
                "to STATIC_TICKER_OVERRIDES, or save the portfolio xlsx that "
                "contains it" % (i, date, description[:50]))

        gross = (qty * price) if (qty and price is not None) else None
        cost = None
        cost_pct = None
        if ttype in TRADE_TYPES and gross and amount is not None:
            cost = round(amount - gross, 2) if ttype == "buy" \
                else round(gross - amount, 2)
            cost_pct = round(cost / gross * 100.0, 4) if gross else None

        row = {
            "date": date,
            "type": ttype,
            "raw_type": raw_type,
            "description": description,
            "ticker": ticker,
            "ticker_match": match,
            "asset_class": asset_class_for(ticker),
            "quantity": qty,
            "price": price,
            "amount_gbp": amount,
            "gross_gbp": round(gross, 2) if gross else None,
            "cost_gbp": cost,
            "cost_pct": cost_pct,
            "cash_impact_gbp": (round(CASH_SIGN.get(ttype, 0) * amount, 2)
                                if amount is not None else None),
            "reference": reference,
            "source_file": os.path.basename(path),
            "source_row": i,
        }
        row["uid"] = _uid(row)
        rows.append(row)

    return rows, warnings


# ---------------------------------------------------------------------------
# Ledger
# ---------------------------------------------------------------------------
def load_ledger(path):
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            led = json.load(f)
        led.setdefault("entries", [])
        led.setdefault("schema_version", LEDGER_SCHEMA_VERSION)
        led.setdefault("_meta", {})
        return led
    return {"schema_version": LEDGER_SCHEMA_VERSION, "_meta": {}, "entries": []}


def save_ledger(led, path):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(led, f, indent=2, ensure_ascii=False)
    os.replace(tmp, path)


def merge_rows(ledger, rows, stamp):
    """Append-and-dedupe on the composite uid. Idempotent by construction, so
    re-importing an overlapping export (or the full-history seed) is safe."""
    existing = set(e["uid"] for e in ledger["entries"])
    added, dupes = 0, 0
    for r in rows:
        if r["uid"] in existing:
            dupes += 1
            continue
        r = dict(r)
        r["first_seen"] = stamp
        ledger["entries"].append(r)
        existing.add(r["uid"])
        added += 1
    ledger["entries"].sort(key=lambda e: (e["date"], e.get("reference") or ""))
    ledger["_meta"] = {
        "last_import": stamp,
        "total_entries": len(ledger["entries"]),
        "earliest": ledger["entries"][0]["date"] if ledger["entries"] else None,
        "latest": ledger["entries"][-1]["date"] if ledger["entries"] else None,
    }
    return {"added": added, "duplicates_skipped": dupes,
            "total": len(ledger["entries"])}


# ---------------------------------------------------------------------------
# Derived analytics
# ---------------------------------------------------------------------------
def fifo_round_trips(entries):
    """Match sells against buys FIFO per ticker -> realised P&L + holding days.

    Feeds: trades-log reconciliation, the min-hold / entry-stability rule
    (audit C-1) and the learning loop's realised-outcome record.
    """
    by_ticker = {}
    for e in sorted(entries, key=lambda x: x["date"]):
        if e["type"] not in TRADE_TYPES or not e.get("ticker"):
            continue
        by_ticker.setdefault(e["ticker"], []).append(e)

    trips = []
    for ticker, evs in by_ticker.items():
        lots = []          # [qty_remaining, unit_cost_gbp, buy_date]
        for e in evs:
            qty = e.get("quantity") or 0
            amt = e.get("amount_gbp") or 0
            if not qty:
                continue
            unit = amt / qty
            if e["type"] == "buy":
                lots.append([qty, unit, e["date"]])
                continue
            remaining = qty
            while remaining > 1e-9 and lots:
                lot = lots[0]
                take = min(remaining, lot[0])
                cost_basis = take * lot[1]
                proceeds = take * unit
                held = (datetime.strptime(e["date"], "%Y-%m-%d")
                        - datetime.strptime(lot[2], "%Y-%m-%d")).days
                trips.append({
                    "ticker": ticker,
                    "buy_date": lot[2],
                    "sell_date": e["date"],
                    "quantity": round(take, 6),
                    "cost_basis_gbp": round(cost_basis, 2),
                    "proceeds_gbp": round(proceeds, 2),
                    "realised_pnl_gbp": round(proceeds - cost_basis, 2),
                    "realised_pct": (round((proceeds / cost_basis - 1) * 100, 2)
                                     if cost_basis else None),
                    "holding_days": held,
                })
                lot[0] -= take
                remaining -= take
                if lot[0] <= 1e-9:
                    lots.pop(0)
            if remaining > 1e-9:
                trips.append({
                    "ticker": ticker, "buy_date": None, "sell_date": e["date"],
                    "quantity": round(remaining, 6), "cost_basis_gbp": None,
                    "proceeds_gbp": round(remaining * unit, 2),
                    "realised_pnl_gbp": None, "realised_pct": None,
                    "holding_days": None,
                    "note": "sell with no matching buy in the ledger window "
                            "(position pre-dates the transaction history)",
                })
    return sorted(trips, key=lambda t: t["sell_date"])


def open_positions(entries):
    """Units and average cost per ticker implied by the ledger. DIAGNOSTIC
    ONLY -- broker truth for HOLDINGS remains the portfolio xlsx."""
    pos = {}
    for e in sorted(entries, key=lambda x: x["date"]):
        t = e.get("ticker")
        qty = e.get("quantity") or 0
        if not t:
            continue
        # Accumulation income moves book cost without moving units: a
        # distribution is reinvested (cost up), equalisation is a return of
        # capital on units bought mid-distribution-period (cost down).
        if e["type"] in ("distribution", "equalisation"):
            amt = e.get("amount_gbp") or 0
            p = pos.setdefault(t, {"quantity": 0.0, "cost_gbp": 0.0})
            p["cost_gbp"] += amt if e["type"] == "distribution" else -amt
            continue
        if not qty:
            continue
        p = pos.setdefault(t, {"quantity": 0.0, "cost_gbp": 0.0})
        if e["type"] == "conversion":
            # A class conversion has two legs sharing one value. The broker
            # REBASES book cost to the conversion value on the surviving class
            # (it does not carry the old accumulated cost across), so mirror
            # that: retire the outgoing class outright, and open the incoming
            # class at the conversion amount.
            if t in CLASS_CONVERSIONS:
                p["quantity"] = 0.0
                p["cost_gbp"] = 0.0
            else:
                p["quantity"] += qty
                p["cost_gbp"] += (e.get("amount_gbp") or 0)
            continue
        if e["type"] in ("buy", "transfer_in"):
            p["quantity"] += qty
            p["cost_gbp"] += (e.get("amount_gbp") or 0)
        elif e["type"] in ("sell", "transfer_out"):
            if p["quantity"] > 0:
                frac = min(qty / p["quantity"], 1.0)
                p["cost_gbp"] -= p["cost_gbp"] * frac
            p["quantity"] -= qty
    out = {}
    for t, p in pos.items():
        if abs(p["quantity"]) < 1e-6:
            continue
        out[t] = {"quantity": round(p["quantity"], 6),
                  "cost_gbp": round(p["cost_gbp"], 2),
                  "avg_cost": (round(p["cost_gbp"] / p["quantity"], 6)
                               if p["quantity"] else None)}
    return out


def cost_calibration(entries):
    """Actual dealing costs vs gross consideration -- calibrates the Step 10.6
    cost estimate instead of leaving it a standing assumption."""
    buckets = {}
    for e in entries:
        if e["type"] not in TRADE_TYPES or e.get("cost_gbp") is None:
            continue
        key = "%s_%s" % (e.get("asset_class") or "unknown", e["type"])
        b = buckets.setdefault(key, {"costs": [], "gross": 0.0})
        b["costs"].append(e["cost_gbp"])
        b["gross"] += e.get("gross_gbp") or 0
    out = {}
    for k, b in buckets.items():
        costs = sorted(b["costs"])
        total = sum(costs)
        out[k] = {
            "n": len(costs),
            "mean_cost_gbp": round(total / len(costs), 2),
            "median_cost_gbp": round(costs[len(costs) // 2], 2),
            "min_cost_gbp": round(costs[0], 2),
            "max_cost_gbp": round(costs[-1], 2),
            "blended_cost_pct": (round(total / b["gross"] * 100, 4)
                                 if b["gross"] else None),
        }
    return out


def reconcile_vs_broker(entries, portfolio_xlsx, tol_qty=0.01, tol_cost=1.00):
    """COMPLETENESS GATE.

    Rebuild holdings from the transaction ledger and compare against the broker
    portfolio file. They must agree. A mismatch means one of:
      * a monthly transaction export was never saved (gap in the ledger)
      * a row failed to parse or map to the right ticker
      * a corporate action the parser does not model
    Cheap to run, and it is the only thing that proves the ledger is COMPLETE
    rather than merely well-formed. Returns a dict; status OK / MISMATCH / SKIP.
    """
    if not portfolio_xlsx or not os.path.exists(portfolio_xlsx):
        return {"status": "SKIP", "reason": "no portfolio xlsx supplied"}
    try:
        wb = openpyxl.load_workbook(portfolio_xlsx, read_only=True,
                                    data_only=True)
    except Exception as exc:
        return {"status": "SKIP", "reason": "unreadable portfolio xlsx: %s" % exc}
    actual = {}
    try:
        for sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                m = re.match(r"^(.*?)\s*\(([A-Z]+):([A-Z0-9\.]+)\)\s*$",
                             str(row[0]))
                if not m:
                    continue
                actual[m.group(3).upper()] = {
                    "quantity": _num(row[1]) if len(row) > 1 else None,
                    "cost_gbp": _num(row[4]) if len(row) > 4 else None,
                }
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # Only compare up to the portfolio's own as-at date -- trades executed
    # AFTER the snapshot are correctly absent from it.
    as_at = None
    try:
        wb = openpyxl.load_workbook(portfolio_xlsx, read_only=True, data_only=True)
        for row in wb[wb.sheetnames[0]].iter_rows(values_only=True):
            if row and len(row) > 11 and isinstance(row[11], datetime):
                as_at = row[11].date().isoformat()
                break
        wb.close()
    except Exception:
        pass
    scoped = [e for e in entries if (as_at is None or e["date"] <= as_at)]
    implied = open_positions(scoped)

    diffs = []
    for ticker in sorted(set(actual) | set(implied)):
        a = actual.get(ticker, {})
        i = implied.get(ticker, {})
        aq, iq = a.get("quantity"), i.get("quantity")
        ac, ic = a.get("cost_gbp"), i.get("cost_gbp")
        qty_bad = (aq is None) != (iq is None) or (
            aq is not None and iq is not None and abs(aq - iq) > tol_qty)
        cost_bad = (ac is not None and ic is not None
                    and abs(ac - ic) > tol_cost)
        if qty_bad or cost_bad:
            diffs.append({"ticker": ticker, "broker_quantity": aq,
                          "ledger_quantity": iq, "broker_cost_gbp": ac,
                          "ledger_cost_gbp": ic,
                          "quantity_mismatch": qty_bad,
                          "cost_mismatch": cost_bad})
    # The check is only meaningful once the ledger spans the period in which
    # the open positions were acquired. On a fresh (unseeded) ledger almost
    # every position is legitimately absent -- report that as a window problem,
    # not as a reconciliation failure, or the real signal gets buried in noise.
    n_total = len(set(actual) | set(implied))
    missing_entirely = sum(1 for d in diffs if d["ledger_quantity"] is None)
    earliest = min((e["date"] for e in scoped), default=None)
    window_short = (earliest is None or (as_at and earliest > as_at)
                    or (n_total and missing_entirely >= n_total / 2.0))
    if diffs and window_short:
        return {"status": "INCOMPLETE_WINDOW",
                "as_at": as_at,
                "ledger_earliest": earliest,
                "n_compared": n_total,
                "n_missing_from_ledger": missing_entirely,
                "note": ("the transaction ledger does not yet span the period in "
                         "which these positions were acquired -- seed it with a "
                         "full transaction-history export "
                         "(extract_transactions.py --seed) before relying on the "
                         "completeness check"),
                "differences": diffs}

    return {"status": "MISMATCH" if diffs else "OK",
            "as_at": as_at,
            "ledger_earliest": earliest,
            "n_compared": n_total,
            "differences": diffs}


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------
def find_latest_monthly(isa_folder):
    """Newest 'Transaction History MM-YYYY.xlsx' by the month in the filename."""
    best, best_key = None, None
    for path in glob.glob(os.path.join(isa_folder, MONTHLY_GLOB)):
        m = MONTHLY_RE.search(os.path.basename(path))
        if not m:
            continue
        key = (int(m.group(2)), int(m.group(1)))
        if best_key is None or key > best_key:
            best, best_key = path, key
    return best, (("%04d-%02d" % best_key) if best_key else None)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def find_latest_portfolio(isa_folder):
    """Newest 'AJ Bell ISA Portfolio dd-mmm-yy.xlsx' (excludes the X-Ray PDF
    and the 'Example' file)."""
    best, best_dt = None, None
    for path in glob.glob(os.path.join(isa_folder,
                                       "AJ Bell ISA Portfolio *.xlsx")):
        base = os.path.basename(path)
        if "X-Ray" in base or "Example" in base:
            continue
        m = re.search(r"(\d{2}-[A-Za-z]{3}-\d{2})\.xlsx$", base)
        if not m:
            continue
        try:
            dt = datetime.strptime(m.group(1), "%d-%b-%y")
        except ValueError:
            continue
        if best_dt is None or dt > best_dt:
            best, best_dt = path, dt
    return best


def run(xlsx_path, isa_folder, ledger_path, out_path, stamp=None, quiet=False,
        portfolio_xlsx=None):
    stamp = stamp or datetime.now().date().isoformat()
    mapping = build_ticker_map(isa_folder)
    rows, warnings = parse_xlsx(xlsx_path, mapping)

    ledger = load_ledger(ledger_path)
    merge_stats = merge_rows(ledger, rows, stamp)
    save_ledger(ledger, ledger_path)

    entries = ledger["entries"]
    month_trades = [r for r in rows if r["type"] in TRADE_TYPES]
    unmapped = sorted(set(r["description"] for r in rows if not r.get("ticker")))

    data = {
        "_meta": {
            "produced_at": datetime.now().isoformat(timespec="seconds"),
            "source_file": os.path.basename(xlsx_path),
            "ledger_file": os.path.basename(ledger_path),
            "schema_version": LEDGER_SCHEMA_VERSION,
            "status": "ERROR" if not rows else ("WARN" if warnings else "OK"),
        },
        "import": merge_stats,
        "warnings": warnings,
        "unmapped_descriptions": unmapped,
        "month_rows": rows,
        "month_summary": {
            "n_rows": len(rows),
            "n_trades": len(month_trades),
            "buys": [r["ticker"] or r["description"]
                     for r in month_trades if r["type"] == "buy"],
            "sells": [r["ticker"] or r["description"]
                      for r in month_trades if r["type"] == "sell"],
            "net_cash_impact_gbp": round(
                sum(r["cash_impact_gbp"] or 0 for r in rows), 2),
            "total_dealing_costs_gbp": round(
                sum(r["cost_gbp"] or 0 for r in month_trades), 2),
            "distributions_gbp": round(
                sum(r["amount_gbp"] or 0 for r in rows
                    if r["type"] in ("distribution", "equalisation")), 2),
        },
        "executed_trades": [
            dict((k, r[k]) for k in ("date", "ticker", "type", "quantity",
                                     "price", "amount_gbp", "cost_gbp",
                                     "cost_pct", "description", "reference"))
            for r in month_trades
        ],
        "round_trips": fifo_round_trips(entries),
        "open_positions_implied": open_positions(entries),
        "cost_calibration": cost_calibration(entries),
        "broker_reconciliation": reconcile_vs_broker(
            entries, portfolio_xlsx or find_latest_portfolio(isa_folder)),
        "ledger_meta": ledger["_meta"],
    }
    if data["broker_reconciliation"]["status"] == "MISMATCH":
        data["_meta"]["status"] = "WARN"
        for d in data["broker_reconciliation"]["differences"]:
            data["warnings"].append(
                "Broker reconciliation mismatch on %s: broker qty %s / ledger "
                "qty %s, broker cost %s / ledger cost %s -- a transaction "
                "export may be missing from the ledger"
                % (d["ticker"], d["broker_quantity"], d["ledger_quantity"],
                   d["broker_cost_gbp"], d["ledger_cost_gbp"]))

    if out_path:
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    if not quiet:
        s = data["month_summary"]
        print("  Source        : %s" % data["_meta"]["source_file"])
        print("  Rows parsed   : %d  (added %d, dupes skipped %d, ledger total %d)"
              % (s["n_rows"], merge_stats["added"],
                 merge_stats["duplicates_skipped"], merge_stats["total"]))
        print("  Trades        : %d  buys=%s sells=%s"
              % (s["n_trades"], s["buys"] or "-", s["sells"] or "-"))
        print("  Dealing costs : GBP %.2f   Net cash impact: GBP %.2f"
              % (s["total_dealing_costs_gbp"], s["net_cash_impact_gbp"]))
        br = data["broker_reconciliation"]
        print("  Broker recon  : %s (%d positions compared, as at %s)"
              % (br["status"], br.get("n_compared", 0), br.get("as_at")))
        if unmapped:
            print("  UNMAPPED      : %s" % "; ".join(unmapped))
        for w in warnings:
            print("  WARN: %s" % w)

    return data


def main():
    ap = argparse.ArgumentParser(
        description="AJ Bell ISA transaction history parser")
    ap.add_argument("--xlsx", help="explicit transaction xlsx path")
    ap.add_argument("--seed", help="one-off full-history xlsx to seed the ledger")
    ap.add_argument("--isa-folder", default=DEFAULT_ISA_FOLDER)
    ap.add_argument("--ledger", help="transaction_ledger.json path")
    ap.add_argument("--out", help="transactions_data_[mmm_yyyy].json path")
    ap.add_argument("--stamp", help="import date stamp (YYYY-MM-DD)")
    ap.add_argument("--report", action="store_true",
                    help="print a ledger report and exit without importing")
    ap.add_argument("--portfolio", help="portfolio xlsx for the completeness "
                                        "check (default: latest in ISA folder)")
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args()

    isa_folder = os.path.abspath(args.isa_folder)
    ledger_path = args.ledger or os.path.join(SCRIPT_DIR,
                                              "transaction_ledger.json")

    if args.report:
        led = load_ledger(ledger_path)
        print(json.dumps({
            "meta": led.get("_meta", {}),
            "round_trips": fifo_round_trips(led["entries"]),
            "open_positions_implied": open_positions(led["entries"]),
            "cost_calibration": cost_calibration(led["entries"]),
        }, indent=2))
        return 0

    xlsx = args.seed or args.xlsx
    if not xlsx:
        xlsx, _ = find_latest_monthly(isa_folder)
    if not xlsx or not os.path.exists(xlsx):
        print("WARN: no transaction history file found in %s (expected "
              "'Transaction History MM-YYYY.xlsx'). Reconciliation falls back "
              "to holdings-delta inference." % isa_folder, file=sys.stderr)
        return 3   # distinct exit code: caller treats as WARN, not ERROR

    out = args.out
    if out is None and not args.seed:
        m = MONTHLY_RE.search(os.path.basename(xlsx))
        if m:
            tag = datetime(int(m.group(2)), int(m.group(1)), 1) \
                .strftime("%b_%Y").lower()
        else:
            tag = datetime.now().strftime("%b_%Y").lower()
        out = os.path.join(SCRIPT_DIR, "transactions_data_%s.json" % tag)

    data = run(xlsx, isa_folder, ledger_path, out, stamp=args.stamp,
               quiet=args.quiet, portfolio_xlsx=args.portfolio)
    return 0 if data["_meta"]["status"] != "ERROR" else 1


if __name__ == "__main__":
    sys.exit(main())
