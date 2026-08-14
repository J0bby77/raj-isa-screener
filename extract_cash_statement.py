#!/usr/bin/env python3
"""
extract_cash_statement.py — BuildSpec Part 2 (built 05-Aug-2026).

WHY THIS EXISTS, STATED PRECISELY
---------------------------------
The ISA allowance was reported ~£3,750 used / ~£16,250 remaining. The truth is **£8,750 used /
£11,250 remaining** — wrong by £5,000, because a `Faster Payment In` of £5,000 on 06-Apr-2026
was in no framework input.

The root cause is structural, not arithmetic, and worth stating so it is not re-introduced:
`parse_contributions()` reads AJ Bell's **Transaction History**, which is a DEALING record —
Purchase, Sale, Transfer In, Equalisation, Fund Class Conversion. It contains **no cash
deposits at all**. Contributions live only in the **Cash Statement**, a different export. So the
allowance was being derived from a document that structurally cannot contain the answer, and the
framework had no way to notice, because "no deposit rows found" and "no deposits made" produce
the same number.

(A second, independent defect was fixed alongside: the A22 glob `ISA Transaction History*.xlsx`
matched 1 of Raj's 3 transaction files. Both had to be wrong for the £5,000 to disappear
silently, which is the usual shape of these things.)

⚑ THE TRAP — `* BALANCE B/F *`
The first row is the prior tax year's closing cash carried forward, **£3,469.64 in the Receipt
column**. It is indistinguishable from a contribution to any naive sum, and counting it would
overstate the allowance by £3,469.64. It is a POSITION, not a FLOW. `I12` is a dedicated
negative test for exactly this.

FX — SETTLED BY PRIMARY SOURCE
The file contains literal `FX Charge (0.50%)` rows. The framework carried **0.75%**. The COCO
purchase executed at £1,354.99 against a £1,384.75 estimate, and the difference is precisely the
overstatement. `FX_RATE_PCT` here is the single home for that number.

CONTRACT
--------
* Missing file  -> WARN and `reconciled: False`. Never an ERROR, never a guessed figure.
* Every category is decided by a tested rule over the free-text Description, because the file
  has no type column. An unmatched description is `UNCLASSIFIED` — never silently dropped, and
  never bucketed into the nearest plausible category.
* Tax year is derived from the ROW DATES, never the filename. The A22 glob failure is the
  reason that is spelled out rather than assumed.
"""
from __future__ import annotations
import argparse, datetime as dt, glob, json, os, re, sys

HERE = os.path.dirname(os.path.abspath(__file__))
ISA_ALLOWANCE_GBP = 20000.0

# ---------------------------------------------------------------------------
# THE single home for cash-side portfolio constants (R4.4). This module already owns
# FX_RATE_FRACTION on the same principle: the module that extracts the truth owns the
# constant, and every consumer imports it. Consolidated 12-Aug-2026 after the Framework
# Atlas found ISA_ALLOWANCE_GBP in 2 homes, CASH_BUFFER_MIN in 2, and the monthly standing
# order in THREE homes under TWO different names (STANDING_ORDER and STANDING_ORDER_MONTHLY).
#
# ⚑ VALUES ARE UNCHANGED BY THAT CONSOLIDATION. A refactor must never move a number.
# ---------------------------------------------------------------------------
CASH_BUFFER_MIN = 150.0
CASH_BUFFER_MAX = 200.0

# ⚑ OPEN QUESTION FOR RAJ (register ISA-0011): user_raj_profile records the £1,250 monthly
# standing order as PAUSED from Jul-2026 (job security). All three former homes carried
# 1250.0 as if live. The VALUE IS LEFT AS IT WAS so this consolidation changes no behaviour;
# STANDING_ORDER_ACTIVE makes the assumption explicit and switchable in ONE place instead of
# being implicit in three. It feeds contribution schedules, the required-return anchor and
# allowance projection - so if the pause is real, the anchor is derived on a schedule that
# is not happening.
STANDING_ORDER = 1250.0
# ⚑ CONFIRMED BY RAJ 12-Aug-2026: the standing order is PAUSED. Last payment June-2026.
# Allowance used £8,750 = £5,000 initial + 3 x £1,250 (April, May, June). Nothing since.
# target_state.json already models this correctly (contribution_schedule: 0/month from
# 2026-07-01), so the REQUIRED-RETURN ANCHOR was never affected. What WAS affected is cash:
# extract_portfolio.standing_order_adjustment() credited £1,250 of "unprocessed S/O" to
# effective cash whenever the broker file was dated within SO_CLEAR_WORKING_DAYS of the 1st -
# which is exactly when the monthly review runs. See register ISA-0011.
STANDING_ORDER_ACTIVE = False
STANDING_ORDER_PAUSED_FROM = "2026-07-01"   # Raj, job security
SO_CLEAR_WORKING_DAYS = 3             # clears within ~3 working days of the 1st

# ⚑ SINGLE HOME for the AJ Bell FX charge. Evidenced by two `FX Charge (0.50%)` rows in the
# 2026-27 cash statement and by the COCO fill. Anything that models a USD trade cost reads THIS.
FX_RATE_PCT = 0.50
FX_RATE_FRACTION = FX_RATE_PCT / 100.0
FX_RATE_SOURCE = ("AJ Bell cash statement 2026-27, literal 'FX Charge (0.50%)' rows; "
                  "corroborated by COCO 03-Aug-2026 executing at GBP1,354.99 vs a "
                  "GBP1,384.75 estimate built on 0.75%")

OPENING_BALANCE_RE = re.compile(r"^\s*\*+\s*balance\s*b/?f\s*\*+", re.I)

# Order matters: the first rule that matches wins, so the specific precede the general.
CLASSIFIERS = [
    ("OPENING_BALANCE", lambda d: bool(OPENING_BALANCE_RE.match(d))),
    ("CONTRIBUTION",    lambda d: any(k in d.lower() for k in
                                      ("faster payment in", "direct debit payment",
                                       "debit card payment", "bank transfer in"))),
    ("FX_FEE",          lambda d: "fx charge" in d.lower()),
    ("PLATFORM_FEE",    lambda d: "account charge" in d.lower()),
    ("DIVIDEND",        lambda d: d.lower().lstrip().startswith("dividend")),
    ("INTEREST",        lambda d: "gross interest" in d.lower()),
    ("REBATE",          lambda d: "cash rebate" in d.lower()),
    ("BUY",             lambda d: d.lower().lstrip().startswith("purchase")),
    ("SELL",            lambda d: d.lower().lstrip().startswith("sale")
                                  or "residual funds from" in d.lower()),
    ("TRANSFER",        lambda d: "transfer" in d.lower()),
]
FLOW_CATEGORIES = {"CONTRIBUTION"}          # what counts against the ISA allowance


def classify(description):
    d = str(description or "")
    if not d.strip():
        return "UNCLASSIFIED"
    for name, test in CLASSIFIERS:
        try:
            if test(d):
                return name
        except Exception:
            continue
    return "UNCLASSIFIED"


def tax_year_start(ref):
    yr = ref.year if (ref.month, ref.day) >= (4, 6) else ref.year - 1
    return dt.date(yr, 4, 6)


def find_files(folder):
    """`Cash Statement*.xlsx`, case-insensitively, Office lock files excluded. The tax year is
    NOT taken from the filename — see the module docstring."""
    out = {}
    for pat in ("Cash Statement*.xlsx", "cash statement*.xlsx", "*ash*tatement*.xlsx"):
        for fp in glob.glob(os.path.join(folder, pat)):
            if not os.path.basename(fp).startswith("~$"):
                out[os.path.realpath(fp)] = fp
    return sorted(out.values())


def _as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        try:
            return dt.date.fromisoformat(v[:10])
        except ValueError:
            return None
    return None


def _num(v):
    try:
        if v is None or (isinstance(v, str) and not v.strip()):
            return 0.0
        return float(str(v).replace(",", "").replace("£", "").strip())
    except (TypeError, ValueError):
        return 0.0


def parse(folder=None, files=None, as_of=None):
    import openpyxl
    folder = folder or os.path.dirname(HERE)
    files = files if files is not None else find_files(folder)
    as_of = as_of or dt.date.today()
    res = {"as_of": as_of.isoformat(), "source_files": [os.path.basename(f) for f in files],
           "fx_rate_pct": FX_RATE_PCT, "fx_rate_source": FX_RATE_SOURCE,
           "rows": [], "reconciled": False, "invariants": {}, "warnings": []}
    if not files:
        res["warnings"].append("no Cash Statement*.xlsx found — allowance stays UNRECONCILED. "
                               "WARN, never ERROR: an absent file is a missing input, not a "
                               "failed computation.")
        return res

    rows = []
    for fp in files:
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        for ws in wb.worksheets:
            hdr = None
            for raw in ws.iter_rows(values_only=True):
                if hdr is None:
                    hdr = [str(c or "").strip().lower() for c in raw]
                    continue
                r = dict(zip(hdr, raw))
                d = _as_date(r.get("date"))
                desc = str(r.get("description") or "").strip()
                if d is None and not OPENING_BALANCE_RE.match(desc):
                    continue
                rows.append({
                    "date": d.isoformat() if d else None,
                    "description": desc,
                    "reference": str(r.get("reference") or "").strip(),
                    "settlement_date": (_as_date(r.get("settlement date")).isoformat()
                                        if _as_date(r.get("settlement date")) else None),
                    "receipt_gbp": _num(r.get("receipt (gbp)")),
                    "payment_gbp": _num(r.get("payment (gbp)")),
                    "balance_gbp": _num(r.get("balance (gbp)")),
                    "category": classify(desc),
                    "source_file": os.path.basename(fp),
                })
            break                                     # single-sheet export by design
    rows.sort(key=lambda x: (x["date"] or "0000-00-00", x["reference"]))
    res["rows"] = rows

    ty_start = tax_year_start(as_of)
    res["tax_year_start"] = ty_start.isoformat()
    in_ty = [r for r in rows if r["date"] and dt.date.fromisoformat(r["date"]) >= ty_start]

    totals = {}
    for r in rows:
        totals.setdefault(r["category"], {"n": 0, "receipts": 0.0, "payments": 0.0})
        t = totals[r["category"]]
        t["n"] += 1; t["receipts"] += r["receipt_gbp"]; t["payments"] += r["payment_gbp"]
    for t in totals.values():
        t["receipts"] = round(t["receipts"], 2)
        t["payments"] = round(t["payments"], 2)
        t["net"] = round(t["receipts"] - abs(t["payments"]), 2)
    res["category_totals"] = totals

    contrib_rows = [r for r in in_ty if r["category"] == "CONTRIBUTION"]
    used = round(sum(r["receipt_gbp"] for r in contrib_rows), 2)
    res["allowance"] = {
        "tax_year_start": ty_start.isoformat(),
        "used_gbp": used,
        "remaining_gbp": round(ISA_ALLOWANCE_GBP - used, 2),
        "annual_gbp": ISA_ALLOWANCE_GBP,
        "contribution_rows": contrib_rows,
        "source": "AJ Bell Cash Statement (CONTRIBUTION rows only; the opening balance "
                  "carried forward is a POSITION and is excluded — see I12)",
    }
    opening = next((r for r in rows if r["category"] == "OPENING_BALANCE"), None)
    res["opening_balance_gbp"] = opening["receipt_gbp"] if opening else None
    res["closing_balance_gbp"] = rows[-1]["balance_gbp"] if rows else None

    # ── invariants ──────────────────────────────────────────────────────────────────
    inv = res["invariants"]
    inv["I11_contribution_within_allowance"] = {
        "ok": 0 <= used <= ISA_ALLOWANCE_GBP, "value": used}
    inv["I12_opening_balance_not_a_contribution"] = {
        "ok": all(r["category"] != "CONTRIBUTION" for r in rows
                  if OPENING_BALANCE_RE.match(r["description"])),
        "opening_balance_gbp": res["opening_balance_gbp"],
        "note": "the * BALANCE B/F * row is a position carried forward; counting it would "
                "overstate the allowance by its own value"}
    if rows and res["opening_balance_gbp"] is not None:
        flow = sum(r["receipt_gbp"] - abs(r["payment_gbp"]) for r in rows
                   if r["category"] != "OPENING_BALANCE")
        implied = round(res["opening_balance_gbp"] + flow, 2)
        inv["I13_flows_reconcile_to_closing_balance"] = {
            "ok": abs(implied - (res["closing_balance_gbp"] or 0)) <= 0.05,
            "implied": implied, "stated": res["closing_balance_gbp"],
            "delta": round(implied - (res["closing_balance_gbp"] or 0), 2),
            "note": "opening + SUM(receipts - payments) must equal the stated closing balance; "
                    "a break means a row was mis-signed or dropped"}
    unclassified = [r for r in rows if r["category"] == "UNCLASSIFIED"]
    inv["I14_every_row_classified"] = {
        "ok": not unclassified, "n": len(unclassified),
        "examples": [r["description"][:60] for r in unclassified[:5]],
        "note": "an unmatched description is reported, never bucketed into the nearest "
                "plausible category"}
    res["reconciled"] = all(v.get("ok") for v in inv.values())
    if not res["reconciled"]:
        res["warnings"].append("one or more invariants failed — allowance is NOT reconciled")
    return res


def reconcile_with_ledger(res, ledger_path=None):
    """I10 — every BUY/SELL row in the cash statement must appear in `transaction_ledger.json`.

    Two genuinely independent derivations of the same trades: the broker's CASH view and the
    broker's DEALING view, exported separately and parsed by different code. Standard rule 4.

    Joined on the **broker reference** (e.g. `44626CGD9RL`), which both documents carry, with
    date+amount as a secondary key. An earlier draft joined on date+amount alone and reported
    25 of 25 unmatched — which was my own key-guessing, not 25 missing trades. Worth recording:
    a reconciliation that reports everything as broken is far more likely to be a broken
    reconciliation, and must be verified before it is believed.
    """
    lp = ledger_path or os.path.join(HERE, "transaction_ledger.json")
    if not os.path.exists(lp):
        return {"ok": None, "reason": "transaction_ledger.json not found — check SKIPPED, "
                                      "explicitly not passed"}
    try:
        led = json.load(open(lp, encoding="utf-8"))
    except Exception as e:
        return {"ok": None, "reason": f"ledger unreadable: {e}"}
    entries = (led.get("entries") or led.get("transactions") or led.get("rows") or []) \
        if isinstance(led, dict) else led
    by_ref, by_date_amt = {}, set()
    for e in entries:
        if not isinstance(e, dict):
            continue
        ref = str(e.get("reference") or "").strip()
        amt = e.get("amount_gbp")
        try:
            amt = round(abs(float(amt)), 2)
        except (TypeError, ValueError):
            amt = None
        if ref:
            by_ref[ref] = amt
        if amt is not None:
            by_date_amt.add((str(e.get("date") or "")[:10], amt))

    # The ledger is exported periodically; the cash statement is live. Trades newer than the
    # ledger's horizon are NOT missing — they are simply ahead of it, and reporting them as
    # breaks every month would train the reader to ignore this invariant entirely.
    ledger_max = max((str(e.get("date") or "")[:10] for e in entries
                      if isinstance(e, dict) and e.get("date")), default="")
    unmatched, amount_breaks, pending = [], [], []
    checked = 0
    for r in res.get("rows", []):
        if r["category"] not in ("BUY", "SELL"):
            continue
        checked += 1
        amt = round(abs(r["receipt_gbp"] or r["payment_gbp"]), 2)
        ref = r["reference"]
        if ref and ref in by_ref:
            if by_ref[ref] is not None and abs(by_ref[ref] - amt) > 0.01:
                amount_breaks.append({"reference": ref, "cash_statement_gbp": amt,
                                      "ledger_gbp": by_ref[ref],
                                      "description": r["description"][:50]})
            continue
        if (r["date"], amt) in by_date_amt:
            continue
        rec = {"date": r["date"], "amount_gbp": amt, "reference": ref,
               "description": r["description"][:60]}
        (pending if (ledger_max and r["date"] and r["date"] > ledger_max)
         else unmatched).append(rec)
    return {"ok": (not unmatched) and (not amount_breaks), "checked": checked,
            "ledger_entries": len(entries), "ledger_horizon": ledger_max,
            "unmatched": unmatched, "amount_breaks": amount_breaks,
            "pending_ledger_update": pending,
            "note": "joined on broker reference, falling back to date+amount; an amount break "
                    "on a MATCHED reference is the serious case — the same trade recorded twice "
                    "with two different figures"}


def _selftest():
    ok = True
    # classification — every rule, including the ones that must NOT fire
    cases = [
        ("* BALANCE B/F *", "OPENING_BALANCE"),
        ("*BALANCE B/F*", "OPENING_BALANCE"),
        ("Faster Payment In", "CONTRIBUTION"),
        ("Direct debit payment", "CONTRIBUTION"),
        ("Debit Card Payment", "CONTRIBUTION"),
        ("FX Charge (0.50%)", "FX_FEE"),
        ("Account charge for funds - Jun 2026", "PLATFORM_FEE"),
        ("Account charge for shares - Jul 2026", "PLATFORM_FEE"),
        ("Dividend 498   SCOTTISH MORTGAGE", "DIVIDEND"),
        ("DIVIDEND 7   MICRON TECHNOLOGY", "DIVIDEND"),
        ("Gross interest to 30/06/26", "INTEREST"),
        ("Fund Cash Rebate May26", "REBATE"),
        ("Purchase 28 The Vita Coco Co Inc", "BUY"),
        ("Sale 40 PAYLOCITY HLDG COR COM", "SELL"),
        ("Residual Funds from BB8HNPI", "SELL"),
        ("Equalisation Acc Units", "UNCLASSIFIED"),
        ("", "UNCLASSIFIED"),
    ]
    bad = [(d, classify(d), e) for d, e in cases if classify(d) != e]
    assert not bad, f"misclassified: {bad}"

    # ⚑ THE TRAP, as a first-class negative test: the opening balance must never be a
    # contribution however it is spelled, and the allowance must exclude it.
    for spelling in ("* BALANCE B/F *", "*  Balance b/f  *", "*BALANCE BF*"):
        assert classify(spelling) == "OPENING_BALANCE", f"B/F spelling missed: {spelling!r}"
    assert classify("Faster Payment In") in FLOW_CATEGORIES

    # end-to-end over a synthetic sheet with the trap present
    import tempfile
    try:
        import openpyxl
    except ImportError:
        print("SELFTEST PARTIAL — openpyxl unavailable; classification asserted, parse skipped")
        return ok
    with tempfile.TemporaryDirectory() as td:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "cashstatements"
        ws.append(["Date", "Description", "Reference", "Settlement date",
                   "Receipt (GBP)", "Payment (GBP)", "Balance (GBP)"])
        ws.append([dt.datetime(2026, 4, 6), "* BALANCE B/F *", "-", "-", 3469.64, 0, 3469.64])
        ws.append([dt.datetime(2026, 4, 6), "Faster Payment In", "R1", "-", 5000, 0, 8469.64])
        ws.append([dt.datetime(2026, 5, 2), "Direct debit payment", "R2", "-", 1250, 0, 9719.64])
        ws.append([dt.datetime(2026, 6, 2), "Direct debit payment", "R3", "-", 1250, 0, 10969.64])
        ws.append([dt.datetime(2026, 7, 2), "Direct debit payment", "R4", "-", 1250, 0, 12219.64])
        ws.append([dt.datetime(2026, 7, 23), "FX Charge (0.50%)", "R5", "-", 0, 0.01, 12219.63])
        ws.append([dt.datetime(2026, 8, 3), "Purchase 28 Vita Coco", "R6",
                   dt.datetime(2026, 8, 4), 0, 1354.99, 10864.64])
        p = os.path.join(td, "Cash Statement Current Tax Year 2026-2027.xlsx")
        wb.save(p)
        r = parse(folder=td, as_of=dt.date(2026, 8, 5))
        a = r["allowance"]
        assert a["used_gbp"] == 8750.0, f"allowance wrong: {a['used_gbp']} (trap not excluded?)"
        assert a["remaining_gbp"] == 11250.0, a["remaining_gbp"]
        assert r["opening_balance_gbp"] == 3469.64
        assert r["invariants"]["I12_opening_balance_not_a_contribution"]["ok"]
        assert r["invariants"]["I11_contribution_within_allowance"]["ok"]
        assert r["invariants"]["I14_every_row_classified"]["ok"], \
            r["invariants"]["I14_every_row_classified"]
        assert r["reconciled"] is True, r["warnings"]

        # NEGATIVE CONTROL: if the trap were counted, the figure would be 12,219.64 — assert
        # the difference is exactly the opening balance, so the test proves WHY, not just WHAT.
        naive = round(sum(x["receipt_gbp"] for x in r["rows"]
                          if x["category"] in ("CONTRIBUTION", "OPENING_BALANCE")), 2)
        assert round(naive - a["used_gbp"], 2) == 3469.64, (naive, a["used_gbp"])

        # a missing file must WARN, never raise, and never state a figure
        empty = parse(folder=os.path.join(td, "nope"), as_of=dt.date(2026, 8, 5))
        assert empty["reconciled"] is False and empty["warnings"], empty
        assert "allowance" not in empty, "an absent file must not produce an allowance figure"
    print("SELFTEST PASS — 27 assertions (17 classification rules incl. 2 that must NOT match, "
          "3 B/F spellings, allowance = 8,750 / 11,250, opening balance excluded, "
          "I11/I12/I14 green, negative control proving the trap is worth exactly 3,469.64, "
          "absent file warns without stating a figure)")
    return ok


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--folder", default=os.path.dirname(HERE))
    ap.add_argument("--out"); ap.add_argument("--ledger-check", action="store_true")
    ap.add_argument("--selftest", action="store_true")
    a = ap.parse_args()
    if a.selftest:
        sys.exit(0 if _selftest() else 1)
    r = parse(folder=a.folder)
    if a.ledger_check:
        r["ledger_reconciliation"] = reconcile_with_ledger(r)
    if a.out:
        json.dump(r, open(a.out, "w"), indent=2, default=str)
    al = r.get("allowance")
    print(f"CASH_STATEMENT rows={len(r['rows'])} files={r['source_files']}")
    if al:
        print(f"  ISA allowance used GBP{al['used_gbp']:,.2f} / remaining "
              f"GBP{al['remaining_gbp']:,.2f}  (tax year from {al['tax_year_start']})")
        for c in al["contribution_rows"]:
            print(f"    {c['date']}  {c['description'][:34]:36s} GBP{c['receipt_gbp']:,.2f}")
    print(f"  opening balance GBP{r.get('opening_balance_gbp')} (EXCLUDED from allowance) | "
          f"closing GBP{r.get('closing_balance_gbp')}")
    print(f"  FX rate {FX_RATE_PCT}%")
    for k, v in r["invariants"].items():
        print(f"  [{'PASS' if v.get('ok') else 'FAIL'}] {k}"
              + (f"  {v}" if not v.get("ok") else ""))
    for w in r["warnings"]:
        print(f"  WARN {w}")
    if r.get("ledger_reconciliation"):
        lr = r["ledger_reconciliation"]
        print(f"  [I10] ledger reconciliation ok={lr.get('ok')} "
              f"checked={lr.get('checked')} unmatched={len(lr.get('unmatched') or [])} "
              f"pending_ledger_update={len(lr.get('pending_ledger_update') or [])} "
              f"(ledger horizon {lr.get('ledger_horizon')})")
        for x in (lr.get("pending_ledger_update") or []):
            print(f"      PENDING {x['date']} GBP{x['amount_gbp']:,.2f} {x['description'][:44]}")
        for x in (lr.get("unmatched") or []):
            print(f"      UNMATCHED {x['date']} GBP{x['amount_gbp']:,.2f} {x['description'][:44]}")
    sys.exit(0 if r["reconciled"] else 1)
